from flask import Flask,render_template,request,jsonify,send_file,redirect,url_for,session,g
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from functools import wraps
from datetime import datetime,timedelta,timezone,date
import os,json,io,hashlib,urllib.request,urllib.error,zipfile,re,csv,base64,hmac,time,secrets
import shutil,threading,smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash,check_password_hash

app=Flask(__name__)
app.secret_key=os.environ.get('SECRET_KEY','rmfacilities2026@prod')
app.config['SQLALCHEMY_DATABASE_URI']='sqlite:///rmfacilities.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS']=False
app.config['PERMANENT_SESSION_LIFETIME']=28800
app.config['MAX_CONTENT_LENGTH']=50*1024*1024


def utcnow():
    return datetime.now(timezone.utc).replace(tzinfo=None)

db=SQLAlchemy(app)

class Usuario(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    nome=db.Column(db.String(100),nullable=False)
    email=db.Column(db.String(150),unique=True,nullable=False)
    senha=db.Column(db.String(256),nullable=False)
    perfil=db.Column(db.String(20),default='admin')
    areas=db.Column(db.Text,default='[]')
    ativo=db.Column(db.Boolean,default=True)
    ultimo_acesso=db.Column(db.DateTime)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def check_senha(self,s): return self.senha==hashlib.sha256(s.encode()).hexdigest()
    def to_dict(self):
        try: a=json.loads(self.areas or '[]')
        except: a=[]
        return {'id':self.id,'nome':self.nome,'email':self.email,'perfil':self.perfil,'ativo':self.ativo,'areas':a}

class Empresa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    nome=db.Column(db.String(200),nullable=False)
    razao=db.Column(db.String(200))
    cnpj=db.Column(db.String(30))
    telefone=db.Column(db.String(30))
    email=db.Column(db.String(150))
    site=db.Column(db.String(200))
    cep=db.Column(db.String(10))
    logradouro=db.Column(db.String(200))
    numero=db.Column(db.String(20))
    complemento=db.Column(db.String(100))
    bairro=db.Column(db.String(100))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    pix=db.Column(db.String(300))
    banco=db.Column(db.String(200))
    agencia=db.Column(db.String(20))
    conta=db.Column(db.String(30))
    contato_nome=db.Column(db.String(150))
    contato_email=db.Column(db.String(150))
    contato_telefone=db.Column(db.String(30))
    logo_url=db.Column(db.String(500))
    boleto=db.Column(db.Text)
    ativa=db.Column(db.Boolean,default=True)
    ordem=db.Column(db.Integer,default=0)
    def end_fmt(self):
        p=[self.logradouro,self.numero,self.complemento,self.bairro]
        e=', '.join(filter(None,p))
        if self.cidade: e+=f' — {self.cidade}/{self.estado or ""}'
        return e
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['end_fmt']=self.end_fmt()
        return d

class Config(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    chave=db.Column(db.String(100),unique=True,nullable=False)
    valor=db.Column(db.Text,default='')

class Cliente(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(10))
    nome=db.Column(db.String(200),nullable=False)
    cnpj=db.Column(db.String(30))
    responsavel=db.Column(db.String(150))
    telefone=db.Column(db.String(30))
    email=db.Column(db.String(150))
    cep=db.Column(db.String(10))
    logradouro=db.Column(db.String(200))
    numero_end=db.Column(db.String(20))
    complemento=db.Column(db.String(100))
    bairro=db.Column(db.String(100))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    status=db.Column(db.String(20),default='Ativo')
    limpeza=db.Column(db.Float,default=0)
    jardinagem=db.Column(db.Float,default=0)
    portaria=db.Column(db.Float,default=0)
    vencimento=db.Column(db.Integer,default=10)
    obs=db.Column(db.Text,default='')
    criado_em=db.Column(db.DateTime,default=utcnow)
    def end_fmt(self):
        p=[self.logradouro,self.numero_end,self.complemento,self.bairro]
        e=', '.join(filter(None,p))
        if self.cidade: e+=f' — {self.cidade}/{self.estado or ""}'
        return e
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['end_fmt']=self.end_fmt()
        return d

class Medicao(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(20))
    tipo=db.Column(db.String(50),default='Medição de Serviços')
    cliente_id=db.Column(db.Integer,db.ForeignKey('cliente.id'),nullable=True)
    cliente_nome=db.Column(db.String(200))
    cliente_cnpj=db.Column(db.String(30))
    cliente_end=db.Column(db.String(300))
    cliente_resp=db.Column(db.String(150))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    empresa_nome=db.Column(db.String(200))
    mes_ref=db.Column(db.String(7))
    dt_emissao=db.Column(db.String(10))
    dt_vencimento=db.Column(db.String(10))
    servicos=db.Column(db.Text)
    valor_bruto=db.Column(db.Float,default=0)
    observacoes=db.Column(db.Text)
    ass_empresa=db.Column(db.String(200))
    ass_cliente=db.Column(db.String(200))
    status=db.Column(db.String(20),default='emitida')
    desconto=db.Column(db.Float,default=0)
    impostos=db.Column(db.Text)
    criado_em=db.Column(db.DateTime,default=utcnow)
    criado_por=db.Column(db.String(100))
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['svcs']=json.loads(self.servicos) if self.servicos else []
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d.setdefault('status','emitida')
        return d

class Funcionario(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    matricula=db.Column(db.String(30),index=True)
    re=db.Column(db.Integer,unique=True,index=True)
    nome=db.Column(db.String(200),nullable=False)
    cpf=db.Column(db.String(20),unique=True)
    email=db.Column(db.String(150))
    telefone=db.Column(db.String(30))
    cargo=db.Column(db.String(120))
    funcao=db.Column(db.String(150))
    cbo=db.Column(db.String(20))
    setor=db.Column(db.String(120))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    data_admissao=db.Column(db.String(10))
    tipo_contrato=db.Column(db.String(60))
    jornada=db.Column(db.String(80))
    status=db.Column(db.String(20),default='Ativo')
    salario=db.Column(db.Float,default=0)
    vale_refeicao=db.Column(db.Float,default=0)
    vale_alimentacao=db.Column(db.Float,default=0)
    vale_transporte=db.Column(db.Float,default=0)
    endereco=db.Column(db.String(250))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    cep=db.Column(db.String(10))
    rg=db.Column(db.String(30))
    orgao_emissor=db.Column(db.String(30))
    pis=db.Column(db.String(30))
    ctps=db.Column(db.String(30))
    titulo_eleitor=db.Column(db.String(30))
    cert_reservista=db.Column(db.String(30))
    cnh=db.Column(db.String(30))
    exame_admissional_data=db.Column(db.String(10))
    docs_admissao_ok=db.Column(db.Boolean,default=False)
    docs_admissao_obs=db.Column(db.Text,default='')
    obs=db.Column(db.Text,default='')
    areas=db.Column(db.Text,default='[]')
    app_senha=db.Column(db.String(256))
    app_ativo=db.Column(db.Boolean,default=True)
    app_ultimo_acesso=db.Column(db.DateTime)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        try: d['areas']=json.loads(self.areas or '[]')
        except: d['areas']=[]
        d.pop('app_senha',None)
        if d.get('app_ativo') is None: d['app_ativo']=True
        return d

class FuncionarioArquivo(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False)
    categoria=db.Column(db.String(40),default='outros')
    competencia=db.Column(db.String(20))
    nome_arquivo=db.Column(db.String(250),nullable=False)
    caminho=db.Column(db.String(500),nullable=False)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['pasta']=os.path.dirname(self.caminho or '')
        return d

class OrdemCompra(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(30),nullable=False)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    solicitante=db.Column(db.String(150))
    fornecedor=db.Column(db.String(200))
    descricao=db.Column(db.Text)
    valor=db.Column(db.Float,default=0)
    status=db.Column(db.String(30),default='Aberta')
    data_emissao=db.Column(db.String(10))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class OperacionalDocumento(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    tipo=db.Column(db.String(80),default='Documento')
    titulo=db.Column(db.String(200),nullable=False)
    descricao=db.Column(db.Text)
    nome_arquivo=db.Column(db.String(250))
    caminho=db.Column(db.String(500))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class FuncionarioAppSessao(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False)
    refresh_hash=db.Column(db.String(256),nullable=False,index=True)
    exp_refresh=db.Column(db.DateTime,nullable=False)
    revogado=db.Column(db.Boolean,default=False)
    ip=db.Column(db.String(60))
    ua=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)
    atualizado_em=db.Column(db.DateTime,default=utcnow,onupdate=utcnow)

class AuthTentativa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    tipo=db.Column(db.String(30),nullable=False)
    identificador=db.Column(db.String(200),nullable=False)
    ip=db.Column(db.String(60))
    ok=db.Column(db.Boolean,default=False)
    motivo=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)

class AuditoriaEvento(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    evento=db.Column(db.String(120),nullable=False)
    ator_tipo=db.Column(db.String(30))
    ator_id=db.Column(db.String(60))
    alvo_tipo=db.Column(db.String(30))
    alvo_id=db.Column(db.String(60))
    ok=db.Column(db.Boolean,default=True)
    ip=db.Column(db.String(60))
    ua=db.Column(db.String(250))
    detalhe=db.Column(db.Text)
    criado_em=db.Column(db.DateTime,default=utcnow)

class MedicaoAnexo(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    medicao_id=db.Column(db.Integer,db.ForeignKey('medicao.id'),nullable=False)
    nome_arquivo=db.Column(db.String(250),nullable=False)
    caminho=db.Column(db.String(500),nullable=False)
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class WhatsAppConversa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(30),nullable=False,index=True)
    nome=db.Column(db.String(200))
    ultima_msg=db.Column(db.DateTime,default=utcnow)
    contexto=db.Column(db.Text,default='{}')
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['ultima_msg']=self.ultima_msg.isoformat() if self.ultima_msg else ''
        d['criado_em']=self.criado_em.isoformat() if self.criado_em else ''
        d['ultima_msg_fmt']=self.ultima_msg.strftime('%d/%m/%Y %H:%M') if self.ultima_msg else ''
        try:
            d['contexto']=json.loads(self.contexto or '{}')
        except:
            d['contexto']={}
        return d

class WhatsAppMensagem(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    conversa_id=db.Column(db.Integer,db.ForeignKey('whats_app_conversa.id'),nullable=False)
    numero=db.Column(db.String(30))
    direcao=db.Column(db.String(10))
    tipo=db.Column(db.String(20),default='texto')
    conteudo=db.Column(db.Text)
    nome_arquivo=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_em']=self.criado_em.isoformat() if self.criado_em else ''
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

_holerite_jobs={}

def hs(s): return hashlib.sha256(s.encode()).hexdigest()

def pw_hash(s):
    return generate_password_hash(s,method='scrypt')

def pw_is_modern(v):
    return str(v or '').startswith(('scrypt:','pbkdf2:'))

def pw_check(stored,plain):
    st=str(stored or '')
    if not st: return False
    if pw_is_modern(st):
        try: return check_password_hash(st,plain)
        except Exception: return False
    return hmac.compare_digest(st,hs(plain))

def token_hash(v):
    return hs('tok:'+str(v or ''))

LOGIN_WINDOW_MIN=15
LOGIN_FAIL_MAX=5
LOGIN_BLOCK_MIN=15

def auth_blocked(tipo,ident,ip):
    lim=utcnow()-timedelta(minutes=LOGIN_WINDOW_MIN)
    q=AuthTentativa.query.filter(AuthTentativa.tipo==tipo,AuthTentativa.identificador==ident,AuthTentativa.criado_em>=lim,AuthTentativa.ok==False)
    fails=q.count()
    if fails<LOGIN_FAIL_MAX: return False
    ult=q.order_by(AuthTentativa.criado_em.desc()).first()
    if not ult: return False
    return ult.criado_em+timedelta(minutes=LOGIN_BLOCK_MIN)>utcnow()

def reg_auth_attempt(tipo,ident,ok,motivo=''):
    ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    db.session.add(AuthTentativa(tipo=tipo,identificador=ident,ip=ip,ok=bool(ok),motivo=(motivo or '')[:250]))
    db.session.commit()

def audit_event(evento,ator_tipo='',ator_id='',alvo_tipo='',alvo_id='',ok=True,det=None):
    try:
        ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
        ua=(request.headers.get('User-Agent') or '')[:250]
        detalhe=det if isinstance(det,str) else json.dumps(det or {},ensure_ascii=False)
        db.session.add(AuditoriaEvento(evento=evento,ator_tipo=str(ator_tipo or '')[:30],ator_id=str(ator_id or '')[:60],alvo_tipo=str(alvo_tipo or '')[:30],alvo_id=str(alvo_id or '')[:60],ok=bool(ok),ip=ip,ua=ua,detalhe=detalhe))
        db.session.commit()
    except Exception:
        db.session.rollback()

def b64u_enc(raw):
    return base64.urlsafe_b64encode(raw).decode().rstrip('=')

def b64u_dec(s):
    pad='='*((4-len(s)%4)%4)
    return base64.urlsafe_b64decode((s+pad).encode())

def app_token_secret():
    return (os.environ.get('APP_TOKEN_SECRET') or app.secret_key or 'rmfacilities-app').encode()

def app_issue_access_token(funcionario_id,sessao_id,ttl=3600):
    now=int(time.time())
    payload={'typ':'access','sid':int(sessao_id),'fid':int(funcionario_id),'iat':now,'exp':now+int(ttl)}
    ptxt=json.dumps(payload,separators=(',',':')).encode()
    p64=b64u_enc(ptxt)
    sig=hmac.new(app_token_secret(),p64.encode(),hashlib.sha256).digest()
    s64=b64u_enc(sig)
    return f'{p64}.{s64}'

def app_issue_refresh_token():
    return secrets.token_urlsafe(48)

def app_parse_token(token):
    try:
        p64,s64=token.split('.',1)
        expected=b64u_enc(hmac.new(app_token_secret(),p64.encode(),hashlib.sha256).digest())
        if not hmac.compare_digest(expected,s64): return None
        payload=json.loads(b64u_dec(p64).decode())
        if int(payload.get('exp',0))<int(time.time()): return None
        return payload
    except Exception:
        return None

def app_func_required(f):
    @wraps(f)
    def w(*a,**k):
        auth=(request.headers.get('Authorization') or '').strip()
        if not auth.lower().startswith('bearer '):
            return jsonify({'erro':'Token ausente'}),401
        tok=auth.split(' ',1)[1].strip()
        payload=app_parse_token(tok)
        if not payload or payload.get('typ')!='access': return jsonify({'erro':'Token invalido ou expirado'}),401
        sid=to_num(payload.get('sid'))
        sessao=FuncionarioAppSessao.query.get(sid)
        if not sessao or sessao.revogado: return jsonify({'erro':'Sessao invalida'}),401
        if sessao.exp_refresh < utcnow(): return jsonify({'erro':'Sessao expirada'}),401
        func=Funcionario.query.get(sessao.funcionario_id)
        if not func: return jsonify({'erro':'Funcionario nao encontrado'}),404
        if to_num(payload.get('fid'))!=func.id: return jsonify({'erro':'Token invalido'}),401
        if func.app_ativo is False: return jsonify({'erro':'Acesso do aplicativo desativado'}),403
        g.app_funcionario=func
        g.app_sessao=sessao
        return f(*a,**k)
    return w

def gc(k,dv=''): c=Config.query.filter_by(chave=k).first(); return c.valor if c else dv

def smtp_cfg():
    return {'host':gc('smtp_host',''),'port':gc('smtp_port','587'),'user':gc('smtp_user',''),'senha':gc('smtp_senha',''),'de':gc('smtp_de',''),'tls':gc('smtp_tls','1')}

def wa_cfg():
    return {'url':gc('wa_url',''),'instancia':gc('wa_instancia',''),'token':gc('wa_token','')}

def wa_backup_cfg():
    return {
        'enabled':gc('wa_backup_enabled','1'),
        'email':gc('wa_backup_email',''),
        'interval_hours':gc('wa_backup_interval_hours','2'),
        'window_hours':gc('wa_backup_window_hours','8'),
        'max_conversas':gc('wa_backup_max_conversas','10'),
        'last_ts':gc('wa_backup_last_ts','0'),
    }

def wa_backup_enabled():
    return str(wa_backup_cfg().get('enabled','0')).strip().lower() in ('1','true','yes','on')

def _wa_backup_root():
    return os.path.join(os.path.dirname(__file__),'instance','wa_backups')

def _to_int(v,dv=0):
    try: return int(float(str(v).strip()))
    except Exception: return dv

def _parse_dt_iso(v):
    s=(v or '').strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace('Z','+00:00')).replace(tzinfo=None)
    except Exception:
        return None

def _cfg_snapshot_dict():
    out={}
    for c in Config.query.all():
        out[c.chave]=c.valor
    return out

def _cfg_apply_snapshot(dct):
    if not isinstance(dct,dict):
        return 0
    n=0
    for k,v in dct.items():
        if not str(k or '').strip():
            continue
        sc_cfg(str(k).strip(),'' if v is None else str(v))
        n+=1
    return n

def _wa_backup_collect(window_hours=8,max_conversas=10):
    janela=max(1,min(168,_to_int(window_hours,8)))
    qtd=max(1,min(50,_to_int(max_conversas,10)))
    corte=utcnow()-timedelta(hours=janela)
    convs=WhatsAppConversa.query.order_by(WhatsAppConversa.ultima_msg.desc()).limit(qtd).all()
    out=[]
    for c in convs:
        q=WhatsAppMensagem.query.filter_by(conversa_id=c.id).filter(WhatsAppMensagem.criado_em>=corte).order_by(WhatsAppMensagem.criado_em.asc())
        msgs=q.all()
        out.append({
            'conversa':c.to_dict(),
            'mensagens':[{
                'id':m.id,
                'numero':m.numero,
                'direcao':m.direcao,
                'tipo':m.tipo,
                'conteudo':m.conteudo,
                'criado_em':m.criado_em.isoformat() if m.criado_em else '',
            } for m in msgs]
        })
    ia=ai_wa_cfg()
    payload={
        'gerado_em':utcnow().isoformat(),
        'versao_backup':'wa-v2',
        'janela_horas':janela,
        'max_conversas':qtd,
        'total_conversas':len(out),
        'total_mensagens':sum(len(c['mensagens']) for c in out),
        'config':_cfg_snapshot_dict(),
        'ia':{
            'enabled':ai_wa_enabled(),
            'provider':ia.get('provider',''),
            'model':ia.get('model',''),
            'temperature':ia.get('temperature',''),
            'max_tokens':ia.get('max_tokens',''),
        },
        'conversas':out,
    }
    return payload

def _wa_backup_store(payload):
    root=_wa_backup_root()
    os.makedirs(root,exist_ok=True)
    nome=f"wa_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    caminho=os.path.join(root,nome)
    with open(caminho,'w',encoding='utf-8') as f:
        json.dump(payload,f,ensure_ascii=False,indent=2)
    return caminho

def _wa_backup_store_txt(payload):
    root=_wa_backup_root()
    os.makedirs(root,exist_ok=True)
    base=datetime.now().strftime('%Y%m%d_%H%M%S')
    nome=f"wa_historico_{base}.txt"
    caminho=os.path.join(root,nome)
    linhas=[]
    linhas.append('Backup de Historico WhatsApp - RM Facilities')
    linhas.append(f"Gerado em: {payload.get('gerado_em','')}")
    linhas.append(f"Janela (horas): {payload.get('janela_horas',0)}")
    linhas.append(f"Conversas: {payload.get('total_conversas',0)}")
    linhas.append(f"Mensagens: {payload.get('total_mensagens',0)}")
    linhas.append('')
    for bloco in (payload.get('conversas') or []):
        conv=bloco.get('conversa') or {}
        nome_conv=conv.get('nome') or conv.get('numero') or 'Sem nome'
        num=conv.get('numero') or ''
        linhas.append('='*72)
        linhas.append(f"Conversa: {nome_conv} ({num})")
        linhas.append('='*72)
        for m in (bloco.get('mensagens') or []):
            lado='Cliente' if m.get('direcao')=='in' else 'Atendente'
            dt=m.get('criado_em') or ''
            txt=(m.get('conteudo') or '').replace('\r','').strip()
            linhas.append(f"[{dt}] {lado}: {txt}")
        linhas.append('')
    with open(caminho,'w',encoding='utf-8') as f:
        f.write('\n'.join(linhas))
    return caminho

def smtp_send_text(dest,assunto,corpo,anexos=None):
    cfg=smtp_cfg()
    if not cfg['host'] or not cfg['user']: raise ValueError('SMTP nao configurado')
    msg=MIMEMultipart()
    msg['From']=cfg['de'] or cfg['user']
    msg['To']=dest
    msg['Subject']=assunto
    msg.attach(MIMEText(corpo or '','plain','utf-8'))
    for a in (anexos or []):
        caminho=a.get('path')
        nome=a.get('name') or os.path.basename(caminho or 'anexo.bin')
        if not caminho or not os.path.isfile(caminho):
            continue
        with open(caminho,'rb') as f:
            part=MIMEBase('application','octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',f'attachment; filename="{nome}"')
        msg.attach(part)
    port=int(cfg['port'] or 587)
    if str(cfg['tls']) in ('1','true','True','yes'):
        with smtplib.SMTP(cfg['host'],port,timeout=20) as s:
            s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
    else:
        with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s:
            s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())

def wa_backup_maybe_send(force=False):
    cfg=wa_backup_cfg()
    if not force and not wa_backup_enabled():
        return {'ok':False,'skip':'desativado'}
    email=(cfg.get('email') or '').strip()
    if not email:
        return {'ok':False,'skip':'sem_email'}
    intervalo=max(1,min(168,_to_int(cfg.get('interval_hours'),2)))
    janela=max(1,min(168,_to_int(cfg.get('window_hours'),8)))
    max_conv=max(1,min(50,_to_int(cfg.get('max_conversas'),10)))
    agora=int(time.time())
    ultimo=max(0,_to_int(cfg.get('last_ts'),0))
    if not force and (agora-ultimo)<(intervalo*3600):
        return {'ok':False,'skip':'intervalo'}
    payload=_wa_backup_collect(janela,max_conv)
    arq=_wa_backup_store(payload)
    arq_txt=_wa_backup_store_txt(payload)
    assunto=f"Backup WhatsApp RM Facilities - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    corpo=(
        f"Backup automatico das conversas WhatsApp.\n\n"
        f"Janela: {payload.get('janela_horas')}h\n"
        f"Conversas: {payload.get('total_conversas')}\n"
        f"Mensagens: {payload.get('total_mensagens')}\n"
        f"Modelo IA: {payload.get('ia',{}).get('model','')}\n"
    )
    smtp_send_text(email,assunto,corpo,anexos=[
        {'path':arq,'name':os.path.basename(arq)},
        {'path':arq_txt,'name':os.path.basename(arq_txt)},
    ])
    sc_cfg('wa_backup_last_ts',str(agora))
    return {'ok':True,'email':email,'arquivo':arq,'total_conversas':payload.get('total_conversas',0),'total_mensagens':payload.get('total_mensagens',0)}

def wa_backup_restore_payload(payload,restore_config=True,restore_conversas=True):
    if not isinstance(payload,dict):
        raise ValueError('Arquivo de backup invalido')
    stat={'configs':0,'conversas':0,'mensagens':0}
    if restore_config:
        stat['configs']=_cfg_apply_snapshot(payload.get('config') or {})
    if restore_conversas:
        blocos=payload.get('conversas') or []
        for bloco in blocos:
            conv=bloco.get('conversa') or {}
            numero=wa_norm_number(conv.get('numero') or '')
            if not wa_is_valid_number(numero):
                continue
            c=WhatsAppConversa.query.filter_by(numero=numero).first()
            if not c:
                c=WhatsAppConversa(numero=numero,nome=(conv.get('nome') or numero))
                db.session.add(c)
                db.session.flush()
                stat['conversas']+=1
            elif (conv.get('nome') or '').strip():
                c.nome=(conv.get('nome') or c.nome)
            ult=_parse_dt_iso(conv.get('ultima_msg') or '')
            if ult and (not c.ultima_msg or ult>c.ultima_msg):
                c.ultima_msg=ult
            for m in (bloco.get('mensagens') or []):
                txt=(m.get('conteudo') or '').strip()
                if not txt:
                    continue
                created=_parse_dt_iso(m.get('criado_em') or '') or utcnow()
                direcao=(m.get('direcao') or 'in').strip().lower()
                tipo=(m.get('tipo') or 'texto').strip().lower() or 'texto'
                existe=WhatsAppMensagem.query.filter_by(conversa_id=c.id,direcao=direcao,tipo=tipo,conteudo=txt,criado_em=created).first()
                if existe:
                    continue
                db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao=direcao,tipo=tipo,conteudo=txt,criado_em=created))
                stat['mensagens']+=1
    db.session.commit()
    return stat

def wa_norm_number(numero):
    n=re.sub(r'\D+','',str(numero or ''))
    if n.startswith('00'): n=n[2:]
    if n.startswith('55') and len(n)>=12: return n
    if len(n) in (10,11): return '55'+n
    return n

def wa_is_valid_number(numero):
    n=re.sub(r'\D+','',str(numero or ''))
    return bool(re.fullmatch(r'\d{12,15}',n))

def wa_phone_matches(a,b):
    da=only_digits(a)
    dbn=only_digits(b)
    if not da or not dbn:
        return False
    na=wa_norm_number(da)
    nb=wa_norm_number(dbn)
    if na and nb and na==nb:
        return True
    for size in (11,10,9,8):
        if len(da)>=size and len(dbn)>=size and da[-size:]==dbn[-size:]:
            return True
    return False

def wa_send_text(numero,mensagem):
    cfg=wa_cfg()
    if not cfg['url'] or not cfg['instancia']: raise ValueError('WhatsApp nao configurado')
    num=wa_norm_number(numero)
    if not wa_is_valid_number(num): raise ValueError(f'Numero WhatsApp invalido: {num or "vazio"}')
    url=f"{cfg['url'].rstrip('/')}/message/sendText/{cfg['instancia']}"
    data=json.dumps({'number':num,'text':mensagem}).encode()
    req=urllib.request.Request(url,data=data,headers={'Content-Type':'application/json','apikey':cfg['token']})
    try:
        with urllib.request.urlopen(req,timeout=15) as r: return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'WhatsApp API {e.code}: {detalhe or e.reason}')

def wa_send_pdf(numero,caminho_abs,nome_arquivo,caption=''):
    cfg=wa_cfg()
    if not cfg['url'] or not cfg['instancia']: raise ValueError('WhatsApp nao configurado')
    num=wa_norm_number(numero)
    if not wa_is_valid_number(num): raise ValueError(f'Numero WhatsApp invalido: {num or "vazio"}')
    with open(caminho_abs,'rb') as f: pdf_b64=base64.b64encode(f.read()).decode()
    url=f"{cfg['url'].rstrip('/')}/message/sendMedia/{cfg['instancia']}"
    data=json.dumps({'number':num,'mediatype':'document','mimetype':'application/pdf','media':pdf_b64,'fileName':nome_arquivo,'caption':caption or nome_arquivo}).encode()
    req=urllib.request.Request(url,data=data,headers={'Content-Type':'application/json','apikey':cfg['token']})
    try:
        with urllib.request.urlopen(req,timeout=30) as r: return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'WhatsApp API {e.code}: {detalhe or e.reason}')

def ai_wa_cfg():
    return {
        'enabled':gc('ia_wa_enabled','0'),
        'provider':(gc('ia_wa_provider','gemini') or 'gemini').strip().lower(),
        'api_key':gc('ia_wa_api_key',''),
        'model':gc('ia_wa_model',''),
        'prompt':gc('ia_wa_prompt',''),
        'temperature':gc('ia_wa_temperature','0.3'),
        'max_tokens':gc('ia_wa_max_tokens','350'),
    }

def ai_provider_norm(v):
    p=(v or 'gemini').strip().lower()
    if p in ('openai','chatgpt','gpt'): return 'openai'
    return 'gemini'

def ai_model_norm(provider,raw_model):
    p=ai_provider_norm(provider)
    s=(raw_model or '').strip()
    if not s:
        return 'gpt-4o-mini' if p=='openai' else 'gemini-1.5-flash'
    low=s.lower()
    if p=='gemini':
        m=re.search(r'models/([a-z0-9._-]+)',low)
        if m: return m.group(1)
        m=re.search(r'(gemini-[a-z0-9._-]+)',low)
        if m: return m.group(1)
        tok=re.split(r'[\s<>{}\[\]"\'\#?&:/]+',low)[0]
        tok=re.sub(r'[^a-z0-9._-]','',tok)
        return tok if tok.startswith('gemini-') else 'gemini-1.5-flash'
    tok=re.split(r'[\s<>{}\[\]"\'\#?&:/]+',s)[0]
    tok=re.sub(r'[^A-Za-z0-9._-]','',tok)
    if tok: return tok
    return 'gpt-4o-mini'

def ai_wa_enabled():
    return str(ai_wa_cfg().get('enabled','0')).strip().lower() in ('1','true','yes','on')

def wa_ai_pause_key(numero):
    n=wa_norm_number(numero)
    return f'wa_ai_pause_until_{n}' if n else ''

def wa_ai_pause_until(numero):
    k=wa_ai_pause_key(numero)
    if not k:
        return None
    return _parse_dt_iso(gc(k,''))

def wa_ai_pause_active(numero,ref=None):
    until=wa_ai_pause_until(numero)
    if not until:
        return False
    now=ref or utcnow()
    return until>now

def wa_ai_pause_for(numero,hours=8):
    n=wa_norm_number(numero)
    if not n:
        return None
    h=max(1,min(168,_to_int(hours,8)))
    until=utcnow()+timedelta(hours=h)
    sc_cfg(wa_ai_pause_key(n),until.isoformat())
    return until

def wa_ai_resume(numero):
    n=wa_norm_number(numero)
    if not n:
        return None
    until=utcnow()-timedelta(seconds=1)
    sc_cfg(wa_ai_pause_key(n),until.isoformat())
    return until

def wa_ai_pause_set(numero,hours=8):
    return wa_ai_pause_for(numero,hours)

def _post_json(url,payload,headers=None,timeout=30):
    h={'Content-Type':'application/json'}
    if headers: h.update(headers)
    req=urllib.request.Request(url,data=json.dumps(payload,ensure_ascii=False).encode('utf-8'),headers=h)
    try:
        with urllib.request.urlopen(req,timeout=timeout) as r:
            raw=r.read().decode('utf-8',errors='ignore')
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'IA API {e.code}: {detalhe or e.reason}')

def _post_multipart(url,fields=None,files=None,headers=None,timeout=60):
    boundary='----rmfacilities'+secrets.token_hex(12)
    body=bytearray()
    for k,v in (fields or {}).items():
        body.extend(f'--{boundary}\r\n'.encode())
        body.extend(f'Content-Disposition: form-data; name="{k}"\r\n\r\n'.encode())
        body.extend(str(v).encode('utf-8'))
        body.extend(b'\r\n')
    for f in (files or []):
        body.extend(f'--{boundary}\r\n'.encode())
        body.extend(
            f'Content-Disposition: form-data; name="{f.get("field","file")}"; filename="{f.get("filename","arquivo.bin")}"\r\n'.encode()
        )
        body.extend(f'Content-Type: {f.get("content_type","application/octet-stream")}\r\n\r\n'.encode())
        body.extend(f.get('data') or b'')
        body.extend(b'\r\n')
    body.extend(f'--{boundary}--\r\n'.encode())
    h={'Content-Type':f'multipart/form-data; boundary={boundary}'}
    if headers:
        h.update(headers)
    req=urllib.request.Request(url,data=bytes(body),headers=h)
    try:
        with urllib.request.urlopen(req,timeout=timeout) as r:
            raw=r.read().decode('utf-8',errors='ignore')
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'IA API {e.code}: {detalhe or e.reason}')

def _http_read_bytes(url,headers=None,timeout=30):
    req=urllib.request.Request(url,headers=headers or {})
    with urllib.request.urlopen(req,timeout=timeout) as r:
        return r.read()

def _audio_ext_from_mime(mime):
    m=(mime or '').split(';')[0].strip().lower()
    return {
        'audio/ogg':'ogg',
        'audio/opus':'ogg',
        'audio/webm':'webm',
        'audio/mpeg':'mp3',
        'audio/mp3':'mp3',
        'audio/mp4':'m4a',
        'audio/x-m4a':'m4a',
        'audio/wav':'wav',
        'audio/x-wav':'wav',
        'audio/aac':'aac',
    }.get(m,'ogg')

def _maybe_b64decode(raw):
    s=str(raw or '').strip()
    if not s:
        return b''
    if ',' in s and ';base64' in s[:80].lower():
        s=s.split(',',1)[1]
    s=re.sub(r'\s+','',s)
    try:
        return base64.b64decode(s)
    except Exception:
        return b''

def _extract_audio_blob(msg_data):
    msg_obj=msg_data.get('message',{}) if isinstance(msg_data.get('message',{}),dict) else {}
    audio_obj=msg_obj.get('audioMessage',{}) if isinstance(msg_obj.get('audioMessage',{}),dict) else {}
    mime=(
        audio_obj.get('mimetype') or
        audio_obj.get('mimeType') or
        msg_data.get('mimetype') or
        msg_data.get('mimeType') or
        'audio/ogg'
    )
    filename=(
        audio_obj.get('fileName') or
        msg_data.get('fileName') or
        f'audio.{_audio_ext_from_mime(mime)}'
    )
    for candidate in (
        audio_obj.get('base64'),
        audio_obj.get('media'),
        audio_obj.get('data'),
        audio_obj.get('fileBase64'),
        msg_data.get('base64'),
        msg_data.get('media'),
        msg_data.get('data'),
    ):
        blob=_maybe_b64decode(candidate)
        if blob:
            return blob,mime,filename

    headers={}
    cfg=wa_cfg()
    if (cfg.get('token') or '').strip():
        headers['apikey']=cfg['token']
        headers['Authorization']=f'Bearer {cfg["token"]}'
    for media_url in (
        audio_obj.get('url'),
        audio_obj.get('mediaUrl'),
        audio_obj.get('directPath'),
        msg_data.get('mediaUrl'),
        msg_data.get('url'),
    ):
        u=str(media_url or '').strip()
        if not u or not u.startswith(('http://','https://')):
            continue
        try:
            blob=_http_read_bytes(u,headers=headers,timeout=45)
            if blob:
                return blob,mime,filename
        except Exception:
            continue
    return b'',mime,filename

def wa_transcribe_audio(msg_data):
    audio_bytes,mime,filename=_extract_audio_blob(msg_data)
    if not audio_bytes:
        return ''
    cfg=ai_wa_cfg()
    key=(cfg.get('api_key') or '').strip()
    if not key:
        return ''
    provider=ai_provider_norm(cfg.get('provider') or 'gemini')
    if key.startswith('AIza'):
        provider='gemini'
    elif key.startswith('sk-'):
        provider='openai'

    mime_base=(mime or 'audio/ogg').split(';')[0].strip().lower() or 'audio/ogg'
    if provider=='openai':
        url=(gc('ia_openai_audio_url','') or 'https://api.openai.com/v1/audio/transcriptions').strip()
        out=_post_multipart(
            url,
            fields={
                'model':'whisper-1',
                'language':'pt',
                'prompt':'Transcreva em portugues do Brasil.',
            },
            files=[{'field':'file','filename':filename,'content_type':mime_base,'data':audio_bytes}],
            headers={'Authorization':f'Bearer {key}'},
            timeout=90,
        )
        return str(out.get('text') or out.get('transcript') or '').strip()

    model=ai_model_norm('gemini',cfg.get('model') or '')
    base=(gc('ia_gemini_url','') or '').strip()
    url=base or f'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}'
    payload={
        'contents':[{ 
            'role':'user',
            'parts':[
                {'text':'Transcreva este audio em portugues do Brasil. Responda apenas com a transcricao, sem observacoes adicionais.'},
                {'inline_data':{'mime_type':mime_base,'data':base64.b64encode(audio_bytes).decode()}}
            ]
        }],
        'generationConfig':{'temperature':0.1,'maxOutputTokens':700}
    }
    out=_post_json(url,payload,timeout=90)
    cand=(out.get('candidates') or [{}])[0]
    parts=((cand.get('content') or {}).get('parts') or [])
    return '\n'.join((p.get('text') or '').strip() for p in parts if (p.get('text') or '').strip()).strip()

def _build_turns(hist,in_role,out_role):
    """Merge consecutive same-role messages into valid chat turns."""
    turns=[]
    for m in hist:
        role=in_role if m.direcao=='in' else out_role
        txt_m=(m.conteudo or '').strip()
        if not txt_m: continue
        if turns and turns[-1]['role']==role:
            turns[-1]['_text']+=('\n'+txt_m)
        else:
            turns.append({'role':role,'_text':txt_m})
    return turns

def _detecta_pedido_holerite(texto):
    """Detecta se o texto contém um pedido de holerite."""
    palavras=['holerite','hollerite','contracheque','contra-cheque','recibo','salario','salário','comprovante de renda','comprovante','vencimento','13o','décimo terceiro']
    txt_lower=texto.lower()
    return any(p in txt_lower for p in palavras)

def funcionario_docs_whatsapp_habilitado(funcionario):
    if not funcionario:
        return False
    try:
        ars=json.loads(funcionario.areas or '[]')
    except Exception:
        ars=[]
    return 'documentos' in ars

def _parse_competencias_holerite(texto,ano_padrao=None):
    """Extrai uma ou mais competências (MM/YYYY) do texto.

    Regras:
    - aceita formatos numéricos (05/2026, 5-2026)
    - aceita nomes de meses (maio, junho)
    - quando o ano não é informado, usa o ano corrente
    """
    if ano_padrao is None:
        ano_padrao=datetime.now().year
    s=(texto or '').strip().lower()
    if not s:
        return []

    ano_global=None
    m_ano=re.search(r'\b(19\d{2}|20\d{2})\b',s)
    if m_ano:
        try:
            ano_global=int(m_ano.group(1))
        except Exception:
            ano_global=None

    meses={
        'janeiro':1,'jan':1,
        'fevereiro':2,'fev':2,
        'marco':3,'março':3,'mar':3,
        'abril':4,'abr':4,
        'maio':5,'mai':5,
        'junho':6,'jun':6,
        'julho':7,'jul':7,
        'agosto':8,'ago':8,
        'setembro':9,'set':9,
        'outubro':10,'out':10,
        'novembro':11,'nov':11,
        'dezembro':12,'dez':12,
    }

    out=[]
    seen=set()

    def add_comp(mes,ano):
        try:
            mi=int(mes)
            ai=int(ano)
        except Exception:
            return
        if not (1<=mi<=12 and 1900<=ai<=2099):
            return
        comp=f"{mi:02d}/{ai}"
        if comp not in seen:
            seen.add(comp)
            out.append(comp)

    for mm,yy in re.findall(r'\b(0?[1-9]|1[0-2])\s*[/-]\s*(\d{4})\b',s):
        add_comp(mm,yy)

    for mt in re.finditer(r'\b(janeiro|jan|fevereiro|fev|março|marco|mar|abril|abr|maio|mai|junho|jun|julho|jul|agosto|ago|setembro|set|outubro|out|novembro|nov|dezembro|dez)\b(?:\s*(?:de|/|-)?\s*(\d{4}))?',s):
        token=mt.group(1)
        ano_txt=mt.group(2)
        ano=int(ano_txt) if ano_txt else (ano_global or ano_padrao)
        add_comp(meses.get(token),ano)

    # Se vier apenas mês numérico sem ano (ex.: "5 e 6"), assume ano corrente.
    for mm in re.findall(r'\b(0?[1-9]|1[0-2])\b',s):
        add_comp(mm,ano_global or ano_padrao)

    return out

def _processa_dialogo_holerite(conversa_id,numero,texto):
    """Processa o diálogo de busca e envio de holerite."""
    conversa=WhatsAppConversa.query.get(conversa_id)
    if not conversa:
        return None
    
    try:
        ctx=json.loads(conversa.contexto or '{}')
    except:
        ctx={}
    
    estado=ctx.get('holerite_estado')
    
    # Estado inicial: solicitar dados
    if not estado:
        ctx['holerite_estado']='aguardando_identificacao'
        ctx['holerite_tentativas']=0
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Ótimo! Para buscar seu holerite, primeiro preciso do seu nome completo.\n\nPor favor, digite seu nome completo."
    
    # Aguardando identificação por nome
    if estado=='aguardando_identificacao':
        texto_clean=texto.strip()
        nome_match=re.sub(r'[\d\-.]','',texto_clean).strip()
        if not nome_match or len(nome_match)<3:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não consegui identificar seu nome completo. Envie 'holerite' para recomeçar."
            return "Não consegui identificar seu nome completo. Digite apenas seu nome e sobrenome."
        
        # Busca o funcionário pelo nome informado.
        funcionarios=Funcionario.query.filter(Funcionario.nome.ilike(f'%{nome_match}%')).all()
        funcionarios_com_tel=[f for f in funcionarios if wa_phone_matches(numero,f.telefone)]
        if funcionarios_com_tel:
            funcionarios=funcionarios_com_tel
        
        if not funcionarios:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não consegui validar seu cadastro com este número de telefone. Verifique o nome informado e se está falando do telefone cadastrado."
            return "Não encontrei um cadastro compatível com seu nome e este número de telefone. Verifique o nome completo e tente novamente."
        
        if len(funcionarios)>1:
            ctx['holerite_estado']='escolhendo_funcionario'
            ctx['holerite_opcoes']=[f.id for f in funcionarios]
            ctx['holerite_tentativas']=0
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            opcoes_txt='\n'.join([f"{i+1}️⃣ {func.nome} (RE {func.re})" for i,func in enumerate(funcionarios)])
            return f"Encontrei múltiplos registros. Qual é você?\n\n{opcoes_txt}"
        
        # Um único funcionário encontrado
        funcionario=funcionarios[0]
        if not wa_phone_matches(numero,funcionario.telefone):
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Por segurança, não consegui validar seu telefone com o cadastro do funcionário. Envie 'holerite' para recomeçar ou fale com o RH."
            return "Por segurança, este telefone não confere com o cadastro do funcionário informado. Tente novamente usando o telefone cadastrado."
        if not funcionario_docs_whatsapp_habilitado(funcionario):
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Seu cadastro não está habilitado para receber documentos pelo WhatsApp. Solicite ao RH a liberação de acesso a documentos."
        ctx['holerite_estado']='aguardando_cpf'
        ctx['holerite_funcionario_id']=funcionario.id
        ctx['holerite_tentativas']=0
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return f"Encontrei seu cadastro, {funcionario.nome}.\n\nPor segurança, informe seu CPF completo (apenas números) para liberar o envio do holerite."
    
    # Escolhendo entre múltiplos funcionários
    if estado=='escolhendo_funcionario':
        try:
            escolha=int(texto.strip())
            opcoes=ctx.get('holerite_opcoes',[])
            if 1<=escolha<=len(opcoes):
                func_id=opcoes[escolha-1]
                funcionario=Funcionario.query.get(func_id)
                if funcionario:
                    if not wa_phone_matches(numero,funcionario.telefone):
                        ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
                        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                        db.session.commit()
                        if ctx['holerite_tentativas']>=3:
                            ctx['holerite_estado']=None
                            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                            db.session.commit()
                            return "Por segurança, este telefone não corresponde ao cadastro do funcionário selecionado. Envie 'holerite' para recomeçar ou fale com o RH."
                        return "Por segurança, este telefone não corresponde ao cadastro do funcionário selecionado."
                    if not funcionario_docs_whatsapp_habilitado(funcionario):
                        ctx['holerite_estado']=None
                        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                        db.session.commit()
                        return "Seu cadastro não está habilitado para receber documentos pelo WhatsApp. Solicite ao RH a liberação de acesso a documentos."
                    ctx['holerite_estado']='aguardando_cpf'
                    ctx['holerite_funcionario_id']=func_id
                    ctx['holerite_tentativas']=0
                    conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                    db.session.commit()
                    return f"Perfeito, {funcionario.nome}.\n\nPor segurança, informe seu CPF completo (apenas números) para liberar o envio do holerite."
        except:
            pass
        
        ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
        if ctx['holerite_tentativas']>=3:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Desculpe, não consegui processar sua escolha. Envie 'holerite' novamente para recomeçar."
        
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Responda com o número (1, 2, 3...) do funcionário desejado."

    # Validação de CPF antes de informar competência
    if estado=='aguardando_cpf':
        func_id=ctx.get('holerite_funcionario_id')
        funcionario=Funcionario.query.get(func_id) if func_id else None
        if not funcionario:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Desculpe, houve um erro ao validar seus dados. Envie 'holerite' para recomeçar."
        if not funcionario_docs_whatsapp_habilitado(funcionario):
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Seu cadastro não está habilitado para receber documentos pelo WhatsApp. Solicite ao RH a liberação de acesso a documentos."
        if not wa_phone_matches(numero,funcionario.telefone):
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Por segurança, este telefone não corresponde ao cadastro do funcionário. Procure o RH para atualizar seu número e tente novamente."

        cpf_informado=only_digits(texto)
        cpf_cadastrado=only_digits(funcionario.cpf or '')
        if len(cpf_informado)!=11:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Não foi possível validar sua identidade. Envie 'holerite' para recomeçar."
            return "CPF inválido. Informe seu CPF completo com 11 dígitos (apenas números)."

        if not cpf_cadastrado or cpf_informado!=cpf_cadastrado:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "CPF não confere com o cadastro. Por segurança, o fluxo foi encerrado. Envie 'holerite' para tentar novamente."
            return "CPF não confere com o cadastro. Tente novamente com o CPF correto."

        ctx['holerite_estado']='aguardando_competencia'
        ctx['holerite_tentativas']=0
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Identidade validada com sucesso.\n\nAgora informe o(s) mês(es) do holerite. Você pode pedir mais de um de uma vez (ex.: 05/2026 e 06/2026, ou maio e junho). Se não informar o ano, assumirei o ano corrente."
    
    # Aguardando competência (mês/ano)
    if estado=='aguardando_competencia':
        competencias=_parse_competencias_holerite(texto,ano_padrao=datetime.now().year)
        if not competencias:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não consegui processar a data. Envie 'holerite' novamente para recomeçar."
            return "Formato inválido. Informe mês/ano (ex: 04/2026) ou nomes de meses (ex: maio e junho). Se não informar o ano, usarei o ano corrente."

        func_id=ctx.get('holerite_funcionario_id')
        funcionario=Funcionario.query.get(func_id) if func_id else None
        
        if not funcionario:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Desculpe, houve um erro ao buscar seus dados. Tente novamente."

        arquivos_por_comp=[]
        nao_encontrados=[]
        for competencia in competencias:
            arquivo=FuncionarioArquivo.query.filter(
                FuncionarioArquivo.funcionario_id==func_id,
                FuncionarioArquivo.categoria=='holerite',
                FuncionarioArquivo.competencia==competencia
            ).first()
            if arquivo:
                arquivos_por_comp.append((competencia,arquivo))
            else:
                nao_encontrados.append(competencia)

        if not arquivos_por_comp:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não encontrei seu holerite. Envie 'holerite' para tentar novamente."
            if nao_encontrados:
                return f"Ops! Não encontrei holerite para: {', '.join(nao_encontrados)}. Tente outro mês ou verifique os dados."
            return "Ops! Não encontrei seu holerite para os meses informados."
        
        # Envia um ou mais holerites
        try:
            enviados=[]
            erros=[]
            for competencia,arquivo in arquivos_por_comp:
                caminho_abs=os.path.join(UPLOAD_ROOT,arquivo.caminho)
                if not os.path.exists(caminho_abs):
                    erros.append(f"{competencia} (arquivo indisponível)")
                    continue
                nome_arquivo=(arquivo.nome_arquivo or os.path.basename(caminho_abs) or f'holerite_{competencia.replace("/", "-")}.pdf')
                wa_send_pdf(numero,caminho_abs,nome_arquivo,f"Holerite {competencia}")
                enviados.append((competencia,nome_arquivo))

            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            conversa.ultima_msg=utcnow()
            if enviados:
                db.session.add(WhatsAppMensagem(
                    conversa_id=conversa.id,
                    numero=numero,
                    direcao='out',
                    tipo='documento',
                    conteudo='Holerites enviados: '+', '.join([f"{c} ({n})" for c,n in enviados])
                ))
            db.session.commit()

            if not enviados:
                base='Não consegui enviar os holerites solicitados agora.'
                if erros:
                    base+=f" Erros: {', '.join(erros)}."
                if nao_encontrados:
                    base+=f" Não encontrados: {', '.join(nao_encontrados)}."
                return base

            comps_enviadas=', '.join([c for c,_ in enviados])
            msg=f"✅ Pronto! Enviei seu(s) holerite(s) de: {comps_enviadas}."
            if nao_encontrados:
                msg+=f"\n\nNão encontrei para: {', '.join(nao_encontrados)}."
            if erros:
                msg+=f"\n\nHouve falha no envio para: {', '.join(erros)}."
            return msg
        except Exception as e:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return f"Desculpe, houve um erro ao enviar seu holerite. Tente novamente mais tarde."
    
    # Padrão: reinicia o diálogo
    ctx['holerite_estado']=None
    conversa.contexto=json.dumps(ctx,ensure_ascii=False)
    db.session.commit()
    return None

def ai_wa_reply(numero,texto,historico=None):
    txt=(texto or '').strip()
    if not txt: return ''
    
    # Verifica se é um pedido de holerite
    conversa=WhatsAppConversa.query.filter_by(numero=numero).first()
    if conversa and _detecta_pedido_holerite(txt):
        resposta_holerite=_processa_dialogo_holerite(conversa.id,numero,txt)
        if resposta_holerite is not None:
            return resposta_holerite
    elif conversa:
        # Verifica se está em um diálogo de holerite já em progresso
        try:
            ctx=json.loads(conversa.contexto or '{}')
            if ctx.get('holerite_estado'):
                resposta_holerite=_processa_dialogo_holerite(conversa.id,numero,txt)
                if resposta_holerite is not None:
                    return resposta_holerite
        except:
            pass
    
    cfg=ai_wa_cfg()
    key=(cfg.get('api_key') or '').strip()
    if not key: raise ValueError('API Key da IA nao configurada')
    provider=ai_provider_norm(cfg.get('provider') or 'gemini')
    model=ai_model_norm(provider,cfg.get('model') or '')
    system=(cfg.get('prompt') or '').strip() or 'Você é um assistente de atendimento da RM Facilities. Responda em português, de forma objetiva e cordial. Se o usuário pedir por holerite, contracheque ou documentos relacionados, responda de forma amigável. Se o usuário pedir mais de um mês e não informar o ano, considere o ano corrente.'
    try: temp=float(cfg.get('temperature') or 0.3)
    except Exception: temp=0.3
    try: max_tk=int(float(cfg.get('max_tokens') or 350))
    except Exception: max_tk=350
    hist=[m for m in (historico or []) if (m.conteudo or '').strip() and m.tipo!='erro']
    if provider=='openai':
        if key.startswith('AIza'):
            raise ValueError('API Key parece ser do Gemini, mas o provedor selecionado é OpenAI.')
        mdl=model or 'gpt-4o-mini'
        url=(gc('ia_openai_url','') or 'https://api.openai.com/v1/chat/completions').strip()
        if hist:
            turns=_build_turns(hist,'user','assistant')
            msgs=[{'role':'system','content':system}]+[{'role':t['role'],'content':t['_text']} for t in turns]
        else:
            msgs=[{'role':'system','content':system},{'role':'user','content':f'Número: {numero}\nMensagem: {txt}'}]
        payload={'model':mdl,'messages':msgs,'temperature':temp,'max_tokens':max_tk}
        out=_post_json(url,payload,headers={'Authorization':f'Bearer {key}'},timeout=45)
        msg=((out.get('choices') or [{}])[0].get('message') or {})
        resp=msg.get('content') or ''
        if isinstance(resp,list):
            # Newer OpenAI payloads may return content blocks instead of a plain string.
            txt_blocks=[]
            for b in resp:
                if isinstance(b,dict):
                    t=(b.get('text') or '').strip()
                    if t:
                        txt_blocks.append(t)
            resp='\n'.join(txt_blocks)
        resp=str(resp).strip()
        if resp:
            return resp
        # Fallback sem historico quando o provider retorna escolha vazia.
        payload_fb={
            'model':mdl,
            'messages':[{'role':'system','content':system},{'role':'user','content':f'Número: {numero}\nMensagem: {txt}'}],
            'temperature':temp,
            'max_tokens':max_tk
        }
        out_fb=_post_json(url,payload_fb,headers={'Authorization':f'Bearer {key}'},timeout=45)
        msg_fb=((out_fb.get('choices') or [{}])[0].get('message') or {})
        resp_fb=msg_fb.get('content') or ''
        if isinstance(resp_fb,list):
            txt_blocks=[]
            for b in resp_fb:
                if isinstance(b,dict):
                    t=(b.get('text') or '').strip()
                    if t:
                        txt_blocks.append(t)
            resp_fb='\n'.join(txt_blocks)
        resp_fb=str(resp_fb).strip()
        if resp_fb:
            return resp_fb
        raise ValueError('IA retornou resposta vazia (OpenAI). Verifique modelo/chave e tente novamente.')
    if provider=='gemini':
        if key.startswith('sk-'):
            raise ValueError('API Key parece ser da OpenAI, mas o provedor selecionado é Gemini.')
        mdl=model or 'gemini-1.5-flash'
        base=(gc('ia_gemini_url','') or '').strip()
        url=base or f'https://generativelanguage.googleapis.com/v1beta/models/{mdl}:generateContent?key={key}'
        if hist:
            turns=_build_turns(hist,'user','model')
            while turns and turns[0]['role']=='model': turns.pop(0)
            contents=[{'role':t['role'],'parts':[{'text':t['_text']}]} for t in turns] if turns else [{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}]
        else:
            contents=[{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}]
        payload={'system_instruction':{'parts':[{'text':system}]},'contents':contents,'generationConfig':{'temperature':temp,'maxOutputTokens':max_tk}}
        out=_post_json(url,payload,timeout=45)
        cand=(out.get('candidates') or [{}])[0]
        parts=((cand.get('content') or {}).get('parts') or [])
        resp='\n'.join((p.get('text') or '').strip() for p in parts if (p.get('text') or '').strip())
        resp=resp.strip()
        if resp:
            return resp
        # Fallback sem historico quando Gemini nao devolve parts/texto.
        payload_fb={
            'system_instruction':{'parts':[{'text':system}]},
            'contents':[{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}],
            'generationConfig':{'temperature':temp,'maxOutputTokens':max_tk}
        }
        out_fb=_post_json(url,payload_fb,timeout=45)
        cand_fb=(out_fb.get('candidates') or [{}])[0]
        parts_fb=((cand_fb.get('content') or {}).get('parts') or [])
        resp_fb='\n'.join((p.get('text') or '').strip() for p in parts_fb if (p.get('text') or '').strip()).strip()
        if resp_fb:
            return resp_fb
        finish=(cand_fb.get('finishReason') or cand.get('finishReason') or '').strip()
        if finish:
            raise ValueError(f'IA retornou resposta vazia (Gemini: {finish}).')
        raise ValueError('IA retornou resposta vazia (Gemini). Verifique modelo/chave e tente novamente.')
    raise ValueError('Provedor de IA inválido. Use gemini ou openai.')

def smtp_send_pdf(dest,nome_dest,caminho_abs,nome_arquivo,competencia='',remetente='RM Facilities'):
    cfg=smtp_cfg()
    if not cfg['host'] or not cfg['user']: raise ValueError('SMTP nao configurado')
    msg=MIMEMultipart()
    msg['From']=f"{remetente} <{cfg['de'] or cfg['user']}>"
    msg['To']=dest; msg['Subject']=f"Holerite {competencia} - {remetente}"
    corpo=f"Ol\u00e1 {nome_dest},\n\nSegue em anexo seu holerite{(' de '+competencia) if competencia else ''}.\n\nAtenciosamente,\n{remetente}"
    msg.attach(MIMEText(corpo,'plain','utf-8'))
    with open(caminho_abs,'rb') as f:
        part=MIMEBase('application','octet-stream'); part.set_payload(f.read())
    encoders.encode_base64(part); part.add_header('Content-Disposition',f'attachment; filename="{nome_arquivo}"'); msg.attach(part)
    port=int(cfg['port'] or 587)
    if str(cfg['tls']) in ('1','true','True','yes'):
        with smtplib.SMTP(cfg['host'],port,timeout=20) as s: s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
    else:
        with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s: s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())

ALLOWED_AREAS=['dashboard','medicoes','historico','clientes','empresas','usuarios','config','rh','operacional','sst','rh-digital','documentos']
UPLOAD_ROOT=os.path.join(os.path.dirname(__file__),'instance','uploads')
DOC_CAT_PATH={
    'aso':'aso',
    'epi':'epi',
    'treinamento':'treinamento',
    'holerite':'holerites',
    'folha_ponto':'folha_ponto',
    'contrato_trabalho':'contrato_trabalho',
    'vale_transporte':'vale_transporte',
    'requisicao_vale_transporte':'requisicao_vale_transporte',
    'uniforme':'uniforme',
    'outros':'outros',
}
DOC_CAT_LABEL={
    'aso':'ASO',
    'epi':'EPI',
    'treinamento':'Treinamento',
    'holerite':'Holerite',
    'folha_ponto':'Folha de Ponto',
    'contrato_trabalho':'Contrato de Trabalho',
    'vale_transporte':'Vale Transporte',
    'requisicao_vale_transporte':'Requisicao de Vale Transporte',
    'uniforme':'Uniforme/Fardamento',
    'outros':'Outros',
}

def jloads(v,dv):
    try: return json.loads(v) if v else dv
    except: return dv

def to_num(s,dec=False):
    if s is None: return 0.0 if dec else 0
    t=str(s).strip().replace('.','').replace(',','.')
    try: return float(t) if dec else int(float(t))
    except: return 0.0 if dec else 0

def next_cli_num():
    max_n=0
    for (n,) in db.session.query(Cliente.numero).all():
        try: max_n=max(max_n,int(str(n or '').strip()))
        except: pass
    return str(max_n+1).zfill(3)

def save_upload(fs,subdir):
    os.makedirs(os.path.join(UPLOAD_ROOT,subdir),exist_ok=True)
    base=secure_filename(fs.filename or 'arquivo.bin')
    nome=f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{base}"
    rel=os.path.join(subdir,nome)
    abs_p=os.path.join(UPLOAD_ROOT,rel)
    fs.save(abs_p)
    return rel,abs_p

def norm_cat(v):
    s=(v or 'outros').strip().lower().replace('-', '_').replace(' ', '_')
    aliases={
        'holerites':'holerite',
        'ficha_de_epi':'epi',
        'fardamento':'uniforme',
        'figurino':'uniforme',
        'req_vt':'requisicao_vale_transporte',
        'requisicao_vt':'requisicao_vale_transporte',
        'req_vale_transporte':'requisicao_vale_transporte',
    }
    s=aliases.get(s,s)
    return s if s in DOC_CAT_PATH else 'outros'

def infer_doc_year(comp=''):
    c=(comp or '').strip()
    m=re.search(r'(19|20)\d{2}',c)
    if m: return m.group(0)
    return str(datetime.now().year)

def func_doc_subdir(funcionario_id,categoria,competencia=''):
    cat=norm_cat(categoria)
    pasta_cat=DOC_CAT_PATH.get(cat,'outros')
    ano=infer_doc_year(competencia)
    return os.path.join('funcionarios',str(funcionario_id),pasta_cat,ano),cat

def prepare_func_doc_dirs(funcionario_id,ano=None):
    y=str(ano or datetime.now().year)
    made=[]
    for _,pasta in DOC_CAT_PATH.items():
        rel=os.path.join('funcionarios',str(funcionario_id),pasta,y)
        ap=os.path.join(UPLOAD_ROOT,rel)
        os.makedirs(ap,exist_ok=True)
        made.append(rel)
    return made

def arq_year_from_path(caminho):
    p=[x for x in str(caminho or '').split(os.sep) if x]
    if p and re.fullmatch(r'(19|20)\d{2}',p[-2] if len(p)>=2 else ''):
        return p[-2]
    m=re.search(r'/(19|20)\d{2}/',('/'+str(caminho or '').replace('\\','/')+'/'))
    return m.group(0).strip('/') if m else str(datetime.now().year)

def can_access_area(area):
    if session.get('perfil')=='dono': return True
    areas=session.get('areas',[]) or []
    # Compatibilidade: usuários antigos sem áreas seguem com acesso até configuração explícita.
    if not areas: return True
    return area in areas

def can_access_request(path,method='GET'):
    area=area_from_path(path)
    if not area: return True
    if can_access_area(area): return True
    p=(path or '').lower()
    m=(method or 'GET').upper()
    if m=='GET' and p.startswith('/api/clientes') and can_access_area('medicoes'):
        return True
    if m=='GET' and (p.startswith('/api/medicoes') or p.startswith('/api/pdf')) and can_access_area('historico'):
        return True
    return False

def area_from_path(path):
    p=(path or '').lower()
    if p.startswith('/api/dashboard'): return 'dashboard'
    if p.startswith('/api/medicoes') or p.startswith('/api/proximo-numero') or p.startswith('/api/pdf'): return 'medicoes'
    if p.startswith('/api/clientes'): return 'clientes'
    if p.startswith('/api/empresas'): return 'empresas'
    if p.startswith('/api/config/smtp') or p.startswith('/api/config/whatsapp') or p.startswith('/api/config/ia-whatsapp') or p.startswith('/api/whatsapp') or p.startswith('/webhook/whatsapp'): return 'rh-digital'
    if p.startswith('/api/config') or p.startswith('/api/backup'): return 'config'
    if p.startswith('/api/funcionarios'): return 'rh'
    if p.startswith('/api/rh/'): return 'rh-digital'
    if p.startswith('/api/ordens-compra') or p.startswith('/api/operacional'): return 'operacional'
    if p.startswith('/api/usuarios'): return 'usuarios'
    return None

def build_func_docs_response(funcionario_id):
    cat_q=(request.args.get('categoria') or '').strip().lower()
    if cat_q and norm_cat(cat_q)=='outros' and cat_q not in ['outros','outro']:
        return {'erro':'Categoria invalida'},400
    cat_filter=norm_cat(cat_q) if cat_q else ''
    ano_filter=(request.args.get('ano') or '').strip()
    if ano_filter and not re.fullmatch(r'(19|20)\d{2}',ano_filter):
        return {'erro':'Ano invalido'},400
    q=(request.args.get('q') or '').strip().lower()
    page=max(1,to_num(request.args.get('page')) or 1)
    per_page=to_num(request.args.get('per_page')) or 50
    per_page=min(max(1,per_page),200)
    formato=(request.args.get('formato') or 'arvore').strip().lower()
    if formato not in ['arvore','lista']:
        return {'erro':'Formato invalido. Use arvore ou lista'},400

    regs=FuncionarioArquivo.query.filter_by(funcionario_id=funcionario_id).order_by(FuncionarioArquivo.criado_em.desc()).all()
    itens=[]
    for a in regs:
        cat=norm_cat(a.categoria)
        ano=arq_year_from_path(a.caminho)
        if cat_filter and cat!=cat_filter:
            continue
        if ano_filter and ano!=ano_filter:
            continue
        cat_label=DOC_CAT_LABEL.get(cat,cat)
        if q and q not in (a.nome_arquivo or '').lower() and q not in (a.competencia or '').lower() and q not in (a.caminho or '').lower() and q not in cat_label.lower():
            continue
        itens.append({
            'id':a.id,
            'categoria':cat,
            'categoria_label':DOC_CAT_LABEL.get(cat,cat),
            'ano':ano,
            'nome_arquivo':a.nome_arquivo,
            'competencia':a.competencia,
            'caminho':a.caminho,
            'criado_em':a.criado_em.isoformat() if a.criado_em else '',
            'criado_fmt':a.criado_em.strftime('%d/%m/%Y %H:%M') if a.criado_em else '',
            'download_url':f'/api/funcionarios/arquivos/{a.id}/download',
            'app_download_url':f'/api/app/funcionario/arquivos/{a.id}/download',
        })

    total_itens=len(itens)
    total_paginas=max(1,(total_itens+per_page-1)//per_page)
    if page>total_paginas: page=total_paginas
    ini=(page-1)*per_page
    fim=ini+per_page
    itens_pag=itens[ini:fim]

    arv={}
    for a in itens_pag:
        cat=a['categoria']
        ano=a['ano']
        if cat not in arv:
            arv[cat]={'categoria':cat,'categoria_label':a['categoria_label'],'anos':{}}
        if ano not in arv[cat]['anos']:
            arv[cat]['anos'][ano]=[]
        arv[cat]['anos'][ano].append(a)

    out=[]
    for cat in sorted(arv.keys(),key=lambda k:DOC_CAT_LABEL.get(k,k)):
        anos=[]
        for ano in sorted(arv[cat]['anos'].keys(),reverse=True):
            anos.append({'ano':ano,'itens':arv[cat]['anos'][ano]})
        out.append({'categoria':cat,'categoria_label':arv[cat]['categoria_label'],'anos':anos})

    resp={
        'ok':True,
        'versao':'v1',
        'funcionario_id':funcionario_id,
        'formato':formato,
        'filtros':{'categoria':cat_filter,'ano':ano_filter,'q':q},
        'paginacao':{
            'page':page,
            'per_page':per_page,
            'total_itens':total_itens,
            'total_paginas':total_paginas,
            'tem_anterior':page>1,
            'tem_proxima':page<total_paginas,
        },
        'categorias':out,
    }
    if formato=='lista':
        resp['itens']=itens_pag
    return resp,200

def read_rows_from_upload(arq):
    nome=(arq.filename or '').lower()
    if nome.endswith('.csv'):
        txt=arq.read().decode('utf-8-sig')
        return list(csv.DictReader(io.StringIO(txt),delimiter=';'))
    if nome.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise ValueError('Dependencia openpyxl nao instalada para XLSX')
        wb=load_workbook(arq,data_only=True,read_only=True)
        ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
        if not rows: return []
        heads=[str(h or '').strip() for h in rows[0]]
        out=[]
        for vals in rows[1:]:
            if not vals or not any(v not in [None,''] for v in vals):
                continue
            out.append({heads[i]:(vals[i] if i < len(vals) else '') for i in range(len(heads))})
        return out
    raise ValueError('Formato inválido. Use CSV ou XLSX')

def only_digits(v):
    return ''.join(ch for ch in str(v or '') if ch.isdigit())

def norm_cpf(v):
    d=only_digits(v)[:11]
    return d or None

def norm_doc(v):
    return only_digits(v)[:14]

def norm_phone(v):
    return only_digits(v)[:11]

def norm_cep(v):
    return only_digits(v)[:8]

def norm_matricula(v):
    return only_digits(v)[:20]

def to_bool(v):
    if isinstance(v,bool):
        return v
    return str(v or '').strip().lower() in ('1','true','sim','yes','on')

def next_func_matricula():
    mx=0
    for (m,) in db.session.query(Funcionario.matricula).all():
        d=only_digits(m)
        if not d:
            continue
        try:
            mx=max(mx,int(d))
        except Exception:
            pass
    return str(mx+1)

def next_func_re():
    mx=299
    for (r,) in db.session.query(Funcionario.re).all():
        if r and isinstance(r,int):
            mx=max(mx,int(r))
    return mx+1

def parse_json_bytes(raw):
    try:
        return json.loads((raw or b'').decode('utf-8','ignore'))
    except Exception:
        return {}

def norm_dt8(v):
    s=''.join(ch for ch in str(v or '') if ch.isdigit())
    if len(s)==8:
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return str(v or '').strip()

def ext_cpf_payload(data):
    base=data if isinstance(data,dict) else {}
    if isinstance(data,list) and data:
        base=data[0] if isinstance(data[0],dict) else {}
    for k in ['data','result','resultado','response']:
        if isinstance(base.get(k),dict):
            base=base[k]
            break
    nome=(base.get('nome') or base.get('name') or base.get('Nome') or '').strip()
    nome_social=(base.get('nome_social') or base.get('NomeSocial') or '').strip()
    cpf=only_digits(base.get('cpf') or base.get('documento') or base.get('Cpf') or '')
    nasc=(base.get('nascimento') or base.get('data_nascimento') or base.get('birth_date') or base.get('DataNascimento') or '').strip()
    mae=(base.get('mae') or base.get('nome_mae') or base.get('mother') or base.get('NomeMae') or '').strip()
    situacao=(base.get('situacao') or base.get('status') or base.get('DescSituacaoCadastral') or base.get('SituacaoCadastral') or '').strip()
    tipo_log=(base.get('tipo_logradouro') or base.get('TipoLogradouro') or '').strip()
    logradouro=(base.get('logradouro') or base.get('Logradouro') or '').strip()
    numero=(base.get('numero_logradouro') or base.get('NumeroLogradouro') or '').strip()
    compl=(base.get('complemento') or base.get('Complemento') or '').strip()
    bairro=(base.get('bairro') or base.get('Bairro') or '').strip()
    municipio=(base.get('municipio') or base.get('Municipio') or '').strip()
    uf=(base.get('uf') or base.get('UF') or '').strip()
    cep=only_digits(base.get('cep') or base.get('Cep') or '')
    ddd=only_digits(base.get('ddd') or base.get('DDD') or '')
    tel=only_digits(base.get('telefone') or base.get('Telefone') or '')
    endereco=' '.join([tipo_log,logradouro]).strip()
    return {
        'nome':nome,
        'nome_social':nome_social,
        'cpf':cpf,
        'nascimento':norm_dt8(nasc),
        'mae':mae,
        'situacao':situacao,
        'endereco':endereco,
        'numero':numero,
        'complemento':compl,
        'bairro':bairro,
        'cidade':municipio,
        'estado':uf,
        'cep':cep,
        'ddd':ddd,
        'telefone':tel,
    }

def lookup_cpf_externo(cpf):
    tpl=(os.environ.get('CPF_LOOKUP_URL') or '').strip()
    if not tpl:
        return {'ok':False,'motivo':'nao_configurado'}
    url=tpl.replace('{cpf}',cpf)
    token=(os.environ.get('CPF_LOOKUP_TOKEN') or '').strip()
    token_hdr=(os.environ.get('CPF_LOOKUP_TOKEN_HEADER') or '').strip()
    headers={'Accept':'application/json'}
    if token:
        headers['Authorization']=f'Bearer {token}'
        if token_hdr:
            headers[token_hdr]=token
    try:
        req=urllib.request.Request(url,headers=headers)
        with urllib.request.urlopen(req,timeout=12) as r:
            data=parse_json_bytes(r.read())
        d=ext_cpf_payload(data)
        if not d.get('nome'):
            return {'ok':False,'motivo':'sem_dados'}
        if not d.get('cpf'):
            d['cpf']=cpf
        return {'ok':True,'fonte':(os.environ.get('CPF_LOOKUP_PROVIDER') or 'API externa CPF'),'dados':d}
    except Exception as e:
        return {'ok':False,'motivo':'erro_consulta','erro':str(e)}

def parse_doc_num(s):
    if not s: return None
    m=re.match(r'\s*(\d+)',str(s))
    return int(m.group(1)) if m else None

def norm_url(v):
    u=(v or '').strip()
    if not u: return ''
    if u.startswith(('http://','https://')): return u
    return f'https://{u}'

def ensure_cols(table,defs):
    cols={r[1] for r in db.session.execute(text(f'PRAGMA table_info({table})')).fetchall()}
    changed=False
    for d in defs:
        name=d.split(' ',1)[0]
        if name not in cols:
            db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {d}'))
            changed=True
    if changed: db.session.commit()

def sc_cfg(k,v):
    c=Config.query.filter_by(chave=k).first()
    if c: c.valor=str(v)
    else: db.session.add(Config(chave=k,valor=str(v)))
    db.session.commit()

def prox_num():
    try: base=int(gc('num_base','100'))
    except: base=100
    try: ultima_cfg=int(gc('num_ultima','0'))
    except: ultima_cfg=0
    ul=Medicao.query.order_by(Medicao.id.desc()).first()
    ultima_db=parse_doc_num(ul.numero) if ul and ul.numero else 0
    prox=max(base,ultima_cfg,ultima_db)+1
    return f"{prox}/{datetime.now().year}"

def lr(f):
    @wraps(f)
    def w(*a,**k):
        if 'uid' not in session: return redirect(url_for('login'))
        if not can_access_request(request.path,request.method): return jsonify({'erro':'Acesso negado'}),403
        return f(*a,**k)
    return w

def dr(f):
    @wraps(f)
    def w(*a,**k):
        if 'uid' not in session: return redirect(url_for('login'))
        if session.get('perfil')!='dono': return jsonify({'erro':'Acesso negado'}),403
        return f(*a,**k)
    return w

def fmt_brl(v):
    try: return 'R$ {:,.2f}'.format(float(v or 0)).replace(',','X').replace('.',',').replace('X','.')
    except: return 'R$ 0,00'

def fmt_data(s):
    if not s: return ''
    try: p=s.split('-'); return f"{p[2]}/{p[1]}/{p[0]}"
    except: return s

def fmt_mes(s):
    if not s: return ''
    ms=['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    try: y,m=s.split('-'); return f"{ms[int(m)-1]}/{y}"
    except: return s

LOGO_PATH=os.path.join(os.path.dirname(__file__),'static','img','logo.png')
LOGO_URL='https://rmfacilities.com.br/wp-content/uploads/2023/08/logo-rm-facilities-1.png'

def get_logo():
    if not os.path.exists(LOGO_PATH):
        try: urllib.request.urlretrieve(LOGO_URL,LOGO_PATH)
        except: pass
    return LOGO_PATH if os.path.exists(LOGO_PATH) else None

@app.route('/login',methods=['GET','POST'])
def login():
    if 'uid' in session: return redirect(url_for('index'))
    erro=None
    if request.method=='POST':
        email=request.form.get('email','').lower().strip()
        senha=request.form.get('senha','')
        if auth_blocked('admin',email,(request.remote_addr or '')):
            erro='Muitas tentativas. Aguarde alguns minutos.'
            audit_event('auth_admin_bloqueado','admin',email,'usuario','',False,{'motivo':'rate_limit'})
            return render_template('login.html',erro=erro)
        u=Usuario.query.filter_by(email=email,ativo=True).first()
        if u and pw_check(u.senha,senha):
            if not pw_is_modern(u.senha):
                u.senha=pw_hash(senha)
            session.permanent=True
            session['uid']=u.id; session['nome']=u.nome; session['perfil']=u.perfil
            session['areas']=jloads(u.areas,[])
            u.ultimo_acesso=utcnow(); db.session.commit()
            reg_auth_attempt('admin',email,True,'ok')
            audit_event('auth_admin_sucesso','usuario',u.id,'usuario',u.id,True,{})
            return redirect(url_for('index'))
        reg_auth_attempt('admin',email,False,'credenciais_invalidas')
        audit_event('auth_admin_falha','admin',email,'usuario','',False,{})
        erro='E-mail ou senha incorretos.'
    return render_template('login.html',erro=erro)

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/')
@lr
def index(): return render_template('app.html',nome=session['nome'],perfil=session['perfil'],areas=json.dumps(session.get('areas',[]),ensure_ascii=False))

@app.route('/api/cnpj/<cnpj>')
@lr
def api_cnpj(cnpj):
    c=''.join(filter(str.isdigit,cnpj))
    if len(c)!=14: return jsonify({'erro':'CNPJ inválido'}),400
    # Tenta Receitaws primeiro e usa BrasilAPI como fallback.
    try:
        req=urllib.request.Request(f'https://receitaws.com.br/v1/cnpj/{c}',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        if d.get('status')!='ERROR':
            return jsonify({'nome':d.get('fantasia') or d.get('nome',''),'razao':d.get('nome',''),'email':d.get('email',''),'telefone':d.get('telefone',''),'cep':d.get('cep','').replace('.','').replace('-','').replace(' ',''),'logradouro':d.get('logradouro',''),'numero':d.get('numero',''),'complemento':d.get('complemento',''),'bairro':d.get('bairro',''),'cidade':d.get('municipio',''),'estado':d.get('uf',''),'situacao':d.get('situacao','')})
    except Exception:
        pass
    try:
        req=urllib.request.Request(f'https://brasilapi.com.br/api/cnpj/v1/{c}',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        return jsonify({'nome':d.get('nome_fantasia') or d.get('razao_social',''),'razao':d.get('razao_social',''),'email':d.get('email',''),'telefone':d.get('ddd_telefone_1') or d.get('ddd_telefone_2',''),'cep':str(d.get('cep','')).replace('-',''),'logradouro':d.get('logradouro',''),'numero':d.get('numero',''),'complemento':d.get('complemento',''),'bairro':d.get('bairro',''),'cidade':d.get('municipio',''),'estado':d.get('uf',''),'situacao':d.get('descricao_situacao_cadastral','')})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/cep/<cep>')
@lr
def api_cep(cep):
    c=''.join(filter(str.isdigit,cep))
    if len(c)!=8: return jsonify({'erro':'CEP inválido'}),400
    try:
        req=urllib.request.Request(f'https://viacep.com.br/ws/{c}/json/',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        if d.get('erro'): return jsonify({'erro':'CEP não encontrado'}),404
        return jsonify({'logradouro':d.get('logradouro',''),'bairro':d.get('bairro',''),'cidade':d.get('localidade',''),'estado':d.get('uf','')})
    except Exception as e: return jsonify({'erro':str(e)}),500

@app.route('/api/empresas',methods=['GET'])
@lr
def api_empresas(): return jsonify([e.to_dict() for e in Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem).all()])

@app.route('/api/empresas/<int:id>',methods=['GET'])
@lr
def api_empresa(id): return jsonify(Empresa.query.get_or_404(id).to_dict())

@app.route('/api/empresas',methods=['POST'])
@dr
def api_criar_empresa():
    d=request.json or {}
    if 'site' in d: d['site']=norm_url(d.get('site'))
    if 'logo_url' in d: d['logo_url']=norm_url(d.get('logo_url'))
    d['cnpj']=norm_doc(d.get('cnpj'))
    d['cep']=norm_cep(d.get('cep'))
    d['telefone']=norm_phone(d.get('telefone'))
    d['contato_telefone']=norm_phone(d.get('contato_telefone'))
    cols=[c.name for c in Empresa.__table__.columns if c.name not in['id','criado_em'] and hasattr(Empresa,c.name)]
    e=Empresa(**{k:d[k] for k in cols if k in d})
    db.session.add(e); db.session.commit(); return jsonify(e.to_dict()),201

@app.route('/api/empresas/<int:id>',methods=['PUT'])
@dr
def api_atualizar_empresa(id):
    e=Empresa.query.get_or_404(id); d=request.json or {}
    if 'site' in d: d['site']=norm_url(d.get('site'))
    if 'logo_url' in d: d['logo_url']=norm_url(d.get('logo_url'))
    if 'cnpj' in d: d['cnpj']=norm_doc(d.get('cnpj'))
    if 'cep' in d: d['cep']=norm_cep(d.get('cep'))
    if 'telefone' in d: d['telefone']=norm_phone(d.get('telefone'))
    if 'contato_telefone' in d: d['contato_telefone']=norm_phone(d.get('contato_telefone'))
    for k in [c.name for c in Empresa.__table__.columns]:
        if k in d and k!='id': setattr(e,k,d[k])
    db.session.commit(); return jsonify(e.to_dict())

@app.route('/api/empresas/<int:id>',methods=['DELETE'])
@dr
def api_deletar_empresa(id):
    e=Empresa.query.get_or_404(id); e.ativa=False; db.session.commit(); return jsonify({'ok':True})

@app.route('/api/config',methods=['GET'])
@lr
def api_get_config(): return jsonify({k:gc(k) for k in ['num_base','num_ultima']})

@app.route('/api/config',methods=['POST'])
@dr
def api_save_config():
    for k,v in request.json.items(): sc_cfg(k,v)
    return jsonify({'ok':True})

@app.route('/api/usuarios',methods=['GET'])
@dr
def api_usuarios(): return jsonify([u.to_dict() for u in Usuario.query.all()])

@app.route('/api/usuarios',methods=['POST'])
@dr
def api_criar_usuario():
    d=request.json
    if Usuario.query.filter_by(email=d['email'].lower()).first(): return jsonify({'erro':'E-mail já cadastrado'}),400
    ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
    u=Usuario(nome=d['nome'],email=d['email'].lower(),senha=hs(d['senha']),perfil=d.get('perfil','admin'),areas=json.dumps(ars,ensure_ascii=False))
    db.session.add(u); db.session.commit(); return jsonify(u.to_dict()),201

@app.route('/api/usuarios/<int:id>',methods=['PUT'])
@dr
def api_atualizar_usuario(id):
    u=Usuario.query.get_or_404(id); d=request.json
    for k in ['nome','perfil','ativo']:
        if k in d: setattr(u,k,d[k])
    if d.get('senha'): u.senha=hs(d['senha'])
    if 'areas' in d:
        ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
        u.areas=json.dumps(ars,ensure_ascii=False)
    db.session.commit(); return jsonify(u.to_dict())

@app.route('/api/usuarios/<int:id>',methods=['DELETE'])
@dr
def api_deletar_usuario(id):
    u=Usuario.query.get_or_404(id)
    if u.perfil=='dono' and Usuario.query.filter_by(perfil='dono').count()<=1: return jsonify({'erro':'Não é possível excluir o único dono'}),400
    db.session.delete(u); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/clientes',methods=['GET'])
@lr
def api_clientes():
    q=request.args.get('q','').lower(); emp=request.args.get('emp',''); st=request.args.get('status','')
    qr=Cliente.query
    if emp: qr=qr.filter_by(empresa_id=int(emp))
    if st: qr=qr.filter_by(status=st)
    lista=qr.order_by(Cliente.nome).all()
    if q:
        qdig=only_digits(q)
        lista=[c for c in lista if (
            q in c.nome.lower() or
            q in (c.cnpj or '').lower() or
            q in (c.numero or '') or
            (qdig and qdig in only_digits(c.cnpj))
        )]
    return jsonify([c.to_dict() for c in lista])

@app.route('/api/clientes',methods=['POST'])
@lr
def api_criar_cliente():
    d=request.json or {}; n=next_cli_num()
    d['cnpj']=norm_doc(d.get('cnpj'))
    d['telefone']=norm_phone(d.get('telefone'))
    d['cep']=norm_cep(d.get('cep'))
    skip=['id','numero','criado_em','end_fmt']
    cols=[c.name for c in Cliente.__table__.columns if c.name not in skip]
    kw={k:d[k] for k in cols if k in d}
    c=Cliente(numero=n,**kw); db.session.add(c); db.session.commit(); return jsonify(c.to_dict()),201

@app.route('/api/clientes/modelo')
@lr
def api_clientes_modelo():
    cab=['nome','cnpj','responsavel','telefone','email','cep','logradouro','numero_end','complemento','bairro','cidade','estado','empresa_id','status','limpeza','jardinagem','portaria','vencimento','obs']
    exemplo=['Condominio Exemplo','12.345.678/0001-90','Maria Silva','(12) 99123-4567','contato@exemplo.com','12246000','Rua Central','100','','Centro','Sao Jose dos Campos','SP','','Ativo','1500,00','300,00','2500,00','10','Contrato mensal']
    buf=io.StringIO(); w=csv.writer(buf,delimiter=';'); w.writerow(cab); w.writerow(exemplo)
    b=io.BytesIO(buf.getvalue().encode('utf-8-sig')); b.seek(0)
    return send_file(b,mimetype='text/csv',as_attachment=True,download_name='modelo_clientes_rmfacilities.csv')

@app.route('/api/clientes/import',methods=['POST'])
@lr
def api_clientes_import():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo nao enviado'}),400
    try:
        linhas=read_rows_from_upload(arq)
    except Exception as e:
        return jsonify({'erro':str(e)}),400
    if not linhas: return jsonify({'erro':'Planilha vazia'}),400
    criados=0; erros=[]; prox=int(next_cli_num())
    for i,row in enumerate(linhas,start=2):
        nome=(row.get('nome') or '').strip()
        if not nome:
            erros.append(f'Linha {i}: nome obrigatorio')
            continue
        try:
            c=Cliente(
                numero=str(prox).zfill(3),
                nome=nome,
                cnpj=norm_doc((row.get('cnpj') or '').strip()),
                responsavel=(row.get('responsavel') or '').strip(),
                telefone=norm_phone((row.get('telefone') or '').strip()),
                email=(row.get('email') or '').strip(),
                cep=norm_cep((row.get('cep') or '').strip()),
                logradouro=(row.get('logradouro') or '').strip(),
                numero_end=(row.get('numero_end') or '').strip(),
                complemento=(row.get('complemento') or '').strip(),
                bairro=(row.get('bairro') or '').strip(),
                cidade=(row.get('cidade') or '').strip(),
                estado=(row.get('estado') or '').strip(),
                empresa_id=to_num(row.get('empresa_id')) or None,
                status=(row.get('status') or 'Ativo').strip() or 'Ativo',
                limpeza=to_num(row.get('limpeza'),dec=True),
                jardinagem=to_num(row.get('jardinagem'),dec=True),
                portaria=to_num(row.get('portaria'),dec=True),
                vencimento=to_num(row.get('vencimento')) or 10,
                obs=(row.get('obs') or '').strip()
            )
            db.session.add(c); prox+=1; criados+=1
        except Exception as e:
            erros.append(f'Linha {i}: {str(e)}')
    db.session.commit()
    return jsonify({'ok':True,'criados':criados,'erros':erros})

@app.route('/api/funcionarios/modelo')
@lr
def api_funcionarios_modelo():
    cab=['matricula','re','nome','cpf','email','telefone','cargo','funcao','cbo','setor','empresa_id','data_admissao','tipo_contrato','jornada','status','salario','vale_refeicao','vale_alimentacao','vale_transporte','cep','endereco','cidade','estado','rg','orgao_emissor','pis','ctps','titulo_eleitor','cert_reservista','cnh','exame_admissional_data','docs_admissao_ok','docs_admissao_obs','obs','areas']
    exemplo=['1001','300','Joao da Silva','123.456.789-00','joao@empresa.com','5512999990000','Auxiliar','Auxiliar de Limpeza','5143-20','Operacional','1','2026-01-10','CLT','44h semanais','Ativo','2500,00','350,00','450,00','220,00','12246000','Rua A, 100','Sao Jose dos Campos','SP','1234567','SSP/SP','12345678901','123456-serie 001','9876543210','123456','AB','2026-01-08','1','Checklist conferido','Exemplo','rh,operacional,sst']
    buf=io.StringIO(); w=csv.writer(buf,delimiter=';'); w.writerow(cab); w.writerow(exemplo)
    b=io.BytesIO(buf.getvalue().encode('utf-8-sig')); b.seek(0)
    return send_file(b,mimetype='text/csv',as_attachment=True,download_name='modelo_funcionarios_rmfacilities.csv')

@app.route('/api/funcionarios/import',methods=['POST'])
@lr
def api_funcionarios_import():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo nao enviado'}),400
    try:
        linhas=read_rows_from_upload(arq)
    except Exception as e:
        return jsonify({'erro':str(e)}),400
    if not linhas: return jsonify({'erro':'Planilha vazia'}),400
    criados=0; erros=[]
    for i,row in enumerate(linhas,start=2):
        nome=(str(row.get('nome',''))).strip()
        if not nome:
            erros.append(f'Linha {i}: nome obrigatorio')
            continue
        try:
            areas_raw=str(row.get('areas','') or '')
            ars=[a.strip().lower() for a in re.split(r'[;,]',areas_raw) if a.strip()]
            ars=[a for a in ars if a in ALLOWED_AREAS]
            mat=norm_matricula(row.get('matricula')) or next_func_matricula()
            f=Funcionario(
                matricula=mat,
                re=to_num(row.get('re')) or next_func_re(),
                nome=nome,
                cpf=norm_cpf(str(row.get('cpf','') or '').strip()),
                email=str(row.get('email','') or '').strip(),
                telefone=wa_norm_number(str(row.get('telefone','') or '').strip()),
                cargo=str(row.get('cargo','') or '').strip(),
                funcao=str(row.get('funcao','') or '').strip(),
                cbo=str(row.get('cbo','') or '').strip(),
                setor=str(row.get('setor','') or '').strip(),
                empresa_id=to_num(row.get('empresa_id')) or None,
                data_admissao=str(row.get('data_admissao','') or '').strip(),
                tipo_contrato=str(row.get('tipo_contrato','') or '').strip(),
                jornada=str(row.get('jornada','') or '').strip(),
                status=str(row.get('status','Ativo') or 'Ativo').strip() or 'Ativo',
                salario=to_num(row.get('salario'),dec=True),
                vale_refeicao=to_num(row.get('vale_refeicao'),dec=True),
                vale_alimentacao=to_num(row.get('vale_alimentacao'),dec=True),
                vale_transporte=to_num(row.get('vale_transporte'),dec=True),
                cep=norm_cep(str(row.get('cep','') or '').strip()),
                endereco=str(row.get('endereco','') or '').strip(),
                cidade=str(row.get('cidade','') or '').strip(),
                estado=str(row.get('estado','') or '').strip(),
                rg=str(row.get('rg','') or '').strip(),
                orgao_emissor=str(row.get('orgao_emissor','') or '').strip(),
                pis=str(row.get('pis','') or '').strip(),
                ctps=str(row.get('ctps','') or '').strip(),
                titulo_eleitor=str(row.get('titulo_eleitor','') or '').strip(),
                cert_reservista=str(row.get('cert_reservista','') or '').strip(),
                cnh=str(row.get('cnh','') or '').strip(),
                exame_admissional_data=str(row.get('exame_admissional_data','') or '').strip(),
                docs_admissao_ok=str(row.get('docs_admissao_ok','') or '').strip().lower() in ('1','true','sim','yes','on'),
                docs_admissao_obs=str(row.get('docs_admissao_obs','') or '').strip(),
                obs=str(row.get('obs','') or '').strip(),
                areas=json.dumps(ars,ensure_ascii=False)
            )
            db.session.add(f); criados+=1
        except Exception as e:
            erros.append(f'Linha {i}: {str(e)}')
    db.session.commit()
    return jsonify({'ok':True,'criados':criados,'erros':erros})

@app.route('/api/clientes/<int:id>',methods=['PUT'])
@lr
def api_atualizar_cliente(id):
    c=Cliente.query.get_or_404(id); d=request.json or {}
    if 'cnpj' in d: d['cnpj']=norm_doc(d.get('cnpj'))
    if 'telefone' in d: d['telefone']=norm_phone(d.get('telefone'))
    if 'cep' in d: d['cep']=norm_cep(d.get('cep'))
    for k in [col.name for col in Cliente.__table__.columns if col.name not in['id','numero','criado_em']]:
        if k in d: setattr(c,k,d[k])
    db.session.commit(); return jsonify(c.to_dict())

@app.route('/api/clientes/<int:id>',methods=['DELETE'])
@lr
def api_deletar_cliente(id):
    db.session.delete(Cliente.query.get_or_404(id)); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/proximo-numero')
@lr
def api_proximo_numero(): return jsonify({'numero':prox_num()})

@app.route('/api/medicoes',methods=['GET'])
@lr
def api_medicoes(): return jsonify([m.to_dict() for m in Medicao.query.order_by(Medicao.criado_em.desc()).all()])

@app.route('/api/medicoes/<int:id>',methods=['GET'])
@lr
def api_medicao(id): return jsonify(Medicao.query.get_or_404(id).to_dict())

@app.route('/api/medicoes',methods=['POST'])
@lr
def api_criar_medicao():
    d=request.json; svcs=d.get('servicos',[])
    numero_auto=prox_num()
    m=Medicao(numero=numero_auto,tipo=d.get('tipo','Medição de Serviços'),
              cliente_id=d.get('cliente_id'),cliente_nome=d.get('cliente_nome',''),
              cliente_cnpj=d.get('cliente_cnpj',''),cliente_end=d.get('cliente_end',''),
              cliente_resp=d.get('cliente_resp',''),empresa_id=d.get('empresa_id'),
              empresa_nome=d.get('empresa_nome',''),mes_ref=d.get('mes_ref',''),
              dt_emissao=d.get('dt_emissao',''),dt_vencimento=d.get('dt_vencimento',''),
              servicos=json.dumps(svcs,ensure_ascii=False),
              valor_bruto=sum(float(s.get('vtot',0)) for s in svcs),
              observacoes=d.get('observacoes',''),ass_empresa=d.get('ass_empresa') or session.get('nome',''),
              ass_cliente=d.get('ass_cliente',''),criado_por=session.get('nome',''))
    db.session.add(m); db.session.commit()
    n=parse_doc_num(numero_auto)
    if n is not None: sc_cfg('num_ultima',n)
    return jsonify(m.to_dict()),201

@app.route('/api/medicoes/<int:id>',methods=['DELETE'])
@lr
def api_deletar_medicao(id):
    db.session.delete(Medicao.query.get_or_404(id)); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/funcionarios',methods=['GET'])
@lr
def api_funcionarios():
    cpf=only_digits(request.args.get('cpf',''))
    if cpf:
        ex_id=to_num(request.args.get('exclude_id'))
        for f in Funcionario.query.all():
            if f.id==ex_id: continue
            if only_digits(f.cpf)==cpf: return jsonify(f.to_dict())
        return jsonify({})
    q=(request.args.get('q','') or '').lower()
    lst=Funcionario.query.order_by(Funcionario.nome).all()
    if q:
        qdig=only_digits(q)
        lst=[f for f in lst if (
            q in (f.matricula or '').lower() or
            q in (f.nome or '').lower() or
            q in (f.cpf or '').lower() or
            q in (f.cargo or '').lower() or
            q in (f.funcao or '').lower() or
            q in (f.telefone or '').lower() or
            (qdig and (qdig in only_digits(f.matricula) or qdig in only_digits(f.cpf) or qdig in only_digits(f.telefone)))
        )]
    return jsonify([f.to_dict() for f in lst])

@app.route('/api/funcionarios/proxima-matricula')
@lr
def api_funcionarios_proxima_matricula():
    return jsonify({'ok':True,'matricula':next_func_matricula()})

@app.route('/api/funcionarios/proxima-re')
@lr
def api_funcionarios_proxima_re():
    return jsonify({'ok':True,'re':next_func_re()})

@app.route('/api/funcionarios/cpf-lookup')
@lr
def api_funcionario_cpf_lookup():
    cpf=only_digits(request.args.get('cpf',''))
    if len(cpf)!=11: return jsonify({'erro':'CPF invalido'}),400
    ex_id=to_num(request.args.get('exclude_id'))
    for f in Funcionario.query.all():
        if f.id==ex_id: continue
        if only_digits(f.cpf)==cpf:
            return jsonify({'ok':True,'origem':'interno','funcionario':f.to_dict()})
    r=lookup_cpf_externo(cpf)
    return jsonify(r)

@app.route('/api/funcionarios/busca-rapida')
@lr
def api_funcionario_busca_rapida():
    re=to_num(request.args.get('re'))
    cpf=only_digits(request.args.get('cpf',''))
    nome=(request.args.get('nome','') or '').strip().lower()
    
    resultados=[]
    
    # Busca por RE se fornecido
    if re and re>0:
        f=Funcionario.query.filter_by(re=re).first()
        if f:
            resultados.append(f.to_dict())
    
    # Busca por CPF se fornecido
    if cpf and len(cpf)==11:
        f=Funcionario.query.filter(Funcionario.cpf.ilike(f'%{cpf}%')).first()
        if f and f.to_dict() not in resultados:
            resultados.append(f.to_dict())
    
    # Busca por nome se fornecido
    if nome:
        funcs=Funcionario.query.filter(Funcionario.nome.ilike(f'%{nome}%')).all()
        for f in funcs:
            if f.to_dict() not in resultados:
                resultados.append(f.to_dict())
    
    return jsonify(resultados if resultados else [])

@app.route('/api/funcionarios',methods=['POST'])
@lr
def api_criar_funcionario():
    d=request.json or {}
    if not d.get('nome'): return jsonify({'erro':'Nome obrigatorio'}),400
    ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
    mat=norm_matricula(d.get('matricula')) or next_func_matricula()
    f=Funcionario(
        matricula=mat,
        re=to_num(d.get('re')) or next_func_re(),
        nome=d.get('nome','').strip(),
        cpf=norm_cpf(d.get('cpf','')),
        email=d.get('email','').strip(),
        telefone=wa_norm_number(d.get('telefone','')),
        cargo=d.get('cargo','').strip(),
        funcao=d.get('funcao','').strip(),
        cbo=d.get('cbo','').strip(),
        setor=d.get('setor','').strip(),
        empresa_id=d.get('empresa_id'),
        data_admissao=d.get('data_admissao',''),
        tipo_contrato=d.get('tipo_contrato','').strip(),
        jornada=d.get('jornada','').strip(),
        status=d.get('status','Ativo'),
        salario=to_num(d.get('salario'),dec=True),
        vale_refeicao=to_num(d.get('vale_refeicao'),dec=True),
        vale_alimentacao=to_num(d.get('vale_alimentacao'),dec=True),
        vale_transporte=to_num(d.get('vale_transporte'),dec=True),
        endereco=d.get('endereco','').strip(),
        cidade=d.get('cidade','').strip(),
        estado=d.get('estado','').strip(),
        cep=norm_cep(d.get('cep','')),
        rg=d.get('rg','').strip(),
        orgao_emissor=d.get('orgao_emissor','').strip(),
        pis=d.get('pis','').strip(),
        ctps=d.get('ctps','').strip(),
        titulo_eleitor=d.get('titulo_eleitor','').strip(),
        cert_reservista=d.get('cert_reservista','').strip(),
        cnh=d.get('cnh','').strip(),
        exame_admissional_data=d.get('exame_admissional_data','').strip(),
        docs_admissao_ok=to_bool(d.get('docs_admissao_ok')),
        docs_admissao_obs=d.get('docs_admissao_obs','').strip(),
        obs=d.get('obs','').strip(),
        areas=json.dumps(ars,ensure_ascii=False)
    )
    db.session.add(f); db.session.commit(); return jsonify(f.to_dict()),201

@app.route('/api/funcionarios/<int:id>',methods=['PUT'])
@lr
def api_atualizar_funcionario(id):
    f=Funcionario.query.get_or_404(id); d=request.json or {}
    for k in ['matricula','re','nome','cpf','email','telefone','cargo','funcao','cbo','setor','empresa_id','data_admissao','tipo_contrato','jornada','status','endereco','cidade','estado','cep','rg','orgao_emissor','pis','ctps','titulo_eleitor','cert_reservista','cnh','exame_admissional_data','docs_admissao_obs','obs']:
        if k in d:
            if k=='cpf': setattr(f,k,norm_cpf(d.get(k)))
            elif k=='matricula': setattr(f,k,norm_matricula(d.get(k)))
            elif k=='re': setattr(f,k,to_num(d.get(k)))
            elif k=='telefone': setattr(f,k,wa_norm_number(d.get(k)))
            elif k=='cep': setattr(f,k,norm_cep(d.get(k)))
            else: setattr(f,k,d[k])
    if 'salario' in d: f.salario=to_num(d.get('salario'),dec=True)
    if 'vale_refeicao' in d: f.vale_refeicao=to_num(d.get('vale_refeicao'),dec=True)
    if 'vale_alimentacao' in d: f.vale_alimentacao=to_num(d.get('vale_alimentacao'),dec=True)
    if 'vale_transporte' in d: f.vale_transporte=to_num(d.get('vale_transporte'),dec=True)
    if 'docs_admissao_ok' in d: f.docs_admissao_ok=to_bool(d.get('docs_admissao_ok'))
    if 'areas' in d:
        ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
        f.areas=json.dumps(ars,ensure_ascii=False)
    db.session.commit(); return jsonify(f.to_dict())

@app.route('/api/funcionarios/<int:id>',methods=['DELETE'])
@lr
def api_deletar_funcionario(id):
    f=Funcionario.query.get_or_404(id)
    arqs=FuncionarioArquivo.query.filter_by(funcionario_id=id).all()
    for a in arqs:
        try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
        except: pass
        db.session.delete(a)
    db.session.delete(f); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/funcionarios/<int:id>/arquivos',methods=['GET'])
@lr
def api_funcionario_arquivos(id):
    Funcionario.query.get_or_404(id)
    return jsonify([a.to_dict() for a in FuncionarioArquivo.query.filter_by(funcionario_id=id).order_by(FuncionarioArquivo.criado_em.desc()).all()])

@app.route('/api/funcionarios/<int:id>/arquivos',methods=['POST'])
@lr
def api_funcionario_upload_arquivo(id):
    Funcionario.query.get_or_404(id)
    fs=request.files.get('arquivo')
    if not fs: return jsonify({'erro':'Arquivo nao enviado'}),400
    cat=(request.form.get('categoria') or 'outros').strip().lower()
    comp=(request.form.get('competencia') or '').strip()
    ano=infer_doc_year(comp)
    prepare_func_doc_dirs(id,ano)
    subdir,cat=func_doc_subdir(id,cat,comp)
    rel,_=save_upload(fs,subdir)
    a=FuncionarioArquivo(funcionario_id=id,categoria=cat,competencia=comp,nome_arquivo=fs.filename,caminho=rel)
    db.session.add(a); db.session.commit()
    audit_event('funcionario_arquivo_upload','usuario',session.get('uid'),'funcionario',id,True,{'arquivo_id':a.id,'categoria':cat,'caminho':rel})
    return jsonify(a.to_dict()),201

@app.route('/api/funcionarios/<int:id>/documentos/preparar',methods=['POST'])
@lr
def api_preparar_pastas_funcionario(id):
    Funcionario.query.get_or_404(id)
    d=request.json or {}
    ano=str((d.get('ano') or request.args.get('ano') or datetime.now().year)).strip()
    if not re.fullmatch(r'(19|20)\d{2}',ano):
        return jsonify({'erro':'Ano invalido'}),400
    pastas=prepare_func_doc_dirs(id,ano)
    return jsonify({'ok':True,'ano':ano,'pastas':pastas})

@app.route('/api/funcionarios/<int:id>/documentos/arvore')
@lr
def api_funcionario_documentos_arvore(id):
    Funcionario.query.get_or_404(id)
    resp,status=build_func_docs_response(id)
    return jsonify(resp),status

@app.route('/api/funcionarios/<int:id>/app-acesso',methods=['PUT'])
@lr
def api_funcionario_app_acesso(id):
    f=Funcionario.query.get_or_404(id)
    d=request.json or {}
    if 'ativo_app' in d:
        f.app_ativo=bool(d.get('ativo_app'))
    senha=(d.get('senha_app') or '').strip()
    if senha:
        if len(senha)<6: return jsonify({'erro':'Senha do app deve ter ao menos 6 caracteres'}),400
        f.app_senha=pw_hash(senha)
    db.session.commit()
    audit_event('funcionario_app_acesso_alterado','usuario',session.get('uid'),'funcionario',f.id,True,{'ativo_app':f.app_ativo,'senha_alterada':bool(senha)})
    return jsonify({'ok':True,'funcionario':f.to_dict()})

@app.route('/api/app/funcionario/login',methods=['POST'])
def api_app_funcionario_login():
    d=request.json or {}
    cpf=norm_cpf(d.get('cpf'))
    senha=(d.get('senha') or '')
    if not cpf or not senha:
        return jsonify({'erro':'CPF e senha obrigatorios'}),400
    if auth_blocked('app',cpf,(request.remote_addr or '')):
        audit_event('auth_app_bloqueado','funcionario',cpf,'funcionario','',False,{'motivo':'rate_limit'})
        return jsonify({'erro':'Muitas tentativas. Aguarde alguns minutos.'}),429
    f=Funcionario.query.filter_by(cpf=cpf).first()
    if not f:
        reg_auth_attempt('app',cpf,False,'credenciais_invalidas')
        audit_event('auth_app_falha','funcionario',cpf,'funcionario','',False,{'motivo':'nao_encontrado'})
        return jsonify({'erro':'Credenciais invalidas'}),401
    if f.app_ativo is False:
        reg_auth_attempt('app',cpf,False,'app_desativado')
        audit_event('auth_app_falha','funcionario',f.id,'funcionario',f.id,False,{'motivo':'app_desativado'})
        return jsonify({'erro':'Acesso do aplicativo desativado'}),403
    if not f.app_senha:
        reg_auth_attempt('app',cpf,False,'app_nao_configurado')
        return jsonify({'erro':'Acesso do aplicativo ainda nao configurado para este funcionario'}),403
    if not pw_check(f.app_senha,senha):
        reg_auth_attempt('app',cpf,False,'credenciais_invalidas')
        audit_event('auth_app_falha','funcionario',f.id,'funcionario',f.id,False,{'motivo':'senha_invalida'})
        return jsonify({'erro':'Credenciais invalidas'}),401
    if not pw_is_modern(f.app_senha):
        f.app_senha=pw_hash(senha)

    refresh=app_issue_refresh_token()
    ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    ua=(request.headers.get('User-Agent') or '')[:250]
    sessao=FuncionarioAppSessao(funcionario_id=f.id,refresh_hash=token_hash(refresh),exp_refresh=utcnow()+timedelta(days=14),revogado=False,ip=ip,ua=ua)
    db.session.add(sessao)
    db.session.flush()
    access=app_issue_access_token(f.id,sessao.id,ttl=3600)
    f.app_ultimo_acesso=utcnow(); db.session.commit()
    reg_auth_attempt('app',cpf,True,'ok')
    audit_event('auth_app_sucesso','funcionario',f.id,'funcionario',f.id,True,{'sessao_id':sessao.id})
    return jsonify({'ok':True,'access_token':access,'refresh_token':refresh,'token_type':'Bearer','expires_in':3600,'refresh_expires_in':1209600,'funcionario':{'id':f.id,'nome':f.nome,'cpf':f.cpf,'cargo':f.cargo,'setor':f.setor,'status':f.status}})

@app.route('/api/app/funcionario/refresh',methods=['POST'])
def api_app_funcionario_refresh():
    d=request.json or {}
    refresh=(d.get('refresh_token') or request.headers.get('X-Refresh-Token') or '').strip()
    if not refresh: return jsonify({'erro':'refresh_token obrigatorio'}),400
    sessao=FuncionarioAppSessao.query.filter_by(refresh_hash=token_hash(refresh),revogado=False).first()
    if not sessao or sessao.exp_refresh<utcnow(): return jsonify({'erro':'Refresh token invalido ou expirado'}),401
    f=Funcionario.query.get(sessao.funcionario_id)
    if not f or f.app_ativo is False: return jsonify({'erro':'Acesso desativado'}),403
    novo_refresh=app_issue_refresh_token()
    sessao.refresh_hash=token_hash(novo_refresh)
    access=app_issue_access_token(f.id,sessao.id,ttl=3600)
    db.session.commit()
    audit_event('auth_app_refresh','funcionario',f.id,'funcionario',f.id,True,{'sessao_id':sessao.id})
    return jsonify({'ok':True,'access_token':access,'refresh_token':novo_refresh,'token_type':'Bearer','expires_in':3600,'refresh_expires_in':max(0,int((sessao.exp_refresh-utcnow()).total_seconds()))})

@app.route('/api/app/funcionario/logout',methods=['POST'])
@app_func_required
def api_app_funcionario_logout():
    f=g.app_funcionario
    d=request.json or {}
    all_devices=bool(d.get('all_devices'))
    if all_devices:
        FuncionarioAppSessao.query.filter_by(funcionario_id=f.id,revogado=False).update({'revogado':True})
    else:
        g.app_sessao.revogado=True
    db.session.commit()
    audit_event('auth_app_logout','funcionario',f.id,'funcionario',f.id,True,{'all_devices':all_devices})
    return jsonify({'ok':True})

@app.route('/api/app/funcionario/me')
@app_func_required
def api_app_funcionario_me():
    f=g.app_funcionario
    return jsonify({'ok':True,'funcionario':{'id':f.id,'nome':f.nome,'cpf':f.cpf,'email':f.email,'telefone':f.telefone,'cargo':f.cargo,'setor':f.setor,'empresa_id':f.empresa_id,'status':f.status}})

@app.route('/api/app/funcionario/me/documentos')
@app_func_required
def api_app_funcionario_me_documentos():
    resp,status=build_func_docs_response(g.app_funcionario.id)
    return jsonify(resp),status

@app.route('/api/app/funcionario/arquivos/<int:id>/download')
@app_func_required
def api_app_funcionario_download_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    if a.funcionario_id!=g.app_funcionario.id:
        return jsonify({'erro':'Acesso negado'}),403
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    audit_event('funcionario_app_arquivo_download','funcionario',g.app_funcionario.id,'funcionario',a.funcionario_id,True,{'arquivo_id':a.id,'caminho':a.caminho})
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

@app.route('/api/app/funcionario/me/senha',methods=['PUT'])
@app_func_required
def api_app_funcionario_me_senha():
    f=g.app_funcionario
    d=request.json or {}
    atual=(d.get('senha_atual') or '')
    nova=(d.get('nova_senha') or '')
    if not atual or not nova:
        return jsonify({'erro':'Senha atual e nova senha sao obrigatorias'}),400
    if not pw_check(f.app_senha,atual): return jsonify({'erro':'Senha atual invalida'}),401
    if len(nova)<6: return jsonify({'erro':'Nova senha deve ter ao menos 6 caracteres'}),400
    f.app_senha=pw_hash(nova)
    db.session.commit()
    audit_event('auth_app_troca_senha','funcionario',f.id,'funcionario',f.id,True,{})
    return jsonify({'ok':True})

@app.route('/api/funcionarios/arquivos/<int:id>',methods=['DELETE'])
@lr
def api_funcionario_delete_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    fid=a.funcionario_id
    cam=a.caminho
    try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
    except: pass
    db.session.delete(a); db.session.commit()
    audit_event('funcionario_arquivo_delete','usuario',session.get('uid'),'funcionario',fid,True,{'arquivo_id':id,'caminho':cam})
    return jsonify({'ok':True})

@app.route('/api/funcionarios/arquivos/<int:id>/download')
@lr
def api_funcionario_download_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    ator_tipo='funcionario_app' if getattr(g,'app_funcionario',None) else 'usuario'
    ator_id=(getattr(g,'app_funcionario',None).id if getattr(g,'app_funcionario',None) else session.get('uid'))
    audit_event('funcionario_arquivo_download',ator_tipo,ator_id,'funcionario',a.funcionario_id,True,{'arquivo_id':a.id,'caminho':a.caminho})
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

@app.route('/api/funcionarios/holerites/upload',methods=['POST'])
@lr
def api_holerites_upload():
    fs=request.files.get('arquivo')
    comp=(request.form.get('competencia') or '').strip()
    if not fs: return jsonify({'erro':'PDF nao enviado'}),400
    try:
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500
    funcs=Funcionario.query.all()
    if not funcs: return jsonify({'erro':'Cadastre funcionarios antes do upload'}),400
    reader=PdfReader(fs)
    enviados=0; sem_match=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=(page.extract_text() or '').lower()
        alvo=None
        for f in funcs:
            nm=(f.nome or '').lower().strip()
            if nm and nm in txt:
                alvo=f; break
        if not alvo:
            sem_match.append(idx)
            continue
        ano=infer_doc_year(comp)
        prepare_func_doc_dirs(alvo.id,ano)
        writer=PdfWriter(); writer.add_page(page)
        fake_name=f"holerite_{idx}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        subdir,_=func_doc_subdir(alvo.id,'holerite',comp)
        rel=os.path.join(subdir,secure_filename(fake_name))
        abs_p=os.path.join(UPLOAD_ROOT,rel)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out: writer.write(out)
        a=FuncionarioArquivo(funcionario_id=alvo.id,categoria='holerite',competencia=comp,nome_arquivo=fake_name,caminho=rel)
        db.session.add(a); enviados+=1
    db.session.commit()
    return jsonify({'ok':True,'arquivos_gerados':enviados,'paginas_sem_funcionario':sem_match})

@app.route('/api/ordens-compra',methods=['GET'])
@lr
def api_ordens_compra():
    return jsonify([o.to_dict() for o in OrdemCompra.query.order_by(OrdemCompra.criado_em.desc()).all()])

@app.route('/api/ordens-compra',methods=['POST'])
@lr
def api_criar_ordem_compra():
    d=request.json or {}
    num=d.get('numero') or f"OC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    o=OrdemCompra(numero=num,empresa_id=d.get('empresa_id'),solicitante=d.get('solicitante',''),fornecedor=d.get('fornecedor',''),descricao=d.get('descricao',''),valor=to_num(d.get('valor'),dec=True),status=d.get('status','Aberta'),data_emissao=d.get('data_emissao',''),criado_por=session.get('nome',''))
    db.session.add(o); db.session.commit(); return jsonify(o.to_dict()),201

@app.route('/api/ordens-compra/<int:id>',methods=['PUT'])
@lr
def api_atualizar_ordem_compra(id):
    o=OrdemCompra.query.get_or_404(id); d=request.json or {}
    for k in ['numero','empresa_id','solicitante','fornecedor','descricao','status','data_emissao']:
        if k in d: setattr(o,k,d[k])
    if 'valor' in d: o.valor=to_num(d.get('valor'),dec=True)
    db.session.commit(); return jsonify(o.to_dict())

@app.route('/api/ordens-compra/<int:id>',methods=['DELETE'])
@lr
def api_deletar_ordem_compra(id):
    db.session.delete(OrdemCompra.query.get_or_404(id)); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/operacional/documentos',methods=['GET'])
@lr
def api_oper_docs():
    return jsonify([d.to_dict() for d in OperacionalDocumento.query.order_by(OperacionalDocumento.criado_em.desc()).all()])

@app.route('/api/operacional/documentos',methods=['POST'])
@lr
def api_criar_oper_doc():
    titulo=(request.form.get('titulo') or '').strip()
    if not titulo: return jsonify({'erro':'Titulo obrigatorio'}),400
    fs=request.files.get('arquivo')
    rel=None
    if fs and fs.filename:
        rel,_=save_upload(fs,'operacional')
    d=OperacionalDocumento(empresa_id=to_num(request.form.get('empresa_id')) or None,tipo=(request.form.get('tipo') or 'Documento').strip(),titulo=titulo,descricao=(request.form.get('descricao') or '').strip(),nome_arquivo=(fs.filename if fs else None),caminho=rel,criado_por=session.get('nome',''))
    db.session.add(d); db.session.commit(); return jsonify(d.to_dict()),201

@app.route('/api/operacional/documentos/<int:id>',methods=['DELETE'])
@lr
def api_del_oper_doc(id):
    d=OperacionalDocumento.query.get_or_404(id)
    if d.caminho:
        try: os.remove(os.path.join(UPLOAD_ROOT,d.caminho))
        except: pass
    db.session.delete(d); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/operacional/documentos/<int:id>/download')
@lr
def api_download_oper_doc(id):
    d=OperacionalDocumento.query.get_or_404(id)
    if not d.caminho: return jsonify({'erro':'Documento sem arquivo'}),404
    abs_p=os.path.join(UPLOAD_ROOT,d.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    return send_file(abs_p,as_attachment=True,download_name=d.nome_arquivo or 'documento.bin')

@app.route('/api/dashboard')
@lr
def api_dashboard():
    ativos=Cliente.query.filter_by(status='Ativo').all()
    receita=sum((c.limpeza or 0)+(c.jardinagem or 0)+(c.portaria or 0) for c in ativos)
    mes=datetime.now().strftime('%Y-%m')
    emitidos={m.cliente_id for m in Medicao.query.filter_by(mes_ref=mes).all() if m.cliente_id}
    return jsonify({'ativos':len(ativos),'receita':receita,'total_med':Medicao.query.count(),
        'med_mes':Medicao.query.filter_by(mes_ref=mes).count(),
        'pendentes':len([c for c in ativos if c.id not in emitidos]),
        'total_cli':Cliente.query.count(),'proximo_num':prox_num(),
        'empresas':[{'id':e.id,'nome':e.nome,'cli':Cliente.query.filter_by(empresa_id=e.id,status='Ativo').count()} for e in Empresa.query.filter_by(ativa=True).all()]})

@app.route('/api/backup')
@dr
def api_backup():
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as z:
        db_p=os.path.join(os.path.dirname(__file__),'instance','rmfacilities.db')
        if os.path.exists(db_p): z.write(db_p,'rmfacilities.db')
        z.writestr('clientes.json',json.dumps([c.to_dict() for c in Cliente.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('medicoes.json',json.dumps([m.to_dict() for m in Medicao.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('empresas.json',json.dumps([e.to_dict() for e in Empresa.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('funcionarios.json',json.dumps([f.to_dict() for f in Funcionario.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('config.json',json.dumps([{'chave':c.chave,'valor':c.valor} for c in Config.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('whatsapp_conversas.json',json.dumps([c.to_dict() for c in WhatsAppConversa.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('whatsapp_mensagens.json',json.dumps([m.to_dict() for m in WhatsAppMensagem.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('funcionario_arquivos.json',json.dumps([a.to_dict() for a in FuncionarioArquivo.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('ordens_compra.json',json.dumps([o.to_dict() for o in OrdemCompra.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('operacional_documentos.json',json.dumps([d.to_dict() for d in OperacionalDocumento.query.all()],default=str,ensure_ascii=False,indent=2))
        if os.path.isdir(UPLOAD_ROOT):
            for root,_,files in os.walk(UPLOAD_ROOT):
                for fn in files:
                    ap=os.path.join(root,fn)
                    rel=os.path.relpath(ap,UPLOAD_ROOT)
                    z.write(ap,os.path.join('uploads',rel))
        z.writestr('info.json',json.dumps({'data':datetime.now().isoformat(),'versao':'3.0'},ensure_ascii=False))
    buf.seek(0)
    return send_file(buf,mimetype='application/zip',as_attachment=True,download_name=f'backup_rm_{datetime.now().strftime("%Y%m%d_%H%M")}.zip')

@app.route('/api/backup/restore',methods=['POST'])
@dr
def api_backup_restore():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo ZIP nao enviado'}),400
    try:
        z=zipfile.ZipFile(arq)
    except Exception:
        return jsonify({'erro':'Arquivo invalido. Envie um ZIP de backup'}),400

    try:
        def jread(name,default):
            try: return json.loads(z.read(name).decode('utf-8'))
            except Exception: return default

        # Limpa dados operacionais antes de restaurar.
        for model in [WhatsAppMensagem,WhatsAppConversa,FuncionarioArquivo,OperacionalDocumento,OrdemCompra,Medicao,Cliente,Funcionario,Empresa,Config]:
            model.query.delete()

        empresas=jread('empresas.json',[])
        clientes=jread('clientes.json',[])
        medicoes=jread('medicoes.json',[])
        funcs=jread('funcionarios.json',[])
        configs=jread('config.json',[])
        wa_convs=jread('whatsapp_conversas.json',[])
        wa_msgs=jread('whatsapp_mensagens.json',[])
        farqs=jread('funcionario_arquivos.json',[])
        ocs=jread('ordens_compra.json',[])
        opdocs=jread('operacional_documentos.json',[])

        def _coerce_value(col,val):
            if val is None:
                return None
            type_name=(col.type.__class__.__name__ or '').lower()
            try:
                py_t=col.type.python_type
            except Exception:
                py_t=None

            if py_t is datetime or 'datetime' in type_name:
                if isinstance(val,datetime):
                    return val
                if isinstance(val,str):
                    s=val.strip()
                    if not s:
                        return None
                    dt=_parse_dt_iso(s)
                    if not dt and ' ' in s and 'T' not in s:
                        dt=_parse_dt_iso(s.replace(' ','T'))
                    if dt:
                        return dt
                    for fmt in ('%Y-%m-%d %H:%M:%S.%f','%Y-%m-%d %H:%M:%S','%Y-%m-%dT%H:%M:%S.%f','%Y-%m-%dT%H:%M:%S'):
                        try:
                            return datetime.strptime(s,fmt)
                        except Exception:
                            pass
                return None

            if py_t is date or (type_name=='date'):
                if isinstance(val,date):
                    return val
                if isinstance(val,str):
                    s=val.strip()
                    if not s:
                        return None
                    for fmt in ('%Y-%m-%d','%d/%m/%Y'):
                        try:
                            return datetime.strptime(s,fmt).date()
                        except Exception:
                            pass
                return None

            if py_t is bool or 'boolean' in type_name:
                if isinstance(val,bool):
                    return val
                if isinstance(val,str):
                    return val.strip().lower() in ('1','true','yes','on','sim')
                return bool(val)

            if py_t is int or 'integer' in type_name:
                if val=='':
                    return None
                try:
                    return int(val)
                except Exception:
                    return None

            if py_t is float or 'float' in type_name:
                if val=='':
                    return None
                try:
                    return float(val)
                except Exception:
                    return None

            return val

        def add_rows(model,rows,conv=None):
            cols_map={c.name:c for c in model.__table__.columns}
            for r in rows:
                d={}
                for k,v in r.items():
                    if k in cols_map:
                        d[k]=_coerce_value(cols_map[k],v)
                if conv: d=conv(d)
                db.session.add(model(**d))

        add_rows(Empresa,empresas)
        add_rows(Cliente,clientes)
        add_rows(Medicao,medicoes,lambda d: ({**d,'servicos':json.dumps(d.get('svcs',[]),ensure_ascii=False)} if 'svcs' in d and 'servicos' not in d else d))
        add_rows(Funcionario,funcs,lambda d: ({**d,'areas':json.dumps(d.get('areas',[]),ensure_ascii=False)} if isinstance(d.get('areas'),list) else d))
        add_rows(Config,configs)
        add_rows(WhatsAppConversa,wa_convs)
        add_rows(WhatsAppMensagem,wa_msgs)
        add_rows(FuncionarioArquivo,farqs)
        add_rows(OrdemCompra,ocs)
        add_rows(OperacionalDocumento,opdocs)
        db.session.commit()

        # Restaura uploads.
        if os.path.isdir(UPLOAD_ROOT): shutil.rmtree(UPLOAD_ROOT)
        os.makedirs(UPLOAD_ROOT,exist_ok=True)
        for n in z.namelist():
            if not n.startswith('uploads/') or n.endswith('/'): continue
            rel=n[len('uploads/'):]
            ap=os.path.join(UPLOAD_ROOT,rel)
            os.makedirs(os.path.dirname(ap),exist_ok=True)
            with open(ap,'wb') as out: out.write(z.read(n))

        return jsonify({'ok':True,'restaurado':{'empresas':len(empresas),'clientes':len(clientes),'medicoes':len(medicoes),'funcionarios':len(funcs),'configs':len(configs),'wa_conversas':len(wa_convs),'wa_mensagens':len(wa_msgs)}})
    except Exception as e:
        app.logger.exception('Falha ao restaurar backup completo')
        db.session.rollback()
        return jsonify({'erro':f'Falha ao restaurar backup: {str(e)}'}),500
    finally:
        try:
            z.close()
        except Exception:
            pass

@app.route('/api/pdf/<int:id>')
@lr
def api_pdf(id):
    m=Medicao.query.get_or_404(id)
    emp=Empresa.query.get(m.empresa_id) if m.empresa_id else None
    d=m.to_dict(); d['empresa']=emp.to_dict() if emp else {}
    return _build_pdf(d)

@app.route('/api/pdf/preview',methods=['POST'])
@lr
def api_pdf_preview():
    d=request.json
    if d.get('empresa_id'):
        emp=Empresa.query.get(d['empresa_id'])
        d['empresa']=emp.to_dict() if emp else {}
    else: d['empresa']={}
    return _build_pdf(d)

def _build_pdf(d):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,HRFlowable,Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT,TA_RIGHT,TA_CENTER

    emp=d.get('empresa',{}); nmed=d.get('numero','001'); tipo=d.get('tipo','Medição de Serviços')
    mes=d.get('mes_ref',''); dtem=d.get('dt_emissao',''); dtvenc=d.get('dt_vencimento','')
    enome=emp.get('razao') or emp.get('nome') or d.get('empresa_nome','RM Facilities')
    eend=emp.get('end_fmt','') or ''; esite=norm_url(emp.get('site','')); epix=emp.get('pix','')
    ebanco=emp.get('banco',''); eagencia=emp.get('agencia',''); econta=emp.get('conta','')
    eboleto=emp.get('boleto','')
    cname=d.get('cliente_nome',''); ccnpj=d.get('cliente_cnpj','')
    cend=d.get('cliente_end',''); cresp=d.get('cliente_resp','')
    obs=d.get('observacoes',''); a1=d.get('ass_empresa',''); a2=d.get('ass_cliente','')
    svcs=d.get('svcs',d.get('servicos',[]));
    if isinstance(svcs,str):
        try: svcs=json.loads(svcs)
        except: svcs=[]
    sub=sum(float(s.get('vtot',0)) for s in svcs)
    por=d.get('criado_por',session.get('nome','')); now=datetime.now()

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=2*cm)
    W=A4[0]-3*cm

    AZ=colors.HexColor('#205d8a'); LJ=colors.HexColor('#f28e34')
    VD=colors.HexColor('#1a7a45'); VDC=colors.HexColor('#d8f0e5'); CI=colors.HexColor('#f5f5f5')

    def ps(nm,**kw):
        b=dict(fontName='Helvetica',fontSize=10,leading=14,textColor=colors.HexColor('#020202'),spaceAfter=0,spaceBefore=0)
        b.update(kw); return ParagraphStyle(nm,**b)

    story=[]

    # Cabeçalho
    lp=get_logo()
    lc=Paragraph(f'<b>{enome}</b>',ps('lg',fontSize=13,textColor=AZ))
    for cand in [emp.get('logo_url'),lp,LOGO_URL]:
        if not cand: continue
        try:
            if isinstance(cand,str) and cand.startswith(('http://','https://')):
                req=urllib.request.Request(cand,headers={'User-Agent':'Mozilla/5.0'})
                with urllib.request.urlopen(req,timeout=8) as r: img_data=r.read()
                lc=Image(io.BytesIO(img_data),width=4.5*cm,height=1.8*cm,kind='proportional')
            elif os.path.exists(cand):
                lc=Image(cand,width=4.5*cm,height=1.8*cm,kind='proportional')
            break
        except Exception:
            continue

    hdr=Table([[lc,Paragraph(f'<font size="13" color="white"><b>{tipo.upper()}</b></font><br/><font size="9" color="#AADDFF">Nº {nmed}</font>',ps('hr',alignment=TA_CENTER))]],colWidths=[W*0.55,W*0.45])
    hdr.setStyle(TableStyle([('BACKGROUND',(1,0),(1,0),AZ),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10),('LEFTPADDING',(1,0),(1,-1),12),('LEFTPADDING',(0,0),(0,-1),0)]))
    story.append(hdr)

    bar=Table([[' ']],colWidths=[W])
    bar.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(bar); story.append(Spacer(1,6))

    def campo(lbl,val):
        return [Paragraph(f'<b><font color="#205d8a">{lbl}</font></b>',ps('lb',fontSize=8)),Paragraph(str(val or '—'),ps('vl',fontSize=9))]

    flat=[]
    for row in [[campo('Cliente:',cname),campo('Mês ref.:',fmt_mes(mes))],[campo('CNPJ/CPF:',ccnpj),campo('Emissão:',fmt_data(dtem))],[campo('Endereço:',cend),campo('Vencimento:',fmt_data(dtvenc))],[campo('Responsável:',cresp),campo('Empresa prestadora:',enome)]]:
        flat.append([row[0][0],row[0][1],row[1][0],row[1][1]])

    inf=Table(flat,colWidths=[W*0.17,W*0.33,W*0.17,W*0.33])
    inf.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('BACKGROUND',(2,0),(2,-1),CI),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
    story.append(inf); story.append(Spacer(1,8))

    story.append(Paragraph('<b><font color="white">  SERVIÇOS REALIZADOS NO PERÍODO</font></b>',ps('sh',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
    story.append(Spacer(1,4))

    th=[Paragraph(f'<b>{h}</b>',ps('th',textColor=colors.white,fontSize=9,alignment=a)) for h,a in [('Descrição',TA_LEFT),('Unid.',TA_CENTER),('Qtd.',TA_CENTER),('Valor unit.',TA_RIGHT),('Valor total',TA_RIGHT)]]
    sr=[th]
    for i,s in enumerate(svcs):
        desc=(s.get('desc','') or '').strip()
        tags=[]
        if bool(s.get('fornece_materiais')):
            tags.append('com materiais')
        if bool(s.get('fornece_equipamentos')):
            tags.append('com equipamentos')
        if tags:
            desc=f"{desc} ({', '.join(tags)})"
        sr.append([Paragraph(desc,ps('td',fontSize=9)),Paragraph(str(s.get('unid','')),ps('tc',fontSize=9,alignment=TA_CENTER)),Paragraph(str(s.get('qtd',1)),ps('tc2',fontSize=9,alignment=TA_CENTER)),Paragraph(fmt_brl(s.get('vun',0)),ps('tr',fontSize=9,alignment=TA_RIGHT)),Paragraph(fmt_brl(s.get('vtot',0)),ps('tv',fontSize=9,alignment=TA_RIGHT,textColor=VD))])

    st2=TableStyle([('BACKGROUND',(0,0),(-1,0),AZ),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),('LINEBELOW',(0,0),(-1,-1),0.3,colors.HexColor('#DDDDDD')),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))])
    for i in range(1,len(sr)):
        if i%2==0: st2.add('BACKGROUND',(0,i),(-1,i),CI)

    svc_tbl=Table(sr,colWidths=[W*0.46,W*0.09,W*0.09,W*0.18,W*0.18])
    svc_tbl.setStyle(st2); story.append(svc_tbl); story.append(Spacer(1,6))

    tot=Table([[Paragraph('<b>VALOR TOTAL A RECEBER:</b>',ps('tl',fontSize=13,textColor=VD)),Paragraph(f'<b>{fmt_brl(sub)}</b>',ps('tv2',fontSize=15,alignment=TA_RIGHT,textColor=VD))]],colWidths=[W*0.68,W*0.32])
    tot.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),VDC),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10),('LEFTPADDING',(0,0),(0,-1),12),('RIGHTPADDING',(-1,0),(-1,-1),12),('BOX',(0,0),(-1,-1),1.5,VD)]))
    story.append(tot); story.append(Spacer(1,10))

    pags=[]
    if epix: pags.append(('Chave Pix:',epix))
    if ebanco: pags.append(('Banco:',f'{ebanco}{" — Ag: "+eagencia if eagencia else ""}{" / Cc: "+econta if econta else ""}'))
    if pags:
        story.append(Paragraph('<b><font color="white">  DADOS PARA PAGAMENTO</font></b>',ps('ph',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        pr=[[Paragraph(f'<b>{l}</b>',ps('pl',fontSize=9,textColor=AZ)),Paragraph(v,ps('pv',fontSize=10))] for l,v in pags]
        pt=Table(pr,colWidths=[W*0.22,W*0.78])
        pt.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD'))]))
        story.append(pt); story.append(Spacer(1,10))

    if eboleto:
        story.append(Paragraph('<b><font color="white">  INFORMAÇÕES DE BOLETO</font></b>',ps('bh',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        bt=Table([[Paragraph(eboleto,ps('bv',fontSize=9,leading=14))]],colWidths=[W])
        bt.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
        story.append(bt); story.append(Spacer(1,10))

    if obs:
        story.append(Paragraph('<b><font color="white">  OBSERVAÇÕES</font></b>',ps('oh',fontSize=10,backColor=LJ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        ot=Table([[Paragraph(obs,ps('ov',fontSize=9,leading=14))]],colWidths=[W])
        ot.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
        story.append(ot); story.append(Spacer(1,16))

    # Assinaturas
    story.append(Spacer(1,24))
    at=Table([[Paragraph(f'_________________________________<br/><font size="8" color="#888888">{a1 or "Empresa prestadora"}</font>',ps('a1',alignment=TA_CENTER,leading=18)),Paragraph('',ps('ae')),Paragraph(f'_________________________________<br/><font size="8" color="#888888">{a2 or "Cliente / Tomador"}</font>',ps('a2',alignment=TA_CENTER,leading=18))]],colWidths=[W*0.42,W*0.16,W*0.42])
    at.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'BOTTOM')]))
    story.append(at); story.append(Spacer(1,14))

    # Rodapé
    bar2=Table([[' ']],colWidths=[W])
    bar2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(bar2); story.append(Spacer(1,4))
    if eend or esite:
        rodape_end=f'<b>{enome}</b>'
        if eend: rodape_end+=f' — {eend}'
        if esite: rodape_end+=f' | {esite}'
        story.append(Paragraph(rodape_end,ps('re',fontSize=8,textColor=colors.HexColor('#555'),alignment=TA_CENTER)))
        story.append(Spacer(1,3))
    story.append(Paragraph(f'Emitido em {now.strftime("%d/%m/%Y")} às {now.strftime("%H:%M")} por {por} — {enome} · {tipo} Nº {nmed} · {fmt_mes(mes)}',ps('rod',fontSize=7,textColor=colors.HexColor('#999'),alignment=TA_CENTER)))

    doc.build(story); buf.seek(0)
    slug=(cname or 'cliente').replace(' ','_')[:20]
    return send_file(buf,mimetype='application/pdf',as_attachment=True,download_name=f'{tipo.replace(" ","_")}_{nmed.replace("/","-")}_{slug}_{mes}.pdf')

@app.route('/api/rh/holerites/processar',methods=['POST'])
@lr
def api_rh_holerites_processar():
    fs=request.files.get('arquivo')
    comp=(request.form.get('competencia') or '').strip()
    if not fs: return jsonify({'erro':'PDF nao enviado'}),400
    nome_arq=(fs.filename or '').strip().lower()
    if nome_arq and not nome_arq.endswith('.pdf'):
        return jsonify({'erro':'Arquivo invalido. Envie um PDF (.pdf).'}),400
    try:
        from pypdf import PdfReader,PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500
    funcs_todos=Funcionario.query.all()
    if not funcs_todos: return jsonify({'erro':'Cadastre funcionarios primeiro'}),400
    pdf_bytes=fs.read()
    if not pdf_bytes or len(pdf_bytes)<8:
        return jsonify({'erro':'Arquivo vazio ou corrompido.'}),400
    if pdf_bytes[:4]!=b'%PDF':
        return jsonify({'erro':'Arquivo invalido. O conteúdo enviado nao parece ser um PDF.'}),400
    try:
        reader=PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return jsonify({'erro':'PDF invalido ou corrompido. Gere/exporte o arquivo novamente e tente de novo.'}),400
    itens=[]; sem_match=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=(page.extract_text() or '').lower()
        alvo=None
        for f in funcs_todos:
            nm=(f.nome or '').lower().strip()
            if nm and nm in txt: alvo=f; break
        if not alvo: sem_match.append(idx); continue
        ano=infer_doc_year(comp); prepare_func_doc_dirs(alvo.id,ano)
        writer=PdfWriter(); writer.add_page(page)
        comp_safe=(comp.replace('/','_') if comp else str(idx))
        ts=datetime.now().strftime('%Y%m%d_%H%M%S%f')[:17]
        fake_name=secure_filename(f"holerite_{comp_safe}_{alvo.nome.replace(' ','_')[:15]}_{ts}.pdf")
        subdir,_=func_doc_subdir(alvo.id,'holerite',comp)
        rel=os.path.join(subdir,fake_name); abs_p=os.path.join(UPLOAD_ROOT,rel)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out: writer.write(out)
        a=FuncionarioArquivo(funcionario_id=alvo.id,categoria='holerite',competencia=comp,nome_arquivo=fake_name,caminho=rel)
        db.session.add(a); db.session.flush()
        whatsapp_num=wa_norm_number(alvo.telefone or '')
        wa_habilitado=funcionario_docs_whatsapp_habilitado(alvo)
        itens.append({'pagina':idx,'funcionario_id':alvo.id,'funcionario_nome':alvo.nome,'arquivo_id':a.id,'nome_arquivo':fake_name,'caminho':rel,'abs_caminho':abs_p,'email':alvo.email or '','whatsapp':(whatsapp_num if wa_is_valid_number(whatsapp_num) else ''),'whatsapp_habilitado':wa_habilitado,'status_envio':None,'erro_envio':None})
    db.session.commit()
    job_id=secrets.token_hex(16)
    _holerite_jobs[job_id]={'id':job_id,'status':'pronto','total_paginas':len(reader.pages),'itens':itens,'sem_match':sem_match,'competencia':comp,'criado_em':utcnow().isoformat()}
    itens_resp=[{k:v for k,v in it.items() if k!='abs_caminho'} for it in itens]
    return jsonify({'ok':True,'job_id':job_id,'total_paginas':len(reader.pages),'separados':len(itens),'sem_match':sem_match,'itens':itens_resp})

@app.route('/api/rh/holerites/job/<job_id>')
@lr
def api_rh_holerites_job(job_id):
    job=_holerite_jobs.get(job_id)
    if not job: return jsonify({'erro':'Job nao encontrado'}),404
    out=dict(job); out['itens']=[{k:v for k,v in it.items() if k!='abs_caminho'} for it in out.get('itens',[])]
    return jsonify({'ok':True,'job':out})

@app.route('/api/rh/holerites/enviar/<job_id>',methods=['POST'])
@lr
def api_rh_holerites_enviar(job_id):
    job=_holerite_jobs.get(job_id)
    if not job: return jsonify({'erro':'Job nao encontrado. Processe o PDF novamente.'}),404
    d=request.json or {}; canal=d.get('canal','email')
    ids_sel=set(int(x) for x in (d.get('arquivo_ids') or []))
    def do_send():
        with app.app_context():
            for item in job['itens']:
                if ids_sel and item.get('arquivo_id') not in ids_sel: continue
                try:
                    abs_p=item.get('abs_caminho') or os.path.join(UPLOAD_ROOT,item['caminho'])
                    comp=job.get('competencia',''); fn=item.get('funcionario_nome','')
                    s_e=False; s_w=False
                    if canal in ('email','ambos') and item.get('email'):
                        smtp_send_pdf(item['email'],fn,abs_p,item['nome_arquivo'],comp); s_e=True
                    if canal in ('whatsapp','ambos') and item.get('whatsapp_habilitado') and item.get('whatsapp'):
                        wa_send_pdf(item['whatsapp'],abs_p,item['nome_arquivo'],f"Holerite {comp} - {fn}"); s_w=True
                    if s_e and s_w: item['status_envio']='enviado_ambos'
                    elif s_e: item['status_envio']='enviado_email'
                    elif s_w: item['status_envio']='enviado_wa'
                    else: item['status_envio']='sem_contato'
                    item['erro_envio']=None
                except Exception as e:
                    item['erro_envio']=str(e); item['status_envio']='erro'
            job['status']='concluido'
    job['status']='enviando'
    threading.Thread(target=do_send,daemon=True).start()
    return jsonify({'ok':True,'job_id':job_id,'status':'enviando'})

@app.route('/api/config/smtp',methods=['GET'])
@dr
def api_smtp_get():
    return jsonify({'host':gc('smtp_host',''),'port':gc('smtp_port','587'),'user':gc('smtp_user',''),'senha':gc('smtp_senha',''),'de':gc('smtp_de',''),'tls':gc('smtp_tls','1')})

@app.route('/api/config/smtp',methods=['POST'])
@dr
def api_smtp_save():
    d=request.json or {}
    for k in ['host','port','user','senha','de','tls']:
        if k in d: sc_cfg(f'smtp_{k}',str(d[k]))
    return jsonify({'ok':True})

@app.route('/api/config/smtp/testar',methods=['POST'])
@dr
def api_smtp_testar():
    d=request.json or {}; dest=(d.get('email') or gc('smtp_user','')).strip()
    if not dest: return jsonify({'erro':'Informe o e-mail de destino'}),400
    try:
        cfg=smtp_cfg()
        if not cfg['host'] or not cfg['user']: return jsonify({'erro':'SMTP nao configurado'}),400
        msg=MIMEMultipart(); msg['From']=cfg['de'] or cfg['user']; msg['To']=dest
        msg['Subject']='Teste de envio — RM Facilities'
        msg.attach(MIMEText('E-mail de teste enviado com sucesso pelo sistema RM Facilities!','plain','utf-8'))
        port=int(cfg['port'] or 587)
        if str(cfg['tls']) in ('1','true','True'):
            with smtplib.SMTP(cfg['host'],port,timeout=20) as s: s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
        else:
            with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s: s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
        return jsonify({'ok':True,'mensagem':f'E-mail enviado para {dest}'})
    except Exception as e: return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp',methods=['GET'])
@dr
def api_wa_cfg_get():
    b=wa_backup_cfg()
    return jsonify({
        'url':gc('wa_url',''),
        'instancia':gc('wa_instancia',''),
        'token':gc('wa_token',''),
        'backup_enabled':b.get('enabled','1'),
        'backup_email':b.get('email',''),
        'backup_interval_hours':b.get('interval_hours','2'),
        'backup_window_hours':b.get('window_hours','8'),
        'backup_max_conversas':b.get('max_conversas','10'),
    })

@app.route('/api/config/whatsapp',methods=['POST'])
@dr
def api_wa_cfg_save():
    d=request.json or {}
    for k in ['url','instancia','token']:
        if k in d: sc_cfg(f'wa_{k}',str(d[k]))
    if 'backup_enabled' in d:
        sc_cfg('wa_backup_enabled','1' if str(d.get('backup_enabled','0')).strip().lower() in ('1','true','yes','on') else '0')
    if 'backup_email' in d:
        sc_cfg('wa_backup_email',str(d.get('backup_email','')).strip())
    if 'backup_interval_hours' in d:
        sc_cfg('wa_backup_interval_hours',str(max(1,min(168,_to_int(d.get('backup_interval_hours'),2)))))
    if 'backup_window_hours' in d:
        sc_cfg('wa_backup_window_hours',str(max(1,min(168,_to_int(d.get('backup_window_hours'),8)))))
    if 'backup_max_conversas' in d:
        sc_cfg('wa_backup_max_conversas',str(max(1,min(50,_to_int(d.get('backup_max_conversas'),10)))))
    return jsonify({'ok':True})

@app.route('/api/config/whatsapp/backup/testar',methods=['POST'])
@dr
def api_wa_backup_testar():
    try:
        r=wa_backup_maybe_send(force=True)
        if not r.get('ok'):
            return jsonify({'erro':f"Backup nao enviado: {r.get('skip','falha')}"}),400
        return jsonify({'ok':True,'mensagem':f"Backup enviado para {r.get('email','')}",'arquivo':r.get('arquivo',''),'total_conversas':r.get('total_conversas',0),'total_mensagens':r.get('total_mensagens',0)})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/backup/gerar',methods=['POST'])
@dr
def api_wa_backup_gerar():
    try:
        cfg=wa_backup_cfg()
        payload=_wa_backup_collect(cfg.get('window_hours','8'),cfg.get('max_conversas','10'))
        arq=_wa_backup_store(payload)
        return jsonify({'ok':True,'arquivo':arq,'payload':payload})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/backup/restaurar',methods=['POST'])
@dr
def api_wa_backup_restaurar():
    arq=request.files.get('arquivo')
    if not arq:
        return jsonify({'erro':'Arquivo JSON do backup nao enviado'}),400
    try:
        payload=json.loads(arq.read().decode('utf-8'))
    except Exception:
        return jsonify({'erro':'JSON invalido'}),400
    restore_config=str(request.form.get('restore_config','1')).strip().lower() in ('1','true','yes','on')
    restore_conversas=str(request.form.get('restore_conversas','1')).strip().lower() in ('1','true','yes','on')
    try:
        st=wa_backup_restore_payload(payload,restore_config=restore_config,restore_conversas=restore_conversas)
        return jsonify({'ok':True,'restaurado':st})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/testar',methods=['POST'])
@dr
def api_wa_testar():
    d=request.json or {}; numero=(d.get('numero') or '').strip()
    if not numero: return jsonify({'erro':'Informe o numero de destino'}),400
    try:
        wa_send_text(numero,'Teste de conexao WhatsApp — RM Facilities')
        return jsonify({'ok':True})
    except Exception as e: return jsonify({'erro':str(e)}),500

@app.route('/api/config/ia-whatsapp',methods=['GET'])
@dr
def api_ia_wa_cfg_get():
    d=ai_wa_cfg()
    return jsonify({'enabled':d['enabled'],'provider':d['provider'],'api_key':d['api_key'],'model':d['model'],'prompt':d['prompt'],'temperature':d['temperature'],'max_tokens':d['max_tokens']})

@app.route('/api/config/ia-whatsapp',methods=['POST'])
@dr
def api_ia_wa_cfg_save():
    d=request.json or {}
    enabled='1' if str(d.get('enabled','0')).strip().lower() in ('1','true','yes','on') else '0'
    provider=ai_provider_norm(d.get('provider','gemini'))
    raw_model=str(d.get('model','') or '').strip()
    model=ai_model_norm(provider,raw_model) if raw_model else (gc('ia_wa_model','') or ai_model_norm(provider,''))
    try:
        temp=max(0.0,min(1.0,float(d.get('temperature',0.3))))
    except Exception:
        temp=0.3
    try:
        max_tokens=max(50,min(2000,int(float(d.get('max_tokens',350)))))
    except Exception:
        max_tokens=350
    sc_cfg('ia_wa_enabled',enabled)
    sc_cfg('ia_wa_provider',provider)
    sc_cfg('ia_wa_api_key',str(d.get('api_key','')))
    sc_cfg('ia_wa_model',model)
    sc_cfg('ia_wa_prompt',str(d.get('prompt','')))
    sc_cfg('ia_wa_temperature',str(temp))
    sc_cfg('ia_wa_max_tokens',str(max_tokens))
    warn=''
    if raw_model and raw_model!=model:
        warn=f'Modelo ajustado automaticamente para: {model}'
    return jsonify({'ok':True,'provider':provider,'model':model,'temperature':temp,'max_tokens':max_tokens,'warning':warn})

@app.route('/api/whatsapp/ia/testar',methods=['POST'])
@dr
def api_wa_ia_testar():
    d=request.json or {}
    txt=(d.get('texto') or '').strip()
    numero=wa_norm_number(d.get('numero') or '5511999999999')
    if not txt: return jsonify({'erro':'Informe o texto para teste'}),400
    try:
        resp=ai_wa_reply(numero,txt)
        if not resp: return jsonify({'erro':'A IA nao retornou resposta'}),400
        return jsonify({'ok':True,'resposta':resp})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/whatsapp/ia/retomar',methods=['POST'])
@lr
def api_wa_ia_retomar():
    d=request.json or {}
    numero=wa_norm_number(d.get('numero') or '')
    if not numero:
        return jsonify({'erro':'numero obrigatorio'}),400
    if not wa_is_valid_number(numero):
        return jsonify({'erro':'numero invalido'}),400
    wa_ai_resume(numero)
    return jsonify({'ok':True,'numero':numero,'ia_pausada':False})

@app.route('/api/whatsapp/ia/pausar',methods=['POST'])
@lr
def api_wa_ia_pausar():
    d=request.json or {}
    numero=wa_norm_number(d.get('numero') or '')
    horas=_to_int(d.get('horas'),8)
    if not numero:
        return jsonify({'erro':'numero obrigatorio'}),400
    if not wa_is_valid_number(numero):
        return jsonify({'erro':'numero invalido'}),400
    pausa_ate=wa_ai_pause_set(numero,max(1,min(168,horas)))
    return jsonify({'ok':True,'numero':numero,'ia_pausada':True,'ia_pausada_ate':(pausa_ate.isoformat() if pausa_ate else '')})

@app.route('/api/whatsapp/conversas')
@lr
def api_wa_conversas():
    lst=WhatsAppConversa.query.order_by(WhatsAppConversa.ultima_msg.desc()).all()
    return jsonify([c.to_dict() for c in lst])

@app.route('/api/whatsapp/conversas/<numero>')
@lr
def api_wa_conversa_msgs(numero):
    c=WhatsAppConversa.query.filter_by(numero=numero).first()
    if not c: return jsonify({'conversa':None,'mensagens':[],'ia_pausada':False,'ia_pausada_ate':''})
    msgs=WhatsAppMensagem.query.filter_by(conversa_id=c.id).order_by(WhatsAppMensagem.criado_em).all()
    pausa_ate=wa_ai_pause_until(numero)
    return jsonify({
        'conversa':c.to_dict(),
        'mensagens':[m.to_dict() for m in msgs],
        'ia_pausada':bool(pausa_ate and pausa_ate>utcnow()),
        'ia_pausada_ate':pausa_ate.isoformat() if pausa_ate else '',
    })

@app.route('/api/whatsapp/send',methods=['POST'])
@lr
def api_wa_send():
    d=request.json or {}; numero=(d.get('numero') or '').strip(); texto=(d.get('texto') or '').strip()
    if not numero or not texto: return jsonify({'erro':'numero e texto obrigatorios'}),400
    try: wa_send_text(numero,texto)
    except Exception as e: return jsonify({'erro':str(e)}),500
    c=WhatsAppConversa.query.filter_by(numero=numero).first()
    if not c:
        c=WhatsAppConversa(numero=numero,nome=d.get('nome') or numero)
        db.session.add(c); db.session.flush()
    c.ultima_msg=utcnow()
    db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=texto))
    db.session.commit()
    wa_ai_pause_for(numero, 2)
    return jsonify({'ok':True})

@app.route('/webhook/whatsapp',methods=['GET','POST'])
def webhook_whatsapp():
    if request.method=='GET':
        return jsonify({'ok':True,'endpoint':'/webhook/whatsapp','metodos':['POST'],'status':'ativo'})
    debug=str(request.args.get('debug','0')).strip().lower() in ('1','true','yes','on')
    diag={'evento':None,'mensagens_recebidas':0,'mensagens_processadas':0,'ia_ativa':ai_wa_enabled(),'respostas_enviadas':0,'erros':[]}
    data=request.json or {}
    try:
        evento=(data.get('event') or '').lower()
        diag['evento']=evento
        raw=data.get('data',{})
        is_message_payload=False
        if isinstance(raw,dict):
            is_message_payload=bool(raw.get('message') or raw.get('body') or raw.get('text'))
        elif isinstance(raw,list):
            for _m in raw:
                if isinstance(_m,dict) and (_m.get('message') or _m.get('body') or _m.get('text')):
                    is_message_payload=True
                    break
        if 'message' in evento or 'upsert' in evento or is_message_payload:
            msgs=[raw] if isinstance(raw,dict) else (raw if isinstance(raw,list) else [])
            diag['mensagens_recebidas']=len(msgs)
            for msg_data in msgs:
                if bool(msg_data.get('key',{}).get('fromMe')) or bool(msg_data.get('fromMe')):
                    continue
                jid=(msg_data.get('key',{}).get('remoteJid') or msg_data.get('sender') or msg_data.get('from') or '')
                if not jid.endswith('@s.whatsapp.net'): continue
                numero=(jid.split('@')[0] if jid else '')
                if not numero: continue
                numero=only_digits(numero) or numero
                if not wa_is_valid_number(numero):
                    diag['erros'].append(f'Numero invalido no webhook: {numero}')
                    continue

                msg_obj=msg_data.get('message',{}) if isinstance(msg_data.get('message',{}),dict) else {}
                audio_obj=msg_obj.get('audioMessage',{}) if isinstance(msg_obj.get('audioMessage',{}),dict) else {}
                tipo_in='audio' if audio_obj else 'texto'
                conteudo=(
                    msg_obj.get('conversation') or
                    msg_obj.get('extendedTextMessage',{}).get('text') or
                    msg_data.get('body') or
                    msg_data.get('text') or
                    msg_data.get('transcription') or
                    msg_data.get('transcript') or
                    audio_obj.get('transcription') or
                    audio_obj.get('transcript') or
                    audio_obj.get('text') or
                    ''
                )
                conteudo=str(conteudo or '').strip()

                if tipo_in=='audio' and not conteudo:
                    try:
                        conteudo=wa_transcribe_audio(msg_data)
                    except Exception as e:
                        diag['erros'].append(f'Falha transcricao audio: {str(e)}')
                        conteudo=''

                if tipo_in=='audio' and not conteudo:
                    # Audio chegou sem transcricao no payload: salva no historico e pede texto.
                    c=WhatsAppConversa.query.filter_by(numero=numero).first()
                    if not c:
                        nome=(msg_data.get('pushName') or msg_data.get('notifyName') or numero)
                        c=WhatsAppConversa(numero=numero,nome=nome)
                        db.session.add(c)
                        db.session.flush()
                    c.ultima_msg=utcnow()
                    db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='in',tipo='audio',conteudo='[audio sem transcricao]'))
                    db.session.commit()

                    cfg_wa=wa_cfg()
                    wa_ready=bool((cfg_wa.get('url') or '').strip() and (cfg_wa.get('instancia') or '').strip())
                    if ai_wa_enabled() and wa_ready and not wa_ai_pause_active(numero):
                        try:
                            aviso='Recebi seu audio, mas ainda nao consegui transcrever automaticamente. Pode me enviar em texto?'
                            wa_send_text(numero,aviso)
                            diag['respostas_enviadas']+=1
                            c.ultima_msg=utcnow()
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=aviso))
                            db.session.commit()
                        except Exception as e:
                            diag['erros'].append(str(e))
                            db.session.rollback()
                    continue

                if not conteudo:
                    continue
                diag['mensagens_processadas']+=1
                c=WhatsAppConversa.query.filter_by(numero=numero).first()
                if not c:
                    nome=(msg_data.get('pushName') or msg_data.get('notifyName') or numero)
                    c=WhatsAppConversa(numero=numero,nome=nome)
                    db.session.add(c); db.session.flush()
                c.ultima_msg=utcnow()
                db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='in',tipo=tipo_in,conteudo=conteudo))
                db.session.commit()
                # Usa as mensagens mais recentes para preservar o estado atual da conversa.
                historico_db=(
                    WhatsAppMensagem.query
                    .filter_by(conversa_id=c.id)
                    .order_by(WhatsAppMensagem.criado_em.desc())
                    .limit(20)
                    .all()
                )
                historico_db.reverse()
                cfg_wa=wa_cfg()
                wa_ready=bool((cfg_wa.get('url') or '').strip() and (cfg_wa.get('instancia') or '').strip())
                if ai_wa_enabled() and not wa_ready:
                    msg_cfg='Auto-reply desativado: WhatsApp nao configurado (url/instancia).'
                    if msg_cfg not in diag['erros']:
                        diag['erros'].append(msg_cfg)
                    app.logger.warning('Auto-reply WhatsApp inativo para %s: configuracao incompleta',numero)
                if ai_wa_enabled() and wa_ready and not wa_ai_pause_active(numero):
                    try:
                        resposta=ai_wa_reply(numero,conteudo,historico=historico_db)
                        if resposta:
                            wa_send_text(numero,resposta)
                            diag['respostas_enviadas']+=1
                            c.ultima_msg=utcnow()
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=resposta))
                            db.session.commit()
                        else:
                            diag['erros'].append('IA nao retornou resposta')
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='erro',conteudo='IA nao retornou resposta.'))
                            db.session.commit()
                    except Exception as e:
                        diag['erros'].append(str(e))
                        db.session.rollback()
                        try:
                            app.logger.exception('Falha no auto-reply WhatsApp para %s',numero)
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='erro',conteudo=('Falha auto-reply: '+str(e))[:700]))
                            db.session.commit()
                        except Exception:
                            db.session.rollback()
        elif debug:
            diag['erros'].append(f'Evento ignorado: {evento or "sem_evento"}')
    except Exception:
        app.logger.exception('Falha no processamento do webhook WhatsApp')
        diag['erros'].append('Falha no processamento do webhook')
        db.session.rollback()
    try:
        bk=wa_backup_maybe_send(force=False)
        if debug and bk.get('ok'):
            diag['backup']=f"enviado para {bk.get('email','')}"
    except Exception as e:
        app.logger.exception('Falha no backup automatico WhatsApp')
        if debug:
            diag['erros'].append(f'Backup automatico: {str(e)}')
    return jsonify({'ok':True,'debug':diag} if debug else {'ok':True})

@app.route('/api/medicoes/<int:id>/status',methods=['PUT'])
@lr
def api_medicao_status(id):
    m=Medicao.query.get_or_404(id)
    d=request.json or {}
    novo=(d.get('status') or '').strip().lower()
    validos=['rascunho','emitida','cancelada']
    if novo not in validos: return jsonify({'erro':f'Status invalido. Use: {validos}'}),400
    m.status=novo; db.session.commit()
    audit_event('medicao_status_alterado','usuario',session.get('uid'),'medicao',m.id,True,{'status':novo,'numero':m.numero})
    return jsonify({'ok':True,'id':m.id,'status':novo})

@app.route('/api/medicoes/<int:id>/anexos',methods=['GET'])
@lr
def api_medicao_anexos(id):
    Medicao.query.get_or_404(id)
    return jsonify([a.to_dict() for a in MedicaoAnexo.query.filter_by(medicao_id=id).order_by(MedicaoAnexo.criado_em.desc()).all()])

@app.route('/api/medicoes/<int:id>/anexos',methods=['POST'])
@lr
def api_medicao_add_anexo(id):
    Medicao.query.get_or_404(id)
    fs=request.files.get('arquivo')
    if not fs: return jsonify({'erro':'Arquivo nao enviado'}),400
    rel,_=save_upload(fs,f'medicoes/{id}')
    a=MedicaoAnexo(medicao_id=id,nome_arquivo=fs.filename,caminho=rel,criado_por=session.get('nome',''))
    db.session.add(a); db.session.commit()
    return jsonify(a.to_dict()),201

@app.route('/api/medicoes/anexos/<int:id>',methods=['DELETE'])
@lr
def api_medicao_del_anexo(id):
    a=MedicaoAnexo.query.get_or_404(id)
    try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
    except: pass
    db.session.delete(a); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/medicoes/anexos/<int:id>/download')
@lr
def api_medicao_download_anexo(id):
    a=MedicaoAnexo.query.get_or_404(id)
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

def seed():
    if Usuario.query.count()==0:
        db.session.add(Usuario(nome='Administrador',email='admin@rmfacilities.com.br',senha=hs('rm@2026'),perfil='dono'))
    if Empresa.query.count()==0:
        db.session.add(Empresa(nome='RM Facilities',razao='RM CONSERVAÇÃO E SERVIÇOS LTDA',site='https://rmfacilities.com.br',cidade='São José dos Campos',estado='SP',ordem=1))
    if not Config.query.filter_by(chave='num_base').first(): db.session.add(Config(chave='num_base',valor='100'))
    if not Config.query.filter_by(chave='num_ultima').first(): db.session.add(Config(chave='num_ultima',valor='0'))
    db.session.commit()

with app.app_context():
    os.makedirs('instance',exist_ok=True)
    os.makedirs(UPLOAD_ROOT,exist_ok=True)
    db.create_all()
    ensure_cols('usuario',[
        'areas TEXT'
    ])
    ensure_cols('empresa',[
        'contato_nome VARCHAR(150)',
        'contato_email VARCHAR(150)',
        'contato_telefone VARCHAR(30)',
        'logo_url VARCHAR(500)',
        'boleto TEXT'
    ])
    ensure_cols('funcionario',[
        're INTEGER',
        'matricula VARCHAR(30)',
        'funcao VARCHAR(150)',
        'cbo VARCHAR(20)',
        'tipo_contrato VARCHAR(60)',
        'jornada VARCHAR(80)',
        'vale_refeicao FLOAT DEFAULT 0',
        'vale_alimentacao FLOAT DEFAULT 0',
        'vale_transporte FLOAT DEFAULT 0',
        'rg VARCHAR(30)',
        'orgao_emissor VARCHAR(30)',
        'pis VARCHAR(30)',
        'ctps VARCHAR(30)',
        'titulo_eleitor VARCHAR(30)',
        'cert_reservista VARCHAR(30)',
        'cnh VARCHAR(30)',
        'exame_admissional_data VARCHAR(10)',
        'docs_admissao_ok BOOLEAN DEFAULT 0',
        'docs_admissao_obs TEXT',
        'app_senha VARCHAR(256)',
        'app_ativo BOOLEAN DEFAULT 1',
        'app_ultimo_acesso DATETIME'
    ])
    db.session.execute(text('CREATE UNIQUE INDEX IF NOT EXISTS ix_funcionario_re ON funcionario(re)'))
    db.session.commit()
    ensure_cols('whats_app_conversa',[
        "contexto TEXT DEFAULT '{}'"
    ])
    ensure_cols('medicao',[
        'status VARCHAR(20) DEFAULT "emitida"',
        'desconto FLOAT DEFAULT 0',
        'impostos TEXT',
    ])
    seed(); get_logo()

if __name__=='__main__':
    app.run(host='0.0.0.0',port=5000,debug=False)
