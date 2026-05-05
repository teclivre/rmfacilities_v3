import io
import urllib.request
import urllib.error
import csv
import mimetypes
import zipfile
import shutil
import base64
import time
import smtplib
import threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
# Corrige NameError: _strict_origin_check
_strict_origin_check = True





from flask import Flask, request, jsonify, redirect, render_template, send_file, Response, url_for, has_request_context

import io
_strict_origin_check = False
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from sqlalchemy.exc import OperationalError, IntegrityError
import os
import re
import math
import unicodedata
import os, json, hashlib, hmac, secrets
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from ponto_module import register_ponto_routes


# Flask app and DB initialization must come first

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _is_prod_hint():
    env = (os.environ.get('APP_ENV') or os.environ.get('FLASK_ENV') or '').strip().lower()
    return env in ('prod', 'production')

def _resolve_data_dir():
    configured = (os.environ.get('DATA_DIR') or '').strip()
    if configured:
        return os.path.abspath(configured)
    # In production we strongly prefer /data (persistent volume in containers).
    if _is_prod_hint():
        try:
            os.makedirs('/data', exist_ok=True)
            return '/data'
        except Exception:
            pass
    return os.path.join(BASE_DIR, 'instance')

DATA_DIR = os.path.abspath(_resolve_data_dir())
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'rmfacilities.db')
DEFAULT_DB_URI = f"sqlite:///{DB_PATH}"
UPLOAD_ROOT = os.path.join(DATA_DIR, 'uploads')

def _migrate_legacy_data_dir():
    # Migra dados locais legados na primeira inicializacao com DATA_DIR.
    if os.environ.get('DATABASE_URL'):
        return
    if not os.path.exists(DB_PATH):
        old_candidates = [
            os.path.join(BASE_DIR, 'instance', 'rmfacilities.db'),
            os.path.join(BASE_DIR, 'instance', 'app.db'),
            os.path.join(BASE_DIR, 'app.db'),
            os.path.join(os.path.dirname(BASE_DIR), 'instance', 'rmfacilities.db'),
            os.path.join(os.path.dirname(BASE_DIR), 'instance', 'app.db'),
        ]
        existing = [p for p in old_candidates if os.path.exists(p) and os.path.abspath(p) != os.path.abspath(DB_PATH)]
        if existing:
            # Prefer rmfacilities.db over app.db, then bigger/newer files.
            existing.sort(
                key=lambda p: (
                    os.path.basename(p) != 'rmfacilities.db',
                    -os.path.getsize(p),
                    -os.path.getmtime(p),
                )
            )
            shutil.copy2(existing[0], DB_PATH)
    legacy_uploads = os.path.join(BASE_DIR, 'instance', 'uploads')
    if not os.path.isdir(UPLOAD_ROOT) and os.path.isdir(legacy_uploads):
        shutil.copytree(legacy_uploads, UPLOAD_ROOT, dirs_exist_ok=True)

_migrate_legacy_data_dir()

def _is_production_env():
    env=(os.environ.get('APP_ENV') or os.environ.get('FLASK_ENV') or '').strip().lower()
    return env in ('prod','production')

def _load_app_secret_key():
    key=(os.environ.get('SECRET_KEY') or '').strip()
    if key:
        return key
    if _is_production_env():
        raise RuntimeError('SECRET_KEY obrigatoria em producao. Configure a variavel de ambiente SECRET_KEY.')
    # Em ambiente local, usa chave efemera para evitar segredo padrao hardcoded no repositorio.
    return secrets.token_urlsafe(48)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
app.secret_key = _load_app_secret_key()
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', DEFAULT_DB_URI)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

from functools import wraps
from flask import session, url_for, g

def _lr_unauth_response():
    if (request.path or '').startswith('/api/'):
        return jsonify({'erro':'Sessao expirada. Faca login novamente.'}),401
    return redirect(url_for('login'))

def lr(f):
    @wraps(f)
    def w(*a,**k):
        if 'uid' not in session: return _lr_unauth_response()
        if not can_access_request(request.path,request.method): return jsonify({'erro':'Acesso negado'}),403
        if request.method in ('POST','PUT','PATCH','DELETE') and not _same_origin_request(request):
            return jsonify({'erro':'Origem da requisição não permitida'}),403
        return f(*a,**k)
    return w

APP_TZ=ZoneInfo('America/Sao_Paulo')

def _same_origin_request(req):
    global _strict_origin_check
    if not _strict_origin_check:
        return True
    host=(req.host_url or '').rstrip('/').lower()
    origin=(req.headers.get('Origin') or '').rstrip('/').lower()
    referer=(req.headers.get('Referer') or '').lower()
    sec_fetch_site=(req.headers.get('Sec-Fetch-Site') or '').lower().strip()
    if sec_fetch_site in ('same-origin', 'none', ''):
        return True
    if origin and host and origin.startswith(host):
        return True
    if not origin and referer and referer.startswith(host):
        return True
    return False

def localnow():
    return datetime.now(APP_TZ).replace(tzinfo=None)

def utcnow():
    # Compatibilidade histórica: a aplicação persiste timestamps como naive datetime,
    # mas o valor correto para o negócio é o horário de Brasília.
    return localnow()



# === ENDPOINT BANCOS-BR ===
@app.route('/api/bancos-br', methods=['GET'])
@lr
def api_bancos_br():
    refresh = request.args.get('refresh') in ('1', 'true', 'True', 'yes')
    bancos = bancos_br_get(refresh=refresh)
    return jsonify({'bancos': bancos})


_strict_origin_check = False  # Corrige NameError para _strict_origin_check


# === MODELOS E ROTAS QUE DEVEM VIR APÓS CRIAÇÃO DO APP E DB ===

# Modelo de marcação de ponto
class PontoMarcacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    funcionario_id = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    tipo = db.Column(db.String(30), nullable=False)
    data_hora = db.Column(db.DateTime, nullable=False)
    origem = db.Column(db.String(30))
    observacao = db.Column(db.String(255))
    criado_por = db.Column(db.String(100))
    ip = db.Column(db.String(60))
    latitude = db.Column(db.Float)
    longitude = db.Column(db.Float)
    precisao_gps = db.Column(db.Float)
    def to_dict(self):
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}

class PontoFechamentoDia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    funcionario_id = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    data_ref = db.Column(db.String(10), nullable=False)  # formato YYYY-MM-DD
    status = db.Column(db.String(50))
    observacao = db.Column(db.Text)
    resumo_json = db.Column(db.Text)
    fechado_por = db.Column(db.String(100))
    fechado_em = db.Column(db.DateTime)
    def to_dict(self):
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}

class PontoAjuste(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    funcionario_id = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    data_ref = db.Column(db.String(10), nullable=False)  # formato YYYY-MM-DD
    motivo = db.Column(db.String(255), nullable=False)
    antes_json = db.Column(db.Text, nullable=False)
    depois_json = db.Column(db.Text, nullable=False)
    criado_por = db.Column(db.String(100))
    criado_em = db.Column(db.DateTime, default=utcnow)
    def to_dict(self):
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}

class JornadaTrabalho(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    descricao = db.Column(db.String(255))
    dias_semana = db.Column(db.String(30), default='1,2,3,4,5')  # 0=dom 1=seg ... 6=sab
    hora_entrada = db.Column(db.String(5), default='08:00')       # HH:MM
    hora_saida = db.Column(db.String(5), default='17:48')         # HH:MM
    hora_intervalo_inicio = db.Column(db.String(5), default='12:00')
    hora_intervalo_fim = db.Column(db.String(5), default='13:00')
    tolerancia_min = db.Column(db.Integer, default=10)            # minutos de tolerância
    ativo = db.Column(db.Boolean, default=True)
    criado_em = db.Column(db.DateTime, default=utcnow)
    def carga_horaria_min(self):
        try:
            he=list(map(int,self.hora_entrada.split(':'))); hs=list(map(int,self.hora_saida.split(':')))
            hi=list(map(int,self.hora_intervalo_inicio.split(':'))); hf=list(map(int,self.hora_intervalo_fim.split(':')))
            trabalho=(hs[0]*60+hs[1])-(he[0]*60+he[1])
            intervalo=(hf[0]*60+hf[1])-(hi[0]*60+hi[1])
            return max(0, trabalho-intervalo)
        except Exception:
            return 480
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        try: d['dias_semana_list']=[int(x) for x in (self.dias_semana or '').split(',') if x.strip().isdigit()]
        except: d['dias_semana_list']=[1,2,3,4,5]
        d['carga_horaria_min']=self.carga_horaria_min()
        d['funcionarios_count']=Funcionario.query.filter_by(jornada_id=self.id).count() if self.id else 0
        return d

class Despesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0)
    categoria = db.Column(db.String(50), default='Outras')
    data = db.Column(db.String(10))
    comprovante = db.Column(db.String(300))
    observacoes = db.Column(db.Text)
    criado_em = db.Column(db.DateTime, default=utcnow)
    def to_dict(self):
        d = {c.name: getattr(self, c.name) for c in self.__table__.columns}
        d['criado_fmt'] = self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

def _parse_date_ymd(raw):
    s=(raw or '').strip()
    if not s:
        return None
    try:
        return datetime.strptime(s,'%Y-%m-%d').date()
    except Exception:
        return None

def _is_true_param(v):
    return str(v or '').strip().lower() in ('1','true','yes','on')

def _to_float_safe(v, default=0.0):
    try:
        return float(str(v).replace(',','.'))
    except Exception:
        return default

def _build_despesas_query(args):
    categoria=(args.get('categoria') or '').strip()
    periodo=(args.get('periodo') or '').strip()  # YYYY-MM
    q=(args.get('q') or '').strip().lower()
    data_ini=_parse_date_ymd(args.get('data_ini'))
    data_fim=_parse_date_ymd(args.get('data_fim'))

    qr=Despesa.query
    if categoria:
        qr=qr.filter(Despesa.categoria==categoria)
    if periodo:
        qr=qr.filter(Despesa.data.like(f'{periodo}%'))
    if q:
        qr=qr.filter(
            Despesa.descricao.ilike(f'%{q}%') |
            Despesa.categoria.ilike(f'%{q}%') |
            Despesa.observacoes.ilike(f'%{q}%')
        )
    if data_ini:
        qr=qr.filter(Despesa.data>=data_ini.strftime('%Y-%m-%d'))
    if data_fim:
        qr=qr.filter(Despesa.data<=data_fim.strftime('%Y-%m-%d'))

    order_by=(args.get('order_by') or 'data').strip().lower()
    order_dir=(args.get('order_dir') or 'desc').strip().lower()
    sort_map={
        'id':Despesa.id,
        'descricao':Despesa.descricao,
        'categoria':Despesa.categoria,
        'data':Despesa.data,
        'valor':Despesa.valor,
        'criado_em':Despesa.criado_em,
    }
    col=sort_map.get(order_by,Despesa.data)
    qr=qr.order_by(col.asc() if order_dir=='asc' else col.desc(),Despesa.id.desc())
    return qr

@app.route('/api/despesas', methods=['GET'])
@lr
def api_despesas_list():
    qr=_build_despesas_query(request.args)
    wants_paged=_is_true_param(request.args.get('paged')) or bool(request.args.get('page') or request.args.get('per_page'))
    if not wants_paged:
        lista=qr.all()
        return jsonify([d.to_dict() for d in lista])

    page=max(1,to_num(request.args.get('page')) or 1)
    per_page=to_num(request.args.get('per_page')) or 20
    per_page=min(max(1,per_page),200)
    total=qr.count()
    itens=qr.offset((page-1)*per_page).limit(per_page).all()
    total_paginas=max(1,(total+per_page-1)//per_page)
    return jsonify({
        'ok':True,
        'itens':[d.to_dict() for d in itens],
        'page':page,
        'per_page':per_page,
        'total':total,
        'total_paginas':total_paginas,
        'tem_anterior':page>1,
        'tem_proxima':page<total_paginas,
        'total_valor':round(sum((x.valor or 0) for x in itens),2),
    })

@app.route('/api/despesas', methods=['POST'])
@lr
def api_despesas_add():
    d=request.form or request.json or {}
    descricao=(d.get('descricao') or '').strip()
    categoria=(d.get('categoria') or 'Outras').strip() or 'Outras'
    data=(d.get('data') or '').strip()
    observacoes=(d.get('observacoes') or '').strip()
    valor=_to_float_safe(d.get('valor'),-1)
    if not descricao:
        return jsonify({'erro':'Descrição é obrigatória.'}),400
    if valor<0:
        return jsonify({'erro':'Valor inválido.'}),400
    if data and not _parse_date_ymd(data):
        return jsonify({'erro':'Data inválida. Use o formato YYYY-MM-DD.'}),400

    comprovante=''
    if 'comprovante' in request.files:
        fs=request.files['comprovante']
        if fs and fs.filename:
            rel,_=save_upload(fs,'despesas')
            comprovante=rel
    try:
        nova=Despesa(descricao=descricao[:200],valor=valor,categoria=categoria[:50],data=data[:10],comprovante=comprovante,observacoes=observacoes)
        db.session.add(nova)
        db.session.commit()
        return jsonify(nova.to_dict()),201
    except Exception:
        db.session.rollback()
        return jsonify({'erro':'Erro ao cadastrar despesa.'}),500

@app.route('/api/despesas/<int:id>', methods=['PUT'])
@lr
def api_despesas_edit(id):
    desp=Despesa.query.get_or_404(id)
    d=request.get_json(silent=True) or request.form or {}
    descricao=(d.get('descricao') if 'descricao' in d else desp.descricao) or ''
    categoria=(d.get('categoria') if 'categoria' in d else desp.categoria) or 'Outras'
    data=(d.get('data') if 'data' in d else desp.data) or ''
    observacoes=(d.get('observacoes') if 'observacoes' in d else desp.observacoes) or ''
    valor=_to_float_safe((d.get('valor') if 'valor' in d else desp.valor),-1)

    descricao=str(descricao).strip()
    categoria=str(categoria).strip() or 'Outras'
    data=str(data).strip()
    observacoes=str(observacoes).strip()
    if not descricao:
        return jsonify({'erro':'Descrição é obrigatória.'}),400
    if valor<0:
        return jsonify({'erro':'Valor inválido.'}),400
    if data and not _parse_date_ymd(data):
        return jsonify({'erro':'Data inválida. Use o formato YYYY-MM-DD.'}),400

    try:
        desp.descricao=descricao[:200]
        desp.categoria=categoria[:50]
        desp.data=data[:10]
        desp.valor=valor
        desp.observacoes=observacoes
        db.session.commit()
        return jsonify(desp.to_dict())
    except Exception:
        db.session.rollback()
        return jsonify({'erro':'Erro ao atualizar despesa.'}),500

@app.route('/api/despesas/<int:id>', methods=['DELETE'])
@lr
def api_despesas_delete(id):
    desp=Despesa.query.get_or_404(id)
    try:
        if (desp.comprovante or '').strip():
            abs_path=os.path.join(UPLOAD_ROOT,desp.comprovante)
            if os.path.exists(abs_path):
                try:
                    os.remove(abs_path)
                except Exception:
                    pass
        db.session.delete(desp)
        db.session.commit()
        return jsonify({'ok':True})
    except Exception:
        db.session.rollback()
        return jsonify({'erro':'Erro ao excluir despesa.'}),500

@app.route('/api/despesas/export.csv', methods=['GET'])
@lr
def api_despesas_export_csv():
    lista=_build_despesas_query(request.args).all()
    out=io.StringIO()
    w=csv.writer(out,delimiter=';')
    w.writerow(['ID','Descricao','Categoria','Data','Valor','Observacoes'])
    for d in lista:
        w.writerow([
            d.id,
            d.descricao or '',
            d.categoria or '',
            d.data or '',
            f'{(d.valor or 0):.2f}'.replace('.',','),
            d.observacoes or '',
        ])
    csv_txt=out.getvalue()
    return Response(
        csv_txt,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition':'attachment; filename=despesas.csv'}
    )

@app.route('/api/medicoes/check-numero')
@lr
def api_check_numero():
    numero=(request.args.get('numero') or '').strip()
    exclude_id=request.args.get('exclude_id',type=int)
    if not numero: return jsonify({'existe':False})
    q=Medicao.query.filter(Medicao.numero==numero)
    if exclude_id: q=q.filter(Medicao.id!=exclude_id)
    return jsonify({'existe':q.first() is not None})

@app.route('/api/medicoes/<int:id>',methods=['GET','DELETE','PUT'])
@lr
def api_medicao_detalhe(id):
    m = Medicao.query.get_or_404(id)
    if request.method=='DELETE':
        # remove anexos do disco antes de deletar
        for a in MedicaoAnexo.query.filter_by(medicao_id=id).all():
            try:
                fp=os.path.join(DATA_DIR,a.caminho) if a.caminho and not os.path.isabs(a.caminho) else (a.caminho or '')
                if fp and os.path.exists(fp): os.remove(fp)
            except Exception: pass
            db.session.delete(a)
        audit_event('medicao_excluida','usuario',session.get('uid'),'medicao',m.id,True,{'numero':m.numero})
        db.session.delete(m)
        db.session.commit()
        return jsonify({'ok':True})
    if request.method=='PUT':
        d=request.json or {}
        for campo in ['numero','tipo','cliente_nome','cliente_cnpj','cliente_end','cliente_resp',
                      'empresa_nome','mes_ref','dt_emissao','dt_vencimento','observacoes',
                      'ass_empresa','ass_cliente','status']:
            if campo in d: setattr(m, campo, d[campo])
        if 'empresa_id' in d: m.empresa_id=to_num(d['empresa_id'])
        if 'cliente_id' in d: m.cliente_id=to_num(d['cliente_id'])
        if 'servicos' in d: m.servicos=json.dumps(d['servicos'],ensure_ascii=False)
        if 'valor_bruto' in d: m.valor_bruto=float(d['valor_bruto'] or 0)
        db.session.commit()
        audit_event('medicao_editada','usuario',session.get('uid'),'medicao',m.id,True,{'numero':m.numero,'status':m.status})
        return jsonify({'ok':True,'id':m.id})
    return jsonify(m.to_dict())

@app.route('/api/proximo-numero')
@lr
def api_proximo_numero():
    return jsonify({'numero': prox_num()})

@app.route('/api/medicoes')
@lr
def api_medicoes():
    mes_ref = request.args.get('mes_ref', '').strip()
    cliente = request.args.get('cliente', '').strip().lower()
    status = request.args.get('status', '').strip().lower()
    cliente_id = to_num(request.args.get('cliente_id'))
    page = max(1, to_num(request.args.get('page', 1)) or 1)
    per_page = max(1, min(200, to_num(request.args.get('per_page', 0)) or 0))
    qr = Medicao.query
    if mes_ref:
        qr = qr.filter(Medicao.mes_ref == mes_ref)
    if cliente:
        qr = qr.filter((Medicao.cliente_nome.ilike(f'%{cliente}%')) | (Medicao.cliente_cnpj.ilike(f'%{cliente}%')))
    if status:
        qr = qr.filter(Medicao.status == status)
    if cliente_id:
        qr = qr.filter(Medicao.cliente_id == cliente_id)
    try:
        qr = qr.order_by(Medicao.dt_emissao.desc(), Medicao.id.desc())
        if per_page:
            total = qr.count()
            lista = qr.offset((page - 1) * per_page).limit(per_page).all()
            return jsonify({'items': [m.to_dict() for m in lista], 'total': total, 'page': page, 'per_page': per_page, 'pages': -(-total // per_page)})
        lista = qr.all()
    except OperationalError as e:
        if not _is_missing_medicao_stamp_error(e):
            raise
        db.session.rollback()
        _ensure_medicao_stamp_cols_runtime(force=True)
        lista = qr.order_by(Medicao.dt_emissao.desc(), Medicao.id.desc()).all()
    return jsonify([m.to_dict() for m in lista])

def _build_faturamento_query(args):
    mes_ref=(args.get('mes_ref') or '').strip()
    cliente=(args.get('cliente') or '').strip().lower()
    status=(args.get('status') or '').strip().lower()
    empresa_id=to_num(args.get('empresa_id'))
    q=(args.get('q') or '').strip().lower()
    dt_ini=(args.get('dt_ini') or '').strip()
    dt_fim=(args.get('dt_fim') or '').strip()
    valor_min=_to_float_safe(args.get('valor_min'),None) if (args.get('valor_min') or '').strip()!='' else None
    valor_max=_to_float_safe(args.get('valor_max'),None) if (args.get('valor_max') or '').strip()!='' else None

    qr=Medicao.query
    if mes_ref:
        qr=qr.filter(Medicao.mes_ref==mes_ref)
    if cliente:
        qr=qr.filter((Medicao.cliente_nome.ilike(f'%{cliente}%')) | (Medicao.cliente_cnpj.ilike(f'%{cliente}%')))
    if status:
        qr=qr.filter(Medicao.status==status)
    if empresa_id:
        qr=qr.filter(Medicao.empresa_id==empresa_id)
    if q:
        qr=qr.filter(
            Medicao.numero.ilike(f'%{q}%') |
            Medicao.cliente_nome.ilike(f'%{q}%') |
            Medicao.cliente_cnpj.ilike(f'%{q}%') |
            Medicao.empresa_nome.ilike(f'%{q}%')
        )
    if dt_ini and _parse_date_ymd(dt_ini):
        qr=qr.filter(Medicao.dt_emissao>=dt_ini)
    if dt_fim and _parse_date_ymd(dt_fim):
        qr=qr.filter(Medicao.dt_emissao<=dt_fim)
    if valor_min is not None:
        qr=qr.filter(Medicao.valor_bruto>=valor_min)
    if valor_max is not None:
        qr=qr.filter(Medicao.valor_bruto<=valor_max)

    order_by=(args.get('order_by') or 'dt_emissao').strip().lower()
    order_dir=(args.get('order_dir') or 'desc').strip().lower()
    sort_map={
        'id':Medicao.id,
        'numero':Medicao.numero,
        'cliente_nome':Medicao.cliente_nome,
        'empresa_nome':Medicao.empresa_nome,
        'mes_ref':Medicao.mes_ref,
        'dt_emissao':Medicao.dt_emissao,
        'dt_vencimento':Medicao.dt_vencimento,
        'valor_bruto':Medicao.valor_bruto,
        'status':Medicao.status,
    }
    col=sort_map.get(order_by,Medicao.dt_emissao)
    qr=qr.order_by(col.asc() if order_dir=='asc' else col.desc(),Medicao.id.desc())
    return qr

@app.route('/api/financeiro/faturamento', methods=['GET'])
@lr
def api_financeiro_faturamento_list():
    try:
        qr=_build_faturamento_query(request.args)
        page=max(1,to_num(request.args.get('page')) or 1)
        per_page=to_num(request.args.get('per_page')) or 20
        per_page=min(max(1,per_page),200)
        total=qr.count()
        itens=qr.offset((page-1)*per_page).limit(per_page).all()
        total_paginas=max(1,(total+per_page-1)//per_page)
        total_valor=0.0
        out=[]
        for m in itens:
            d=m.to_dict()
            total_valor+=(d.get('valor_bruto') or 0)
            out.append(d)
        return jsonify({
            'ok':True,
            'itens':out,
            'page':page,
            'per_page':per_page,
            'total':total,
            'total_paginas':total_paginas,
            'tem_anterior':page>1,
            'tem_proxima':page<total_paginas,
            'total_valor':round(total_valor,2),
        })
    except OperationalError as e:
        if not _is_missing_medicao_stamp_error(e):
            raise
        db.session.rollback()
        _ensure_medicao_stamp_cols_runtime(force=True)
        return api_financeiro_faturamento_list()

@app.route('/api/financeiro/faturamento/export.csv', methods=['GET'])
@lr
def api_financeiro_faturamento_export_csv():
    try:
        lista=_build_faturamento_query(request.args).all()
    except OperationalError as e:
        if not _is_missing_medicao_stamp_error(e):
            raise
        db.session.rollback()
        _ensure_medicao_stamp_cols_runtime(force=True)
        lista=_build_faturamento_query(request.args).all()

    out=io.StringIO()
    w=csv.writer(out,delimiter=';')
    w.writerow(['ID','Numero','Cliente','CNPJ','Empresa','Mes Ref','Emissao','Vencimento','Valor','Status'])
    for m in lista:
        d=m.to_dict()
        w.writerow([
            d.get('id') or '',
            d.get('numero') or '',
            d.get('cliente_nome') or '',
            d.get('cliente_cnpj') or '',
            d.get('empresa_nome') or '',
            d.get('mes_ref') or '',
            d.get('dt_emissao') or '',
            d.get('dt_vencimento') or '',
            f"{(d.get('valor_bruto') or 0):.2f}".replace('.',','),
            d.get('status') or '',
        ])
    csv_txt=out.getvalue()
    return Response(
        csv_txt,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition':'attachment; filename=faturamento.csv'}
    )

class Usuario(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    nome=db.Column(db.String(100),nullable=False)
    email=db.Column(db.String(150),unique=True,nullable=False)
    telefone=db.Column(db.String(30))
    senha=db.Column(db.String(256),nullable=False)
    perfil=db.Column(db.String(20),default='admin')
    twofa_ativo=db.Column(db.Boolean,default=False)
    cert_arquivo=db.Column(db.String(500))
    cert_nome_arquivo=db.Column(db.String(255))
    cert_senha=db.Column(db.String(255))
    cert_ativo=db.Column(db.Boolean,default=False)
    cert_assunto=db.Column(db.String(255))
    cert_validade_fim=db.Column(db.String(30))
    areas=db.Column(db.Text,default='[]')
    permissoes=db.Column(db.Text,default='{}')
    ativo=db.Column(db.Boolean,default=True)
    ultimo_acesso=db.Column(db.DateTime)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def check_senha(self, s):
        return pw_check(self.senha, s)
    def to_dict(self):
        try: a=json.loads(self.areas or '[]')
        except: a=[]
        try: p=json.loads(self.permissoes or '{}')
        except: p={}
        if not isinstance(p,dict): p={}
        return {
            'id':self.id,
            'nome':self.nome,
            'email':self.email,
            'telefone':self.telefone or '',
            'perfil':self.perfil,
            'twofa_ativo':bool(self.twofa_ativo if self.twofa_ativo is not None else True),
            'ativo':self.ativo,
            'areas':a,
            'permissoes':p,
            'rbac_actions_ativo':bool(p),
            'cert_configurado':bool((self.cert_arquivo or '').strip()),
            'cert_nome_arquivo':self.cert_nome_arquivo or '',
            'cert_ativo':bool(self.cert_ativo if self.cert_ativo is not None else False),
            'cert_assunto':self.cert_assunto or '',
            'cert_validade_fim':self.cert_validade_fim or ''
        }

class Empresa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    nome=db.Column(db.String(200),nullable=False)
    razao=db.Column(db.String(200))
    cnpj=db.Column(db.String(30))
    telefone=db.Column(db.String(30))
    email=db.Column(db.String(150))
    site=db.Column(db.String(200))
    cep=db.Column(db.String(10))
    logradouro=db.Column(db.String(200))
    numero=db.Column(db.String(20))
    complemento=db.Column(db.String(100))
    bairro=db.Column(db.String(100))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    pix=db.Column(db.String(300))
    banco=db.Column(db.String(200))
    agencia=db.Column(db.String(20))
    conta=db.Column(db.String(30))
    contato_nome=db.Column(db.String(150))
    contato_email=db.Column(db.String(150))
    contato_telefone=db.Column(db.String(30))
    logo_url=db.Column(db.String(500))
    cert_arquivo=db.Column(db.String(500))
    cert_nome_arquivo=db.Column(db.String(255))
    cert_senha=db.Column(db.String(255))
    cert_ativo=db.Column(db.Boolean,default=False)
    cert_assunto=db.Column(db.String(255))
    cert_validade_fim=db.Column(db.String(30))
    boleto=db.Column(db.Text)
    ativa=db.Column(db.Boolean,default=True)
    ordem=db.Column(db.Integer,default=0)
    def end_fmt(self):
        p=[self.logradouro,self.numero,self.complemento,self.bairro]
        e=', '.join(filter(None,p))
        if self.cidade: e+=f' — {self.cidade}/{self.estado or ""}'
        return e
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['end_fmt']=self.end_fmt()
        d['cert_configurado']=bool((self.cert_arquivo or '').strip())
        d['cert_nome_arquivo']=self.cert_nome_arquivo or ''
        d['cert_ativo']=bool(self.cert_ativo if self.cert_ativo is not None else False)
        d['cert_assunto']=self.cert_assunto or ''
        d['cert_validade_fim']=self.cert_validade_fim or ''
        d.pop('cert_senha',None)
        return d

class Config(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    chave=db.Column(db.String(100),unique=True,nullable=False)
    valor=db.Column(db.Text,default='')

class Cliente(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(10))
    nome=db.Column(db.String(200),nullable=False)
    cnpj=db.Column(db.String(30))
    responsavel=db.Column(db.String(150))
    telefone=db.Column(db.String(30))
    email=db.Column(db.String(150))
    cep=db.Column(db.String(10))
    logradouro=db.Column(db.String(200))
    numero_end=db.Column(db.String(20))
    complemento=db.Column(db.String(100))
    bairro=db.Column(db.String(100))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    numero_contrato=db.Column(db.String(60))
    qtd_funcionarios_posto=db.Column(db.Integer,default=0)
    status=db.Column(db.String(20),default='Ativo')
    limpeza=db.Column(db.Float,default=0)
    jardinagem=db.Column(db.Float,default=0)
    portaria=db.Column(db.Float,default=0)
    materiais_equip_locacao=db.Column(db.Float,default=0)
    vencimento=db.Column(db.Integer,default=10)
    dia_faturamento=db.Column(db.Integer,default=1)
    dias_faturamento=db.Column(db.Integer,default=30)
    dt_contrato_vencimento=db.Column(db.String(10))
    reajuste_percentual=db.Column(db.Float,default=0)
    reajuste_data_base=db.Column(db.String(10))
    ultimo_reajuste_em=db.Column(db.String(10))
    geo_lat=db.Column(db.Float)
    geo_lon=db.Column(db.Float)
    geofence_raio_m=db.Column(db.Float,default=150)
    obs=db.Column(db.Text,default='')
    criado_em=db.Column(db.DateTime,default=utcnow)
    def end_fmt(self):
        p=[self.logradouro,self.numero_end,self.complemento,self.bairro]
        e=', '.join(filter(None,p))
        if self.cidade: e+=f' — {self.cidade}/{self.estado or ""}'
        return e
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['end_fmt']=self.end_fmt()
        return d

class Contrato(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    cliente_id=db.Column(db.Integer,db.ForeignKey('cliente.id'),nullable=False)
    numero=db.Column(db.String(60))
    status=db.Column(db.String(20),default='Ativo')
    dt_inicio=db.Column(db.String(10))
    dt_vencimento=db.Column(db.String(10))
    qtd_funcionarios_posto=db.Column(db.Integer,default=0)
    limpeza=db.Column(db.Float,default=0)
    jardinagem=db.Column(db.Float,default=0)
    portaria=db.Column(db.Float,default=0)
    materiais_equip_locacao=db.Column(db.Float,default=0)
    dia_faturamento=db.Column(db.Integer,default=1)
    dias_faturamento=db.Column(db.Integer,default=30)
    reajuste_percentual=db.Column(db.Float,default=0)
    reajuste_data_base=db.Column(db.String(10))
    ultimo_reajuste_em=db.Column(db.String(10))
    obs=db.Column(db.Text,default='')
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        return {c.name:getattr(self,c.name) for c in self.__table__.columns}

class Medicao(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(20))
    tipo=db.Column(db.String(50),default='Medição de Serviços')
    cliente_id=db.Column(db.Integer,db.ForeignKey('cliente.id'),nullable=True)
    cliente_nome=db.Column(db.String(200))
    cliente_cnpj=db.Column(db.String(30))
    cliente_end=db.Column(db.String(300))
    cliente_resp=db.Column(db.String(150))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    empresa_nome=db.Column(db.String(200))
    mes_ref=db.Column(db.String(7))
    dt_emissao=db.Column(db.String(10))
    dt_vencimento=db.Column(db.String(10))
    servicos=db.Column(db.Text)
    valor_bruto=db.Column(db.Float,default=0)
    observacoes=db.Column(db.Text)
    ass_empresa=db.Column(db.String(200))
    ass_cliente=db.Column(db.String(200))
    status=db.Column(db.String(20),default='emitida')
    desconto=db.Column(db.Float,default=0)
    impostos=db.Column(db.Text)
    assinatura_status=db.Column(db.String(20),default='nao_solicitada')
    assinatura_token=db.Column(db.String(120))
    assinatura_expira_em=db.Column(db.DateTime)
    assinatura_nome=db.Column(db.String(200))
    assinatura_cpf=db.Column(db.String(20))
    assinatura_cargo=db.Column(db.String(120))
    assinatura_ip=db.Column(db.String(60))
    assinatura_em=db.Column(db.DateTime)
    assinatura_codigo=db.Column(db.String(120))
    assinatura_otp_hash=db.Column(db.String(256))
    assinatura_otp_expira_em=db.Column(db.DateTime)
    assinatura_otp_tentativas=db.Column(db.Integer,default=0)
    assinatura_doc_hash=db.Column(db.String(128))
    assinatura_crypto_ok=db.Column(db.Boolean,default=False)
    assinatura_cert_subject=db.Column(db.String(255))
    dt_pagamento=db.Column(db.String(10))
    forma_pagamento=db.Column(db.String(50))
    valor_juros=db.Column(db.Float,default=0)
    valor_multa=db.Column(db.Float,default=0)
    criado_em=db.Column(db.DateTime,default=utcnow)
    criado_por=db.Column(db.String(100))
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['svcs']=json.loads(self.servicos) if self.servicos else []
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d.setdefault('status','emitida')
        # Corrige empresa_nome vazia ou com texto de placeholder para registros antigos
        nome=d.get('empresa_nome') or ''
        if (not nome or nome.lower().startswith('selecione')) and self.empresa_id:
            emp=db.session.get(Empresa, self.empresa_id)
            d['empresa_nome']=emp.nome if emp else ''
        return d

class OrdemCompra(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(30),nullable=False,unique=True)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    solicitante=db.Column(db.String(200))
    fornecedor=db.Column(db.String(200),nullable=False)
    descricao=db.Column(db.Text)
    valor=db.Column(db.Float,default=0)
    status=db.Column(db.String(50),default='Aberta')
    data_emissao=db.Column(db.String(10))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    ass_assinatura_img=db.Column(db.Text)
    def to_dict(self):
        return {c.name:getattr(self,c.name) for c in self.__table__.columns}

class Funcionario(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    matricula=db.Column(db.String(30),index=True)
    re=db.Column(db.Integer,unique=True,index=True)
    nome=db.Column(db.String(200),nullable=False)
    cpf=db.Column(db.String(20),unique=True)
    email=db.Column(db.String(150))
    telefone=db.Column(db.String(30))
    cargo=db.Column(db.String(120))
    funcao=db.Column(db.String(150))
    cbo=db.Column(db.String(20))
    setor=db.Column(db.String(120))
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    data_admissao=db.Column(db.String(10))
    tipo_contrato=db.Column(db.String(60))
    jornada=db.Column(db.String(80))
    jornada_id=db.Column(db.Integer,db.ForeignKey('jornada_trabalho.id'),nullable=True)
    status=db.Column(db.String(20),default='Ativo')
    salario=db.Column(db.Float,default=0)
    vale_refeicao=db.Column(db.Float,default=0)
    vale_alimentacao=db.Column(db.Float,default=0)
    vale_transporte=db.Column(db.Float,default=0)
    opta_vt=db.Column(db.Boolean,default=True)
    opta_vr=db.Column(db.Boolean,default=True)
    opta_va=db.Column(db.Boolean,default=True)
    opta_premio_prod=db.Column(db.Boolean,default=False)
    opta_vale_gasolina=db.Column(db.Boolean,default=False)
    opta_cesta_natal=db.Column(db.Boolean,default=False)
    premio_produtividade=db.Column(db.Float,default=0)
    vale_gasolina=db.Column(db.Float,default=0)
    cesta_natal=db.Column(db.Float,default=0)
    posto_operacional=db.Column(db.String(150))
    posto_cliente_id=db.Column(db.Integer,db.ForeignKey('cliente.id'),nullable=True)
    endereco=db.Column(db.String(250))
    cidade=db.Column(db.String(100))
    estado=db.Column(db.String(2))
    cep=db.Column(db.String(10))
    endereco_numero=db.Column(db.String(20))
    endereco_complemento=db.Column(db.String(120))
    endereco_bairro=db.Column(db.String(120))
    banco_codigo=db.Column(db.String(3))
    banco_nome=db.Column(db.String(150))
    banco_agencia=db.Column(db.String(30))
    banco_conta=db.Column(db.String(40))
    banco_tipo_conta=db.Column(db.String(20))
    banco_pix=db.Column(db.String(150))
    rg=db.Column(db.String(30))
    orgao_emissor=db.Column(db.String(30))
    pis=db.Column(db.String(30))
    ctps=db.Column(db.String(30))
    titulo_eleitor=db.Column(db.String(30))
    cert_reservista=db.Column(db.String(30))
    cnh=db.Column(db.String(30))
    exame_admissional_data=db.Column(db.String(10))
    docs_admissao_ok=db.Column(db.Boolean,default=False)
    docs_admissao_obs=db.Column(db.Text,default='')
    obs=db.Column(db.Text,default='')
    areas=db.Column(db.Text,default='[]')
    app_senha=db.Column(db.String(256))
    app_ativo=db.Column(db.Boolean,default=True)
    app_ultimo_acesso=db.Column(db.DateTime)
    app_otp_hash=db.Column(db.String(256))
    app_otp_expira_em=db.Column(db.DateTime)
    app_otp_tentativas=db.Column(db.Integer,default=0)
    app_push_token=db.Column(db.String(300))
    app_lat=db.Column(db.Float)
    app_lon=db.Column(db.Float)
    app_localizacao_em=db.Column(db.DateTime)
    foto_perfil=db.Column(db.String(500))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        try: d['areas']=json.loads(self.areas or '[]')
        except: d['areas']=[]
        return d

class FuncionarioArquivo(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False,index=True)
    categoria=db.Column(db.String(30),nullable=False,index=True)
    competencia=db.Column(db.String(20),index=True)
    nome_arquivo=db.Column(db.String(250),nullable=False)
    caminho=db.Column(db.String(500),nullable=False)
    ass_status=db.Column(db.String(20),default='nao_solicitada')
    ass_token=db.Column(db.String(120))
    ass_expira_em=db.Column(db.DateTime)
    ass_codigo=db.Column(db.String(120))
    ass_nome=db.Column(db.String(200))
    ass_cargo=db.Column(db.String(120))
    ass_cpf=db.Column(db.String(20))
    ass_ip=db.Column(db.String(60))
    ass_em=db.Column(db.DateTime)
    ass_otp_hash=db.Column(db.String(256))
    ass_otp_expira_em=db.Column(db.DateTime)
    ass_otp_tentativas=db.Column(db.Integer,default=0)
    ass_doc_hash=db.Column(db.String(128))
    ass_crypto_ok=db.Column(db.Boolean,default=False)
    ass_cert_subject=db.Column(db.String(255))
    ass_canal_envio=db.Column(db.String(20))
    ass_enviado_em=db.Column(db.DateTime)
    ass_recebido_em=db.Column(db.DateTime)
    ass_aberto_em=db.Column(db.DateTime)
    ass_wa_status=db.Column(db.String(20),default='nao_enviado')
    ass_wa_enviado_em=db.Column(db.DateTime)
    ass_wa_recebido_em=db.Column(db.DateTime)
    ass_email_status=db.Column(db.String(20),default='nao_enviado')
    ass_email_enviado_em=db.Column(db.DateTime)
    ass_email_recebido_em=db.Column(db.DateTime)
    ass_lembretes_enviados=db.Column(db.Integer,default=0)
    ass_ultimo_lembrete_em=db.Column(db.DateTime)
    ass_prazo_em=db.Column(db.DateTime)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['ass_em_fmt']=self.ass_em.strftime('%d/%m/%Y %H:%M') if self.ass_em else ''
        d['ass_enviado_fmt']=self.ass_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_enviado_em else ''
        d['ass_recebido_fmt']=self.ass_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_recebido_em else ''
        d['ass_aberto_fmt']=self.ass_aberto_em.strftime('%d/%m/%Y %H:%M') if self.ass_aberto_em else ''
        d['ass_wa_enviado_fmt']=self.ass_wa_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_wa_enviado_em else ''
        d['ass_wa_recebido_fmt']=self.ass_wa_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_wa_recebido_em else ''
        d['ass_email_enviado_fmt']=self.ass_email_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_email_enviado_em else ''
        d['ass_email_recebido_fmt']=self.ass_email_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_email_recebido_em else ''
        return d

class OperacionalDocumento(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True,index=True)
    tipo=db.Column(db.String(50),default='Documento')
    titulo=db.Column(db.String(200),nullable=False)
    descricao=db.Column(db.Text)
    nome_arquivo=db.Column(db.String(250))
    caminho=db.Column(db.String(500))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class BeneficioMensal(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False,index=True)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True,index=True)
    competencia=db.Column(db.String(7),nullable=False,index=True)
    dias_trabalhados=db.Column(db.Integer,default=0)
    dias_vt=db.Column(db.Integer,default=0)
    dias_vr=db.Column(db.Integer,default=0)
    dias_va=db.Column(db.Integer,default=0)
    dias_vg=db.Column(db.Integer,default=0)
    salario=db.Column(db.Float,default=0)
    vale_refeicao=db.Column(db.Float,default=0)
    vale_alimentacao=db.Column(db.Float,default=0)
    vale_transporte=db.Column(db.Float,default=0)
    pp_falta=db.Column(db.Boolean,default=False)
    premio_produtividade=db.Column(db.Float)
    vale_gasolina=db.Column(db.Float)
    cesta_natal=db.Column(db.Float)
    criado_em=db.Column(db.DateTime,default=utcnow)
    atualizado_em=db.Column(db.DateTime,default=utcnow,onupdate=utcnow)
    __table_args__=(db.UniqueConstraint('funcionario_id','competencia',name='uq_beneficio_func_comp'),)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['atualizado_fmt']=self.atualizado_em.strftime('%d/%m/%Y %H:%M') if self.atualizado_em else ''
        return d

class FuncionarioAppSessao(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False)
    refresh_hash=db.Column(db.String(256),nullable=False,index=True)
    exp_refresh=db.Column(db.DateTime,nullable=False)
    revogado=db.Column(db.Boolean,default=False)
    ip=db.Column(db.String(60))
    ua=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)
    atualizado_em=db.Column(db.DateTime,default=utcnow,onupdate=utcnow)

class FuncionarioAlteracaoSolicitacao(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False,index=True)
    payload=db.Column(db.Text,default='{}')
    observacao=db.Column(db.Text,default='')
    status=db.Column(db.String(20),default='pendente')
    motivo_admin=db.Column(db.Text,default='')
    analisado_por=db.Column(db.Integer,db.ForeignKey('usuario.id'),nullable=True)
    solicitado_em=db.Column(db.DateTime,default=utcnow)
    analisado_em=db.Column(db.DateTime)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['solicitado_fmt']=self.solicitado_em.strftime('%d/%m/%Y %H:%M') if self.solicitado_em else ''
        d['analisado_fmt']=self.analisado_em.strftime('%d/%m/%Y %H:%M') if self.analisado_em else ''
        d['payload']=jloads(self.payload,{})
        return d

class AuthTentativa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    tipo=db.Column(db.String(30),nullable=False)
    identificador=db.Column(db.String(200),nullable=False)
    ip=db.Column(db.String(60))
    ok=db.Column(db.Boolean,default=False)
    motivo=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)

class AuditoriaEvento(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    evento=db.Column(db.String(120),nullable=False)
    ator_tipo=db.Column(db.String(30))
    ator_id=db.Column(db.String(60))
    alvo_tipo=db.Column(db.String(30))
    alvo_id=db.Column(db.String(60))
    ok=db.Column(db.Boolean,default=True)
    ip=db.Column(db.String(60))
    ua=db.Column(db.String(250))
    detalhe=db.Column(db.Text)
    criado_em=db.Column(db.DateTime,default=utcnow)

class MedicaoAnexo(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    medicao_id=db.Column(db.Integer,db.ForeignKey('medicao.id'),nullable=False)
    nome_arquivo=db.Column(db.String(250),nullable=False)
    caminho=db.Column(db.String(500),nullable=False)
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class CobrangaLog(db.Model):
    __tablename__='cobranca_log'
    id=db.Column(db.Integer,primary_key=True)
    medicao_id=db.Column(db.Integer,db.ForeignKey('medicao.id'),nullable=False)
    tipo=db.Column(db.String(20))   # D-5, D-1, D+3, manual
    enviado_em=db.Column(db.DateTime,default=utcnow)
    status=db.Column(db.String(20),default='ok')  # ok | erro
    dest_email=db.Column(db.String(150))
    erro=db.Column(db.Text)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['enviado_fmt']=self.enviado_em.strftime('%d/%m/%Y %H:%M') if self.enviado_em else ''
        return d

class ConciliacaoLote(db.Model):
    __tablename__='conciliacao_lote'
    id=db.Column(db.Integer,primary_key=True)
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    arquivo_nome=db.Column(db.String(250))
    importado_em=db.Column(db.DateTime,default=utcnow)
    total_transacoes=db.Column(db.Integer,default=0)
    importado_por=db.Column(db.String(100))
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['importado_fmt']=self.importado_em.strftime('%d/%m/%Y %H:%M') if self.importado_em else ''
        return d

class ConciliacaoTransacao(db.Model):
    __tablename__='conciliacao_transacao'
    id=db.Column(db.Integer,primary_key=True)
    lote_id=db.Column(db.Integer,db.ForeignKey('conciliacao_lote.id'),nullable=False)
    data_mov=db.Column(db.String(10))
    valor=db.Column(db.Float,default=0)
    historico=db.Column(db.Text)
    num_doc=db.Column(db.String(80))
    tipo=db.Column(db.String(10),default='C')  # C=crédito D=débito
    medicao_id=db.Column(db.Integer,db.ForeignKey('medicao.id'),nullable=True)
    conciliado_em=db.Column(db.DateTime)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        return d

class AssinaturaEnvelope(db.Model):
    __tablename__='assinatura_envelope'
    id=db.Column(db.Integer,primary_key=True)
    titulo=db.Column(db.String(200),nullable=False)
    descricao=db.Column(db.Text)
    tipo=db.Column(db.String(20),default='avulso')  # funcionario|cliente|avulso
    empresa_id=db.Column(db.Integer,db.ForeignKey('empresa.id'),nullable=True)
    ref_id=db.Column(db.Integer)
    status=db.Column(db.String(20),default='rascunho')  # rascunho|pendente|parcial|concluido|cancelado
    codigo=db.Column(db.String(120))
    nome_documento_assinado=db.Column(db.String(255))
    destino_salvar_tipo=db.Column(db.String(30),default='envelope')  # envelope|funcionario
    destino_funcionario_id=db.Column(db.Integer)
    destino_categoria=db.Column(db.String(40),default='outros')
    destino_competencia=db.Column(db.String(20))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    expira_em=db.Column(db.DateTime)
    assinatura_doc_hash=db.Column(db.String(128))
    assinatura_crypto_ok=db.Column(db.Boolean,default=False)
    assinatura_cert_subject=db.Column(db.String(255))
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['expira_fmt']=self.expira_em.strftime('%d/%m/%Y') if self.expira_em else ''
        return d

class AssinaturaEnvelopeArquivo(db.Model):
    __tablename__='assinatura_envelope_arquivo'
    id=db.Column(db.Integer,primary_key=True)
    envelope_id=db.Column(db.Integer,db.ForeignKey('assinatura_envelope.id'),nullable=False)
    origem=db.Column(db.String(10),default='upload')  # sistema|upload
    func_arquivo_id=db.Column(db.Integer)
    nome_arquivo=db.Column(db.String(250),nullable=False)
    caminho=db.Column(db.String(500))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class AssinaturaEnvelopeSignatario(db.Model):
    __tablename__='assinatura_envelope_signatario'
    id=db.Column(db.Integer,primary_key=True)
    envelope_id=db.Column(db.Integer,db.ForeignKey('assinatura_envelope.id'),nullable=False)
    nome=db.Column(db.String(200))
    email=db.Column(db.String(150))
    telefone=db.Column(db.String(30))
    cpf=db.Column(db.String(20))
    cargo=db.Column(db.String(120))
    tipo=db.Column(db.String(20),default='externo')  # funcionario|cliente|externo
    ref_id=db.Column(db.Integer)
    token=db.Column(db.String(120))
    status=db.Column(db.String(20),default='pendente')  # pendente|assinado
    ass_ip=db.Column(db.String(60))
    ass_em=db.Column(db.DateTime)
    ass_codigo=db.Column(db.String(120))
    ass_cpf_informado=db.Column(db.String(20))
    ass_otp_hash=db.Column(db.String(256))
    ass_otp_expira_em=db.Column(db.DateTime)
    ass_otp_tentativas=db.Column(db.Integer,default=0)
    ass_canal_envio=db.Column(db.String(20))
    ass_enviado_em=db.Column(db.DateTime)
    ass_recebido_em=db.Column(db.DateTime)
    ass_aberto_em=db.Column(db.DateTime)
    ass_wa_status=db.Column(db.String(20),default='nao_enviado')
    ass_wa_enviado_em=db.Column(db.DateTime)
    ass_wa_recebido_em=db.Column(db.DateTime)
    ass_email_status=db.Column(db.String(20),default='nao_enviado')
    ass_email_enviado_em=db.Column(db.DateTime)
    ass_email_recebido_em=db.Column(db.DateTime)
    ordem=db.Column(db.Integer,default=0)
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['ass_em_fmt']=self.ass_em.strftime('%d/%m/%Y %H:%M') if self.ass_em else ''
        d['ass_enviado_fmt']=self.ass_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_enviado_em else ''
        d['ass_recebido_fmt']=self.ass_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_recebido_em else ''
        d['ass_aberto_fmt']=self.ass_aberto_em.strftime('%d/%m/%Y %H:%M') if self.ass_aberto_em else ''
        d['ass_wa_enviado_fmt']=self.ass_wa_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_wa_enviado_em else ''
        d['ass_wa_recebido_fmt']=self.ass_wa_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_wa_recebido_em else ''
        d['ass_email_enviado_fmt']=self.ass_email_enviado_em.strftime('%d/%m/%Y %H:%M') if self.ass_email_enviado_em else ''
        d['ass_email_recebido_fmt']=self.ass_email_recebido_em.strftime('%d/%m/%Y %H:%M') if self.ass_email_recebido_em else ''
        return d

class WhatsAppConversa(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    numero=db.Column(db.String(30),nullable=False,index=True)
    nome=db.Column(db.String(200))
    ultima_msg=db.Column(db.DateTime,default=utcnow)
    contexto=db.Column(db.Text,default='{}')
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['ultima_msg']=self.ultima_msg.isoformat() if self.ultima_msg else ''
        d['criado_em']=self.criado_em.isoformat() if self.criado_em else ''
        d['ultima_msg_fmt']=self.ultima_msg.strftime('%d/%m/%Y %H:%M') if self.ultima_msg else ''
        try:
            d['contexto']=json.loads(self.contexto or '{}')
        except:
            d['contexto']={}
        return d

class WhatsAppMensagem(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    conversa_id=db.Column(db.Integer,db.ForeignKey('whats_app_conversa.id'),nullable=False)
    numero=db.Column(db.String(30))
    direcao=db.Column(db.String(10))
    tipo=db.Column(db.String(20),default='texto')
    conteudo=db.Column(db.Text)
    nome_arquivo=db.Column(db.String(250))
    criado_em=db.Column(db.DateTime,default=utcnow)
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['criado_em']=self.criado_em.isoformat() if self.criado_em else ''
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        return d

class ShortLink(db.Model):
    __tablename__='short_link'
    id=db.Column(db.Integer,primary_key=True)
    codigo=db.Column(db.String(12),unique=True,nullable=False,index=True)
    destino=db.Column(db.String(1000),nullable=False)
    criado_em=db.Column(db.DateTime,default=utcnow)

def _short_link_criar(destino):
    """Cria ou reutiliza link curto para o destino dado."""
    ex=ShortLink.query.filter_by(destino=destino).first()
    if ex:
        return ex.codigo
    for _ in range(8):
        codigo=secrets.token_urlsafe(6)[:8]
        if not ShortLink.query.filter_by(codigo=codigo).first():
            sl=ShortLink(codigo=codigo,destino=destino)
            db.session.add(sl)
            try:
                db.session.commit()
                return codigo
            except Exception:
                db.session.rollback()
    return None

@app.route('/s/<codigo>')
def short_link_redirect(codigo):
    sl=ShortLink.query.filter_by(codigo=codigo).first_or_404()
    return redirect(sl.destino)

class ComunicadoApp(db.Model):
    __tablename__='comunicado_app'
    id=db.Column(db.Integer,primary_key=True)
    titulo=db.Column(db.String(200),nullable=False)
    conteudo=db.Column(db.Text,nullable=False)
    # None = para todos; int = para funcionario específico
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=True)
    # None = todos os postos; string = apenas esse posto
    posto_operacional=db.Column(db.String(150))
    criado_por=db.Column(db.String(100))
    criado_em=db.Column(db.DateTime,default=utcnow)
    ativo=db.Column(db.Boolean,default=True)
    # JSON list of funcionario_ids que leram
    lidos_por_json=db.Column(db.Text,default='[]')
    def lidos_por(self):
        try: return json.loads(self.lidos_por_json or '[]')
        except: return []
    def marcar_lido(self,fid):
        lst=self.lidos_por()
        if fid not in lst:
            lst.append(fid)
            self.lidos_por_json=json.dumps(lst)
    def to_dict(self,funcionario_id=None):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns if c.name!='lidos_por_json'}
        d['criado_fmt']=self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else ''
        d['lido']=funcionario_id in self.lidos_por() if funcionario_id else False
        d['lidos_count']=len(self.lidos_por())
        return d

class MensagemApp(db.Model):
    __tablename__='mensagem_app'
    id=db.Column(db.Integer,primary_key=True)
    funcionario_id=db.Column(db.Integer,db.ForeignKey('funcionario.id'),nullable=False)
    de_rh=db.Column(db.Boolean,default=False)  # True = RH→funcionário; False = funcionário→RH
    conteudo=db.Column(db.Text,nullable=False)
    enviado_em=db.Column(db.DateTime,default=utcnow)
    lida=db.Column(db.Boolean,default=False)
    enviado_por=db.Column(db.String(100))  # nome do usuário RH ou 'funcionario'
    tipo=db.Column(db.String(20),default='texto')  # 'texto' | 'arquivo'
    arquivo_nome=db.Column(db.String(300))
    arquivo_caminho=db.Column(db.String(500))
    def to_dict(self):
        d={c.name:getattr(self,c.name) for c in self.__table__.columns}
        d['enviado_fmt']=self.enviado_em.strftime('%d/%m/%Y %H:%M') if self.enviado_em else ''
        if self.arquivo_caminho:
            d['arquivo_url']=f'/api/app/funcionario/mensagens/{self.id}/arquivo'
        else:
            d['arquivo_url']=None
        return d

_holerite_jobs={}

def hs(s): return hashlib.sha256(s.encode()).hexdigest()

def pw_hash(s):
    return generate_password_hash(s,method='scrypt')

def pw_is_modern(v):
    return str(v or '').startswith(('scrypt:','pbkdf2:'))

def pw_check(stored,plain):
    st=str(stored or '')
    if not st: return False
    if pw_is_modern(st):
        try: return check_password_hash(st,plain)
        except Exception: return False
    return hmac.compare_digest(st,hs(plain))

def token_hash(v):
    return hs('tok:'+str(v or ''))

LOGIN_WINDOW_MIN=15
LOGIN_FAIL_MAX=5
LOGIN_BLOCK_MIN=15

def auth_blocked(tipo,ident,ip):
    lim=utcnow()-timedelta(minutes=LOGIN_WINDOW_MIN)
    q=AuthTentativa.query.filter(AuthTentativa.tipo==tipo,AuthTentativa.identificador==ident,AuthTentativa.criado_em>=lim,AuthTentativa.ok==False)
    fails=q.count()
    if fails<LOGIN_FAIL_MAX: return False
    ult=q.order_by(AuthTentativa.criado_em.desc()).first()
    if not ult: return False
    return ult.criado_em+timedelta(minutes=LOGIN_BLOCK_MIN)>utcnow()

def reg_auth_attempt(tipo,ident,ok,motivo=''):
    ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    db.session.add(AuthTentativa(tipo=tipo,identificador=ident,ip=ip,ok=bool(ok),motivo=(motivo or '')[:250]))
    db.session.commit()

def _mask_email(v):
    s=(v or '').strip()
    if '@' not in s:
        return s
    usr,dom=s.split('@',1)
    if len(usr)<=2:
        usr_mask='*'*len(usr)
    else:
        usr_mask=usr[0]+'*'*(len(usr)-2)+usr[-1]
    return f'{usr_mask}@{dom}'

def _mask_phone(v):
    d=only_digits(v)
    if len(d)<4:
        return '***'
    return f'*** *** {d[-4:]}'

def _admin_needs_2fa(u):
    if not u:
        return False
    if (u.perfil or '').strip().lower() not in ('admin','dono'):
        return False
    return bool(u.twofa_ativo)

def _send_admin_2fa_code(u,codigo,contexto='login'):
    tel=norm_phone(getattr(u,'telefone','') or '')
    email=(getattr(u,'email','') or '').strip()
    if contexto=='recuperacao':
        msg=(
            f'RM Facilities - Recuperação de acesso\n'
            f'Código: {codigo}\n'
            'Validade: 10 minutos.\n'
            'Se não solicitou, ignore esta mensagem.'
        )
    else:
        msg=(
            f'RM Facilities - Verificação em duas etapas\n'
            f'Código: {codigo}\n'
            'Validade: 10 minutos.\n'
            'Não compartilhe este código.'
        )
    ultimo_erro=''
    if len(tel)>=10:
        try:
            wa_send_text(tel,msg)
            return
        except Exception as ex:
            ultimo_erro=str(ex)
    if email:
        try:
            smtp_send_text(email,'Código de segurança RM Facilities',msg)
            return
        except Exception as ex:
            ultimo_erro=str(ex)
    raise ValueError(ultimo_erro or 'Não foi possível enviar o código de segurança por celular/e-mail.')

def _certs_base_dir():
    p=os.path.join(UPLOAD_ROOT,'certificados')
    os.makedirs(p,exist_ok=True)
    return p

def _cert_rel_to_abs(rel_path):
    raw=(rel_path or '').strip()
    if not raw:
        return ''
    if os.path.isabs(raw):
        return raw if os.path.exists(raw) else ''
    cands=[
        os.path.join(UPLOAD_ROOT,raw),
        os.path.join(_get_uploads_base(),raw) if '_get_uploads_base' in globals() else '',
    ]
    for p in cands:
        if p and os.path.exists(p):
            return p
    return ''

def _cert_store_file(fs,scope,obj_id):
    if not fs or not fs.filename:
        raise ValueError('Arquivo de certificado não enviado.')
    name=secure_filename(fs.filename)
    ext=os.path.splitext(name)[1].lower()
    if ext not in ('.p12','.pfx'):
        raise ValueError('Formato inválido. Envie certificado .p12 ou .pfx.')
    base=os.path.join(_certs_base_dir(),scope,str(obj_id))
    os.makedirs(base,exist_ok=True)
    final_name=f'certificado{ext}'
    abs_path=os.path.join(base,final_name)
    fs.save(abs_path)
    rel=os.path.relpath(abs_path,UPLOAD_ROOT).replace('\\','/')
    return rel,name

def _cert_inspect_pkcs12(abs_path,senha=''):
    try:
        from cryptography.hazmat.primitives.serialization import pkcs12
    except Exception as ex:
        raise ValueError(f'Biblioteca de certificado não disponível: {str(ex)}')
    with open(abs_path,'rb') as f:
        raw=f.read()
    pwd=(senha or '').encode('utf-8') if (senha or '').strip() else None
    key,cert,_=pkcs12.load_key_and_certificates(raw,pwd)
    if not key or not cert:
        raise ValueError('Certificado inválido ou sem chave privada.')
    assunto=''
    validade=''
    try:
        assunto=cert.subject.rfc4514_string()
    except Exception:
        assunto=''
    try:
        dt=getattr(cert,'not_valid_after_utc',None) or cert.not_valid_after
        validade=dt.strftime('%Y-%m-%d') if dt else ''
    except Exception:
        validade=''
    return {'assunto':assunto[:255],'validade_fim':validade}

def _get_cert_context(empresa_id=None,usuario_id=None):
    if empresa_id:
        emp=Empresa.query.get(empresa_id)
        if emp and bool(emp.cert_ativo if emp.cert_ativo is not None else False):
            abs_cert=_cert_rel_to_abs(emp.cert_arquivo)
            if abs_cert and (emp.cert_senha or '').strip():
                return {
                    'cert_path':abs_cert,
                    'cert_pass':emp.cert_senha,
                    'cert_subject':emp.cert_assunto or '',
                    'source':'empresa',
                }
    if usuario_id:
        usr=Usuario.query.get(usuario_id)
        if usr and bool(usr.cert_ativo if usr.cert_ativo is not None else False):
            abs_cert=_cert_rel_to_abs(usr.cert_arquivo)
            if abs_cert and (usr.cert_senha or '').strip():
                return {
                    'cert_path':abs_cert,
                    'cert_pass':usr.cert_senha,
                    'cert_subject':usr.cert_assunto or '',
                    'source':'usuario',
                }
    return None

def _sha256_bytes(raw):
    h=hashlib.sha256()
    h.update(raw or b'')
    return h.hexdigest().upper()

def _sha256_file(path):
    h=hashlib.sha256()
    with open(path,'rb') as f:
        for chunk in iter(lambda: f.read(1024*1024),b''):
            h.update(chunk)
    return h.hexdigest().upper()

def _otp_new_code():
    return ''.join(secrets.choice('0123456789') for _ in range(6))

def _send_signature_otp(codigo,nome_dest='',telefone='',email='',contexto='assinatura'):
    nome=(nome_dest or 'assinante').strip()
    if contexto=='medicao':
        assunto='Código de confirmação de assinatura da medição'
        titulo='assinatura da medição'
    elif contexto=='envelope':
        assunto='Código de confirmação de assinatura do envelope'
        titulo='assinatura do envelope'
    else:
        assunto='Código de confirmação de assinatura de documento'
        titulo='assinatura do documento'
    msg=(
        f'RM Facilities - Confirmação de {titulo}\n'
        f'Destinatário: {nome}\n'
        f'Código OTP: {codigo}\n'
        'Validade: 10 minutos.\n'
        'Não compartilhe este código.'
    )
    tel=wa_norm_number(telefone or '')
    ultimo_erro=''
    if wa_is_valid_number(tel):
        try:
            wa_send_text(tel,msg)
            return {'canal':'whatsapp','destino':_mask_phone(tel)}
        except Exception as ex:
            ultimo_erro=str(ex)
    if (email or '').strip():
        try:
            smtp_send_text((email or '').strip(),assunto,msg)
            return {'canal':'email','destino':_mask_email(email)}
        except Exception as ex:
            ultimo_erro=str(ex)
    raise ValueError(ultimo_erro or 'Não foi possível enviar o código OTP para confirmação da assinatura.')

def _send_app_login_otp(codigo,funcionario):
    """Envia OTP de login do app exclusivamente por WhatsApp."""
    f=funcionario
    nome=(f.nome or 'funcionario').strip()
    msg=(
        'RM Facilities - Codigo de acesso do aplicativo\n'
        f'Funcionario: {nome}\n'
        f'Codigo OTP: {codigo}\n'
        'Validade: 10 minutos. Nao compartilhe este codigo.'
    )
    tel=wa_norm_number(f.telefone or '')
    if not wa_is_valid_number(tel):
        raise ValueError('Telefone WhatsApp invalido ou nao cadastrado para este funcionario.')
    try:
        wa_send_text(tel,msg)
        return {'canal':'whatsapp','destino':_mask_phone(tel)}
    except Exception as ex:
        raise ValueError(str(ex) or 'Nao foi possivel enviar OTP por WhatsApp.')

def _assinatura_json_base(**extra):
    base={
        'ok':False,
        'mensagem':'',
        'erro':'',
        'otp_required':False,
        'validacao_link':'',
        'signed_pdf_link':'',
        'whatsapp_enviado':False,
        'codigo':'',
    }
    base.update(extra or {})
    return base

def _assinatura_json_erro(msg,status=400,**extra):
    payload=_assinatura_json_base(ok=False,erro=(msg or 'Não foi possível concluir sua solicitação.'),**extra)
    return jsonify(payload),status

def _assinatura_json_otp(mensagem,canal='',destino='',**extra):
    payload=_assinatura_json_base(ok=False,otp_required=True,mensagem=(mensagem or ''),canal=(canal or ''),destino=(destino or ''),**extra)
    return jsonify(payload)

def _assinatura_json_ok(mensagem='',**extra):
    payload=_assinatura_json_base(ok=True,mensagem=(mensagem or 'Assinatura concluída com sucesso.'),**extra)
    return jsonify(payload)

def _ass_track_channel(src,default='link'):
    s=(src or '').strip().lower()
    if s in ('wa','wpp','whatsapp'):
        return 'whatsapp'
    if s in ('mail','email','e-mail'):
        return 'email'
    if s in ('app','aplicativo','push'):
        return 'app'
    if s in ('link','manual','direto'):
        return 'link'
    return (default or 'link')

def _ass_track_mark_sent(obj,channel):
    ch=_ass_track_channel(channel)
    now=utcnow()
    changed=False
    if getattr(obj,'ass_canal_envio',None)!=ch:
        obj.ass_canal_envio=ch
        changed=True
    if not getattr(obj,'ass_enviado_em',None):
        obj.ass_enviado_em=now
        changed=True
    if ch=='whatsapp':
        if (getattr(obj,'ass_wa_status',None) or '')!='enviado':
            obj.ass_wa_status='enviado'
            changed=True
        if not getattr(obj,'ass_wa_enviado_em',None):
            obj.ass_wa_enviado_em=now
            changed=True
    elif ch=='email':
        if (getattr(obj,'ass_email_status',None) or '')!='enviado':
            obj.ass_email_status='enviado'
            changed=True
        if not getattr(obj,'ass_email_enviado_em',None):
            obj.ass_email_enviado_em=now
            changed=True
    return changed

def _ass_track_mark_received(obj,channel):
    ch=_ass_track_channel(channel,getattr(obj,'ass_canal_envio',None) or 'link')
    now=utcnow()
    changed=False
    if not getattr(obj,'ass_recebido_em',None):
        obj.ass_recebido_em=now
        changed=True
    if ch=='whatsapp':
        if (getattr(obj,'ass_wa_status',None) or '')!='recebido':
            obj.ass_wa_status='recebido'
            changed=True
        if not getattr(obj,'ass_wa_recebido_em',None):
            obj.ass_wa_recebido_em=now
            changed=True
    elif ch=='email':
        if (getattr(obj,'ass_email_status',None) or '')!='recebido':
            obj.ass_email_status='recebido'
            changed=True
        if not getattr(obj,'ass_email_recebido_em',None):
            obj.ass_email_recebido_em=now
            changed=True
    return changed

def _ass_track_mark_opened(obj,channel):
    now=utcnow()
    changed=_ass_track_mark_received(obj,channel)
    if not getattr(obj,'ass_aberto_em',None):
        obj.ass_aberto_em=now
        changed=True
    return changed

def _try_sign_pdf_bytes_crypto(pdf_bytes,empresa_id=None,usuario_id=None):
    cert_ctx=_get_cert_context(empresa_id=empresa_id,usuario_id=usuario_id)
    p12_path=(cert_ctx.get('cert_path') if cert_ctx else '') or (os.environ.get('PDF_SIGN_P12_PATH') or '').strip()
    p12_pass=(cert_ctx.get('cert_pass') if cert_ctx else '') or (os.environ.get('PDF_SIGN_P12_PASS') or '').strip()
    if not p12_path or not p12_pass or not os.path.exists(p12_path):
        return {'ok':False,'bytes':pdf_bytes,'reason':'crypto_not_configured'}
    try:
        from pyhanko.sign import signers
        from pyhanko.pdf_utils.incremental_writer import IncrementalPdfFileWriter
    except Exception as ex:
        return {'ok':False,'bytes':pdf_bytes,'reason':f'pyhanko_unavailable:{str(ex)}'}
    try:
        meta=signers.PdfSignatureMetadata(
            field_name='Signature1',
            reason=(os.environ.get('PDF_SIGN_REASON') or 'Assinatura eletrônica RM Facilities'),
            location=(os.environ.get('PDF_SIGN_LOCATION') or 'Brasil')
        )
        signer=signers.SimpleSigner.load_pkcs12(
            pfx_file=p12_path,
            passphrase=p12_pass.encode('utf-8')
        )
        out=io.BytesIO()
        writer=IncrementalPdfFileWriter(io.BytesIO(pdf_bytes))
        signers.PdfSigner(signature_meta=meta,signer=signer).sign_pdf(writer,output=out)
        cert=''
        try:
            cert=signer.signing_cert.subject.human_friendly
        except Exception:
            cert=cert_ctx.get('cert_subject','') if cert_ctx else ''
        return {'ok':True,'bytes':out.getvalue(),'cert_subject':cert}
    except Exception as ex:
        return {'ok':False,'bytes':pdf_bytes,'reason':str(ex)}

def _try_sign_pdf_file_crypto(abs_path,empresa_id=None,usuario_id=None):
    if not abs_path or not os.path.exists(abs_path):
        return {'ok':False,'reason':'file_not_found'}
    with open(abs_path,'rb') as f:
        raw=f.read()
    rs=_try_sign_pdf_bytes_crypto(raw,empresa_id=empresa_id,usuario_id=usuario_id)
    if rs.get('ok'):
        with open(abs_path,'wb') as f:
            f.write(rs.get('bytes') or b'')
    return rs

def audit_event(evento,ator_tipo='',ator_id='',alvo_tipo='',alvo_id='',ok=True,det=None):
    try:
        ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
        ua=(request.headers.get('User-Agent') or '')[:250]
        detalhe=det if isinstance(det,str) else json.dumps(det or {},ensure_ascii=False)
        db.session.add(AuditoriaEvento(evento=evento,ator_tipo=str(ator_tipo or '')[:30],ator_id=str(ator_id or '')[:60],alvo_tipo=str(alvo_tipo or '')[:30],alvo_id=str(alvo_id or '')[:60],ok=bool(ok),ip=ip,ua=ua,detalhe=detalhe))
        db.session.commit()
    except Exception:
        db.session.rollback()

def b64u_enc(raw):
    return base64.urlsafe_b64encode(raw).decode().rstrip('=')

def b64u_dec(s):
    pad='='*((4-len(s)%4)%4)
    return base64.urlsafe_b64decode((s+pad).encode())

def app_token_secret():
    tok=(os.environ.get('APP_TOKEN_SECRET') or '').strip()
    if tok:
        return tok.encode()
    if _is_production_env():
        raise RuntimeError('APP_TOKEN_SECRET obrigatoria em producao. Configure a variavel de ambiente APP_TOKEN_SECRET.')
    sk=app.secret_key or ''
    return (sk if isinstance(sk,str) else str(sk)).encode()

def app_issue_access_token(funcionario_id,sessao_id,ttl=3600):
    now=int(time.time())
    payload={'typ':'access','sid':int(sessao_id),'fid':int(funcionario_id),'iat':now,'exp':now+int(ttl)}
    ptxt=json.dumps(payload,separators=(',',':')).encode()
    p64=b64u_enc(ptxt)
    sig=hmac.new(app_token_secret(),p64.encode(),hashlib.sha256).digest()
    s64=b64u_enc(sig)
    return f'{p64}.{s64}'

def app_issue_refresh_token():
    return secrets.token_urlsafe(48)

def app_parse_token(token):
    try:
        p64,s64=token.split('.',1)
        expected=b64u_enc(hmac.new(app_token_secret(),p64.encode(),hashlib.sha256).digest())
        if not hmac.compare_digest(expected,s64): return None
        payload=json.loads(b64u_dec(p64).decode())
        if int(payload.get('exp',0))<int(time.time()): return None
        return payload
    except Exception:
        return None

def app_func_required(f):
    @wraps(f)
    def w(*a,**k):
        auth=(request.headers.get('Authorization') or '').strip()
        if not auth.lower().startswith('bearer '):
            return jsonify({'erro':'Token ausente'}),401
        tok=auth.split(' ',1)[1].strip()
        payload=app_parse_token(tok)
        if not payload or payload.get('typ')!='access': return jsonify({'erro':'Token invalido ou expirado'}),401
        sid=to_num(payload.get('sid'))
        sessao=FuncionarioAppSessao.query.get(sid)
        if not sessao or sessao.revogado: return jsonify({'erro':'Sessao invalida'}),401
        if sessao.exp_refresh < utcnow(): return jsonify({'erro':'Sessao expirada'}),401
        func=Funcionario.query.get(sessao.funcionario_id)
        if not func: return jsonify({'erro':'Funcionario nao encontrado'}),404
        if to_num(payload.get('fid'))!=func.id: return jsonify({'erro':'Token invalido'}),401
        if func.app_ativo is False: return jsonify({'erro':'Acesso do aplicativo desativado'}),403
        g.app_funcionario=func
        g.app_sessao=sessao
        return f(*a,**k)
    return w

def _norm_nome_login(v):
    s=unicodedata.normalize('NFKD',str(v or ''))
    s=''.join(ch for ch in s if not unicodedata.combining(ch))
    s=re.sub(r'\s+',' ',s).strip().lower()
    return s

def _nome_confere_funcionario(nome_informado,nome_cadastro):
    ni=_norm_nome_login(nome_informado)
    nc=_norm_nome_login(nome_cadastro)
    if not ni or not nc:
        return False
    if ni==nc:
        return True
    # Aceita conferir primeiro e último nome para reduzir rejeição por nome composto.
    ni_parts=[p for p in ni.split(' ') if p]
    nc_parts=[p for p in nc.split(' ') if p]
    if len(ni_parts)>=2 and len(nc_parts)>=2:
        return ni_parts[0]==nc_parts[0] and ni_parts[-1]==nc_parts[-1]
    return False

def _app_issue_session_tokens(funcionario):
    refresh=app_issue_refresh_token()
    ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    ua=(request.headers.get('User-Agent') or '')[:250]
    sessao=FuncionarioAppSessao(
        funcionario_id=funcionario.id,
        refresh_hash=token_hash(refresh),
        exp_refresh=utcnow()+timedelta(days=14),
        revogado=False,
        ip=ip,
        ua=ua
    )
    db.session.add(sessao)
    db.session.flush()
    access=app_issue_access_token(funcionario.id,sessao.id,ttl=3600)
    funcionario.app_ultimo_acesso=utcnow()
    return {
        'access_token':access,
        'refresh_token':refresh,
        'token_type':'Bearer',
        'expires_in':3600,
        'refresh_expires_in':1209600,
        'sessao_id':sessao.id,
    }

def gc(k,dv=''): c=Config.query.filter_by(chave=k).first(); return c.valor if c else dv

def smtp_cfg():
    return {'host':gc('smtp_host',''),'port':gc('smtp_port','587'),'user':gc('smtp_user',''),'senha':gc('smtp_senha',''),'de':gc('smtp_de',''),'tls':gc('smtp_tls','1')}

def wa_cfg():
    return {'url':gc('wa_url',''),'instancia':gc('wa_instancia',''),'token':gc('wa_token','')}

def wa_backup_cfg():
    return {
        'enabled':gc('wa_backup_enabled','1'),
        'email':gc('wa_backup_email',''),
        'interval_hours':gc('wa_backup_interval_hours','2'),
        'window_hours':gc('wa_backup_window_hours','8'),
        'max_conversas':gc('wa_backup_max_conversas','10'),
        'last_ts':gc('wa_backup_last_ts','0'),
    }

def wa_backup_enabled():
    return str(wa_backup_cfg().get('enabled','0')).strip().lower() in ('1','true','yes','on')

def _wa_backup_root():
    return os.path.join(DATA_DIR, 'wa_backups')

def _to_int(v,dv=0):
    try: return int(float(str(v).strip()))
    except Exception: return dv

def _parse_dt_iso(v):
    s=(v or '').strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace('Z','+00:00')).replace(tzinfo=None)
    except Exception:
        return None

def _cfg_snapshot_dict():
    out={}
    for c in Config.query.all():
        out[c.chave]=c.valor
    return out

def _cfg_apply_snapshot(dct):
    if not isinstance(dct,dict):
        return 0
    n=0
    for k,v in dct.items():
        if not str(k or '').strip():
            continue
        sc_cfg(str(k).strip(),'' if v is None else str(v))
        n+=1
    return n

def _wa_backup_collect(window_hours=8,max_conversas=10):
    janela=max(1,min(168,_to_int(window_hours,8)))
    qtd=max(1,min(50,_to_int(max_conversas,10)))
    corte=utcnow()-timedelta(hours=janela)
    convs=WhatsAppConversa.query.order_by(WhatsAppConversa.ultima_msg.desc()).limit(qtd).all()
    out=[]
    for c in convs:
        q=WhatsAppMensagem.query.filter_by(conversa_id=c.id).filter(WhatsAppMensagem.criado_em>=corte).order_by(WhatsAppMensagem.criado_em.asc())
        msgs=q.all()
        out.append({
            'conversa':c.to_dict(),
            'mensagens':[{
                'id':m.id,
                'numero':m.numero,
                'direcao':m.direcao,
                'tipo':m.tipo,
                'conteudo':m.conteudo,
                'criado_em':m.criado_em.isoformat() if m.criado_em else '',
            } for m in msgs]
        })
    ia=ai_wa_cfg()
    payload={
        'gerado_em':utcnow().isoformat(),
        'versao_backup':'wa-v2',
        'janela_horas':janela,
        'max_conversas':qtd,
        'total_conversas':len(out),
        'total_mensagens':sum(len(c['mensagens']) for c in out),
        'config':_cfg_snapshot_dict(),
        'ia':{
            'enabled':ai_wa_enabled(),
            'provider':ia.get('provider',''),
            'model':ia.get('model',''),
            'temperature':ia.get('temperature',''),
            'max_tokens':ia.get('max_tokens',''),
        },
        'conversas':out,
    }
    return payload

def _wa_backup_store(payload):
    root=_wa_backup_root()
    os.makedirs(root,exist_ok=True)
    nome=f"wa_backup_{localnow().strftime('%Y%m%d_%H%M%S')}.json"
    caminho=os.path.join(root,nome)
    with open(caminho,'w',encoding='utf-8') as f:
        json.dump(payload,f,ensure_ascii=False,indent=2)
    return caminho

def _wa_backup_store_txt(payload):
    root=_wa_backup_root()
    os.makedirs(root,exist_ok=True)
    base=localnow().strftime('%Y%m%d_%H%M%S')
    nome=f"wa_historico_{base}.txt"
    caminho=os.path.join(root,nome)
    linhas=[]
    linhas.append('Backup de Historico WhatsApp - RM Facilities')
    linhas.append(f"Gerado em: {payload.get('gerado_em','')}")
    linhas.append(f"Janela (horas): {payload.get('janela_horas',0)}")
    linhas.append(f"Conversas: {payload.get('total_conversas',0)}")
    linhas.append(f"Mensagens: {payload.get('total_mensagens',0)}")
    linhas.append('')
    for bloco in (payload.get('conversas') or []):
        conv=bloco.get('conversa') or {}
        nome_conv=conv.get('nome') or conv.get('numero') or 'Sem nome'
        num=conv.get('numero') or ''
        linhas.append('='*72)
        linhas.append(f"Conversa: {nome_conv} ({num})")
        linhas.append('='*72)
        for m in (bloco.get('mensagens') or []):
            lado='Cliente' if m.get('direcao')=='in' else 'Atendente'
            dt=m.get('criado_em') or ''
            txt=(m.get('conteudo') or '').replace('\r','').strip()
            linhas.append(f"[{dt}] {lado}: {txt}")
        linhas.append('')
    with open(caminho,'w',encoding='utf-8') as f:
        f.write('\n'.join(linhas))
    return caminho

def smtp_send_text(dest,assunto,corpo,anexos=None):
    cfg=smtp_cfg()
    if not cfg['host'] or not cfg['user']: raise ValueError('SMTP nao configurado')
    msg=MIMEMultipart()
    msg['From']=cfg['de'] or cfg['user']
    msg['To']=dest
    msg['Subject']=assunto
    msg.attach(MIMEText(corpo or '','plain','utf-8'))
    for a in (anexos or []):
        caminho=a.get('path')
        nome=a.get('name') or os.path.basename(caminho or 'anexo.bin')
        if not caminho or not os.path.isfile(caminho):
            continue
        with open(caminho,'rb') as f:
            part=MIMEBase('application','octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',f'attachment; filename="{nome}"')
        msg.attach(part)
    port=int(cfg['port'] or 587)
    if str(cfg['tls']) in ('1','true','True','yes'):
        with smtplib.SMTP(cfg['host'],port,timeout=20) as s:
            s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
    else:
        with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s:
            s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())

def wa_backup_maybe_send(force=False):
    cfg=wa_backup_cfg()
    if not force and not wa_backup_enabled():
        return {'ok':False,'skip':'desativado'}
    email=(cfg.get('email') or '').strip()
    if not email:
        return {'ok':False,'skip':'sem_email'}
    intervalo=max(1,min(168,_to_int(cfg.get('interval_hours'),2)))
    janela=max(1,min(168,_to_int(cfg.get('window_hours'),8)))
    max_conv=max(1,min(50,_to_int(cfg.get('max_conversas'),10)))
    agora=int(time.time())
    ultimo=max(0,_to_int(cfg.get('last_ts'),0))
    if not force and (agora-ultimo)<(intervalo*3600):
        return {'ok':False,'skip':'intervalo'}
    payload=_wa_backup_collect(janela,max_conv)
    arq=_wa_backup_store(payload)
    arq_txt=_wa_backup_store_txt(payload)
    assunto=f"Backup WhatsApp RM Facilities - {localnow().strftime('%d/%m/%Y %H:%M')}"
    corpo=(
        f"Backup automatico das conversas WhatsApp.\n\n"
        f"Janela: {payload.get('janela_horas')}h\n"
        f"Conversas: {payload.get('total_conversas')}\n"
        f"Mensagens: {payload.get('total_mensagens')}\n"
        f"Modelo IA: {payload.get('ia',{}).get('model','')}\n"
    )
    smtp_send_text(email,assunto,corpo,anexos=[
        {'path':arq,'name':os.path.basename(arq)},
        {'path':arq_txt,'name':os.path.basename(arq_txt)},
    ])
    sc_cfg('wa_backup_last_ts',str(agora))
    return {'ok':True,'email':email,'arquivo':arq,'total_conversas':payload.get('total_conversas',0),'total_mensagens':payload.get('total_mensagens',0)}

def wa_backup_restore_payload(payload,restore_config=True,restore_conversas=True):
    if not isinstance(payload,dict):
        raise ValueError('Arquivo de backup invalido')
    stat={'configs':0,'conversas':0,'mensagens':0}
    if restore_config:
        stat['configs']=_cfg_apply_snapshot(payload.get('config') or {})
    if restore_conversas:
        blocos=payload.get('conversas') or []
        for bloco in blocos:
            conv=bloco.get('conversa') or {}
            numero=wa_norm_number(conv.get('numero') or '')
            if not wa_is_valid_number(numero):
                continue
            c=WhatsAppConversa.query.filter_by(numero=numero).first()
            if not c:
                c=WhatsAppConversa(numero=numero,nome=(conv.get('nome') or numero))
                db.session.add(c)
                db.session.flush()
                stat['conversas']+=1
            elif (conv.get('nome') or '').strip():
                c.nome=(conv.get('nome') or c.nome)
            ult=_parse_dt_iso(conv.get('ultima_msg') or '')
            if ult and (not c.ultima_msg or ult>c.ultima_msg):
                c.ultima_msg=ult
            for m in (bloco.get('mensagens') or []):
                txt=(m.get('conteudo') or '').strip()
                if not txt:
                    continue
                created=_parse_dt_iso(m.get('criado_em') or '') or utcnow()
                direcao=(m.get('direcao') or 'in').strip().lower()
                tipo=(m.get('tipo') or 'texto').strip().lower() or 'texto'
                existe=WhatsAppMensagem.query.filter_by(conversa_id=c.id,direcao=direcao,tipo=tipo,conteudo=txt,criado_em=created).first()
                if existe:
                    continue
                db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao=direcao,tipo=tipo,conteudo=txt,criado_em=created))
                stat['mensagens']+=1
    db.session.commit()
    return stat

def wa_norm_number(numero):
    n=re.sub(r'\D+','',str(numero or ''))
    if n.startswith('00'): n=n[2:]
    if n.startswith('55') and len(n)>=12: return n
    if len(n) in (10,11): return '55'+n
    return n

def wa_is_valid_number(numero):
    n=re.sub(r'\D+','',str(numero or ''))
    return bool(re.fullmatch(r'\d{12,15}',n))

def wa_phone_matches(a,b):
    da=only_digits(a)
    dbn=only_digits(b)
    if not da or not dbn:
        return False
    na=wa_norm_number(da)
    nb=wa_norm_number(dbn)
    if na and nb and na==nb:
        return True
    for size in (11,10,9,8):
        if len(da)>=size and len(dbn)>=size and da[-size:]==dbn[-size:]:
            return True
    return False

def _peek_upload_bytes(fs,size=8):
    if not fs:
        return b''
    try:
        stream=getattr(fs,'stream',None)
        if stream is not None:
            pos=stream.tell()
            data=stream.read(size) or b''
            stream.seek(pos)
            return data
    except Exception:
        pass
    try:
        data=fs.read(size) or b''
        try:
            fs.seek(0)
        except Exception:
            pass
        return data
    except Exception:
        return b''

def _upload_is_pdf(fs):
    return _peek_upload_bytes(fs,5).startswith(b'%PDF-')

def _funcionario_por_whatsapp(numero):
    num=only_digits(numero)
    if not num:
        return None
    for funcionario in Funcionario.query.all():
        if wa_phone_matches(num,funcionario.telefone or ''):
            return funcionario
    return None

def _funcionario_por_cpf(cpf_digits):
    alvo=only_digits(cpf_digits)
    if len(alvo)!=11:
        return None
    for funcionario in Funcionario.query.all():
        if only_digits(funcionario.cpf or '')==alvo:
            return funcionario
    return None

def _valida_cpf(cpf):
    """Valida CPF pelo algoritmo dos dígitos verificadores."""
    d=only_digits(cpf)
    if len(d)!=11 or len(set(d))==1:
        return False
    s=sum(int(d[i])*(10-i) for i in range(9))
    r=(s*10)%11; r=0 if r==10 else r
    if r!=int(d[9]): return False
    s=sum(int(d[i])*(11-i) for i in range(10))
    r=(s*10)%11; r=0 if r==10 else r
    return r==int(d[10])

def _telefone_final_3(numero):
    dig=only_digits(numero)
    if len(dig)<3:
        return ''
    return dig[-3:]

def _normaliza_telefone_destino(texto):
    num=wa_norm_number(texto)
    if not wa_is_valid_number(num):
        return ''
    return num

def _resposta_sim_nao(texto):
    t=(texto or '').strip().lower()
    if t in ('s','sim','ok','confirmo','confirma','isso','correto'):
        return True
    if t in ('n','nao','não','negativo'):
        return False
    return None

DEFAULT_IA_WA_PROMPT="""PERSONA
Você é o Rômulo, assistente virtual oficial da RM Facilities.
Responda sempre em português do Brasil, com tom cordial, objetivo e profissional.

REGRAS DE COMUNICAÇÃO
Formato WhatsApp: sempre uma linha em branco entre blocos.
Use negrito com um asterisco em cada lado da palavra.
Áudios entram como texto; a resposta deve ser sempre por texto.
Nunca invente dados, políticas, valores ou prazos.
Faça apenas uma pergunta por mensagem.
Não repita perguntas já respondidas.
Não reinicie o fluxo sem pedido explícito do usuário.

HORÁRIO
Atendimento humano: segunda a sexta, 08h00 às 18h00.
Fora do expediente: avise isso logo no início, mas continue a coleta normalmente.

COBERTURA SPOT
Para serviços pontuais/avulsos (SPOT), atendemos somente:

São José dos Campos
Jacareí
Caçapava

Se for SPOT fora dessas cidades: informar sem cobertura e encerrar apenas esse fluxo.
Se for serviço fixo/recorrente: seguir normalmente.

IDENTIFICAÇÃO POR TELEFONE (PRIORIDADE MÁXIMA)
Antes de qualquer triagem, verificar se o telefone do WhatsApp está vinculado a funcionário cadastrado.

Se estiver vinculado:

Não pedir nome.
Ir direto para atendimento de funcionário, perguntando o que deseja.
Para liberar documento, exigir CPF.
Pedir confirmação do telefone de destino mostrando apenas os 3 últimos dígitos.
Permitir informar outro número para envio após a validação do CPF.

Se não estiver vinculado:

Seguir triagem normal.

ESTADO DA CONVERSA
Estados possíveis:

TRIAGEM
CLIENTE_COLETA
FUNCIONARIO_ATENDIMENTO
FUNCIONARIO_DOC_CPF
CANDIDATO
FORNECEDOR
OUVIDORIA_COLETA
ENCERRADO

Regras:

Mostrar menu apenas quando estado for TRIAGEM.
Não repetir menu se já houver fluxo ativo.
Não voltar ao início sem comando do usuário.
Se o usuário enviar vários dados em uma mensagem, aproveite os válidos e pergunte apenas o próximo campo faltante.

TRIAGEM PADRÃO
Não pedir nome completo para clientes.
Quando não houver fluxo ativo:

Olá! Seja bem-vindo à RM Facilities. Como posso te ajudar hoje?

1 - Cliente — Solicitar orçamento
2 - Funcionário — Assuntos de RH ou operacional
3 - Candidato a vaga
4 - Fornecedor / Compras
5 - Ouvidoria

Se a primeira mensagem já indicar o assunto, iniciar direto no fluxo correspondente, sem repetir menu.

FLUXOS
1) CLIENTE (ORÇAMENTO)
Coletar, nesta ordem:

Nome
Tipo de serviço
Cidade (aplicar regra SPOT)
Escala ou metragem
CNPJ ou CPF
E-mail

2) FUNCIONÁRIO (RH/OPERACIONAL)
Se identificado por telefone:

Não pedir nome.
Perguntar diretamente o que deseja.
Se pedir documento, entrar em validação de CPF.
Liberar documento mesmo se o telefone não for o mesmo do cadastro após validar o CPF.

Se não identificado:

Pedir CPF primeiro como identificação principal.
Depois confirmar o telefone mostrando apenas os 3 últimos dígitos do cadastro.
Perguntar o assunto.
Pedir o detalhamento.

3) CANDIDATO
Responder com o e-mail de currículo: trabalheconosco@rmfacilities.com.br

4) FORNECEDOR / COMPRAS
Responder com o e-mail: compras@rmfacilities.com.br

5) OUVIDORIA
Pedir:

O que aconteceu
Quando ocorreu
Local/contrato
Como prefere retorno

SEGURANÇA PARA DOCUMENTOS (OBRIGATÓRIA)
Para holerite, contracheque e documentos:

Sempre pedir CPF completo antes de liberar.
Sem CPF validado, não enviar.
CPF divergente: negar envio e orientar contato com RH.
Mesmo com telefone identificado, CPF continua obrigatório.

Mensagem padrão:
Por segurança, confirme seu CPF completo (apenas números) para liberar o documento.

Telefone para envio:
Aceitar telefone com ou sem máscara, por exemplo: (12) 99775-2283 ou 12997752283.

CPF:
Aceitar CPF com ou sem máscara, por exemplo: 273.962.528-89 ou 27396252889.

MÚLTIPLOS DOCUMENTOS E ANO CORRENTE
Quando o usuário pedir mais de um documento no mesmo pedido, processar todos os meses solicitados.

Exemplos válidos:

maio e junho
05/2026 e 06/2026
maio, junho e julho

Regra obrigatória:

Se o mês vier sem ano, considerar automaticamente o ano corrente.
Ao final, informar quais meses foram enviados e quais não foram encontrados, se houver.

ENCERRAMENTO
Ao finalizar fluxos com coleta, gerar resumo técnico em tópicos e encerrar cordialmente."""

def wa_send_text(numero,mensagem):
    cfg=wa_cfg()
    if not cfg['url'] or not cfg['instancia']: raise ValueError('WhatsApp nao configurado')
    num=wa_norm_number(numero)
    if not wa_is_valid_number(num): raise ValueError(f'Numero WhatsApp invalido: {num or "vazio"}')
    url=f"{cfg['url'].rstrip('/')}/message/sendText/{cfg['instancia']}"
    data=json.dumps({'number':num,'text':mensagem}).encode()
    req=urllib.request.Request(url,data=data,headers={'Content-Type':'application/json','apikey':cfg['token']})
    try:
        with urllib.request.urlopen(req,timeout=15) as r: return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'WhatsApp API {e.code}: {detalhe or e.reason}')

def _fcm_send_to_token(token,titulo,corpo,data=None):
    """Envio opcional via FCM. Se firebase-admin nao estiver configurado, ignora com segurança."""
    if not token:
        return False
    cred_val=(os.environ.get('FIREBASE_CREDENTIALS_JSON') or '').strip()
    cred_b64=(os.environ.get('FIREBASE_CREDENTIALS_B64') or '').strip()
    cred_path=(os.environ.get('FIREBASE_CREDENTIALS_FILE') or '').strip()
    if not cred_val:
        if cred_b64:
            try:
                cred_val=base64.b64decode(cred_b64).decode('utf-8').strip()
            except Exception as e:
                app.logger.error(f'[fcm] FIREBASE_CREDENTIALS_B64 inválido: {e}')
                return False
        elif cred_path:
            cred_val=cred_path
        else:
            app.logger.warning('[fcm] FIREBASE_CREDENTIALS_JSON ausente; push ignorado')
            return False
    try:
        import firebase_admin
        from firebase_admin import credentials, messaging
    except Exception as e:
        app.logger.exception(f'[fcm] falha ao importar firebase_admin: {e}')
        return False
    try:
        firebase_admin.get_app()
    except Exception:
        try:
            # Aceita JSON inline (string) ou caminho de arquivo
            if cred_val.startswith('{'):
                import json as _json
                import ast as _ast
                import tempfile, atexit
                payload=cred_val
                # Remove aspas externas extras comuns em variáveis mal copiadas
                if (payload.startswith('"') and payload.endswith('"')) or (payload.startswith("'") and payload.endswith("'")):
                    payload=payload[1:-1].strip()
                try:
                    _json.loads(payload)
                except Exception:
                    # Suporta dicionário em formato Python com aspas simples
                    try:
                        parsed=_ast.literal_eval(payload)
                        if isinstance(parsed,dict):
                            payload=_json.dumps(parsed)
                        else:
                            raise ValueError('Formato de credencial inválido (não é objeto JSON).')
                    except Exception as pe:
                        raise ValueError(f'JSON de credencial inválido: {pe}')
                _tmp=tempfile.NamedTemporaryFile(mode='w',suffix='.json',delete=False)
                _tmp.write(payload); _tmp.flush(); _tmp.close()
                atexit.register(lambda p=_tmp.name: os.path.exists(p) and os.remove(p))
                firebase_admin.initialize_app(credentials.Certificate(_tmp.name))
            else:
                if not os.path.exists(cred_val):
                    app.logger.warning(f'[fcm] arquivo de credencial nao encontrado: {cred_val}')
                    return False
                firebase_admin.initialize_app(credentials.Certificate(cred_val))
        except Exception as e:
            app.logger.error(f'[fcm] falha ao inicializar firebase app: {e}')
            return False
    try:
        msg=messaging.Message(
            token=token,
            notification=messaging.Notification(title=titulo,body=corpo),
            data={k:str(v) for k,v in (data or {}).items()},
            android=messaging.AndroidConfig(
                priority='high',
                notification=messaging.AndroidNotification(
                    channel_id='rmf_documentos',
                    sound='default',
                    visibility=messaging.AndroidNotificationVisibility.PUBLIC,
                    default_vibrate_timings=True,
                    default_sound=True,
                    default_light_settings=True,
                )
            ),
            apns=messaging.APNSConfig(
                payload=messaging.APNSPayload(
                    aps=messaging.Aps(badge=1,sound='default')
                )
            ),
        )
        messaging.send(msg)
        return True
    except Exception as e:
        app.logger.exception(f'[fcm] falha ao enviar para token: {e}')
        return False

def _push_notify_funcionario(fid,titulo,corpo,data=None):
    f=Funcionario.query.get(fid)
    if not f or not (f.app_push_token or '').strip():
        if f:
            app.logger.info(f'[fcm] funcionario {fid} sem app_push_token salvo')
        return False
    token=f.app_push_token.strip()
    ok=_fcm_send_to_token(token,titulo,corpo,data=data)
    if ok:
        return True
    try:
        import firebase_admin
        from firebase_admin import messaging
        try:
            firebase_admin.get_app()
        except Exception:
            return False
        msg=messaging.Message(
            token=token,
            data={k:str(v) for k,v in (data or {}).items()},
        )
        messaging.send(msg)
        return True
    except Exception as e:
        msg=(str(e) or '').lower()
        app.logger.exception(f'[fcm] segunda tentativa falhou para funcionario {fid}: {e}')
        if (
            'unregistered' in msg or
            'registration-token-not-registered' in msg or
            'requested entity was not found' in msg
        ):
            f.app_push_token=None
            db.session.commit()
            app.logger.warning(f'[fcm] token invalido removido para funcionario {fid}')
        return False

def wa_media_meta(nome_arquivo,mimetype=''):
    mime=(mimetype or '').split(';')[0].strip().lower()
    if not mime:
        mime=(mimetypes.guess_type(nome_arquivo or '')[0] or '').strip().lower()
    if mime.startswith('image/'):
        return 'image',mime
    if mime.startswith('audio/'):
        return 'audio',mime
    if mime.startswith('video/'):
        return 'video',mime
    if not mime:
        mime='application/octet-stream'
    return 'document',mime

def wa_send_media_bytes(numero,arquivo_bytes,nome_arquivo,mimetype='',caption=''):
    cfg=wa_cfg()
    if not cfg['url'] or not cfg['instancia']: raise ValueError('WhatsApp nao configurado')
    num=wa_norm_number(numero)
    if not wa_is_valid_number(num): raise ValueError(f'Numero WhatsApp invalido: {num or "vazio"}')
    media_type,mime=wa_media_meta(nome_arquivo,mimetype)
    media_b64=base64.b64encode(arquivo_bytes).decode()
    url=f"{cfg['url'].rstrip('/')}/message/sendMedia/{cfg['instancia']}"
    data=json.dumps({'number':num,'mediatype':media_type,'mimetype':mime,'media':media_b64,'fileName':nome_arquivo,'caption':caption or nome_arquivo}).encode()
    req=urllib.request.Request(url,data=data,headers={'Content-Type':'application/json','apikey':cfg['token']})
    try:
        with urllib.request.urlopen(req,timeout=30) as r: return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'WhatsApp API {e.code}: {detalhe or e.reason}')

def wa_send_pdf(numero,caminho_abs,nome_arquivo,caption=''):
    with open(caminho_abs,'rb') as f:
        return wa_send_media_bytes(numero,f.read(),nome_arquivo,'application/pdf',caption)

def ai_wa_cfg():
    return {
        'enabled':gc('ia_wa_enabled','0'),
        'provider':(gc('ia_wa_provider','gemini') or 'gemini').strip().lower(),
        'api_key':gc('ia_wa_api_key',''),
        'model':gc('ia_wa_model',''),
        'prompt':gc('ia_wa_prompt',''),
        'temperature':gc('ia_wa_temperature','0.3'),
        'max_tokens':gc('ia_wa_max_tokens','350'),
    }

def ai_provider_norm(v):
    p=(v or 'gemini').strip().lower()
    if p in ('openai','chatgpt','gpt'): return 'openai'
    return 'gemini'

def ai_model_norm(provider,raw_model):
    p=ai_provider_norm(provider)
    s=(raw_model or '').strip()
    if not s:
        return 'gpt-4o-mini' if p=='openai' else 'gemini-1.5-flash'
    low=s.lower()
    if p=='gemini':
        m=re.search(r'models/([a-z0-9._-]+)',low)
        if m: return m.group(1)
        m=re.search(r'(gemini-[a-z0-9._-]+)',low)
        if m: return m.group(1)
        tok=re.split(r'[\s<>{}\[\]"\'\#?&:/]+',low)[0]
        tok=re.sub(r'[^a-z0-9._-]','',tok)
        return tok if tok.startswith('gemini-') else 'gemini-1.5-flash'
    tok=re.split(r'[\s<>{}\[\]"\'\#?&:/]+',s)[0]
    tok=re.sub(r'[^A-Za-z0-9._-]','',tok)
    if tok: return tok
    return 'gpt-4o-mini'

def ai_wa_enabled():
    return str(ai_wa_cfg().get('enabled','0')).strip().lower() in ('1','true','yes','on')

def wa_ai_pause_key(numero):
    n=wa_norm_number(numero)
    return f'wa_ai_pause_until_{n}' if n else ''

def wa_ai_pause_until(numero):
    k=wa_ai_pause_key(numero)
    if not k:
        return None
    return _parse_dt_iso(gc(k,''))

def wa_ai_pause_active(numero,ref=None):
    until=wa_ai_pause_until(numero)
    if not until:
        return False
    now=ref or utcnow()
    return until>now

def wa_ai_pause_for(numero,hours=8):
    n=wa_norm_number(numero)
    if not n:
        return None
    h=max(1,min(168,_to_int(hours,8)))
    until=utcnow()+timedelta(hours=h)
    sc_cfg(wa_ai_pause_key(n),until.isoformat())
    return until

def wa_ai_resume(numero):
    n=wa_norm_number(numero)
    if not n:
        return None
    until=utcnow()-timedelta(seconds=1)
    sc_cfg(wa_ai_pause_key(n),until.isoformat())
    return until

def wa_ai_pause_set(numero,hours=8):
    return wa_ai_pause_for(numero,hours)

def _post_json(url,payload,headers=None,timeout=30):
    h={'Content-Type':'application/json'}
    if headers: h.update(headers)
    req=urllib.request.Request(url,data=json.dumps(payload,ensure_ascii=False).encode('utf-8'),headers=h)
    try:
        with urllib.request.urlopen(req,timeout=timeout) as r:
            raw=r.read().decode('utf-8',errors='ignore')
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'IA API {e.code}: {detalhe or e.reason}')

def _post_multipart(url,fields=None,files=None,headers=None,timeout=60):
    boundary='----rmfacilities'+secrets.token_hex(12)
    body=bytearray()
    for k,v in (fields or {}).items():
        body.extend(f'--{boundary}\r\n'.encode())
        body.extend(f'Content-Disposition: form-data; name="{k}"\r\n\r\n'.encode())
        body.extend(str(v).encode('utf-8'))
        body.extend(b'\r\n')
    for f in (files or []):
        body.extend(f'--{boundary}\r\n'.encode())
        body.extend(
            f'Content-Disposition: form-data; name="{f.get("field","file")}"; filename="{f.get("filename","arquivo.bin")}"\r\n'.encode()
        )
        body.extend(f'Content-Type: {f.get("content_type","application/octet-stream")}\r\n\r\n'.encode())
        body.extend(f.get('data') or b'')
        body.extend(b'\r\n')
    body.extend(f'--{boundary}--\r\n'.encode())
    h={'Content-Type':f'multipart/form-data; boundary={boundary}'}
    if headers:
        h.update(headers)
    req=urllib.request.Request(url,data=bytes(body),headers=h)
    try:
        with urllib.request.urlopen(req,timeout=timeout) as r:
            raw=r.read().decode('utf-8',errors='ignore')
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detalhe=e.read().decode(errors='ignore')
        raise ValueError(f'IA API {e.code}: {detalhe or e.reason}')

def _http_read_bytes(url,headers=None,timeout=30):
    req=urllib.request.Request(url,headers=headers or {})
    with urllib.request.urlopen(req,timeout=timeout) as r:
        return r.read()

def _audio_ext_from_mime(mime):
    m=(mime or '').split(';')[0].strip().lower()
    return {
        'audio/ogg':'ogg',
        'audio/opus':'ogg',
        'audio/webm':'webm',
        'audio/mpeg':'mp3',
        'audio/mp3':'mp3',
        'audio/mp4':'m4a',
        'audio/x-m4a':'m4a',
        'audio/wav':'wav',
        'audio/x-wav':'wav',
        'audio/aac':'aac',
    }.get(m,'ogg')

def _maybe_b64decode(raw):
    s=str(raw or '').strip()
    if not s:
        return b''
    if ',' in s and ';base64' in s[:80].lower():
        s=s.split(',',1)[1]
    s=re.sub(r'\s+','',s)
    try:
        return base64.b64decode(s)
    except Exception:
        return b''

def _extract_audio_blob(msg_data):
    msg_obj=msg_data.get('message',{}) if isinstance(msg_data.get('message',{}),dict) else {}
    audio_obj=msg_obj.get('audioMessage',{}) if isinstance(msg_obj.get('audioMessage',{}),dict) else {}
    mime=(
        audio_obj.get('mimetype') or
        audio_obj.get('mimeType') or
        msg_data.get('mimetype') or
        msg_data.get('mimeType') or
        'audio/ogg'
    )
    filename=(
        audio_obj.get('fileName') or
        msg_data.get('fileName') or
        f'audio.{_audio_ext_from_mime(mime)}'
    )
    for candidate in (
        audio_obj.get('base64'),
        audio_obj.get('media'),
        audio_obj.get('data'),
        audio_obj.get('fileBase64'),
        msg_data.get('base64'),
        msg_data.get('media'),
        msg_data.get('data'),
    ):
        blob=_maybe_b64decode(candidate)
        if blob:
            return blob,mime,filename

    headers={}
    cfg=wa_cfg()
    if (cfg.get('token') or '').strip():
        headers['apikey']=cfg['token']
        headers['Authorization']=f'Bearer {cfg["token"]}'
    for media_url in (
        audio_obj.get('url'),
        audio_obj.get('mediaUrl'),
        audio_obj.get('directPath'),
        msg_data.get('mediaUrl'),
        msg_data.get('url'),
    ):
        u=str(media_url or '').strip()
        if not u or not u.startswith(('http://','https://')):
            continue
        try:
            blob=_http_read_bytes(u,headers=headers,timeout=45)
            if blob:
                return blob,mime,filename
        except Exception:
            continue
    return b'',mime,filename

def wa_transcribe_audio(msg_data):
    audio_bytes,mime,filename=_extract_audio_blob(msg_data)
    if not audio_bytes:
        return ''
    cfg=ai_wa_cfg()
    key=(cfg.get('api_key') or '').strip()
    if not key:
        return ''
    provider=ai_provider_norm(cfg.get('provider') or 'gemini')
    if key.startswith('AIza'):
        provider='gemini'
    elif key.startswith('sk-'):
        provider='openai'

    mime_base=(mime or 'audio/ogg').split(';')[0].strip().lower() or 'audio/ogg'
    if provider=='openai':
        url=(gc('ia_openai_audio_url','') or 'https://api.openai.com/v1/audio/transcriptions').strip()
        out=_post_multipart(
            url,
            fields={
                'model':'whisper-1',
                'language':'pt',
                'prompt':'Transcreva em portugues do Brasil.',
            },
            files=[{'field':'file','filename':filename,'content_type':mime_base,'data':audio_bytes}],
            headers={'Authorization':f'Bearer {key}'},
            timeout=90,
        )
        return str(out.get('text') or out.get('transcript') or '').strip()

    model=ai_model_norm('gemini',cfg.get('model') or '')
    base=(gc('ia_gemini_url','') or '').strip()
    url=base or f'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}'
    payload={
        'contents':[{ 
            'role':'user',
            'parts':[
                {'text':'Transcreva este audio em portugues do Brasil. Responda apenas com a transcricao, sem observacoes adicionais.'},
                {'inline_data':{'mime_type':mime_base,'data':base64.b64encode(audio_bytes).decode()}}
            ]
        }],
        'generationConfig':{'temperature':0.1,'maxOutputTokens':700}
    }
    out=_post_json(url,payload,timeout=90)
    cand=(out.get('candidates') or [{}])[0]
    parts=((cand.get('content') or {}).get('parts') or [])
    return '\n'.join((p.get('text') or '').strip() for p in parts if (p.get('text') or '').strip()).strip()

def _build_turns(hist,in_role,out_role):
    """Merge consecutive same-role messages into valid chat turns."""
    turns=[]
    for m in hist:
        role=in_role if m.direcao=='in' else out_role
        txt_m=(m.conteudo or '').strip()
        if not txt_m: continue
        if turns and turns[-1]['role']==role:
            turns[-1]['_text']+=('\n'+txt_m)
        else:
            turns.append({'role':role,'_text':txt_m})
    return turns

def _detecta_pedido_holerite(texto):
    """Detecta se o texto contém um pedido de holerite."""
    palavras=['holerite','hollerite','contracheque','contra-cheque','recibo','salario','salário','comprovante de renda','comprovante','vencimento','13o','décimo terceiro']
    txt_lower=texto.lower()
    return any(p in txt_lower for p in palavras)

def funcionario_docs_whatsapp_habilitado(funcionario):
    if not funcionario:
        return False
    # O envio de holerite/documentos para o próprio colaborador não deve depender de áreas.
    return True

def _parse_competencias_holerite(texto,ano_padrao=None):
    """Extrai uma ou mais competências (MM/YYYY) do texto.

    Regras:
    - aceita formatos numéricos (05/2026, 5-2026)
    - aceita nomes de meses (maio, junho)
    - quando o ano não é informado, usa o ano corrente
    """
    if ano_padrao is None:
        ano_padrao=localnow().year
    s=(texto or '').strip().lower()
    if not s:
        return []

    # Versao normalizada sem acentos/pontuacao para capturar textos com OCR irregular.
    s_norm=_norm_text_match(s)

    ano_global=None
    m_ano=re.search(r'\b(19\d{2}|20\d{2})\b',s)
    if m_ano:
        try:
            ano_global=int(m_ano.group(1))
        except Exception:
            ano_global=None

    meses={
        'janeiro':1,'jan':1,
        'fevereiro':2,'fev':2,
        'marco':3,'março':3,'mar':3,
        'abril':4,'abr':4,
        'maio':5,'mai':5,
        'junho':6,'jun':6,
        'julho':7,'jul':7,
        'agosto':8,'ago':8,
        'setembro':9,'set':9,
        'outubro':10,'out':10,
        'novembro':11,'nov':11,
        'dezembro':12,'dez':12,
    }

    out=[]
    seen=set()

    def add_comp(mes,ano):
        try:
            mi=int(mes)
            ai=int(ano)
        except Exception:
            return
        if not (1<=mi<=12 and 1900<=ai<=2099):
            return
        comp=f"{mi:02d}/{ai}"
        if comp not in seen:
            seen.add(comp)
            out.append(comp)

    for mm,yy in re.findall(r'\b(0?[1-9]|1[0-2])\s*[/-]\s*(\d{4})\b',s):
        add_comp(mm,yy)

    for mt in re.finditer(r'\b(janeiro|jan|fevereiro|fev|março|marco|mar|abril|abr|maio|mai|junho|jun|julho|jul|agosto|ago|setembro|set|outubro|out|novembro|nov|dezembro|dez)\b(?:\s*(?:de|/|-)?\s*(\d{4}))?',s):
        token=mt.group(1)
        ano_txt=mt.group(2)
        ano=int(ano_txt) if ano_txt else (ano_global or ano_padrao)
        add_comp(meses.get(token),ano)

    # Padrao em texto normalizado (sem acentos): "janeiro de 2025" etc.
    for mt in re.finditer(r'\b(janeiro|jan|fevereiro|fev|marco|mar|abril|abr|maio|mai|junho|jun|julho|jul|agosto|ago|setembro|set|outubro|out|novembro|nov|dezembro|dez)\b(?:\s*(?:de|/|-)?\s*(\d{4}))?',s_norm):
        token=mt.group(1)
        ano_txt=mt.group(2)
        ano=int(ano_txt) if ano_txt else (ano_global or ano_padrao)
        add_comp(meses.get(token),ano)

    # Mes sem ano so e aceito quando vier com marcador de competencia/mes.
    for mm in re.findall(r'(?:\bcompet[eê]ncia\b|\bm[eê]s\b|\bperiodo\b)\D{0,8}(0?[1-9]|1[0-2])\b',s):
        add_comp(mm,ano_global or ano_padrao)

    return out

def _competencia_mes_atual():
    now=localnow()
    return f"{now.month:02d}/{now.year}"

def _competencia_from_texto_nome(texto='',nome_arquivo=''):
    base=f"{texto or ''}\n{nome_arquivo or ''}"
    comps=_parse_competencias_holerite(base,ano_padrao=localnow().year)
    return comps[0] if comps else ''

def _resolver_competencia_envio(comp_in='',texto='',nome_arquivo=''):
    """Resolve competência e informa a origem usada.
    Origem: manual | texto_pdf | nome_arquivo | mes_atual
    """
    comp_manual=(comp_in or '').strip()
    if comp_manual:
        return comp_manual,'manual'

    comp_txt=''
    if (texto or '').strip():
        comp_txt=_parse_competencias_holerite(texto,ano_padrao=localnow().year)
        comp_txt=comp_txt[0] if comp_txt else ''
    if comp_txt:
        return comp_txt,'texto_pdf'

    comp_nome=''
    if (nome_arquivo or '').strip():
        comp_nome=_parse_competencias_holerite(nome_arquivo,ano_padrao=localnow().year)
        comp_nome=comp_nome[0] if comp_nome else ''
    if comp_nome:
        return comp_nome,'nome_arquivo'

    return _competencia_mes_atual(),'mes_atual'

def _nome_candidatos_holerite(page_text):
    """Extrai candidatos de nome em layouts comuns de holerite (ex.: Dominio)."""
    txt=str(page_text or '')
    if not txt.strip():
        return []
    out=[]
    seen=set()
    termos_ruido={
        'salario','rescisao','ferias','faltas','referencia','vencimentos','descontos','descricao','descricao',
        'codigo','funcionario','assinatura','liquido','inss','vale','transporte','adiantamento','desconto',
        'familia','insalubridade','provisao','prov','acumulo','funcao','normal','normais','dias','horas',
        'reflexo','noturno','dsr','cbo','departamento','filial','custo','cnpj','mensalista','folha'
    }

    def add_nome(raw):
        s=re.sub(r'\s+',' ',str(raw or '')).strip(' -:\t')
        # Remove sufixos numéricos comuns do layout (CBO, depto, filial etc.).
        s=re.sub(r'\s+\d[\d\s./-]*$','',s).strip()
        # Remove eventual código no início da linha.
        s=re.sub(r'^\d{1,8}\s+','',s).strip()
        if not s:
            return
        # Se ainda sobrou dígito no meio, é provável ruído de OCR/metadado.
        if any(ch.isdigit() for ch in s):
            return
        if len(s)<6:
            return
        partes=[p for p in _norm_text_match(s).split(' ') if p]
        if partes:
            hits_ruido=sum(1 for p in partes if p in termos_ruido)
            # Evita capturar linhas de verbas/descontos como se fossem nome.
            if hits_ruido>=2:
                return
        n=_norm_text_match(s)
        if n and n not in seen:
            seen.add(n)
            out.append(n)

    # Linha com codigo + nome (nome normalmente em caixa alta no holerite).
    for ln in txt.splitlines():
        m=re.match(r'^\s*\d{1,6}\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\'\.-]{3,120}?)(?:\s+\d.*)?\s*$',(ln or '').strip())
        if m:
            add_nome(m.group(1))

    # Extração em linha única/colunas coladas: "Código MARIA ... Nome do Funcionário"
    for m in re.finditer(r'c[oó]digo\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\'\.-]{4,140}?)(?=\s*nome\s+do\s+funcion[aá]rio\b)',txt,re.I):
        add_nome(m.group(1))
    # Variação comum quando o código numérico vem antes de "Código" no texto extraído.
    for m in re.finditer(r'\b(\d{1,8})\s*c[oó]digo\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\'\.-]{4,140}?)(?=\s*nome\s+do\s+funcion[aá]rio\b|\s+cbo\b|\s+departamento\b|\s+filial\b|$)',txt,re.I):
        add_nome(m.group(2))

    # Ancora em "Nome do Funcionário" e usa a proxima linha nao vazia.
    linhas=[(ln or '').strip() for ln in txt.splitlines()]
    for i,ln in enumerate(linhas):
        if re.search(r'nome\s+do\s+funcion[aá]rio',ln,re.I):
            for j in range(i+1,min(i+5,len(linhas))):
                cand=linhas[j]
                if not cand:
                    continue
                cand=re.sub(r'^\d{1,6}\s+','',cand)
                cand=re.sub(r'\s{2,}.*$','',cand)
                add_nome(cand)
                break

    # Fallback de topo: em alguns PDFs o marcador não vem íntegro,
    # mas o nome aparece nas primeiras linhas em caixa alta.
    topo=linhas[:28]
    stop_words=('folha mensal','mensalista','admissao','admissão','codigo','código','descricao','descrição',
                'vencimentos','descontos','referencia','referência','cbo','departamento','filial','cc geral','cnpj')
    for ln in topo:
        l=(ln or '').strip()
        if not l:
            continue
        low=l.lower()
        if any(sw in low for sw in stop_words):
            continue
        # remove metadados numéricos comuns
        l=re.sub(r'^\d{1,8}\s+','',l)
        l=re.sub(r'\s+\d[\d\s./-]*$','',l).strip()
        if not l:
            continue
        words=[w for w in re.split(r'\s+',l) if w]
        if len(words)<3:
            continue
        letters=[c for c in l if c.isalpha()]
        if not letters:
            continue
        upper_ratio=sum(1 for c in letters if c.isupper())/len(letters)
        if upper_ratio>=0.75:
            add_nome(l)
    return out

def _indicadores_pdf_funcionario(page_text):
    """Extrai indicadores úteis para matching em PDFs (CPF, matrícula/RE)."""
    txt=str(page_text or '')
    txt_norm=_norm_text_match(txt)
    txt_digits=only_digits(txt)

    cpfs=set(re.findall(r'\d{11}',txt_digits or ''))
    mats=set()
    for m in re.finditer(r'\b(?:matricula|matr|codigo|cod|re)\s*[:\-]?\s*(\d{1,8})\b',txt_norm):
        v=(m.group(1) or '').strip()
        if v:
            mats.add((v.lstrip('0') or '0'))

    # Só aceita número no início quando a linha parece realmente nome de funcionário.
    linhas=[(ln or '').strip() for ln in txt.splitlines()]
    stop_linha=('salario','rescisao','ferias','faltas','inss','vale','desconto','adiantamento','codigo descricao')
    for ln in linhas:
        if not ln:
            continue
        mm=re.match(r'^\s*(\d{1,8})\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\'\.-]{6,120})\s*$',ln)
        if not mm:
            continue
        nome_raw=(mm.group(2) or '').strip()
        nome_norm=_norm_text_match(nome_raw)
        if any(sw in nome_norm for sw in stop_linha):
            continue
        partes=[p for p in nome_norm.split(' ') if p]
        if len(partes)<2:
            continue
        mats.add((mm.group(1).lstrip('0') or '0'))

    return {'txt_norm':txt_norm,'txt_digits':txt_digits,'cpfs':cpfs,'mats':mats}

def _codigo_holerite_candidatos(page_text):
    """Extrai códigos de funcionário no bloco 'Código / Nome do Funcionário'."""
    txt=str(page_text or '')
    if not txt.strip():
        return []
    out=[]
    seen=set()
    linhas=[(ln or '').strip() for ln in txt.splitlines()]
    for i,ln in enumerate(linhas):
        if re.search(r'codigo\s+nome\s+do\s+funcion[aá]rio',ln,re.I):
            for j in range(i+1,min(i+7,len(linhas))):
                lj=linhas[j]
                if not lj:
                    continue
                m=re.match(r'^(\d{1,8})\b',lj)
                if not m:
                    # Alguns PDFs extraem em colunas com espaços iniciais estranhos.
                    m=re.search(r'\b(\d{1,8})\b',lj)
                if m:
                    code=(m.group(1).lstrip('0') or '0')
                    if code not in seen:
                        seen.add(code)
                        out.append(code)
                    break

    # Padrão de extração sem quebra de linha entre colunas.
    # Ex.: "... C.Custo: 54Código BENEDITO ... Nome do Funcionário ..."
    for m in re.finditer(r'(?<![\d,.])(\d{1,8})\s*c[oó]digo\s+[A-ZÀ-Ú]',txt,re.I):
        code=(m.group(1).lstrip('0') or '0')
        if code=='0':
            continue
        if code not in seen:
            seen.add(code)
            out.append(code)

    # Ex.: "Código 54 Nome do Funcionário" (variação por ordem de colunas)
    for m in re.finditer(r'c[oó]digo\s*(\d{1,8})\b(?:\s+nome\s+do\s+funcion[aá]rio)?',txt,re.I):
        code=(m.group(1).lstrip('0') or '0')
        if code=='0':
            continue
        if code not in seen:
            seen.add(code)
            out.append(code)
    return out

def _extract_pdf_page_text(page):
    """Extrai texto de uma página PDF com fallback para modo layout."""
    txt=''
    try:
        txt=(page.extract_text() or '')
    except Exception:
        txt=''
    txt=txt or ''

    # Fallback útil para PDFs com fonte/encoding ruim no modo padrão.
    txt_layout=''
    if len(txt.strip())<40:
        try:
            txt_layout=(page.extract_text(extraction_mode='layout') or '')
        except Exception:
            txt_layout=''

    if txt_layout:
        if txt and txt_layout!=txt:
            return (txt+'\n'+txt_layout).strip()
        return txt_layout.strip()
    return txt.strip()

def _extract_pdf_sample_text(reader,max_pages=5):
    partes=[]
    try:
        total=len(reader.pages)
    except Exception:
        total=0
    limite=min(max_pages,max(0,total))
    for i in range(limite):
        try:
            t=_extract_pdf_page_text(reader.pages[i])
        except Exception:
            t=''
        if t:
            partes.append(t)
    return ' '.join(partes)

def _extract_pdf_competencia_text(reader,max_pages=30):
    """Texto de amostra ampliado para detectar competência em PDFs longos."""
    partes=[]
    try:
        total=len(reader.pages)
    except Exception:
        total=0
    if total<=0:
        return ''
    # Captura início e fim do documento (muitos layouts repetem cabeçalho no fim).
    head=min(max_pages//2,total)
    tail=min(max_pages-head,max(0,total-head))
    idxs=list(range(head)) + list(range(max(0,total-tail),total))
    used=set()
    for i in idxs:
        if i in used:
            continue
        used.add(i)
        try:
            t=_extract_pdf_page_text(reader.pages[i])
        except Exception:
            t=''
        if t:
            partes.append(t)
    return ' '.join(partes)

def _processa_dialogo_holerite(conversa_id,numero,texto):
    """Processa o diálogo de busca e envio de holerite."""
    conversa=WhatsAppConversa.query.get(conversa_id)
    if not conversa:
        return None
    
    try:
        ctx=json.loads(conversa.contexto or '{}')
    except:
        ctx={}
    
    estado=ctx.get('holerite_estado')
    
    # Estado inicial: pedir CPF diretamente.
    if not estado:
        reiniciou=ctx.pop('_reiniciou_inatividade',False)
        funcionario_vinculado=_funcionario_por_whatsapp(numero)
        numero_destino=_normaliza_telefone_destino(numero)
        ctx['holerite_tentativas']=0
        ctx['holerite_funcionario_id']=(funcionario_vinculado.id if funcionario_vinculado else None)
        ctx['holerite_numero_envio']=numero_destino or only_digits(numero)
        ctx['holerite_telefone_final3']=_telefone_final_3(numero_destino or numero)
        ctx['holerite_estado']='aguardando_cpf'
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        aviso="Olá! O atendimento anterior foi encerrado por inatividade. Vamos iniciar um novo atendimento.\n\n" if reiniciou else ""
        return aviso+"Por segurança, confirme seu CPF completo (apenas números) para liberar o documento."

    if estado in ('aguardando_identificacao','escolhendo_funcionario'):
        ctx['holerite_estado']='aguardando_cpf'
        ctx['holerite_tentativas']=0
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Por segurança, confirme seu CPF completo (apenas números) para liberar o documento."

    # Validação de CPF antes de informar competência
    if estado=='aguardando_cpf':
        func_id=ctx.get('holerite_funcionario_id')
        funcionario=Funcionario.query.get(func_id) if func_id else None
        cpf_informado=only_digits(texto)
        if len(cpf_informado)!=11 or not _valida_cpf(cpf_informado):
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Não foi possível validar sua identidade após 3 tentativas. Envie 'holerite' para recomeçar."
            restantes=3-ctx['holerite_tentativas']
            return f"CPF inválido ou não reconhecido. Informe seu CPF completo com 11 dígitos. Você ainda tem {restantes} tentativa(s)."

        funcionario_cpf=_funcionario_por_cpf(cpf_informado)
        if funcionario and only_digits(funcionario.cpf or '')==cpf_informado and funcionario_cpf is None:
            funcionario_cpf=funcionario
        if funcionario_cpf:
            funcionario=funcionario_cpf
            ctx['holerite_funcionario_id']=funcionario.id

        cpf_cadastrado=only_digits(funcionario.cpf or '') if funcionario else ''
        if not funcionario or not cpf_cadastrado or cpf_informado!=cpf_cadastrado:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "CPF não confere com o cadastro. Por segurança, o atendimento foi encerrado após 3 tentativas. Envie 'holerite' para tentar novamente."
            restantes=3-ctx['holerite_tentativas']
            return f"CPF não confere com o cadastro. Tente novamente com o CPF correto. Você ainda tem {restantes} tentativa(s)."

        if not funcionario_docs_whatsapp_habilitado(funcionario):
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Seu cadastro não está habilitado para receber documentos pelo WhatsApp. Solicite ao RH a liberação de acesso a documentos."

        numero_destino=ctx.get('holerite_numero_envio') or _normaliza_telefone_destino(numero) or wa_norm_number(funcionario.telefone or '')
        final3=_telefone_final_3(numero_destino)
        if final3:
            ctx['holerite_estado']='confirmando_telefone'
            ctx['holerite_telefone_final3']=final3
            ctx['holerite_numero_envio']=numero_destino
            ctx['holerite_tentativas']=0
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return f"CPF validado com sucesso.\n\nPosso enviar para o número com final *{final3}*? Se preferir outro, digite o número completo com DDD."

        ctx['holerite_estado']='aguardando_competencia'
        ctx['holerite_tentativas']=0
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Identidade validada com sucesso.\n\nAgora informe o(s) mês(es) do holerite. Você pode pedir mais de um de uma vez (ex.: 05/2026 e 06/2026, ou maio e junho). Se não informar o ano, assumirei o ano corrente."

    if estado=='confirmando_telefone':
        confirmado=_resposta_sim_nao(texto)
        if confirmado is True:
            ctx['holerite_estado']='aguardando_competencia'
            ctx['holerite_tentativas']=0
            ctx['holerite_telefone_confirmado']='sim'
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Perfeito.\n\nAgora informe o(s) mês(es) do holerite. Você pode pedir mais de um de uma vez (ex.: 05/2026 e 06/2026, ou maio e junho). Se não informar o ano, assumirei o ano corrente."

        novo_numero=_normaliza_telefone_destino(texto)
        if confirmado is False and not novo_numero:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Não consegui confirmar o telefone de envio. Envie 'holerite' para recomeçar."
            return "Digite o número completo com DDD para envio do documento."

        if not novo_numero and confirmado is None:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Não consegui confirmar o telefone de envio. Envie 'holerite' para recomeçar."
            return "Responda SIM ou informe outro número com DDD para envio."

        ctx['holerite_estado']='aguardando_competencia'
        ctx['holerite_tentativas']=0
        ctx['holerite_telefone_confirmado']='sim' if confirmado else 'outro'
        ctx['holerite_numero_envio']=novo_numero or ctx.get('holerite_numero_envio') or _normaliza_telefone_destino(numero)
        conversa.contexto=json.dumps(ctx,ensure_ascii=False)
        db.session.commit()
        return "Número confirmado.\n\nAgora informe o(s) mês(es) do holerite. Você pode pedir mais de um de uma vez (ex.: 05/2026 e 06/2026, ou maio e junho). Se não informar o ano, assumirei o ano corrente."
    
    # Aguardando competência (mês/ano)
    if estado=='aguardando_competencia':
        competencias=_parse_competencias_holerite(texto,ano_padrao=localnow().year)
        if not competencias:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não consegui processar a data. Envie 'holerite' novamente para recomeçar."
            return "Formato inválido. Informe mês/ano (ex: 04/2026) ou nomes de meses (ex: maio e junho). Se não informar o ano, usarei o ano corrente."

        func_id=ctx.get('holerite_funcionario_id')
        funcionario=Funcionario.query.get(func_id) if func_id else None
        
        if not funcionario:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return "Desculpe, houve um erro ao buscar seus dados. Tente novamente."

        arquivos_por_comp=[]
        nao_encontrados=[]
        for competencia in competencias:
            arquivo=FuncionarioArquivo.query.filter(
                FuncionarioArquivo.funcionario_id==func_id,
                FuncionarioArquivo.categoria=='holerite',
                FuncionarioArquivo.competencia==competencia
            ).first()
            if arquivo:
                arquivos_por_comp.append((competencia,arquivo))
            else:
                nao_encontrados.append(competencia)

        if not arquivos_por_comp:
            ctx['holerite_tentativas']=ctx.get('holerite_tentativas',0)+1
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            if ctx['holerite_tentativas']>=3:
                ctx['holerite_estado']=None
                conversa.contexto=json.dumps(ctx,ensure_ascii=False)
                db.session.commit()
                return "Desculpe, não encontrei seu holerite. Envie 'holerite' para tentar novamente."
            if nao_encontrados:
                return f"Ops! Não encontrei holerite para: {', '.join(nao_encontrados)}. Tente outro mês ou verifique os dados."
            return "Ops! Não encontrei seu holerite para os meses informados."
        
        # Envia um ou mais holerites
        try:
            enviados=[]
            erros=[]
            numero_envio=_normaliza_telefone_destino(ctx.get('holerite_numero_envio') or numero)
            if not numero_envio:
                raise ValueError('Número de envio inválido.')
            for competencia,arquivo in arquivos_por_comp:
                caminho_abs=os.path.join(UPLOAD_ROOT,arquivo.caminho)
                if not os.path.exists(caminho_abs):
                    erros.append(f"{competencia} (arquivo indisponível)")
                    continue
                nome_arquivo=(arquivo.nome_arquivo or os.path.basename(caminho_abs) or f'holerite_{competencia.replace("/", "-")}.pdf')
                wa_send_pdf(numero_envio,caminho_abs,nome_arquivo,f"Holerite {competencia}")
                enviados.append((competencia,nome_arquivo))

            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            conversa.ultima_msg=utcnow()
            if enviados:
                db.session.add(WhatsAppMensagem(
                    conversa_id=conversa.id,
                    numero=numero,
                    direcao='out',
                    tipo='documento',
                    conteudo='Holerites enviados: '+', '.join([f"{c} ({n})" for c,n in enviados])
                ))
            db.session.commit()

            if not enviados:
                base='Não consegui enviar os holerites solicitados agora.'
                if erros:
                    base+=f" Erros: {', '.join(erros)}."
                if nao_encontrados:
                    base+=f" Não encontrados: {', '.join(nao_encontrados)}."
                return base

            comps_enviadas=', '.join([c for c,_ in enviados])
            msg=f"✅ Pronto! Enviei seu(s) holerite(s) de: {comps_enviadas}."
            if nao_encontrados:
                msg+=f"\n\nNão encontrei para: {', '.join(nao_encontrados)}."
            if erros:
                msg+=f"\n\nHouve falha no envio para: {', '.join(erros)}."
            return msg
        except Exception as e:
            ctx['holerite_estado']=None
            conversa.contexto=json.dumps(ctx,ensure_ascii=False)
            db.session.commit()
            return f"Desculpe, houve um erro ao enviar seu holerite. Tente novamente mais tarde."
    
    # Padrão: reinicia o diálogo
    ctx['holerite_estado']=None
    conversa.contexto=json.dumps(ctx,ensure_ascii=False)
    db.session.commit()
    return None

def ai_wa_reply(numero,texto,historico=None):
    txt=(texto or '').strip()
    if not txt: return ''
    
    # Verifica se é um pedido de holerite
    conversa=WhatsAppConversa.query.filter_by(numero=numero).first()
    if conversa and _detecta_pedido_holerite(txt):
        resposta_holerite=_processa_dialogo_holerite(conversa.id,numero,txt)
        if resposta_holerite is not None:
            return resposta_holerite
    elif conversa:
        # Verifica se está em um diálogo de holerite já em progresso
        try:
            ctx=json.loads(conversa.contexto or '{}')
            if ctx.get('holerite_estado'):
                resposta_holerite=_processa_dialogo_holerite(conversa.id,numero,txt)
                if resposta_holerite is not None:
                    return resposta_holerite
        except:
            pass
    
    cfg=ai_wa_cfg()
    key=(cfg.get('api_key') or '').strip()
    if not key: raise ValueError('API Key da IA nao configurada')
    provider=ai_provider_norm(cfg.get('provider') or 'gemini')
    model=ai_model_norm(provider,cfg.get('model') or '')
    system=(cfg.get('prompt') or '').strip() or DEFAULT_IA_WA_PROMPT
    # Verifica reinício por inatividade para personalizar saudação na resposta da IA.
    if conversa:
        try:
            ctx_ia=json.loads(conversa.contexto or '{}')
            if ctx_ia.pop('_reiniciou_inatividade',False):
                conversa.contexto=json.dumps(ctx_ia,ensure_ascii=False)
                db.session.commit()
                system+='\n\nINSTRUÇÃO: O atendimento anterior encerrou por mais de 2 horas de inatividade. Inicie sua resposta cumprimentando o usuário e informando que um novo atendimento foi iniciado.'
        except Exception:
            pass
    funcionario_vinculado=_funcionario_por_whatsapp(numero)
    if funcionario_vinculado:
        final3=_telefone_final_3(funcionario_vinculado.telefone or numero)
        extra='\n\nCONTEXTO INTERNO AUTOMÁTICO\nO telefone atual está vinculado a um funcionário cadastrado. Não peça nome. Se o assunto for documento, peça CPF antes de liberar qualquer arquivo. O documento pode ser enviado para qualquer número confirmado pelo usuário após a validação do CPF.'
        if final3:
            extra+=f' Ao confirmar telefone, mostre somente os 3 últimos dígitos: {final3}.'
        system+=extra
    try: temp=float(cfg.get('temperature') or 0.3)
    except Exception: temp=0.3
    try: max_tk=int(float(cfg.get('max_tokens') or 350))
    except Exception: max_tk=350
    hist=[m for m in (historico or []) if (m.conteudo or '').strip() and m.tipo!='erro']
    if provider=='openai':
        if key.startswith('AIza'):
            raise ValueError('API Key parece ser do Gemini, mas o provedor selecionado é OpenAI.')
        mdl=model or 'gpt-4o-mini'
        url=(gc('ia_openai_url','') or 'https://api.openai.com/v1/chat/completions').strip()
        if hist:
            turns=_build_turns(hist,'user','assistant')
            msgs=[{'role':'system','content':system}]+[{'role':t['role'],'content':t['_text']} for t in turns]
        else:
            msgs=[{'role':'system','content':system},{'role':'user','content':f'Número: {numero}\nMensagem: {txt}'}]
        payload={'model':mdl,'messages':msgs,'temperature':temp,'max_tokens':max_tk}
        out=_post_json(url,payload,headers={'Authorization':f'Bearer {key}'},timeout=45)
        msg=((out.get('choices') or [{}])[0].get('message') or {})
        resp=msg.get('content') or ''
        if isinstance(resp,list):
            # Newer OpenAI payloads may return content blocks instead of a plain string.
            txt_blocks=[]
            for b in resp:
                if isinstance(b,dict):
                    t=(b.get('text') or '').strip()
                    if t:
                        txt_blocks.append(t)
            resp='\n'.join(txt_blocks)
        resp=str(resp).strip()
        if resp:
            return resp
        # Fallback sem historico quando o provider retorna escolha vazia.
        payload_fb={
            'model':mdl,
            'messages':[{'role':'system','content':system},{'role':'user','content':f'Número: {numero}\nMensagem: {txt}'}],
            'temperature':temp,
            'max_tokens':max_tk
        }
        out_fb=_post_json(url,payload_fb,headers={'Authorization':f'Bearer {key}'},timeout=45)
        msg_fb=((out_fb.get('choices') or [{}])[0].get('message') or {})
        resp_fb=msg_fb.get('content') or ''
        if isinstance(resp_fb,list):
            txt_blocks=[]
            for b in resp_fb:
                if isinstance(b,dict):
                    t=(b.get('text') or '').strip()
                    if t:
                        txt_blocks.append(t)
            resp_fb='\n'.join(txt_blocks)
        resp_fb=str(resp_fb).strip()
        if resp_fb:
            return resp_fb
        raise ValueError('IA retornou resposta vazia (OpenAI). Verifique modelo/chave e tente novamente.')
    if provider=='gemini':
        if key.startswith('sk-'):
            raise ValueError('API Key parece ser da OpenAI, mas o provedor selecionado é Gemini.')
        mdl=model or 'gemini-1.5-flash'
        base=(gc('ia_gemini_url','') or '').strip()
        url=base or f'https://generativelanguage.googleapis.com/v1beta/models/{mdl}:generateContent?key={key}'
        if hist:
            turns=_build_turns(hist,'user','model')
            while turns and turns[0]['role']=='model': turns.pop(0)
            contents=[{'role':t['role'],'parts':[{'text':t['_text']}]} for t in turns] if turns else [{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}]
        else:
            contents=[{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}]
        payload={'system_instruction':{'parts':[{'text':system}]},'contents':contents,'generationConfig':{'temperature':temp,'maxOutputTokens':max_tk}}
        out=_post_json(url,payload,timeout=45)
        cand=(out.get('candidates') or [{}])[0]
        parts=((cand.get('content') or {}).get('parts') or [])
        resp='\n'.join((p.get('text') or '').strip() for p in parts if (p.get('text') or '').strip())
        resp=resp.strip()
        if resp:
            return resp
        # Fallback sem historico quando Gemini nao devolve parts/texto.
        payload_fb={
            'system_instruction':{'parts':[{'text':system}]},
            'contents':[{'role':'user','parts':[{'text':f'Número: {numero}\nMensagem: {txt}'}]}],
            'generationConfig':{'temperature':temp,'maxOutputTokens':max_tk}
        }
        out_fb=_post_json(url,payload_fb,timeout=45)
        cand_fb=(out_fb.get('candidates') or [{}])[0]
        parts_fb=((cand_fb.get('content') or {}).get('parts') or [])
        resp_fb='\n'.join((p.get('text') or '').strip() for p in parts_fb if (p.get('text') or '').strip()).strip()
        if resp_fb:
            return resp_fb
        finish=(cand_fb.get('finishReason') or cand.get('finishReason') or '').strip()
        if finish:
            raise ValueError(f'IA retornou resposta vazia (Gemini: {finish}).')
        raise ValueError('IA retornou resposta vazia (Gemini). Verifique modelo/chave e tente novamente.')
    raise ValueError('Provedor de IA inválido. Use gemini ou openai.')

def smtp_send_pdf(dest,nome_dest,caminho_abs,nome_arquivo,competencia='',remetente='RM Facilities'):
    cfg=smtp_cfg()
    if not cfg['host'] or not cfg['user']: raise ValueError('SMTP nao configurado')
    msg=MIMEMultipart()
    msg['From']=f"{remetente} <{cfg['de'] or cfg['user']}>"
    msg['To']=dest; msg['Subject']=f"Holerite {competencia} - {remetente}"
    corpo=f"Ol\u00e1 {nome_dest},\n\nSegue em anexo seu holerite{(' de '+competencia) if competencia else ''}.\n\nAtenciosamente,\n{remetente}"
    msg.attach(MIMEText(corpo,'plain','utf-8'))
    with open(caminho_abs,'rb') as f:
        part=MIMEBase('application','octet-stream'); part.set_payload(f.read())
    encoders.encode_base64(part); part.add_header('Content-Disposition',f'attachment; filename="{nome_arquivo}"'); msg.attach(part)
    port=int(cfg['port'] or 587)
    if str(cfg['tls']) in ('1','true','True','yes'):
        with smtplib.SMTP(cfg['host'],port,timeout=20) as s: s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
    else:
        with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s: s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())


def smtp_send_link_assinatura(dest, nome_dest, titulo_envelope, link, remetente='RM Facilities', eh_lembrete=False):
    """Envia link de assinatura de documento por e-mail."""
    cfg=smtp_cfg()
    if not cfg['host'] or not cfg['user']: raise ValueError('SMTP não configurado')
    msg=MIMEMultipart('alternative')
    msg['From']=f"{remetente} <{cfg['de'] or cfg['user']}>"
    msg['To']=dest
    assunto_prefixo='Lembrete: ' if eh_lembrete else ''
    texto_intro='Este e um lembrete de um envio anterior.\n\n' if eh_lembrete else ''
    html_intro=(
        "<p style='color:#8a5c00;font-size:13px;font-weight:700;margin-top:8px'>"
        "🔔 Este e um lembrete de um envio anterior.</p>"
        if eh_lembrete else ''
    )
    msg['Subject']=f"{assunto_prefixo}Documento para assinar — {titulo_envelope}"
    corpo_txt=(
        f"Olá {nome_dest},\n\n"
        f"{texto_intro}"
        f"Você recebeu um documento para assinar eletronicamente.\n\n"
        f"Documento: {titulo_envelope}\n\n"
        f"Acesse o link abaixo para assinar:\n{link}\n\n"
        f"Este link é exclusivo para você. Não compartilhe.\n\n"
        f"Atenciosamente,\n{remetente}"
    )
    corpo_html=(
        f"<div style='font-family:Arial,sans-serif;max-width:520px;margin:0 auto'>"
        f"<div style='background:#205d8a;padding:18px 24px;border-radius:8px 8px 0 0'>"
        f"<span style='color:#fff;font-size:18px;font-weight:700'>✍ {remetente}</span></div>"
        f"<div style='background:#fff;border:1px solid #dde5f0;border-top:none;padding:24px;border-radius:0 0 8px 8px'>"
        f"<p style='color:#333;font-size:15px'>Olá <strong>{nome_dest}</strong>,</p>"
        f"{html_intro}"
        f"<p style='color:#555;font-size:14px;margin-top:10px'>Você recebeu um documento para assinar eletronicamente.</p>"
        f"<div style='background:#f5f9ff;border:1px solid #c5d9f0;border-radius:8px;padding:14px;margin:18px 0'>"
        f"<span style='font-size:13px;color:#205d8a;font-weight:600'>📄 Documento:</span><br>"
        f"<span style='font-size:15px;color:#1a2b3c;font-weight:700'>{titulo_envelope}</span></div>"
        f"<div style='text-align:center;margin:22px 0'>"
        f"<a href='{link}' style='background:#205d8a;color:#fff;text-decoration:none;padding:13px 32px;"
        f"border-radius:8px;font-size:15px;font-weight:600;display:inline-block'>Assinar documento agora</a></div>"
        f"<p style='color:#999;font-size:11px;margin-top:18px'>Este link é exclusivo para você. Não compartilhe.<br>"
        f"Link: <a href='{link}' style='color:#205d8a'>{link}</a></p>"
        f"</div></div>"
    )
    msg.attach(MIMEText(corpo_txt,'plain','utf-8'))
    msg.attach(MIMEText(corpo_html,'html','utf-8'))
    port=int(cfg['port'] or 587)
    if str(cfg['tls']) in ('1','true','True','yes'):
        with smtplib.SMTP(cfg['host'],port,timeout=20) as s: s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
    else:
        with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s: s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())

ALLOWED_AREAS=['dashboard','medicoes','historico','clientes','empresas','usuarios','config','rh','operacional','compras','sst','rh-digital','documentos']
ALLOWED_ACTIONS=['view','create','edit','delete','approve','export']
DOC_CAT_PATH={
    'aso':'aso',
    'atestado':'atestado',
    'epi':'epi',
    'treinamento':'treinamento',
    'holerite':'holerites',
    'folha_ponto':'folha_ponto',
    'recibo_ferias':'recibo_ferias',
    'contrato_trabalho':'contrato_trabalho',
    'vale_transporte':'vale_transporte',
    'requisicao_vale_transporte':'requisicao_vale_transporte',
    'uniforme':'uniforme',
    'outros':'outros',
}
DOC_CAT_LABEL={
    'aso':'ASO',
    'atestado':'Atestado',
    'epi':'EPI',
    'treinamento':'Treinamento',
    'holerite':'Holerite',
    'folha_ponto':'Folha de Ponto',
    'recibo_ferias':'Recibo de Ferias',
    'contrato_trabalho':'Contrato de Trabalho',
    'vale_transporte':'Vale Transporte',
    'requisicao_vale_transporte':'Requisicao de Vale Transporte',
    'uniforme':'Uniforme/Fardamento',
    'outros':'Outros',
}

def jloads(v,dv):
    try: return json.loads(v) if v else dv
    except: return dv

def to_num(s,dec=False):
    if s is None:
        return 0.0 if dec else 0
    t=str(s).strip().replace(' ','')
    if not t:
        return 0.0 if dec else 0
    neg=t.startswith('-')
    if neg:
        t=t[1:]
    if ',' in t and '.' in t:
        if t.rfind(',')>t.rfind('.'):
            t=t.replace('.','').replace(',','.')
        else:
            t=t.replace(',','')
    elif ',' in t:
        t=t.replace('.','').replace(',','.')
    else:
        t=t.replace(',','')
    try:
        v=float(t)
        if neg:
            v=-v
        return v if dec else int(v)
    except Exception:
        return 0.0 if dec else 0

def norm_competencia(v=''):
    s=str(v or '').strip()
    if not s:
        return localnow().strftime('%Y-%m')
    m=re.search(r'^(\d{4})[-/](\d{2})$',s)
    if m:
        y,mm=m.group(1),m.group(2)
        if 1<=int(mm)<=12:
            return f'{y}-{mm}'
    m=re.search(r'^(\d{2})[-/](\d{4})$',s)
    if m:
        mm,y=m.group(1),m.group(2)
        if 1<=int(mm)<=12:
            return f'{y}-{mm}'
    return localnow().strftime('%Y-%m')

def next_cli_num():
    max_n=0
    for (n,) in db.session.query(Cliente.numero).all():
        try: max_n=max(max_n,int(str(n or '').strip()))
        except: pass
    return str(max_n+1).zfill(3)

def save_upload(fs,subdir):
    os.makedirs(os.path.join(UPLOAD_ROOT,subdir),exist_ok=True)
    base=secure_filename(fs.filename or 'arquivo.bin')
    nome=f"{localnow().strftime('%Y%m%d_%H%M%S')}_{base}"
    rel=os.path.join(subdir,nome)
    abs_p=os.path.join(UPLOAD_ROOT,rel)
    fs.save(abs_p)
    return rel,abs_p

def norm_cat(v):
    s=(v or 'outros').strip().lower().replace('-', '_').replace(' ', '_')
    aliases={
        'holerites':'holerite',
        'recibo_de_ferias':'recibo_ferias',
        'recibo_ferias':'recibo_ferias',
        'ficha_de_epi':'epi',
        'fardamento':'uniforme',
        'figurino':'uniforme',
        'req_vt':'requisicao_vale_transporte',
        'requisicao_vt':'requisicao_vale_transporte',
        'req_vale_transporte':'requisicao_vale_transporte',
    }
    s=aliases.get(s,s)
    return s if s in DOC_CAT_PATH else 'outros'

def _norm_text_match(v):
    s=str(v or '').strip().lower()
    if not s:
        return ''
    s=''.join(ch for ch in unicodedata.normalize('NFKD',s) if not unicodedata.combining(ch))
    s=re.sub(r'[^a-z0-9]+',' ',s)
    return re.sub(r'\s+',' ',s).strip()

def _match_conf_level(score):
    try:
        s=int(score or 0)
    except Exception:
        s=0
    if s>=100:
        return 'alta'
    if s>=70:
        return 'media'
    return 'baixa'

def _rank_funcionarios_in_text(page_text,funcs,limit=5):
    from difflib import SequenceMatcher

    indic=_indicadores_pdf_funcionario(page_text)
    txt_norm=indic.get('txt_norm') or ''
    txt_digits=indic.get('txt_digits') or ''
    cpfs_pdf=indic.get('cpfs') or set()
    mats_pdf=indic.get('mats') or set()
    cods_holerite=_codigo_holerite_candidatos(page_text)
    nome_cands=_nome_candidatos_holerite(page_text)
    txt_pad=f' {txt_norm} '
    ranking=[]

    for f in funcs:
        nome_norm=_norm_text_match(getattr(f,'nome',None))
        if not nome_norm:
            continue
        score=0
        motivos=[]

        mat_f=(only_digits(getattr(f,'matricula',None)).lstrip('0') or '0') if only_digits(getattr(f,'matricula',None)) else ''
        re_f=(only_digits(getattr(f,'re',None)).lstrip('0') or '0') if only_digits(getattr(f,'re',None)) else ''
        cpf_f=only_digits(getattr(f,'cpf',None))

        if cods_holerite and ((mat_f and mat_f in cods_holerite) or (re_f and re_f in cods_holerite)):
            score+=180
            motivos.append('codigo_holerite')
        elif mats_pdf and ((mat_f and mat_f in mats_pdf) or (re_f and re_f in mats_pdf)):
            score+=120
            motivos.append('codigo_pdf')

        if f' {nome_norm} ' in txt_pad:
            score+=100
            motivos.append('nome_exato')
        else:
            partes=[p for p in nome_norm.split(' ') if len(p)>=3]
            if len(partes)>=2:
                if f' {partes[0]} ' in txt_pad and f' {partes[-1]} ' in txt_pad:
                    score+=60
                    motivos.append('primeiro_ultimo_nome')
                hits=sum(1 for p in partes if f' {p} ' in txt_pad)
                if hits:
                    score+=min(30,hits*10)
                    motivos.append(f'partes_nome:{hits}')
            elif len(partes)==1 and f' {partes[0]} ' in txt_pad:
                score+=25
                motivos.append('nome_unico')

        for cand in nome_cands:
            if cand==nome_norm:
                score+=110
                motivos.append('candidato_nome_exato')
                break
            ratio_full=SequenceMatcher(None,nome_norm,cand).ratio()
            if ratio_full>=0.94:
                score+=95
                motivos.append(f'fuzzy:{ratio_full:.2f}')
                break
            if ratio_full>=0.88:
                score+=65
                motivos.append(f'fuzzy:{ratio_full:.2f}')
                break
            nome_parts=[p for p in nome_norm.split(' ') if len(p)>=3]
            if nome_parts:
                hits=sum(1 for p in nome_parts if p in cand.split(' '))
                ratio=hits/max(1,len(nome_parts))
                if ratio>=0.75:
                    score+=75
                    motivos.append(f'partes_candidato:{ratio:.2f}')
                    break
                if ratio>=0.5:
                    score+=40
                    motivos.append(f'partes_candidato:{ratio:.2f}')
                    break

        if mat_f and len(mat_f)>=4 and mat_f in txt_digits:
            score+=25
            motivos.append('matricula_texto')
        if re_f and len(re_f)>=4 and re_f in txt_digits:
            score+=20
            motivos.append('re_texto')
        if len(cpf_f)==11:
            if cpf_f in cpfs_pdf:
                score+=140
                motivos.append('cpf_exato')
            elif cpf_f in txt_digits:
                score+=80
                motivos.append('cpf_texto')

        ranking.append({
            'funcionario_id':getattr(f,'id',None),
            'nome':getattr(f,'nome','') or '',
            'matricula':getattr(f,'matricula','') or '',
            're':getattr(f,'re',None),
            'score':int(score),
            'motivos':motivos,
        })

    ranking.sort(key=lambda x:(-int(x.get('score') or 0),(x.get('nome') or '').lower()))
    return {
        'indicadores':{
            'codigos_holerite':sorted(list(cods_holerite)),
            'codigos_pdf':sorted(list(mats_pdf)),
            'cpfs':sorted(list(cpfs_pdf))[:3],
            'nomes_candidatos':nome_cands[:5],
            'texto_vazio':(not txt_norm),
        },
        'top':ranking[:max(1,int(limit or 5))]
    }

def find_funcionario_in_text(page_text,funcs,return_meta=False):
    from difflib import SequenceMatcher

    indic=_indicadores_pdf_funcionario(page_text)
    txt_norm=indic.get('txt_norm') or ''
    txt_digits=indic.get('txt_digits') or ''
    cpfs_pdf=indic.get('cpfs') or set()
    mats_pdf=indic.get('mats') or set()
    cods_holerite=_codigo_holerite_candidatos(page_text)
    nome_cands=_nome_candidatos_holerite(page_text)
    meta={'score':0,'second_score':0,'confianca':'baixa'}
    if not txt_norm:
        return (None,meta) if return_meta else None

    # Atalho mais forte: código identificado no cabeçalho padrão de holerite.
    if cods_holerite:
        candidatos=[]
        for f in funcs:
            mat_f=(only_digits(getattr(f,'matricula',None)).lstrip('0') or '0') if only_digits(getattr(f,'matricula',None)) else ''
            re_f=(only_digits(getattr(f,'re',None)).lstrip('0') or '0') if only_digits(getattr(f,'re',None)) else ''
            if (mat_f and mat_f in cods_holerite) or (re_f and re_f in cods_holerite):
                candidatos.append(f)
        if len(candidatos)==1:
            meta={'score':230,'second_score':0,'confianca':'alta'}
            return (candidatos[0],meta) if return_meta else candidatos[0]

    # Atalho determinístico: RE/matricula único encontrado na página.
    if mats_pdf:
        candidatos=[]
        for f in funcs:
            mat_f=(only_digits(getattr(f,'matricula',None)).lstrip('0') or '0') if only_digits(getattr(f,'matricula',None)) else ''
            re_f=(only_digits(getattr(f,'re',None)).lstrip('0') or '0') if only_digits(getattr(f,'re',None)) else ''
            if (mat_f and mat_f in mats_pdf) or (re_f and re_f in mats_pdf):
                candidatos.append(f)
        # Se houver apenas um candidato por código, prioriza identificação dessa página.
        if len(candidatos)==1:
            meta={'score':180,'second_score':0,'confianca':'alta'}
            return (candidatos[0],meta) if return_meta else candidatos[0]

    txt_pad=f' {txt_norm} '
    best=None
    best_score=0
    second_score=0
    for f in funcs:
        nome_norm=_norm_text_match(getattr(f,'nome',None))
        if not nome_norm:
            continue
        score=0
        if f' {nome_norm} ' in txt_pad:
            score+=100
        else:
            partes=[p for p in nome_norm.split(' ') if len(p)>=3]
            if len(partes)>=2:
                if f' {partes[0]} ' in txt_pad and f' {partes[-1]} ' in txt_pad:
                    score+=60
                hits=sum(1 for p in partes if f' {p} ' in txt_pad)
                score+=min(30,hits*10)
            elif len(partes)==1 and f' {partes[0]} ' in txt_pad:
                score+=25
        if nome_cands:
            for cand in nome_cands:
                if cand==nome_norm:
                    score+=110
                    break
                ratio_full=SequenceMatcher(None,nome_norm,cand).ratio()
                if ratio_full>=0.94:
                    score+=95
                    break
                if ratio_full>=0.88:
                    score+=65
                    break
                nome_parts=[p for p in nome_norm.split(' ') if len(p)>=3]
                if not nome_parts:
                    continue
                hits=sum(1 for p in nome_parts if p in cand.split(' '))
                ratio=hits/max(1,len(nome_parts))
                if ratio>=0.75:
                    score+=75
                    break
                if ratio>=0.5:
                    score+=40
                    break
        mat=only_digits(getattr(f,'matricula',None))
        if len(mat)>=4 and mat in txt_digits:
            score+=25
        if mat:
            mat_n=(mat.lstrip('0') or '0')
            if mat_n in mats_pdf:
                score+=55
        re_num=only_digits(getattr(f,'re',None))
        if len(re_num)>=4 and re_num in txt_digits:
            score+=20
        if re_num:
            re_n=(re_num.lstrip('0') or '0')
            if re_n in mats_pdf:
                score+=50
        cpf=only_digits(getattr(f,'cpf',None))
        if len(cpf)==11:
            if cpf in cpfs_pdf:
                score+=140
            elif cpf in txt_digits:
                score+=80
        if score>best_score:
            second_score=best_score
            best_score=score
            best=f
        elif score>second_score:
            second_score=score
    meta={'score':best_score,'second_score':second_score,'confianca':_match_conf_level(best_score)}
    # Regra mais permissiva para reduzir falsos negativos em holerite de lote.
    if best_score<32 and not (best_score>=24 and (best_score-second_score)>=8):
        return (None,meta) if return_meta else None
    return (best,meta) if return_meta else best

def infer_doc_year(comp=''):
    c=(comp or '').strip()
    m=re.search(r'(19|20)\d{2}',c)
    if m: return m.group(0)
    return str(localnow().year)

def func_doc_subdir(funcionario_id,categoria,competencia=''):
    cat=norm_cat(categoria)
    pasta_cat=DOC_CAT_PATH.get(cat,'outros')
    ano=infer_doc_year(competencia)
    return os.path.join('funcionarios',str(funcionario_id),pasta_cat,ano),cat

def prepare_func_doc_dirs(funcionario_id,ano=None):
    y=str(ano or localnow().year)
    made=[]
    for _,pasta in DOC_CAT_PATH.items():
        rel=os.path.join('funcionarios',str(funcionario_id),pasta,y)
        ap=os.path.join(UPLOAD_ROOT,rel)
        os.makedirs(ap,exist_ok=True)
        made.append(rel)
    return made

def holerite_comp_label(comp=''):
    c=(comp or '').strip()
    if not c:
        return localnow().strftime('%m-%Y')
    m=re.match(r'^(\d{4})[-/](\d{2})$',c)
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    m=re.match(r'^(\d{2})[-/](\d{4})$',c)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return re.sub(r'[\\/]+','-',c)

def _clean_file_part(v,max_len=80,fallback='item'):
    s=str(v or '').strip()
    s=re.sub(r'[\\/:*?"<>|]+','-',s)
    s=re.sub(r'\s+',' ',s).strip(' .-_')
    if not s:
        s=fallback
    return s[:max_len]

def holerite_batch_filename(funcionario,competencia=''):
    nome=_clean_file_part(getattr(funcionario,'nome','') or '',80,'Colaborador')
    matricula=(getattr(funcionario,'matricula',None) or '').strip()
    if not matricula:
        re_val=getattr(funcionario,'re',None)
        matricula=str(re_val) if re_val not in [None,''] else str(getattr(funcionario,'id','SEM-MAT'))
    mat=_clean_file_part(matricula,30,'SEM-MAT')
    comp=holerite_comp_label(competencia)
    return f"{nome} - {mat} - {comp}.pdf"

def unique_rel_filename(subdir,filename):
    name=str(filename or '').strip() or 'arquivo.pdf'
    base,ext=os.path.splitext(name)
    if not ext:
        ext='.pdf'
    idx=1
    cand=f"{base}{ext}"
    while True:
        rel=os.path.join(subdir,cand)
        abs_p=os.path.join(UPLOAD_ROOT,rel)
        if not os.path.exists(abs_p):
            return rel,abs_p,cand
        idx+=1
        cand=f"{base} ({idx}){ext}"

def arq_year_from_path(caminho):
    p=[x for x in str(caminho or '').split(os.sep) if x]
    if p and re.fullmatch(r'(19|20)\d{2}',p[-2] if len(p)>=2 else ''):
        return p[-2]
    m=re.search(r'/(19|20)\d{2}/',('/'+str(caminho or '').replace('\\','/')+'/'))
    return m.group(0).strip('/') if m else str(localnow().year)

def is_owner_user():
    perfil=(session.get('perfil') or '').strip().lower()
    if perfil=='dono':
        return True
    uid=session.get('uid')
    if not uid:
        return False
    try:
        u=Usuario.query.get(int(uid))
        if u and (u.perfil or '').strip().lower()=='dono':
            session['perfil']='dono'
            return True
    except Exception:
        pass
    return False

def can_access_area(area):
    if is_owner_user(): return True
    areas=session.get('areas',[]) or []
    # Compatibilidade: usuários antigos sem áreas seguem com acesso até configuração explícita.
    if not areas: return True
    return area in areas

def action_from_request(path,method='GET'):
    p=(path or '').lower()
    m=(method or 'GET').upper()
    if 'export' in p or p.endswith('.csv'):
        return 'export'
    if any(x in p for x in ['/aprovar','/aprovar/','/reprovar','/rejeitar','/autorizar']):
        return 'approve'
    if m=='GET': return 'view'
    if m=='POST': return 'create'
    if m in ('PUT','PATCH'): return 'edit'
    if m=='DELETE': return 'delete'
    return 'view'

def can_access_action(area,action='view'):
    if is_owner_user():
        return True
    if not session.get('rbac_actions_ativo'):
        return None
    perms=session.get('permissoes',{}) or {}
    if not isinstance(perms,dict):
        return False
    acts=perms.get(area)
    if acts is None:
        acts=perms.get('*',[])
    if isinstance(acts,str):
        acts=[acts]
    if not isinstance(acts,list):
        return False
    acts=[str(a).strip().lower() for a in acts if str(a).strip()]
    return ('*' in acts) or (str(action or 'view').lower() in acts)

def can_access_scope(area,action='view'):
    chk=can_access_action(area,action)
    if chk is None:
        return can_access_area(area)
    return chk

def can_access_request(path,method='GET'):
    area=area_from_path(path)
    if not area: return True
    action=action_from_request(path,method)
    if can_access_scope(area,action): return True
    p=(path or '').lower()
    m=(method or 'GET').upper()
    if p.startswith('/api/config/whatsapp') or p.startswith('/api/config/ia-whatsapp') or p.startswith('/api/whatsapp/ia/'):
        if can_access_scope('config',action):
            return True
    if p.startswith('/api/backup'):
        return True
    if m=='GET' and p.startswith('/api/clientes') and can_access_scope('medicoes','view'):
        return True
    if m=='GET' and p.startswith('/api/empresas') and can_access_scope('medicoes','view'):
        return True
    if m in ('GET','DELETE') and (p.startswith('/api/medicoes') or p.startswith('/api/pdf')) and can_access_scope('historico',action):
        return True
    return False

def area_from_path(path):
    p=(path or '').lower()
    if p.startswith('/api/dashboard'): return 'dashboard'
    if p.startswith('/api/financeiro'): return 'medicoes'
    if p.startswith('/api/medicoes') or p.startswith('/api/proximo-numero') or p.startswith('/api/pdf'): return 'medicoes'
    if p.startswith('/api/clientes'): return 'clientes'
    if p.startswith('/api/empresas'): return 'empresas'
    if p.startswith('/api/config/smtp') or p.startswith('/api/config/whatsapp') or p.startswith('/api/config/ia-whatsapp') or p.startswith('/api/whatsapp') or p.startswith('/webhook/whatsapp'): return 'rh-digital'
    if p.startswith('/api/config') or p.startswith('/api/backup'): return 'config'
    if p.startswith('/api/funcionarios'): return 'rh'
    if p.startswith('/api/rh/'): return 'rh-digital'
    if p.startswith('/api/ordens-compra'): return 'compras'
    if p.startswith('/api/operacional'): return 'operacional'
    if p.startswith('/api/usuarios'): return 'usuarios'
    return None

def build_func_docs_response(funcionario_id):
    cat_q=(request.args.get('categoria') or '').strip().lower()
    if cat_q and norm_cat(cat_q)=='outros' and cat_q not in ['outros','outro']:
        return {'erro':'Categoria invalida'},400
    cat_filter=norm_cat(cat_q) if cat_q else ''
    ano_filter=(request.args.get('ano') or '').strip()
    if ano_filter and not re.fullmatch(r'(19|20)\d{2}',ano_filter):
        return {'erro':'Ano invalido'},400
    q=(request.args.get('q') or '').strip().lower()
    page=max(1,to_num(request.args.get('page')) or 1)
    per_page=to_num(request.args.get('per_page')) or 50
    per_page=min(max(1,per_page),200)
    formato=(request.args.get('formato') or 'arvore').strip().lower()
    if formato not in ['arvore','lista']:
        return {'erro':'Formato invalido. Use arvore ou lista'},400

    from sqlalchemy import or_ as _or_
    regs=FuncionarioArquivo.query.filter(
        FuncionarioArquivo.funcionario_id==funcionario_id,
        _or_(FuncionarioArquivo.ass_status==None,FuncionarioArquivo.ass_status!='pendente')
    ).order_by(FuncionarioArquivo.criado_em.desc()).all()
    itens=[]
    for a in regs:
        cat=norm_cat(a.categoria)
        ano=arq_year_from_path(a.caminho)
        if cat_filter and cat!=cat_filter:
            continue
        if ano_filter and ano!=ano_filter:
            continue
        cat_label=DOC_CAT_LABEL.get(cat,cat)
        if q and q not in (a.nome_arquivo or '').lower() and q not in (a.competencia or '').lower() and q not in (a.caminho or '').lower() and q not in cat_label.lower():
            continue
        itens.append({
            'id':a.id,
            'categoria':cat,
            'categoria_label':DOC_CAT_LABEL.get(cat,cat),
            'ano':ano,
            'nome_arquivo':a.nome_arquivo,
            'competencia':a.competencia,
            'ass_status':a.ass_status or 'nao_solicitada',
            'ass_em_fmt':a.ass_em.strftime('%d/%m/%Y %H:%M') if a.ass_em else '',
            'can_assinar':(a.ass_status or '').lower()!='concluida',
            'caminho':a.caminho,
            'criado_em':a.criado_em.isoformat() if a.criado_em else '',
            'criado_fmt':a.criado_em.strftime('%d/%m/%Y %H:%M') if a.criado_em else '',
            'download_url':f'/api/funcionarios/arquivos/{a.id}/download',
            'app_download_url':f'/api/app/funcionario/arquivos/{a.id}/download',
        })

    total_itens=len(itens)
    total_paginas=max(1,(total_itens+per_page-1)//per_page)
    if page>total_paginas: page=total_paginas
    ini=(page-1)*per_page
    fim=ini+per_page
    itens_pag=itens[ini:fim]

    arv={}
    for a in itens_pag:
        cat=a['categoria']
        ano=a['ano']
        if cat not in arv:
            arv[cat]={'categoria':cat,'categoria_label':a['categoria_label'],'anos':{}}
        if ano not in arv[cat]['anos']:
            arv[cat]['anos'][ano]=[]
        arv[cat]['anos'][ano].append(a)

    out=[]
    for cat in sorted(arv.keys(),key=lambda k:DOC_CAT_LABEL.get(k,k)):
        anos=[]
        for ano in sorted(arv[cat]['anos'].keys(),reverse=True):
            anos.append({'ano':ano,'itens':arv[cat]['anos'][ano]})
        out.append({'categoria':cat,'categoria_label':arv[cat]['categoria_label'],'anos':anos})

    resp={
        'ok':True,
        'versao':'v1',
        'funcionario_id':funcionario_id,
        'formato':formato,
        'filtros':{'categoria':cat_filter,'ano':ano_filter,'q':q},
        'paginacao':{
            'page':page,
            'per_page':per_page,
            'total_itens':total_itens,
            'total_paginas':total_paginas,
            'tem_anterior':page>1,
            'tem_proxima':page<total_paginas,
        },
        'categorias':out,
    }
    if formato=='lista':
        resp['itens']=itens_pag
    return resp,200

def read_rows_from_upload(arq):
    nome=(arq.filename or '').lower()
    if nome.endswith('.csv'):
        txt=arq.read().decode('utf-8-sig')
        return list(csv.DictReader(io.StringIO(txt),delimiter=';'))
    if nome.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise ValueError('Dependencia openpyxl nao instalada para XLSX')
        wb=load_workbook(arq,data_only=True,read_only=True)
        ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
        if not rows: return []
        heads=[str(h or '').strip() for h in rows[0]]
        out=[]
        for vals in rows[1:]:
            if not vals or not any(v not in [None,''] for v in vals):
                continue
            out.append({heads[i]:(vals[i] if i < len(vals) else '') for i in range(len(heads))})
        return out
    raise ValueError('Formato inválido. Use CSV ou XLSX')

def only_digits(v):
    return ''.join(ch for ch in str(v or '') if ch.isdigit())

def norm_cpf(v):
    d=only_digits(v)[:11]
    return d or None

def norm_doc(v):
    return only_digits(v)[:14]

def norm_phone(v):
    return only_digits(v)[:11]

def norm_cep(v):
    return only_digits(v)[:8]

def norm_bank_code(v):
    return only_digits(v)[:3]

def norm_uf(v):
    return str(v or '').strip().upper()[:2]

def norm_matricula(v):
    return only_digits(v)[:20]

def to_bool(v):
    if isinstance(v,bool):
        return v
    return str(v or '').strip().lower() in ('1','true','sim','yes','on')

def next_func_matricula():
    mx=0
    for (m,) in db.session.query(Funcionario.matricula).all():
        d=only_digits(m)
        if not d:
            continue
        try:
            mx=max(mx,int(d))
        except Exception:
            pass
    return str(mx+1)

def next_func_re():
    mx=299
    for (r,) in db.session.query(Funcionario.re).all():
        if r and isinstance(r,int):
            mx=max(mx,int(r))
    return mx+1

def parse_json_bytes(raw):
    try:
        return json.loads((raw or b'').decode('utf-8','ignore'))
    except Exception:
        return {}

_BANCOS_BR_CACHE=os.path.join(DATA_DIR, 'bancos_br.json')
_BANCOS_BR_FALLBACK=[
    {'codigo':'001','nome':'Banco do Brasil S.A.'},
    {'codigo':'003','nome':'Banco da Amazonia S.A.'},
    {'codigo':'004','nome':'Banco do Nordeste do Brasil S.A.'},
    {'codigo':'007','nome':'Banco Nacional de Desenvolvimento Economico e Social (BNDES)'},
    {'codigo':'010','nome':'Credicoamo'},
    {'codigo':'021','nome':'BANESTES S.A. Banco do Estado do Espirito Santo'},
    {'codigo':'033','nome':'Banco Santander (Brasil) S.A.'},
    {'codigo':'041','nome':'Banco do Estado do Rio Grande do Sul S.A. (Banrisul)'},
    {'codigo':'047','nome':'Banco do Estado de Sergipe S.A. (Banese)'},
    {'codigo':'070','nome':'BRB - Banco de Brasilia S.A.'},
    {'codigo':'077','nome':'Banco Inter S.A.'},
    {'codigo':'084','nome':'Uniprime Norte do Parana'},
    {'codigo':'085','nome':'Cooperativa Central de Credito Urbano (Ailos)'},
    {'codigo':'104','nome':'Caixa Economica Federal'},
    {'codigo':'107','nome':'Banco Bocom BBM S.A.'},
    {'codigo':'121','nome':'Agibank S.A.'},
    {'codigo':'133','nome':'Cresol Confederacao'},
    {'codigo':'136','nome':'Unicred do Brasil'},
    {'codigo':'197','nome':'Stone Pagamentos S.A.'},
    {'codigo':'208','nome':'Banco BTG Pactual S.A.'},
    {'codigo':'212','nome':'Banco Original S.A.'},
    {'codigo':'218','nome':'Banco BS2 S.A.'},
    {'codigo':'237','nome':'Banco Bradesco S.A.'},
    {'codigo':'246','nome':'Banco ABC Brasil S.A.'},
    {'codigo':'260','nome':'Nu Pagamentos S.A. (Nubank)'},
    {'codigo':'290','nome':'PagSeguro Internet S.A.'},
    {'codigo':'318','nome':'Banco BMG S.A.'},
    {'codigo':'336','nome':'Banco C6 S.A. (C6 Bank)'},
    {'codigo':'341','nome':'Itau Unibanco S.A.'},
    {'codigo':'364','nome':'Gerencianet S.A.'},
    {'codigo':'380','nome':'PicPay Bank - Banco Multiplo S.A.'},
    {'codigo':'389','nome':'Banco Mercantil do Brasil S.A.'},
    {'codigo':'422','nome':'Banco Safra S.A.'},
    {'codigo':'623','nome':'Banco Pan S.A.'},
    {'codigo':'633','nome':'Banco Rendimento S.A.'},
    {'codigo':'637','nome':'Sofisa S.A.'},
    {'codigo':'655','nome':'Banco Votorantim S.A.'},
    {'codigo':'707','nome':'Banco Daycoval S.A.'},
    {'codigo':'735','nome':'Neon Pagamentos S.A.'},
    {'codigo':'748','nome':'Sicredi'},
    {'codigo':'756','nome':'Banco Cooperativo do Brasil S.A. (Sicoob)'}
]

def _bancos_br_normalize(items):
    out=[]
    for it in (items or []):
        if not isinstance(it,dict):
            continue
        cod=norm_bank_code(it.get('codigo') or it.get('code') or it.get('compe') or '')
        nome=(it.get('nome') or it.get('name') or it.get('fullName') or it.get('full_name') or '').strip()
        if not cod or not nome:
            continue
        cod=cod.zfill(3)
        out.append({'codigo':cod,'nome':nome,'label':f'{cod} - {nome}'})
    seen={}
    for b in sorted(out,key=lambda x:(x['codigo'],x['nome'].lower())):
        seen[b['codigo']]=b
    return list(seen.values())

def _bancos_br_fetch_remote():
    req=urllib.request.Request('https://brasilapi.com.br/api/banks/v1',headers={'User-Agent':'Mozilla/5.0'})
    with urllib.request.urlopen(req,timeout=15) as r:
        data=parse_json_bytes(r.read())
    if not isinstance(data,list):
        return []
    return _bancos_br_normalize(data)

def _bancos_br_load_cached():
    try:
        if not os.path.isfile(_BANCOS_BR_CACHE):
            return []
        with open(_BANCOS_BR_CACHE,'r',encoding='utf-8') as f:
            data=json.load(f)
        return _bancos_br_normalize(data)
    except Exception:
        return []

def _bancos_br_save_cache(items):
    try:
        os.makedirs(os.path.dirname(_BANCOS_BR_CACHE),exist_ok=True)
        with open(_BANCOS_BR_CACHE,'w',encoding='utf-8') as f:
            json.dump(items,f,ensure_ascii=False,indent=2)
    except Exception:
        pass

def bancos_br_get(refresh=False):
    if not refresh:
        cached=_bancos_br_load_cached()
        if cached:
            return cached
    try:
        remote=_bancos_br_fetch_remote()
        if remote:
            _bancos_br_save_cache(remote)
            return remote
    except Exception:
        pass
    cached=_bancos_br_load_cached()
    if cached:
        return cached
    return _bancos_br_normalize(_BANCOS_BR_FALLBACK)

def norm_dt8(v):
    s=''.join(ch for ch in str(v or '') if ch.isdigit())
    if len(s)==8:
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return str(v or '').strip()

def ext_cpf_payload(data):
    base=data if isinstance(data,dict) else {}
    if isinstance(data,list) and data:
        base=data[0] if isinstance(data[0],dict) else {}
    for k in ['data','result','resultado','response']:
        if isinstance(base.get(k),dict):
            base=base[k]
            break
    nome=(base.get('nome') or base.get('name') or base.get('Nome') or '').strip()
    nome_social=(base.get('nome_social') or base.get('NomeSocial') or '').strip()
    cpf=only_digits(base.get('cpf') or base.get('documento') or base.get('Cpf') or '')
    nasc=(base.get('nascimento') or base.get('data_nascimento') or base.get('birth_date') or base.get('DataNascimento') or '').strip()
    mae=(base.get('mae') or base.get('nome_mae') or base.get('mother') or base.get('NomeMae') or '').strip()
    situacao=(base.get('situacao') or base.get('status') or base.get('DescSituacaoCadastral') or base.get('SituacaoCadastral') or '').strip()
    tipo_log=(base.get('tipo_logradouro') or base.get('TipoLogradouro') or '').strip()
    logradouro=(base.get('logradouro') or base.get('Logradouro') or '').strip()
    numero=(base.get('numero_logradouro') or base.get('NumeroLogradouro') or '').strip()
    compl=(base.get('complemento') or base.get('Complemento') or '').strip()
    bairro=(base.get('bairro') or base.get('Bairro') or '').strip()
    municipio=(base.get('municipio') or base.get('Municipio') or '').strip()
    uf=(base.get('uf') or base.get('UF') or '').strip()
    cep=only_digits(base.get('cep') or base.get('Cep') or '')
    ddd=only_digits(base.get('ddd') or base.get('DDD') or '')
    tel=only_digits(base.get('telefone') or base.get('Telefone') or '')
    endereco=' '.join([tipo_log,logradouro]).strip()
    return {
        'nome':nome,
        'nome_social':nome_social,
        'cpf':cpf,
        'nascimento':norm_dt8(nasc),
        'mae':mae,
        'situacao':situacao,
        'endereco':endereco,
        'numero':numero,
        'complemento':compl,
        'bairro':bairro,
        'cidade':municipio,
        'estado':uf,
        'cep':cep,
        'ddd':ddd,
        'telefone':tel,
    }

def lookup_cpf_externo(cpf):
    tpl=(os.environ.get('CPF_LOOKUP_URL') or '').strip()
    if not tpl:
        return {'ok':False,'motivo':'nao_configurado'}
    url=tpl.replace('{cpf}',cpf)
    token=(os.environ.get('CPF_LOOKUP_TOKEN') or '').strip()
    token_hdr=(os.environ.get('CPF_LOOKUP_TOKEN_HEADER') or '').strip()
    headers={'Accept':'application/json'}
    if token:
        headers['Authorization']=f'Bearer {token}'
        if token_hdr:
            headers[token_hdr]=token
    try:
        req=urllib.request.Request(url,headers=headers)
        with urllib.request.urlopen(req,timeout=12) as r:
            data=parse_json_bytes(r.read())
        d=ext_cpf_payload(data)
        if not d.get('nome'):
            return {'ok':False,'motivo':'sem_dados'}
        if not d.get('cpf'):
            d['cpf']=cpf
        return {'ok':True,'fonte':(os.environ.get('CPF_LOOKUP_PROVIDER') or 'API externa CPF'),'dados':d}
    except Exception as e:
        return {'ok':False,'motivo':'erro_consulta','erro':str(e)}

def parse_doc_num(s):
    if not s: return None
    m=re.match(r'\s*(\d+)',str(s))
    return int(m.group(1)) if m else None

def norm_url(v):
    u=(v or '').strip()
    if not u: return ''
    if u.startswith(('http://','https://')): return u
    return f'https://{u}'

def ensure_cols(table,defs):
    cols={r[1] for r in db.session.execute(text(f'PRAGMA table_info({table})')).fetchall()}
    changed=False
    for d in defs:
        name=d.split(' ',1)[0]
        if name not in cols:
            try:
                db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {d}'))
                changed=True
            except Exception as e:
                if 'duplicate column' in str(e).lower():
                    pass  # coluna já existe, ignorar
                else:
                    raise
    if changed: db.session.commit()

def sc_cfg(k,v):
    c=Config.query.filter_by(chave=k).first()
    if c: c.valor=str(v)
    else: db.session.add(Config(chave=k,valor=str(v)))
    db.session.commit()

def prox_num():
    try: base=int(gc('num_base','100'))
    except: base=100
    try: ultima_cfg=int(gc('num_ultima','0'))
    except: ultima_cfg=0
    ul=Medicao.query.order_by(Medicao.id.desc()).first()
    ultima_db=parse_doc_num(ul.numero) if ul and ul.numero else 0
    prox=max(base,ultima_cfg,ultima_db)+1
    return f"{prox}/{localnow().year}"

def lr(f):
    @wraps(f)
    def w(*a,**k):
        if 'uid' not in session: return _lr_unauth_response()
        if not can_access_request(request.path,request.method): return jsonify({'erro':'Acesso negado'}),403
        if request.method in ('POST','PUT','PATCH','DELETE') and not _same_origin_request(request):
            return jsonify({'erro':'Origem da requisição não permitida'}),403
        return f(*a,**k)
    return w

def dr(f):
    @wraps(f)
    def w(*a,**k):
        if 'uid' not in session: return redirect(url_for('login'))
        if not is_owner_user(): return jsonify({'erro':'Acesso negado'}),403
        if request.method in ('POST','PUT','PATCH','DELETE') and not _same_origin_request(request):
            return jsonify({'erro':'Origem da requisição não permitida'}),403
        return f(*a,**k)
    return w

def fmt_brl(v):
    try: return 'R$ {:,.2f}'.format(float(v or 0)).replace(',','X').replace('.',',').replace('X','.')
    except: return 'R$ 0,00'

def fmt_data(s):
    if not s: return ''
    try: p=s.split('-'); return f"{p[2]}/{p[1]}/{p[0]}"
    except: return s

def fmt_mes(s):
    if not s: return ''
    ms=['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    try: y,m=s.split('-'); return f"{ms[int(m)-1]}/{y}"
    except: return s

LOGO_PATH=os.path.join(os.path.dirname(__file__),'static','img','logo.png')
LOGO_URL='https://rmfacilities.com.br/wp-content/uploads/2023/08/logo-rm-facilities-1.png'
LOGO_URL_SECUNDARIO='https://rmfacilities.com.br/wp-content/uploads/2023/08/logo-rm-facilities-1.png'

def get_logo():
    if not os.path.exists(LOGO_PATH):
        try: urllib.request.urlretrieve(LOGO_URL,LOGO_PATH)
        except: pass
    return LOGO_PATH if os.path.exists(LOGO_PATH) else None

def _pdf_companies_for_header(empresa_obj=None,empresa_dict=None,limit=2):
    itens=[]
    seen=set()

    def _push(nome,cnpj,logo_url,emp_id=None):
        if len(itens)>=limit:
            return
        key=(str(emp_id or '').strip(),str(nome or '').strip().lower(),str(cnpj or '').strip())
        if key in seen:
            return
        seen.add(key)
        logos=[]
        for cand in [logo_url,get_logo(),LOGO_URL,LOGO_URL_SECUNDARIO]:
            c=(cand or '').strip() if isinstance(cand,str) else cand
            if not c:
                continue
            if c not in logos:
                logos.append(c)
        itens.append({
            'nome':(nome or 'RM Facilities').strip(),
            'cnpj':(cnpj or '').strip(),
            'logos':logos,
        })

    if empresa_obj is not None:
        _push(getattr(empresa_obj,'razao',None) or getattr(empresa_obj,'nome',None),getattr(empresa_obj,'cnpj',None),getattr(empresa_obj,'logo_url',None),getattr(empresa_obj,'id',None))
    elif isinstance(empresa_dict,dict):
        _push(empresa_dict.get('razao') or empresa_dict.get('nome') or empresa_dict.get('empresa_nome'),empresa_dict.get('cnpj'),empresa_dict.get('logo_url'),empresa_dict.get('id'))

    for e in Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem,Empresa.id).all():
        _push(getattr(e,'razao',None) or getattr(e,'nome',None),getattr(e,'cnpj',None),getattr(e,'logo_url',None),getattr(e,'id',None))
        if len(itens)>=limit:
            break

    if not itens:
        _push('RM Facilities','',None,None)
    return itens

register_ponto_routes(
    app,
    db=db,
    utcnow=utcnow,
    to_num=to_num,
    lr=lr,
    audit_event=audit_event,
    Funcionario=Funcionario,
    PontoMarcacao=PontoMarcacao,
    PontoAjuste=PontoAjuste,
    PontoFechamentoDia=PontoFechamentoDia,
    Empresa=Empresa,
    get_logo=get_logo,
)

@app.route('/login',methods=['GET','POST'])
def login():
    recuperar=request.args.get('recuperar')=='1'
    if request.method=='GET' and request.args.get('cancel2fa'):
        session.pop('login_2fa_uid',None)
        session.pop('login_2fa_email',None)
        session.pop('login_2fa_code_hash',None)
        session.pop('login_2fa_exp',None)
        session.pop('login_2fa_attempts',None)
    if 'uid' in session: return redirect(url_for('index'))
    erro=None
    if request.method=='POST':
        etapa=(request.form.get('etapa') or '').strip().lower()
        if etapa=='2fa':
            uid=session.get('login_2fa_uid')
            if not uid:
                erro='Sessão de verificação expirada. Faça login novamente.'
                return render_template('login.html',erro=erro)
            u=Usuario.query.get(uid)
            if not u or not u.ativo:
                erro='Usuário inválido para verificação.'
                return render_template('login.html',erro=erro)
            if request.form.get('acao')=='reenviar':
                codigo=f'{secrets.randbelow(1000000):06d}'
                try:
                    _send_admin_2fa_code(u,codigo,'login')
                except Exception as ex:
                    erro=f'Não foi possível reenviar o código: {str(ex)}'
                    return render_template('login.html',erro=erro,etapa='2fa',email=u.email,email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone))
                session['login_2fa_code_hash']=token_hash(codigo)
                session['login_2fa_exp']=int(time.time())+600
                session['login_2fa_attempts']=0
                return render_template('login.html',ok='Novo código enviado para seu celular.',etapa='2fa',email=u.email,email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone))
            codigo=(request.form.get('codigo') or '').strip()
            if not re.fullmatch(r'\d{6}',codigo):
                return render_template('login.html',erro='Informe o código de 6 dígitos.',etapa='2fa',email=u.email,email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone))
            if int(session.get('login_2fa_exp',0) or 0)<int(time.time()):
                erro='Código expirado. Faça login novamente para gerar outro.'
                session.pop('login_2fa_uid',None)
                session.pop('login_2fa_email',None)
                session.pop('login_2fa_code_hash',None)
                session.pop('login_2fa_exp',None)
                session.pop('login_2fa_attempts',None)
                return render_template('login.html',erro=erro)
            tent=int(session.get('login_2fa_attempts',0) or 0)
            if tent>=5:
                session.pop('login_2fa_uid',None)
                session.pop('login_2fa_email',None)
                session.pop('login_2fa_code_hash',None)
                session.pop('login_2fa_exp',None)
                session.pop('login_2fa_attempts',None)
                return render_template('login.html',erro='Muitas tentativas inválidas. Faça login novamente.')
            if not hmac.compare_digest(token_hash(codigo),str(session.get('login_2fa_code_hash') or '')):
                session['login_2fa_attempts']=tent+1
                return render_template('login.html',erro='Código inválido.',etapa='2fa',email=u.email,email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone))
            session.permanent=True
            session['uid']=u.id; session['nome']=u.nome; session['perfil']=u.perfil
            session['areas']=jloads(u.areas,[])
            perms=jloads(getattr(u,'permissoes','{}'),{})
            if not isinstance(perms,dict): perms={}
            session['permissoes']=perms
            session['rbac_actions_ativo']=bool(perms)
            u.ultimo_acesso=utcnow(); db.session.commit()
            reg_auth_attempt('admin',u.email,True,'ok_2fa')
            audit_event('auth_admin_sucesso_2fa','usuario',u.id,'usuario',u.id,True,{})
            session.pop('login_2fa_uid',None)
            session.pop('login_2fa_email',None)
            session.pop('login_2fa_code_hash',None)
            session.pop('login_2fa_exp',None)
            session.pop('login_2fa_attempts',None)
            return redirect(url_for('index'))

        email=request.form.get('email','').lower().strip()
        senha=request.form.get('senha','')
        if auth_blocked('admin',email,(request.remote_addr or '')):
            erro='Muitas tentativas. Aguarde alguns minutos.'
            audit_event('auth_admin_bloqueado','admin',email,'usuario','',False,{'motivo':'rate_limit'})
            return render_template('login.html',erro=erro,recuperar=recuperar)
        u=Usuario.query.filter_by(email=email,ativo=True).first()
        if u and pw_check(u.senha,senha):
            if not pw_is_modern(u.senha):
                u.senha=pw_hash(senha)
            if _admin_needs_2fa(u):
                codigo=f'{secrets.randbelow(1000000):06d}'
                try:
                    _send_admin_2fa_code(u,codigo,'login')
                except Exception as ex:
                    erro=f'Falha ao enviar código de verificação: {str(ex)}'
                    audit_event('auth_admin_2fa_envio_falha','usuario',u.id,'usuario',u.id,False,{'erro':str(ex)[:200]})
                    return render_template('login.html',erro=erro,recuperar=recuperar)
                session['login_2fa_uid']=u.id
                session['login_2fa_email']=u.email
                session['login_2fa_code_hash']=token_hash(codigo)
                session['login_2fa_exp']=int(time.time())+600
                session['login_2fa_attempts']=0
                audit_event('auth_admin_2fa_desafio','usuario',u.id,'usuario',u.id,True,{})
                return render_template('login.html',etapa='2fa',email=u.email,email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone),ok='Código de 6 dígitos enviado para seu celular.')
            session.permanent=True
            session['uid']=u.id; session['nome']=u.nome; session['perfil']=u.perfil
            session['areas']=jloads(u.areas,[])
            perms=jloads(getattr(u,'permissoes','{}'),{})
            if not isinstance(perms,dict): perms={}
            session['permissoes']=perms
            session['rbac_actions_ativo']=bool(perms)
            u.ultimo_acesso=utcnow(); db.session.commit()
            reg_auth_attempt('admin',email,True,'ok')
            audit_event('auth_admin_sucesso','usuario',u.id,'usuario',u.id,True,{})
            return redirect(url_for('index'))
        reg_auth_attempt('admin',email,False,'credenciais_invalidas')
        audit_event('auth_admin_falha','admin',email,'usuario','',False,{})
        erro='E-mail ou senha incorretos.'
    return render_template('login.html',erro=erro,recuperar=recuperar)


    return render_template('login.html',recuperar=True,erro='Sessão de recuperação expirada. Solicite novo código.')
    codigo=(request.form.get('codigo') or '').strip()
    nova_senha=(request.form.get('nova_senha') or '').strip()
    if not re.fullmatch(r'\d{6}',codigo):
        return render_template('login.html',recuperar=True,rec_etapa='codigo',email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone),erro='Código inválido. Use 6 dígitos.')
    if len(nova_senha)<8:
        return render_template('login.html',recuperar=True,rec_etapa='codigo',email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone),erro='A nova senha deve ter ao menos 8 caracteres.')
    if int(session.get('rec_exp',0) or 0)<int(time.time()):
        session.pop('rec_uid',None); session.pop('rec_code_hash',None); session.pop('rec_exp',None); session.pop('rec_attempts',None)
        return render_template('login.html',recuperar=True,erro='Código expirado. Solicite outro.')
    tent=int(session.get('rec_attempts',0) or 0)
    if tent>=5:
        session.pop('rec_uid',None); session.pop('rec_code_hash',None); session.pop('rec_exp',None); session.pop('rec_attempts',None)
        return render_template('login.html',recuperar=True,erro='Muitas tentativas inválidas. Solicite novo código.')
    if not hmac.compare_digest(token_hash(codigo),str(session.get('rec_code_hash') or '')):
        session['rec_attempts']=tent+1
        return render_template('login.html',recuperar=True,rec_etapa='codigo',email_mask=_mask_email(u.email),telefone_mask=_mask_phone(u.telefone),erro='Código incorreto.')

    u.senha=pw_hash(nova_senha)
    db.session.commit()
    audit_event('auth_recuperacao_senha','usuario',u.id,'usuario',u.id,True,{})
    session.pop('rec_uid',None); session.pop('rec_code_hash',None); session.pop('rec_exp',None); session.pop('rec_attempts',None)
    return render_template('login.html',ok=f'Acesso recuperado. Usuário: {_mask_email(u.email)}. Faça login com a nova senha.')

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/privacidade')
@app.route('/politica-de-privacidade')
@app.route('/privacy-policy')
def pagina_privacidade_publica():
    return render_template('privacidade_publica.html',atualizado_em='05/05/2026')

@app.route('/')
@lr
def index(): return render_template('app.html',nome=session['nome'],perfil=session['perfil'],areas=json.dumps(session.get('areas',[]),ensure_ascii=False))

@app.route('/api/cnpj/<cnpj>')
@lr
def api_cnpj(cnpj):
    c=''.join(filter(str.isdigit,cnpj))
    if len(c)!=14: return jsonify({'erro':'CNPJ inválido'}),400
    # Tenta Receitaws primeiro e usa BrasilAPI como fallback.
    try:
        req=urllib.request.Request(f'https://receitaws.com.br/v1/cnpj/{c}',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        if d.get('status')!='ERROR':
            return jsonify({'nome':d.get('fantasia') or d.get('nome',''),'razao':d.get('nome',''),'email':d.get('email',''),'telefone':d.get('telefone',''),'cep':d.get('cep','').replace('.','').replace('-','').replace(' ',''),'logradouro':d.get('logradouro',''),'numero':d.get('numero',''),'complemento':d.get('complemento',''),'bairro':d.get('bairro',''),'cidade':d.get('municipio',''),'estado':d.get('uf',''),'situacao':d.get('situacao','')})
    except Exception:
        pass
    try:
        req=urllib.request.Request(f'https://brasilapi.com.br/api/cnpj/v1/{c}',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        return jsonify({'nome':d.get('nome_fantasia') or d.get('razao_social',''),'razao':d.get('razao_social',''),'email':d.get('email',''),'telefone':d.get('ddd_telefone_1') or d.get('ddd_telefone_2',''),'cep':str(d.get('cep','')).replace('-',''),'logradouro':d.get('logradouro',''),'numero':d.get('numero',''),'complemento':d.get('complemento',''),'bairro':d.get('bairro',''),'cidade':d.get('municipio',''),'estado':d.get('uf',''),'situacao':d.get('descricao_situacao_cadastral','')})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/cep/<cep>')
@lr
def api_cep(cep):
    c=''.join(filter(str.isdigit,cep))
    if len(c)!=8: return jsonify({'erro':'CEP inválido'}),400
    try:
        req=urllib.request.Request(f'https://viacep.com.br/ws/{c}/json/',headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req,timeout=8) as r: d=json.loads(r.read().decode())
        if d.get('erro'): return jsonify({'erro':'CEP não encontrado'}),404
        return jsonify({'logradouro':d.get('logradouro',''),'bairro':d.get('bairro',''),'cidade':d.get('localidade',''),'estado':d.get('uf','')})
    except Exception as e: return jsonify({'erro':str(e)}),500


@app.route('/api/empresas',methods=['GET'])
@lr
def api_empresas():
    # Compatibilidade: registros antigos podem ter ativa nulo.
    qr=Empresa.query.filter((Empresa.ativa==True) | (Empresa.ativa.is_(None))).order_by(Empresa.ordem,Empresa.nome)
    return jsonify([e.to_dict() for e in qr.all()])

@app.route('/api/empresas/<int:id>',methods=['GET'])
@lr
def api_empresa(id): return jsonify(Empresa.query.get_or_404(id).to_dict())

@app.route('/api/empresas',methods=['POST'])
@lr
def api_criar_empresa():
    d=request.json or {}
    if 'site' in d: d['site']=norm_url(d.get('site'))
    if 'logo_url' in d: d['logo_url']=norm_url(d.get('logo_url'))
    d['cnpj']=norm_doc(d.get('cnpj'))
    d['cep']=norm_cep(d.get('cep'))
    d['telefone']=norm_phone(d.get('telefone'))
    d['contato_telefone']=norm_phone(d.get('contato_telefone'))
    cols=[c.name for c in Empresa.__table__.columns if c.name not in['id','criado_em'] and hasattr(Empresa,c.name)]
    e=Empresa(**{k:d[k] for k in cols if k in d})
    db.session.add(e)
    db.session.commit()
    return jsonify(e.to_dict()),201

@app.route('/api/empresas/<int:id>',methods=['PUT'])
@lr
def api_editar_empresa(id):
    e=Empresa.query.get_or_404(id)
    d=request.json or {}
    if 'site' in d: d['site']=norm_url(d.get('site'))
    if 'logo_url' in d: d['logo_url']=norm_url(d.get('logo_url'))
    if 'cnpj' in d: d['cnpj']=norm_doc(d.get('cnpj'))
    if 'cep' in d: d['cep']=norm_cep(d.get('cep'))
    if 'telefone' in d: d['telefone']=norm_phone(d.get('telefone'))
    if 'contato_telefone' in d: d['contato_telefone']=norm_phone(d.get('contato_telefone'))
    cols=[c.name for c in Empresa.__table__.columns if c.name not in['id','criado_em'] and hasattr(Empresa,c.name)]
    for k in cols:
        if k in d: setattr(e,k,d[k])
    db.session.commit()
    return jsonify(e.to_dict())

@app.route('/api/empresas/<int:id>',methods=['DELETE'])
@lr
def api_remover_empresa(id):
    e=Empresa.query.get_or_404(id)
    db.session.delete(e)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/empresas/<int:id>/certificado',methods=['DELETE'])
@lr
def api_empresa_cert_delete(id):
    e=Empresa.query.get_or_404(id)
    abs_old=_cert_rel_to_abs(e.cert_arquivo)
    e.cert_arquivo=None
    e.cert_nome_arquivo=None
    e.cert_senha=None
    e.cert_ativo=False
    e.cert_assunto=None
    e.cert_validade_fim=None
    db.session.commit()
    if abs_old and os.path.exists(abs_old):
        try: os.remove(abs_old)
        except Exception: pass
    return jsonify({'ok':True,'empresa':e.to_dict()})

@app.route('/api/config',methods=['GET'])
@lr
def api_get_config(): return jsonify({k:gc(k) for k in ['num_base','num_ultima']})

@app.route('/api/config',methods=['POST'])
@dr
def api_save_config():
    for k,v in request.json.items(): sc_cfg(k,v)
    return jsonify({'ok':True})

@app.route('/api/usuarios',methods=['GET'])
@lr
def api_usuarios(): return jsonify([u.to_dict() for u in Usuario.query.all()])

@app.route('/api/usuarios',methods=['POST'])
@lr
def api_criar_usuario():
    d=request.json or {}
    if Usuario.query.filter_by(email=(d.get('email','').lower())).first(): return jsonify({'erro':'E-mail já cadastrado'}),400
    tel=norm_phone(d.get('telefone'))
    perfil=(d.get('perfil','admin') or 'admin').strip().lower()
    if perfil=='dono' and not is_owner_user():
        return jsonify({'erro':'Apenas dono pode criar usuário dono.'}),403
    if perfil in ('admin','dono') and len(tel)<10:
        return jsonify({'erro':'Telefone é obrigatório para usuários admin/dono (2FA).'}),400
    ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
    perms_in=d.get('permissoes') if isinstance(d.get('permissoes'),dict) else {}
    perms={}
    for area,acts in perms_in.items():
        if area!='*' and area not in ALLOWED_AREAS:
            continue
        if isinstance(acts,str): acts=[acts]
        if not isinstance(acts,list):
            continue
        limpos=[str(a).strip().lower() for a in acts if str(a).strip().lower() in ALLOWED_ACTIONS or str(a).strip()=='*']
        if limpos:
            perms[area]=sorted(set(limpos))
    u=Usuario(
        nome=d['nome'],
        email=d['email'].lower(),
        telefone=tel,
        senha=pw_hash(d['senha']),
        perfil=perfil,
        twofa_ativo=bool(d.get('twofa_ativo',True)),
        areas=json.dumps(ars,ensure_ascii=False),
        permissoes=json.dumps(perms,ensure_ascii=False)
    )
    db.session.add(u); db.session.commit(); return jsonify(u.to_dict()),201

@app.route('/api/usuarios/<int:id>',methods=['PUT'])
@lr
def api_atualizar_usuario(id):
    u=Usuario.query.get_or_404(id); d=request.json or {}
    perfil_novo=(d.get('perfil',u.perfil) or u.perfil or '').strip().lower()
    perfil_atual=(u.perfil or '').strip().lower()
    if (not is_owner_user()) and (perfil_atual=='dono' or perfil_novo=='dono'):
        return jsonify({'erro':'Apenas dono pode alterar usuário dono.'}),403
    for k in ['nome','perfil','ativo']:
        if k in d: setattr(u,k,d[k])
    if 'telefone' in d:
        u.telefone=norm_phone(d.get('telefone'))
    if 'twofa_ativo' in d:
        u.twofa_ativo=bool(d.get('twofa_ativo'))
    if d.get('senha'): u.senha=pw_hash(d['senha'])
    if 'areas' in d:
        ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
        u.areas=json.dumps(ars,ensure_ascii=False)
    if 'permissoes' in d:
        perms_in=d.get('permissoes') if isinstance(d.get('permissoes'),dict) else {}
        perms={}
        for area,acts in perms_in.items():
            if area!='*' and area not in ALLOWED_AREAS:
                continue
            if isinstance(acts,str): acts=[acts]
            if not isinstance(acts,list):
                continue
            limpos=[str(a).strip().lower() for a in acts if str(a).strip().lower() in ALLOWED_ACTIONS or str(a).strip()=='*']
            if limpos:
                perms[area]=sorted(set(limpos))
        u.permissoes=json.dumps(perms,ensure_ascii=False)
    if (u.perfil or '').strip().lower() in ('admin','dono') and len(norm_phone(u.telefone))<10:
        return jsonify({'erro':'Telefone é obrigatório para usuários admin/dono (2FA).'}),400
    db.session.commit(); return jsonify(u.to_dict())

@app.route('/api/usuarios/<int:id>',methods=['DELETE'])
@lr
def api_deletar_usuario(id):
    u=Usuario.query.get_or_404(id)
    if (u.perfil or '').strip().lower()=='dono' and (not is_owner_user()):
        return jsonify({'erro':'Apenas dono pode excluir usuário dono.'}),403
    if u.perfil=='dono' and Usuario.query.filter_by(perfil='dono').count()<=1: return jsonify({'erro':'Não é possível excluir o único dono'}),400
    db.session.delete(u); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/usuarios/<int:id>/certificado',methods=['POST'])
@lr
def api_usuario_cert_upload(id):
    u=Usuario.query.get_or_404(id)
    fs=request.files.get('arquivo')
    senha=(request.form.get('senha') or '').strip()
    ativo=str(request.form.get('ativo','1')).strip().lower() in ('1','true','yes','on')
    if not fs:
        return jsonify({'erro':'Arquivo do certificado não enviado.'}),400
    if not senha:
        return jsonify({'erro':'Informe a senha do certificado do usuário.'}),400
    old_abs=_cert_rel_to_abs(u.cert_arquivo)
    try:
        rel,name=_cert_store_file(fs,'usuario',id)
        abs_path=_cert_rel_to_abs(rel)
        info=_cert_inspect_pkcs12(abs_path,senha)
    except Exception as ex:
        return jsonify({'erro':str(ex)}),400
    u.cert_arquivo=rel
    u.cert_nome_arquivo=name
    u.cert_senha=senha
    u.cert_ativo=ativo
    u.cert_assunto=info.get('assunto','')
    u.cert_validade_fim=info.get('validade_fim','')
    db.session.commit()
    if old_abs and old_abs != abs_path and os.path.exists(old_abs):
        try:
            os.remove(old_abs)
        except Exception:
            pass
    return jsonify({'ok':True,'usuario':u.to_dict()})


_medicao_stamp_ready=False

def _ensure_medicao_stamp_cols_runtime(force=False):
    """Auto-recupera colunas de carimbo em medicao quando faltar no SQLite."""
    global _medicao_stamp_ready
    if _medicao_stamp_ready and not force:
        return
    ensure_cols('medicao',[
        'stamp_habilitado BOOLEAN DEFAULT 0',
        'stamp_pagina INTEGER DEFAULT 1',
        'stamp_x_pct REAL DEFAULT 60.0',
        'stamp_y_pct REAL DEFAULT 10.0',
    ])
    db.session.commit()
    _medicao_stamp_ready=True

def _is_missing_medicao_stamp_error(err):
    msg=str(err or '').lower()
    return 'no such column' in msg and 'medicao.stamp_' in msg

@app.route('/api/usuarios/<int:id>/certificado',methods=['DELETE'])
@lr
def api_usuario_cert_delete(id):
    u=Usuario.query.get_or_404(id)
    abs_old=_cert_rel_to_abs(u.cert_arquivo)
    u.cert_arquivo=None
    u.cert_nome_arquivo=None
    u.cert_senha=None
    u.cert_ativo=False
    u.cert_assunto=None
    u.cert_validade_fim=None
    db.session.commit()
    if abs_old and os.path.exists(abs_old):
        try: os.remove(abs_old)
        except Exception: pass
    return jsonify({'ok':True,'usuario':u.to_dict()})

@app.route('/api/clientes',methods=['GET'])
@lr
def api_clientes():
    q=request.args.get('q','').lower(); emp=request.args.get('emp',''); st=request.args.get('status','')
    qr=Cliente.query
    if emp: qr=qr.filter_by(empresa_id=int(emp))
    if st: qr=qr.filter_by(status=st)
    lista=qr.order_by(Cliente.nome).all()
    if q:
        qdig=only_digits(q)
        lista=[c for c in lista if (
            q in c.nome.lower() or
            q in (c.cnpj or '').lower() or
            q in (c.numero or '') or
            (qdig and qdig in only_digits(c.cnpj))
        )]
    return jsonify([c.to_dict() for c in lista])

@app.route('/api/clientes',methods=['POST'])
@lr
def api_criar_cliente():
    d=request.json or {}; n=next_cli_num()
    d['cnpj']=norm_doc(d.get('cnpj'))
    d['telefone']=norm_phone(d.get('telefone'))
    d['cep']=norm_cep(d.get('cep'))
    if 'qtd_funcionarios_posto' in d:
        d['qtd_funcionarios_posto']=max(0,to_num(d.get('qtd_funcionarios_posto')))
    skip=['id','numero','criado_em','end_fmt']
    cols=[c.name for c in Cliente.__table__.columns if c.name not in skip]
    kw={k:d[k] for k in cols if k in d}
    c=Cliente(numero=n,**kw); db.session.add(c); db.session.commit(); return jsonify(c.to_dict()),201

@app.route('/api/clientes/<int:id>',methods=['PUT'])
@lr
def api_atualizar_cliente(id):
    c=Cliente.query.get_or_404(id)
    d=request.json or {}
    d['cnpj']=norm_doc(d.get('cnpj'))
    d['telefone']=norm_phone(d.get('telefone'))
    d['cep']=norm_cep(d.get('cep'))
    if 'qtd_funcionarios_posto' in d:
        d['qtd_funcionarios_posto']=max(0,to_num(d.get('qtd_funcionarios_posto')))
    skip={'id','numero','criado_em','end_fmt'}
    cols={col.name for col in Cliente.__table__.columns}
    for k,v in d.items():
        if k in skip or k not in cols:
            continue
        setattr(c,k,v)
    db.session.commit()
    return jsonify(c.to_dict())

@app.route('/api/clientes/<int:id>/reajuste',methods=['POST'])
@lr
def api_cliente_reajuste(id):
    c=Cliente.query.get_or_404(id)
    d=request.json or {}
    pct=to_num(d.get('percentual'),dec=True)
    if pct is None:
        pct=to_num(c.reajuste_percentual,dec=True) or 0
    try:
        pct=float(pct or 0)
    except Exception:
        return jsonify({'erro':'Percentual de reajuste inválido.'}),400
    if pct<=-100:
        return jsonify({'erro':'Percentual deve ser maior que -100%.'}),400

    fator=(100.0+pct)/100.0
    campos=['limpeza','jardinagem','portaria','materiais_equip_locacao']
    antes={k:float(getattr(c,k) or 0) for k in campos}
    for k in campos:
        setattr(c,k,round((float(getattr(c,k) or 0)*fator),2))
    depois={k:float(getattr(c,k) or 0) for k in campos}

    hoje=localnow().strftime('%Y-%m-%d')
    c.ultimo_reajuste_em=hoje
    if d.get('atualizar_padrao',True):
        c.reajuste_percentual=pct
        c.reajuste_data_base=hoje
    if d.get('anotar_obs',True):
        linha=f"[Reajuste {hoje}] {pct:+.2f}% aplicado no contrato"
        c.obs=(f"{(c.obs or '').strip()}\n{linha}").strip()

    db.session.commit()
    audit_event('cliente_reajuste_contrato','usuario',session.get('uid'),'cliente',c.id,True,
                {'percentual':pct,'antes_total':sum(antes.values()),'depois_total':sum(depois.values())})
    return jsonify({
        'ok':True,
        'cliente':c.to_dict(),
        'percentual':pct,
        'antes':antes,
        'depois':depois,
        'antes_total':round(sum(antes.values()),2),
        'depois_total':round(sum(depois.values()),2)
    })

@app.route('/api/clientes/<int:id>',methods=['DELETE'])
@lr
def api_deletar_cliente(id):
    c=Cliente.query.get_or_404(id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok':True})

# ── CONTRATOS ────────────────────────────────────────────────────────────────
@app.route('/api/contratos',methods=['GET'])
@lr
def api_listar_contratos():
    cliente_id=request.args.get('cliente_id','')
    st=request.args.get('status','')
    qr=Contrato.query
    if cliente_id: qr=qr.filter_by(cliente_id=int(cliente_id))
    if st: qr=qr.filter_by(status=st)
    lista=qr.order_by(Contrato.cliente_id,Contrato.id).all()
    return jsonify([c.to_dict() for c in lista])

@app.route('/api/clientes/<int:cid>/contratos',methods=['POST'])
@lr
def api_criar_contrato(cid):
    Cliente.query.get_or_404(cid)
    d=request.json or {}
    skip=['id','criado_em']
    cols=[c.name for c in Contrato.__table__.columns if c.name not in skip]
    kw={k:d[k] for k in cols if k in d}
    kw['cliente_id']=cid
    ct=Contrato(**kw)
    db.session.add(ct); db.session.commit()
    return jsonify(ct.to_dict()),201

@app.route('/api/contratos/<int:id>',methods=['GET'])
@lr
def api_get_contrato(id):
    return jsonify(Contrato.query.get_or_404(id).to_dict())

@app.route('/api/contratos/<int:id>',methods=['PUT'])
@lr
def api_atualizar_contrato(id):
    ct=Contrato.query.get_or_404(id)
    d=request.json or {}
    skip={'id','cliente_id','criado_em'}
    cols={col.name for col in Contrato.__table__.columns}
    for k,v in d.items():
        if k in skip or k not in cols: continue
        setattr(ct,k,v)
    db.session.commit()
    return jsonify(ct.to_dict())

@app.route('/api/contratos/<int:id>',methods=['DELETE'])
@lr
def api_deletar_contrato(id):
    ct=Contrato.query.get_or_404(id)
    db.session.delete(ct); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/contratos/<int:id>/reajuste',methods=['POST'])
@lr
def api_contrato_reajuste(id):
    ct=Contrato.query.get_or_404(id)
    d=request.json or {}
    pct=to_num(d.get('percentual'),dec=True)
    if pct is None:
        pct=to_num(ct.reajuste_percentual,dec=True) or 0
    try:
        pct=float(pct or 0)
    except Exception:
        return jsonify({'erro':'Percentual de reajuste inválido.'}),400
    if pct<=-100:
        return jsonify({'erro':'Percentual deve ser maior que -100%.'}),400
    fator=(100.0+pct)/100.0
    campos=['limpeza','jardinagem','portaria','materiais_equip_locacao']
    antes={k:float(getattr(ct,k) or 0) for k in campos}
    for k in campos:
        setattr(ct,k,round(float(getattr(ct,k) or 0)*fator,2))
    depois={k:float(getattr(ct,k) or 0) for k in campos}
    hoje=localnow().strftime('%Y-%m-%d')
    ct.ultimo_reajuste_em=hoje
    if d.get('atualizar_padrao',True):
        ct.reajuste_percentual=pct
        ct.reajuste_data_base=hoje
    if d.get('anotar_obs',True):
        linha=f"[Reajuste {hoje}] {pct:+.2f}% aplicado"
        ct.obs=(f"{(ct.obs or '').strip()}\n{linha}").strip()
    db.session.commit()
    audit_event('contrato_reajuste','usuario',session.get('uid'),'contrato',ct.id,True,
                {'percentual':pct,'antes_total':sum(antes.values()),'depois_total':sum(depois.values())})
    return jsonify({'ok':True,'contrato':ct.to_dict(),'percentual':pct,
                    'antes':antes,'depois':depois,
                    'antes_total':round(sum(antes.values()),2),
                    'depois_total':round(sum(depois.values()),2)})

@app.route('/api/clientes/modelo')
@lr
def api_clientes_modelo():
    import pandas as pd
    cab=['nome','cnpj','responsavel','telefone','email','cep','logradouro','numero_end','complemento','bairro','cidade','estado','empresa_id','numero_contrato','qtd_funcionarios_posto','status','limpeza','jardinagem','portaria','materiais_equip_locacao','dia_faturamento','dias_faturamento','reajuste_percentual','reajuste_data_base','obs']
    exemplo=['Condominio Exemplo','12.345.678/0001-90','Maria Silva','(12) 99123-4567','contato@exemplo.com','12246000','Rua Central','100','','Centro','Sao Jose dos Campos','SP','','CT-2026-001','8','Ativo','1500,00','300,00','2500,00','900,00','1','21','5,00','2026-04-01','Contrato mensal']
    df = pd.DataFrame([exemplo], columns=cab)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='modelo_clientes_rmfacilities.xlsx')

@app.route('/api/clientes/import',methods=['POST'])
@lr
def api_clientes_import():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo nao enviado'}),400
    try:
        linhas=read_rows_from_upload(arq)
    except Exception as e:
        return jsonify({'erro':str(e)}),400
    if not linhas: return jsonify({'erro':'Planilha vazia'}),400
    criados=0; erros=[]; prox=int(next_cli_num())
    for i,row in enumerate(linhas,start=2):
        nome=(row.get('nome') or '').strip()
        if not nome:
            erros.append(f'Linha {i}: nome obrigatorio')
            continue
        try:
            c=Cliente(
                numero=str(prox).zfill(3),
                nome=nome,
                cnpj=norm_doc((row.get('cnpj') or '').strip()),
                responsavel=(row.get('responsavel') or '').strip(),
                telefone=norm_phone((row.get('telefone') or '').strip()),
                email=(row.get('email') or '').strip(),
                cep=norm_cep((row.get('cep') or '').strip()),
                logradouro=(row.get('logradouro') or '').strip(),
                numero_end=(row.get('numero_end') or '').strip(),
                complemento=(row.get('complemento') or '').strip(),
                bairro=(row.get('bairro') or '').strip(),
                cidade=(row.get('cidade') or '').strip(),
                estado=(row.get('estado') or '').strip(),
                empresa_id=to_num(row.get('empresa_id')) or None,
                numero_contrato=(row.get('numero_contrato') or '').strip(),
                qtd_funcionarios_posto=max(0,to_num(row.get('qtd_funcionarios_posto'))),
                status=(row.get('status') or 'Ativo').strip() or 'Ativo',
                limpeza=to_num(row.get('limpeza'),dec=True),
                jardinagem=to_num(row.get('jardinagem'),dec=True),
                portaria=to_num(row.get('portaria'),dec=True),
                materiais_equip_locacao=to_num(row.get('materiais_equip_locacao'),dec=True),
                vencimento=to_num(row.get('vencimento')) or 10,
                dia_faturamento=to_num(row.get('dia_faturamento')) or 1,
                dias_faturamento=to_num(row.get('dias_faturamento')) or 30,
                reajuste_percentual=to_num(row.get('reajuste_percentual'),dec=True),
                reajuste_data_base=(row.get('reajuste_data_base') or '').strip() or None,
                obs=(row.get('obs') or '').strip()
            )
            db.session.add(c); prox+=1; criados+=1
        except Exception as e:
            erros.append(f'Linha {i}: {str(e)}')
    db.session.commit()
    return jsonify({'ok':True,'criados':criados,'erros':erros})

@app.route('/api/funcionarios/modelo')
@lr
def api_funcionarios_modelo():
    import pandas as pd
    cab=['matricula','re','nome','cpf','email','telefone','cargo','funcao','cbo','setor','empresa_id','data_admissao','tipo_contrato','jornada','status','salario','vale_refeicao','vale_alimentacao','vale_transporte','cep','endereco','endereco_numero','endereco_complemento','endereco_bairro','cidade','estado','banco_codigo','banco_nome','banco_agencia','banco_conta','banco_tipo_conta','banco_pix','rg','orgao_emissor','pis','ctps','titulo_eleitor','cert_reservista','cnh','exame_admissional_data','docs_admissao_ok','docs_admissao_obs','obs','areas']
    exemplo=['1001','300','Joao da Silva','123.456.789-00','joao@empresa.com','5512999990000','Auxiliar','Auxiliar de Limpeza','5143-20','Operacional','1','2026-01-10','CLT','44h semanais','Ativo','2500,00','350,00','450,00','220,00','12246000','Rua A','100','Apto 12','Centro','Sao Jose dos Campos','SP','077','Banco Inter S.A.','0001','12345-6','corrente','joao@pix.com','1234567','SSP/SP','12345678901','123456-serie 001','9876543210','123456','AB','2026-01-08','1','Checklist conferido','Exemplo','rh,operacional,sst']
    df = pd.DataFrame([exemplo], columns=cab)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='modelo_funcionarios_rmfacilities.xlsx')

@app.route('/api/funcionarios/import',methods=['POST'])
@lr
def api_funcionarios_import():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo nao enviado'}),400
    modo=(request.form.get('modo') or 'adicionar').strip().lower()
    atualizar_existentes=modo in ('atualizar','upsert','merge')

    try:
        linhas=read_rows_from_upload(arq)
    except Exception as e:
        return jsonify({'erro':str(e)}),400
    if not linhas:
        return jsonify({'erro':'Planilha vazia'}),400
    criados=0
    atualizados=0
    ignorados=0
    erros=[]

    try:
        for i, row in enumerate(linhas, start=2):
            nome = (str(row.get('nome', ''))).strip()
            if not nome:
                erros.append(f'Linha {i}: nome obrigatorio')
                continue
            try:
                areas_raw = str(row.get('areas', '') or '')
                ars = [a.strip().lower() for a in re.split(r'[;,]', areas_raw) if a.strip()]
                ars = [a for a in ars if a in ALLOWED_AREAS]
                mat = norm_matricula(row.get('matricula')) or next_func_matricula()
                re_num = to_num(row.get('re')) or next_func_re()
                # Verifica se já existe funcionário com esse RE
                f = Funcionario.query.filter_by(re=re_num).first()
                if f:
                    if not atualizar_existentes:
                        ignorados += 1
                        erros.append(f'Linha {i}: RE {re_num} já existe (ignorado para evitar sobrescrever).')
                        continue
                    # Atualiza campos do funcionário existente
                    f.matricula = mat
                    f.nome = nome
                    f.cpf = norm_cpf(str(row.get('cpf', '') or '').strip())
                    f.email = str(row.get('email', '') or '').strip()
                    f.telefone = wa_norm_number(str(row.get('telefone', '') or '').strip())
                    f.cargo = str(row.get('cargo', '') or '').strip()
                    f.funcao = str(row.get('funcao', '') or '').strip()
                    f.cbo = str(row.get('cbo', '') or '').strip()
                    f.setor = str(row.get('setor', '') or '').strip()
                    f.empresa_id = to_num(row.get('empresa_id')) or None
                    f.data_admissao = str(row.get('data_admissao', '') or '').strip()
                    f.tipo_contrato = str(row.get('tipo_contrato', '') or '').strip()
                    f.jornada = str(row.get('jornada', '') or '').strip()
                    f.status = str(row.get('status', 'Ativo') or 'Ativo').strip() or 'Ativo'
                    f.posto_operacional = 'Reserva tecnica'
                    f.salario = to_num(row.get('salario'), dec=True)
                    f.vale_refeicao = to_num(row.get('vale_refeicao'), dec=True)
                    f.vale_alimentacao = to_num(row.get('vale_alimentacao'), dec=True)
                    f.vale_transporte = to_num(row.get('vale_transporte'), dec=True)
                    f.cep = norm_cep(str(row.get('cep', '') or '').strip())
                    f.endereco = str(row.get('endereco', '') or '').strip()
                    f.endereco_numero = str(row.get('endereco_numero', '') or '').strip()
                    f.endereco_complemento = str(row.get('endereco_complemento', '') or '').strip()
                    f.endereco_bairro = str(row.get('endereco_bairro', '') or '').strip()
                    f.cidade = str(row.get('cidade', '') or '').strip()
                    f.estado = norm_uf(row.get('estado', ''))
                    f.banco_codigo = norm_bank_code(row.get('banco_codigo'))
                    f.banco_nome = str(row.get('banco_nome', '') or '').strip()
                    f.banco_agencia = str(row.get('banco_agencia', '') or '').strip()
                    f.banco_conta = str(row.get('banco_conta', '') or '').strip()
                    f.banco_tipo_conta = str(row.get('banco_tipo_conta', '') or '').strip()
                    f.banco_pix = str(row.get('banco_pix', '') or '').strip()
                    f.rg = str(row.get('rg', '') or '').strip()
                    f.orgao_emissor = str(row.get('orgao_emissor', '') or '').strip()
                    f.pis = str(row.get('pis', '') or '').strip()
                    f.ctps = str(row.get('ctps', '') or '').strip()
                    f.titulo_eleitor = str(row.get('titulo_eleitor', '') or '').strip()
                    f.cert_reservista = str(row.get('cert_reservista', '') or '').strip()
                    f.cnh = str(row.get('cnh', '') or '').strip()
                    f.exame_admissional_data = str(row.get('exame_admissional_data', '') or '').strip()
                    f.docs_admissao_ok = to_bool(row.get('docs_admissao_ok'))
                    f.docs_admissao_obs = str(row.get('docs_admissao_obs', '') or '').strip()
                    f.obs = str(row.get('obs', '') or '').strip()
                    f.areas = json.dumps(ars, ensure_ascii=False)
                    atualizados += 1
                else:
                    f = Funcionario(
                        matricula=mat,
                        re=re_num,
                        nome=nome,
                        cpf=norm_cpf(str(row.get('cpf', '') or '').strip()),
                        email=str(row.get('email', '') or '').strip(),
                        telefone=wa_norm_number(str(row.get('telefone', '') or '').strip()),
                        cargo=str(row.get('cargo', '') or '').strip(),
                        funcao=str(row.get('funcao', '') or '').strip(),
                        cbo=str(row.get('cbo', '') or '').strip(),
                        setor=str(row.get('setor', '') or '').strip(),
                        empresa_id=to_num(row.get('empresa_id')) or None,
                        data_admissao=str(row.get('data_admissao', '') or '').strip(),
                        tipo_contrato=str(row.get('tipo_contrato', '') or '').strip(),
                        jornada=str(row.get('jornada', '') or '').strip(),
                        status=str(row.get('status', 'Ativo') or 'Ativo').strip() or 'Ativo',
                        posto_operacional='Reserva tecnica',
                        salario=to_num(row.get('salario'), dec=True),
                        vale_refeicao=to_num(row.get('vale_refeicao'), dec=True),
                        vale_alimentacao=to_num(row.get('vale_alimentacao'), dec=True),
                        vale_transporte=to_num(row.get('vale_transporte'), dec=True),
                        cep=norm_cep(str(row.get('cep', '') or '').strip()),
                        endereco=str(row.get('endereco', '') or '').strip(),
                        endereco_numero=str(row.get('endereco_numero', '') or '').strip(),
                        endereco_complemento=str(row.get('endereco_complemento', '') or '').strip(),
                        endereco_bairro=str(row.get('endereco_bairro', '') or '').strip(),
                        cidade=str(row.get('cidade', '') or '').strip(),
                        estado=norm_uf(row.get('estado', '')),
                        banco_codigo=norm_bank_code(row.get('banco_codigo')),
                        banco_nome=str(row.get('banco_nome', '') or '').strip(),
                        banco_agencia=str(row.get('banco_agencia', '') or '').strip(),
                        banco_conta=str(row.get('banco_conta', '') or '').strip(),
                        banco_tipo_conta=str(row.get('banco_tipo_conta', '') or '').strip(),
                        banco_pix=str(row.get('banco_pix', '') or '').strip(),
                        rg=str(row.get('rg', '') or '').strip(),
                        orgao_emissor=str(row.get('orgao_emissor', '') or '').strip(),
                        pis=str(row.get('pis', '') or '').strip(),
                        ctps=str(row.get('ctps', '') or '').strip(),
                        titulo_eleitor=str(row.get('titulo_eleitor', '') or '').strip(),
                        cert_reservista=str(row.get('cert_reservista', '') or '').strip(),
                        cnh=str(row.get('cnh', '') or '').strip(),
                        exame_admissional_data=str(row.get('exame_admissional_data', '') or '').strip(),
                        docs_admissao_ok=to_bool(row.get('docs_admissao_ok')),
                        docs_admissao_obs=str(row.get('docs_admissao_obs', '') or '').strip(),
                        obs=str(row.get('obs', '') or '').strip(),
                        areas=json.dumps(ars, ensure_ascii=False)
                    )
                    db.session.add(f)
                    criados += 1
            except Exception as e:
                erros.append(f'Linha {i}: {str(e)}')
        db.session.commit()
        return jsonify({'ok': True, 'criados': criados, 'atualizados': atualizados, 'ignorados': ignorados, 'modo': ('atualizar' if atualizar_existentes else 'adicionar'), 'erros': erros})
    except Exception as e:
        return jsonify({'erro': str(e)}), 400

# --- Fim da função de importação de funcionários ---

@app.route('/assinatura/<token>')
def assinatura_publica(token):
    m=Medicao.query.filter_by(assinatura_token=token).first()
    if not m:
        return render_template('assinatura.html',ok=False,mensagem='Link de assinatura invalido.',medicao=None)
    if (m.assinatura_status or '')=='assinado':
        return render_template('assinatura.html',ok=False,mensagem='Este documento ja foi assinado.',medicao=m)
    if m.assinatura_expira_em and m.assinatura_expira_em<utcnow():
        m.assinatura_status='expirado'
        db.session.commit()
        return render_template('assinatura.html',ok=False,mensagem='Link expirado. Solicite um novo link.',medicao=m)
    return render_template('assinatura.html',ok=True,mensagem='',medicao=m)

@app.route('/api/assinatura/<token>/enviar-otp',methods=['GET','POST'])
def api_assinatura_enviar_otp(token):
    m=Medicao.query.filter_by(assinatura_token=token).first()
    if not m:
        return _assinatura_json_erro('Link inválido.',404)
    if (m.assinatura_status or '')=='assinado':
        return _assinatura_json_erro('Documento já assinado.',400)
    if m.assinatura_expira_em and m.assinatura_expira_em<utcnow():
        return _assinatura_json_erro('Link expirado.',400)
    cli=Cliente.query.get(m.cliente_id) if m.cliente_id else None
    tel=wa_norm_number((cli.telefone if cli else '') or '')
    email=((getattr(cli,'email','') or '').strip() if cli else '')
    if not tel and not email:
        return _assinatura_json_erro('Nenhum telefone ou e-mail cadastrado para envio do OTP.',400)
    codigo=_otp_new_code()
    m.assinatura_otp_hash=token_hash(codigo)
    m.assinatura_otp_expira_em=utcnow()+timedelta(minutes=10)
    m.assinatura_otp_tentativas=0
    try:
        envio=_send_signature_otp(codigo,nome_dest=(cli.nome if cli else ''),telefone=tel,email=email,contexto='medicao')
        db.session.commit()
        return _assinatura_json_ok(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )
    except Exception as ex:
        db.session.rollback()
        return _assinatura_json_erro(f'Falha ao enviar OTP: {str(ex)}',500)

@app.route('/assinatura/validar/<codigo>')
def assinatura_validar_publica(codigo):
    cod=(codigo or '').strip()
    if not cod:
        return render_template('assinatura_validacao.html',ok=False,mensagem='Codigo de validacao invalido.',medicao=None)
    m=Medicao.query.filter_by(assinatura_codigo=cod).first()
    if not m:
        return render_template('assinatura_validacao.html',ok=False,mensagem='Assinatura nao encontrada para o codigo informado.',medicao=None)
    status=(m.assinatura_status or '').strip().lower()
    if status!='assinado':
        msg='A assinatura ainda nao foi concluida.' if status=='pendente' else ('O link de assinatura expirou.' if status=='expirado' else 'Documento sem assinatura valida no momento.')
        return render_template('assinatura_validacao.html',ok=False,mensagem=msg,medicao=m)
    return render_template('assinatura_validacao.html',ok=True,mensagem='Assinatura valida.',medicao=m)

@app.route('/api/assinatura/<token>/confirmar',methods=['POST'])
def api_assinatura_confirmar(token):
    m=Medicao.query.filter_by(assinatura_token=token).first()
    if not m:
        return _assinatura_json_erro('Link inválido.',404)
    if m.assinatura_expira_em and m.assinatura_expira_em<utcnow():
        m.assinatura_status='expirado'
        db.session.commit()
        return _assinatura_json_erro('Link expirado. Solicite um novo link.',400)
    d=request.json or {}
    nome=(d.get('nome') or '').strip()
    cargo=(d.get('cargo') or '').strip()
    cpf=(only_digits(d.get('cpf') or '') or '').strip()
    otp=(only_digits(d.get('otp') or '') or '').strip()
    aceite=bool(d.get('aceite'))
    if not nome:
        return _assinatura_json_erro('Informe o nome completo para assinar.',400)
    if not cpf or len(cpf)!=11 or not _valida_cpf(cpf):
        return _assinatura_json_erro('Informe um CPF válido (11 dígitos) para assinar.',400)
    if not aceite:
        return _assinatura_json_erro('Confirme o aceite para concluir a assinatura.',400)

    cli=Cliente.query.get(m.cliente_id) if m.cliente_id else None
    tel=wa_norm_number((cli.telefone if cli else '') or '')
    email=((getattr(cli,'email','') or '').strip() if cli else '')

    if not otp:
        codigo=_otp_new_code()
        m.assinatura_otp_hash=token_hash(codigo)
        m.assinatura_otp_expira_em=utcnow()+timedelta(minutes=10)
        m.assinatura_otp_tentativas=0
        try:
            envio=_send_signature_otp(codigo,nome_dest=nome,telefone=tel,email=email,contexto='medicao')
        except Exception as ex:
            db.session.rollback()
            return _assinatura_json_erro(f'Falha ao enviar OTP de confirmação: {str(ex)}',400)
        db.session.commit()
        return _assinatura_json_otp(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )

    if not (m.assinatura_otp_hash or '').strip() or not m.assinatura_otp_expira_em:
        return _assinatura_json_erro('Solicite um novo código OTP para concluir a assinatura.',400)
    if m.assinatura_otp_expira_em<utcnow():
        return _assinatura_json_erro('Código OTP expirado. Solicite um novo código.',400)
    tent=int(m.assinatura_otp_tentativas or 0)
    if tent>=5:
        return _assinatura_json_erro('Limite de tentativas de OTP excedido. Solicite um novo código.',400)
    if not hmac.compare_digest(token_hash(otp),str(m.assinatura_otp_hash or '')):
        m.assinatura_otp_tentativas=tent+1
        db.session.commit()
        return _assinatura_json_erro('Código OTP inválido.',400)

    m.assinatura_status='assinado'
    if not (m.assinatura_codigo or '').strip():
        m.assinatura_codigo=secrets.token_urlsafe(10)
    m.assinatura_nome=nome
    m.assinatura_cpf=cpf
    m.assinatura_cargo=cargo
    m.assinatura_ip=(request.headers.get('X-Forwarded-For','') or request.remote_addr or '').split(',')[0].strip()[:60]
    m.assinatura_em=utcnow()
    m.assinatura_otp_hash=None
    m.assinatura_otp_expira_em=None
    m.assinatura_otp_tentativas=0
    m.ass_cliente=f"{nome}{(' - '+cargo) if cargo else ''}".strip()
    m.assinatura_token=None
    db.session.commit()

    # Envia cópia da medição assinada para o assinante via WhatsApp, quando houver telefone válido.
    enviado_wa=False
    try:
        cli=Cliente.query.get(m.cliente_id) if m.cliente_id else None
        tel=wa_norm_number((cli.telefone if cli else '') or '')
        if tel and wa_is_valid_number(tel):
            emp=Empresa.query.get(m.empresa_id) if m.empresa_id else None
            d=m.to_dict()
            d['empresa']=emp.to_dict() if emp else {}
            pdf_resp=_build_pdf(d)
            pdf_bytes=b''
            try:
                pdf_resp.direct_passthrough=False
                pdf_bytes=pdf_resp.get_data()
            except Exception:
                pdf_bytes=b''
            if pdf_bytes:
                tmp_dir=os.path.join(UPLOAD_ROOT,'tmp_ass')
                os.makedirs(tmp_dir,exist_ok=True)
                nome_pdf=f"medicao_{(m.numero or m.id)}_assinada.pdf".replace('/','-')
                tmp_file=os.path.join(tmp_dir,nome_pdf)
                with open(tmp_file,'wb') as fp:
                    fp.write(pdf_bytes)
                validacao_link=f"{request.url_root.rstrip('/')}/assinatura/validar/{m.assinatura_codigo}"
                wa_send_pdf(tel,tmp_file,nome_pdf,
                    f"✅ Medição assinada com sucesso.\nCódigo: {m.assinatura_codigo}\nValidar: {validacao_link}")
                try:
                    os.remove(tmp_file)
                except Exception:
                    pass
                enviado_wa=True
    except Exception:
        enviado_wa=False

    audit_event('medicao_assinatura_confirmada','externo',None,'medicao',m.id,True,{'numero':m.numero,'nome':nome})
    validacao_link=f"{request.url_root.rstrip('/')}/assinatura/validar/{m.assinatura_codigo}"
    return _assinatura_json_ok(
        mensagem='Assinatura concluída com sucesso.',
        validacao_link=validacao_link,
        whatsapp_enviado=enviado_wa,
        codigo=(m.assinatura_codigo or '')
    )

@app.route('/api/funcionarios',methods=['GET'])
@lr
def api_funcionarios():
    cpf=only_digits(request.args.get('cpf',''))
    if cpf:
        ex_id=to_num(request.args.get('exclude_id'))
        for f in Funcionario.query.all():
            if f.id==ex_id: continue
            if only_digits(f.cpf)==cpf: return jsonify(f.to_dict())
        return jsonify({})
    q=(request.args.get('q','') or '').lower()
    lst=Funcionario.query.order_by(Funcionario.nome).all()
    if q:
        qdig=only_digits(q)
        lst=[f for f in lst if (
            q in (f.matricula or '').lower() or
            q in (f.nome or '').lower() or
            q in (f.cpf or '').lower() or
            q in (f.cargo or '').lower() or
            q in (f.funcao or '').lower() or
            q in (f.posto_operacional or '').lower() or
            q in (f.telefone or '').lower() or
            (qdig and (qdig in only_digits(f.matricula) or qdig in only_digits(f.cpf) or qdig in only_digits(f.telefone)))
        )]
    return jsonify([f.to_dict() for f in lst])

def _funcionarios_ativos_filtrados_export():
    empresa_id=to_num(request.args.get('empresa_id')) or None

    postos_raw=[]
    csv_postos=(request.args.get('postos') or '').strip()
    if csv_postos:
        postos_raw.extend([x.strip() for x in csv_postos.split(',') if x.strip()])
    postos_raw.extend([(x or '').strip() for x in request.args.getlist('posto') if (x or '').strip()])

    postos_norm=[]
    seen=set()
    for p in postos_raw:
        k=p.lower()
        if k and k not in seen:
            seen.add(k)
            postos_norm.append(k)

    q=Funcionario.query.filter_by(status='Ativo')
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    lst=q.order_by(Funcionario.nome).all()

    if postos_norm:
        set_postos=set(postos_norm)
        lst=[f for f in lst if ((f.posto_operacional or 'Reserva tecnica').strip().lower() in set_postos)]

    return lst

def _export_funcionarios_ativos_xlsx(funcs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb=Workbook()
    ws=wb.active
    ws.title='Colaboradores ativos'

    headers=['RE','Matrícula','Nome','CPF','Telefone','Cargo','Função','Posto','Empresa','Status']
    ws.append(['Relatório de colaboradores ativos'])
    ws.append([f'Gerado em: {localnow().strftime("%d/%m/%Y %H:%M") }'])
    ws.append([])
    ws.append(headers)

    emps_map={e.id:e.nome for e in Empresa.query.all()}
    for f in funcs:
        ws.append([
            f.re or '',
            f.matricula or '',
            f.nome or '',
            f.cpf or '',
            f.telefone or '',
            f.cargo or '',
            f.funcao or '',
            f.posto_operacional or 'Reserva tecnica',
            emps_map.get(f.empresa_id,'') if f.empresa_id else '',
            f.status or 'Ativo',
        ])

    header_fill=PatternFill('solid',fgColor='205D8A')
    header_font=Font(bold=True,color='FFFFFF')
    center=Alignment(horizontal='center',vertical='center')
    left=Alignment(horizontal='left',vertical='center')
    thin=Side(style='thin',color='D0D7DE')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    hrow=4
    for c in range(1,len(headers)+1):
        cell=ws.cell(row=hrow,column=c)
        cell.fill=header_fill
        cell.font=header_font
        cell.alignment=center
        cell.border=border

    for r in range(hrow+1,ws.max_row+1):
        for c in range(1,len(headers)+1):
            cell=ws.cell(row=r,column=c)
            cell.border=border
            cell.alignment=left

    widths=[10,12,34,18,16,20,20,24,26,10]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome=f'colaboradores_ativos_{localnow().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=nome)

def _export_funcionarios_ativos_pdf(funcs):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
    from reportlab.lib.styles import ParagraphStyle

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.2*cm,rightMargin=1.2*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    st_t=ParagraphStyle('t',fontName='Helvetica-Bold',fontSize=12,leading=14,textColor=colors.HexColor('#123B60'))
    st_s=ParagraphStyle('s',fontName='Helvetica',fontSize=9,leading=11,textColor=colors.HexColor('#4A5A6A'))
    st_c=ParagraphStyle('c',fontName='Helvetica',fontSize=8,leading=10)

    emps_map={e.id:e.nome for e in Empresa.query.all()}
    data=[[Paragraph('<b>RE</b>',st_c),Paragraph('<b>Nome</b>',st_c),Paragraph('<b>Posto</b>',st_c),Paragraph('<b>Empresa</b>',st_c),Paragraph('<b>Telefone</b>',st_c)]]
    for f in funcs:
        data.append([
            Paragraph(str(f.re or '—'),st_c),
            Paragraph((f.nome or '—')[:80],st_c),
            Paragraph((f.posto_operacional or 'Reserva tecnica')[:60],st_c),
            Paragraph((emps_map.get(f.empresa_id,'') if f.empresa_id else '—')[:60],st_c),
            Paragraph(str(f.telefone or '—'),st_c),
        ])

    table=Table(data,colWidths=[1.8*cm,5.9*cm,4.6*cm,4.7*cm,3.0*cm],repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#205D8A')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#D0D7DE')),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,1),(-1,-1),8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F8FBFF')]),
        ('LEFTPADDING',(0,0),(-1,-1),4),
        ('RIGHTPADDING',(0,0),(-1,-1),4),
        ('TOPPADDING',(0,0),(-1,-1),3),
        ('BOTTOMPADDING',(0,0),(-1,-1),3),
    ]))

    elementos=[
        Paragraph('Relatório de colaboradores ativos',st_t),
        Paragraph(f'Gerado em {localnow().strftime("%d/%m/%Y %H:%M")} · Total: {len(funcs)} colaborador(es)',st_s),
        Spacer(1,0.3*cm),
        table
    ]
    doc.build(elementos)
    buf.seek(0)
    nome=f'colaboradores_ativos_{localnow().strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(buf,mimetype='application/pdf',as_attachment=False,download_name=nome)

@app.route('/api/funcionarios/ativos/exportar')
@lr
def api_funcionarios_ativos_exportar():
    formato=(request.args.get('formato') or 'xlsx').strip().lower()
    if formato not in ('xlsx','pdf'):
        return jsonify({'erro':'Formato inválido. Use xlsx ou pdf.'}),400

    funcs=_funcionarios_ativos_filtrados_export()
    if not funcs:
        return jsonify({'erro':'Nenhum colaborador ativo encontrado para os filtros informados.'}),404

    if formato=='pdf':
        return _export_funcionarios_ativos_pdf(funcs)
    return _export_funcionarios_ativos_xlsx(funcs)

@app.route('/api/funcionarios/proxima-matricula')
@lr
def api_funcionarios_proxima_matricula():
    return jsonify({'ok':True,'matricula':next_func_matricula()})

@app.route('/api/funcionarios/proxima-re')
@lr
def api_funcionarios_proxima_re():
    return jsonify({'ok':True,'re':next_func_re()})

@app.route('/api/funcionarios/cpf-lookup')
@lr
def api_funcionario_cpf_lookup():
    cpf=only_digits(request.args.get('cpf',''))
    if len(cpf)!=11: return jsonify({'erro':'CPF invalido'}),400
    ex_id=to_num(request.args.get('exclude_id'))
    for f in Funcionario.query.all():
        if f.id==ex_id: continue
        if only_digits(f.cpf)==cpf:
            return jsonify({'ok':True,'origem':'interno','funcionario':f.to_dict()})
    r=lookup_cpf_externo(cpf)
    return jsonify(r)

@app.route('/api/funcionarios/busca-rapida')
@lr
def api_funcionario_busca_rapida():
    re=to_num(request.args.get('re'))
    cpf=only_digits(request.args.get('cpf',''))
    nome=(request.args.get('nome','') or '').strip().lower()
    
    resultados=[]
    
    # Busca por RE se fornecido
    if re and re>0:
        f=Funcionario.query.filter_by(re=re).first()
        if f:
            resultados.append(f.to_dict())
    
    # Busca por CPF se fornecido
    if cpf and len(cpf)==11:
        f=Funcionario.query.filter(Funcionario.cpf.ilike(f'%{cpf}%')).first()
        if f and f.to_dict() not in resultados:
            resultados.append(f.to_dict())
    
    # Busca por nome se fornecido
    if nome:
        funcs=Funcionario.query.filter(Funcionario.nome.ilike(f'%{nome}%')).all()
        for f in funcs:
            if f.to_dict() not in resultados:
                resultados.append(f.to_dict())
    
    return jsonify(resultados if resultados else [])

@app.route('/api/funcionarios',methods=['POST'])
@lr
def api_criar_funcionario():
    d=request.json or {}
    if not d.get('nome'): return jsonify({'erro':'Nome obrigatorio'}),400
    ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
    mat=next_func_matricula()
    f=Funcionario(
        matricula=mat,
        re=to_num(d.get('re')) or next_func_re(),
        nome=d.get('nome','').strip(),
        cpf=norm_cpf(d.get('cpf','')),
        email=d.get('email','').strip(),
        telefone=wa_norm_number(d.get('telefone','')),
        cargo=d.get('cargo','').strip(),
        funcao=d.get('funcao','').strip(),
        cbo=d.get('cbo','').strip(),
        setor=d.get('setor','').strip(),
        empresa_id=d.get('empresa_id'),
        data_admissao=d.get('data_admissao',''),
        tipo_contrato=d.get('tipo_contrato','').strip(),
        jornada=d.get('jornada','').strip(),
        status=d.get('status','Ativo'),
        posto_operacional='Reserva tecnica',
        salario=to_num(d.get('salario'),dec=True),
        vale_refeicao=to_num(d.get('vale_refeicao'),dec=True),
        vale_alimentacao=to_num(d.get('vale_alimentacao'),dec=True),
        vale_transporte=to_num(d.get('vale_transporte'),dec=True),
        endereco=d.get('endereco','').strip(),
        cidade=d.get('cidade','').strip(),
        estado=norm_uf(d.get('estado','')),
        cep=norm_cep(d.get('cep','')),
        endereco_numero=d.get('endereco_numero','').strip(),
        endereco_complemento=d.get('endereco_complemento','').strip(),
        endereco_bairro=d.get('endereco_bairro','').strip(),
        banco_codigo=norm_bank_code(d.get('banco_codigo','')),
        banco_nome=d.get('banco_nome','').strip(),
        banco_agencia=d.get('banco_agencia','').strip(),
        banco_conta=d.get('banco_conta','').strip(),
        banco_tipo_conta=d.get('banco_tipo_conta','').strip(),
        banco_pix=d.get('banco_pix','').strip(),
        rg=d.get('rg','').strip(),
        orgao_emissor=d.get('orgao_emissor','').strip(),
        pis=d.get('pis','').strip(),
        ctps=d.get('ctps','').strip(),
        titulo_eleitor=d.get('titulo_eleitor','').strip(),
        cert_reservista=d.get('cert_reservista','').strip(),
        cnh=d.get('cnh','').strip(),
        exame_admissional_data=d.get('exame_admissional_data','').strip(),
        docs_admissao_ok=to_bool(d.get('docs_admissao_ok')),
        docs_admissao_obs=d.get('docs_admissao_obs','').strip(),
        obs=d.get('obs','').strip(),
        areas=json.dumps(ars,ensure_ascii=False)
    )
    try:
        db.session.add(f)
        db.session.commit()
        return jsonify(f.to_dict()),201
    except IntegrityError as e:
        db.session.rollback()
        msg=str(e).lower()
        if 'funcionario.re' in msg:
            return jsonify({'erro':'RE já cadastrado. Informe outro RE.'}),400
        if 'funcionario.cpf' in msg:
            return jsonify({'erro':'CPF já cadastrado para outro funcionário.'}),400
        if 'funcionario.matricula' in msg:
            return jsonify({'erro':'Matrícula já cadastrada. Informe outra matrícula.'}),400
        return jsonify({'erro':'Não foi possível salvar o funcionário (dados duplicados).'}),400

@app.route('/api/funcionarios/<int:id>',methods=['PUT'])
@lr
def api_atualizar_funcionario(id):
    f=Funcionario.query.get_or_404(id); d=request.json or {}
    for k in ['re','nome','cpf','email','telefone','cargo','funcao','cbo','setor','empresa_id','data_admissao','tipo_contrato','jornada','status','endereco','endereco_numero','endereco_complemento','endereco_bairro','cidade','estado','cep','banco_codigo','banco_nome','banco_agencia','banco_conta','banco_tipo_conta','banco_pix','rg','orgao_emissor','pis','ctps','titulo_eleitor','cert_reservista','cnh','exame_admissional_data','docs_admissao_obs','obs']:
        if k in d:
            if k=='cpf': setattr(f,k,norm_cpf(d.get(k)))
            elif k=='re': setattr(f,k,to_num(d.get(k)))
            elif k=='telefone': setattr(f,k,wa_norm_number(d.get(k)))
            elif k=='cep': setattr(f,k,norm_cep(d.get(k)))
            elif k=='estado': setattr(f,k,norm_uf(d.get(k)))
            elif k=='banco_codigo': setattr(f,k,norm_bank_code(d.get(k)))
            else: setattr(f,k,d[k])
    if 'salario' in d: f.salario=to_num(d.get('salario'),dec=True)
    if 'vale_refeicao' in d: f.vale_refeicao=to_num(d.get('vale_refeicao'),dec=True)
    if 'vale_alimentacao' in d: f.vale_alimentacao=to_num(d.get('vale_alimentacao'),dec=True)
    if 'vale_transporte' in d: f.vale_transporte=to_num(d.get('vale_transporte'),dec=True)
    if 'opta_vt' in d: f.opta_vt=to_bool(d.get('opta_vt'))
    if 'opta_vr' in d: f.opta_vr=to_bool(d.get('opta_vr'))
    if 'opta_va' in d: f.opta_va=to_bool(d.get('opta_va'))
    if 'opta_premio_prod' in d: f.opta_premio_prod=to_bool(d.get('opta_premio_prod'))
    if 'opta_vale_gasolina' in d: f.opta_vale_gasolina=to_bool(d.get('opta_vale_gasolina'))
    if 'opta_cesta_natal' in d: f.opta_cesta_natal=to_bool(d.get('opta_cesta_natal'))
    if 'premio_produtividade' in d: f.premio_produtividade=to_num(d.get('premio_produtividade'),dec=True)
    if 'vale_gasolina' in d: f.vale_gasolina=to_num(d.get('vale_gasolina'),dec=True)
    if 'cesta_natal' in d: f.cesta_natal=to_num(d.get('cesta_natal'),dec=True)
    if 'docs_admissao_ok' in d: f.docs_admissao_ok=to_bool(d.get('docs_admissao_ok'))
    if 'areas' in d:
        ars=[a for a in d.get('areas',[]) if a in ALLOWED_AREAS]
        f.areas=json.dumps(ars,ensure_ascii=False)
    try:
        db.session.commit()
        return jsonify(f.to_dict())
    except IntegrityError as e:
        db.session.rollback()
        msg=str(e).lower()
        if 'funcionario.re' in msg:
            return jsonify({'erro':'RE já cadastrado. Informe outro RE.'}),400
        if 'funcionario.cpf' in msg:
            return jsonify({'erro':'CPF já cadastrado para outro funcionário.'}),400
        if 'funcionario.matricula' in msg:
            return jsonify({'erro':'Matrícula já cadastrada. Informe outra matrícula.'}),400
        return jsonify({'erro':'Não foi possível atualizar o funcionário (dados duplicados).'}),400

@app.route('/api/funcionarios/<int:id>',methods=['DELETE'])
@lr
def api_deletar_funcionario(id):
    f=Funcionario.query.get_or_404(id)
    arqs=FuncionarioArquivo.query.filter_by(funcionario_id=id).all()
    for a in arqs:
        try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
        except: pass
        db.session.delete(a)
    PontoMarcacao.query.filter_by(funcionario_id=id).delete()
    PontoAjuste.query.filter_by(funcionario_id=id).delete()
    PontoFechamentoDia.query.filter_by(funcionario_id=id).delete()
    db.session.delete(f); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/funcionarios/<int:id>/arquivos',methods=['GET'])
@lr
def api_funcionario_arquivos(id):
    Funcionario.query.get_or_404(id)
    return jsonify([a.to_dict() for a in FuncionarioArquivo.query.filter_by(funcionario_id=id).order_by(FuncionarioArquivo.criado_em.desc()).all()])

@app.route('/api/funcionarios/<int:id>/arquivos',methods=['POST'])
@lr
def api_funcionario_upload_arquivo(id):
    f=Funcionario.query.get_or_404(id)
    fs=request.files.get('arquivo')
    if not fs: return jsonify({'erro':'Arquivo nao enviado'}),400
    cat=(request.form.get('categoria') or 'outros').strip().lower()
    comp_in=(request.form.get('competencia') or '').strip()
    canal_ass=(request.form.get('canal_assinatura') or 'whatsapp').strip().lower()
    if canal_ass not in ('whatsapp','link','nao','app'):
        canal_ass='whatsapp'
    prazo_dias_raw=request.form.get('prazo_dias') or ''
    try:
        prazo_dias=max(1,min(90,int(prazo_dias_raw))) if prazo_dias_raw.strip() else None
    except (ValueError, TypeError):
        prazo_dias=None
    texto=''
    if not comp_in and str((fs.filename or '')).lower().endswith('.pdf') and _upload_is_pdf(fs):
        try:
            from pypdf import PdfReader
            import io
            blob=fs.read()
            try:
                fs.seek(0)
            except Exception:
                pass
            reader=PdfReader(io.BytesIO(blob))
            texto=' '.join((p.extract_text() or '') for p in reader.pages[:5])
        except Exception:
            texto=''
    comp,comp_origem=_resolver_competencia_envio(comp_in=comp_in,texto=texto,nome_arquivo=(fs.filename or ''))
    ano=infer_doc_year(comp)
    prepare_func_doc_dirs(id,ano)
    subdir,cat=func_doc_subdir(id,cat,comp)
    rel,_=save_upload(fs,subdir)
    a=FuncionarioArquivo(funcionario_id=id,categoria=cat,competencia=comp,nome_arquivo=fs.filename,caminho=rel)
    db.session.add(a); db.session.commit()
    assinatura_auto={'status':'nao_solicitada'}
    if canal_ass=='app':
        a.ass_status='pendente'
        if prazo_dias:
            a.ass_prazo_em=utcnow()+timedelta(days=prazo_dias)
        _ass_track_mark_sent(a,'app')
        db.session.commit()
        _push_notify_funcionario(
            f.id,
            'Documento para assinar',
            f'{a.nome_arquivo} aguarda sua assinatura no app.',
            {'tipo':'documento_assinar','arquivo_id':str(a.id)}
        )
        assinatura_auto={'status':'app_pendente','canal':'app'}
    elif canal_ass in ('whatsapp','link'):
        try:
            rs=_solicitar_assinatura_arquivo_funcionario(a,f,canal=canal_ass,commit_now=True)
            if rs.get('ok'):
                assinatura_auto={
                    'status':('solicitada' if canal_ass=='whatsapp' else 'link_gerado'),
                    'link':(rs.get('link_curto') or rs.get('link','')),
                    'canal':canal_ass,
                }
            else:
                assinatura_auto={'status':'erro','erro':rs.get('erro',''),'link':rs.get('link','')}
        except Exception as e:
            assinatura_auto={'status':'erro','erro':str(e)}
        # Notificar funcionário de novo documento (quando não é 'app', pois 'app' já notifica acima)
        try:
            _push_notify_funcionario(
                f.id,
                'Novo documento disponível',
                f'{a.nome_arquivo} foi adicionado ao seu perfil.',
                {'tipo':'novo_documento','arquivo_id':str(a.id)}
            )
        except Exception:
            pass
    elif canal_ass=='nao':
        # Sem assinatura — apenas notificar sobre o novo documento
        try:
            _push_notify_funcionario(
                f.id,
                'Novo documento disponível',
                f'{a.nome_arquivo} foi adicionado ao seu perfil.',
                {'tipo':'novo_documento','arquivo_id':str(a.id)}
            )
        except Exception:
            pass
    audit_event('funcionario_arquivo_upload','usuario',session.get('uid'),'funcionario',id,True,{'arquivo_id':a.id,'categoria':cat,'caminho':rel})
    out=a.to_dict(); out['assinatura_auto']=assinatura_auto; out['competencia_origem']=comp_origem
    return jsonify(out),201


def _solicitar_assinatura_arquivo_funcionario(arquivo,funcionario,canal='link',dias_validade=7,commit_now=True,forcar_novo_token=True,eh_lembrete=False):
    if not arquivo:
        return {'ok':False,'erro':'Arquivo invalido.'}
    if not funcionario:
        funcionario=Funcionario.query.get(arquivo.funcionario_id)
    if (arquivo.ass_status or '')=='assinado':
        return {'ok':False,'erro':'Documento ja assinado.'}
    canal=(canal or 'link').strip().lower()
    if canal not in ('link','whatsapp','email','app'):
        canal='link'
    tel=''
    email=''
    if canal=='whatsapp':
        tel=wa_norm_number((funcionario.telefone if funcionario else '') or '')
        if not wa_is_valid_number(tel):
            return {'ok':False,'erro':'Funcionario sem WhatsApp valido cadastrado.'}
    if canal=='email':
        email=((funcionario.email if funcionario else '') or '').strip()
        if not email or '@' not in email:
            return {'ok':False,'erro':'Funcionario sem e-mail valido cadastrado.'}
    if canal=='app':
        if not funcionario:
            return {'ok':False,'erro':'Funcionario invalido para envio no app.'}
    if not arquivo.ass_codigo:
        arquivo.ass_codigo=secrets.token_urlsafe(10)
    if forcar_novo_token or not (arquivo.ass_token or '').strip():
        arquivo.ass_token=secrets.token_urlsafe(24)
    arquivo.ass_status='pendente'
    arquivo.ass_expira_em=utcnow()+timedelta(days=max(1,int(dias_validade or 7)))
    _ass_track_mark_sent(arquivo,canal)

    src_q=_ass_track_channel(canal)
    if has_request_context():
        base_url=request.url_root.rstrip('/')
    else:
        base_url=(
            (os.environ.get('PUBLIC_BASE_URL') or '').strip() or
            (gc('public_base_url','') or '').strip() or
            (os.environ.get('APP_BASE_URL') or '').strip() or
            'https://portal.grupormfacilities.com.br'
        ).rstrip('/')
    link=f"{base_url}/doc/assinar/{arquivo.ass_token}?src={src_q}"
    try:
        sc=_short_link_criar(link)
        link_curto=(f"{base_url}/s/{sc}" if sc else link)
    except Exception:
        link_curto=link
    enviado_wa=False
    enviado_email=False
    enviado_app=False
    erro_envio=''
    if canal=='whatsapp':
        nome_func=(funcionario.nome if funcionario else 'colaborador')
        if eh_lembrete:
            msg=(f"🔔 Lembrete de envio anterior\n"
                 f"Olá, {nome_func}! Este e um lembrete para assinatura do documento "
                 f"'{arquivo.nome_arquivo}'. O link expira em 7 dias: {link_curto}")
        else:
            msg=(f"Olá, {nome_func}! Segue o link para assinatura do documento "
                 f"'{arquivo.nome_arquivo}'. O link expira em 7 dias: {link_curto}")
        try:
            wa_send_text(tel,msg)
            enviado_wa=True
        except Exception as ex:
            # Mantem assinatura pendente com link ativo mesmo se o WhatsApp falhar.
            erro_envio=str(ex)
    elif canal=='email':
        nome_func=(funcionario.nome if funcionario else 'colaborador')
        try:
            smtp_send_link_assinatura(
                email,
                nome_func,
                arquivo.nome_arquivo or 'Documento',
                link_curto,
                eh_lembrete=eh_lembrete,
            )
            enviado_email=True
        except Exception as ex:
            erro_envio=str(ex)
    elif canal=='app':
        try:
            titulo_push='Lembrete: documento para assinar' if eh_lembrete else 'Documento para assinar'
            corpo_push=(
                f"Lembrete de envio anterior: {arquivo.nome_arquivo} ainda aguarda sua assinatura no app."
                if eh_lembrete else
                f"{arquivo.nome_arquivo} aguarda sua assinatura no app."
            )
            enviado_app=bool(_push_notify_funcionario(
                funcionario.id,
                titulo_push,
                corpo_push,
                {'tipo':'documento_assinar','arquivo_id':str(arquivo.id)}
            ))
            if not enviado_app:
                erro_envio='Falha ao enviar notificacao push para o aplicativo.'
        except Exception as ex:
            erro_envio=str(ex)

    if commit_now:
        db.session.commit()
    else:
        db.session.flush()
    return {
        'ok':True,
        'link':link,
        'link_curto':link_curto,
        'canal':canal,
        'expira_em':(arquivo.ass_expira_em.isoformat() if arquivo.ass_expira_em else ''),
        'enviado_wa':enviado_wa,
        'enviado_email':enviado_email,
        'enviado_app':enviado_app,
        'erro_envio':erro_envio,
    }

@app.route('/api/funcionarios/<int:id>/documentos/preparar',methods=['POST'])
@lr
def api_preparar_pastas_funcionario(id):
    Funcionario.query.get_or_404(id)
    d=request.json or {}
    ano=str((d.get('ano') or request.args.get('ano') or localnow().year)).strip()
    if not re.fullmatch(r'(19|20)\d{2}',ano):
        return jsonify({'erro':'Ano invalido'}),400
    pastas=prepare_func_doc_dirs(id,ano)
    return jsonify({'ok':True,'ano':ano,'pastas':pastas})

@app.route('/api/funcionarios/<int:id>/documentos/arvore')
@lr
def api_funcionario_documentos_arvore(id):
    Funcionario.query.get_or_404(id)
    resp,status=build_func_docs_response(id)
    return jsonify(resp),status

@app.route('/api/funcionarios/<int:id>/app-acesso',methods=['PUT'])
@lr
def api_funcionario_app_acesso(id):
    f=Funcionario.query.get_or_404(id)
    d=request.json or {}
    if 'ativo_app' in d:
        f.app_ativo=bool(d.get('ativo_app'))
    senha=(d.get('senha_app') or '').strip()
    if senha:
        if len(senha)<6: return jsonify({'erro':'Senha do app deve ter ao menos 6 caracteres'}),400
        f.app_senha=pw_hash(senha)
    db.session.commit()
    audit_event('funcionario_app_acesso_alterado','usuario',session.get('uid'),'funcionario',f.id,True,{'ativo_app':f.app_ativo,'senha_alterada':bool(senha)})
    return jsonify({'ok':True,'funcionario':f.to_dict()})

@app.route('/api/funcionarios/<int:id>/push-teste',methods=['POST'])
@app.route('/api/funcionarios/<int:id>/push-validar',methods=['POST'])
@lr
def api_funcionario_push_teste(id):
    f=Funcionario.query.get_or_404(id)
    token_antes=(f.app_push_token or '').strip()
    tem_token=bool(token_antes)
    if not token_antes:
        return jsonify({'ok':False,'erro':'Funcionario sem token push salvo'}),400
    ok=_push_notify_funcionario(
        f.id,
        'Teste de notificacao',
        'Se voce recebeu esta mensagem, o push de documentos esta funcionando.',
        {'tipo':'documento_assinar','arquivo_id':'0','origem':'teste_push_admin'}
    )
    f2=Funcionario.query.get(f.id)
    token_depois=((f2.app_push_token or '').strip() if f2 else '')
    token_removido=bool(token_antes and not token_depois)
    status='enviado' if ok else ('token_invalido_removido' if token_removido else 'falha_envio')
    return jsonify({
        'ok':ok,
        'status':status,
        'funcionario_id':f.id,
        'tem_token':tem_token,
        'token_valido':bool(token_depois),
        'token_removido':token_removido,
    })

@app.route('/api/app/versao')
def api_app_versao():
    """Retorna versão mínima e atual do app. Configurável por variável de ambiente APP_VERSION_CODE."""
    import os
    versao_minima=int(os.environ.get('APP_VERSION_MINIMA','14'))
    versao_atual=int(os.environ.get('APP_VERSION_ATUAL','14'))
    download_url=os.environ.get('APP_DOWNLOAD_URL','')
    return jsonify({'versao_minima':versao_minima,'versao_atual':versao_atual,'download_url':download_url})

@app.route('/api/app/funcionario/login',methods=['POST'])
def api_app_funcionario_login():
    d=request.json or {}
    cpf=norm_cpf(d.get('cpf'))
    senha=(d.get('senha') or '')
    if not cpf or not senha:
        return jsonify({'erro':'CPF e senha obrigatorios'}),400
    if auth_blocked('app',cpf,(request.remote_addr or '')):
        audit_event('auth_app_bloqueado','funcionario',cpf,'funcionario','',False,{'motivo':'rate_limit'})
        return jsonify({'erro':'Muitas tentativas. Aguarde alguns minutos.'}),429
    f=Funcionario.query.filter_by(cpf=cpf).first()
    if not f:
        reg_auth_attempt('app',cpf,False,'credenciais_invalidas')
        audit_event('auth_app_falha','funcionario',cpf,'funcionario','',False,{'motivo':'nao_encontrado'})
        return jsonify({'erro':'Credenciais invalidas'}),401
    if f.app_ativo is False:
        reg_auth_attempt('app',cpf,False,'app_desativado')
        audit_event('auth_app_falha','funcionario',f.id,'funcionario',f.id,False,{'motivo':'app_desativado'})
        return jsonify({'erro':'Acesso do aplicativo desativado'}),403
    if not f.app_senha:
        reg_auth_attempt('app',cpf,False,'app_nao_configurado')
        return jsonify({'erro':'Acesso do aplicativo ainda nao configurado para este funcionario'}),403
    if not pw_check(f.app_senha,senha):
        reg_auth_attempt('app',cpf,False,'credenciais_invalidas')
        audit_event('auth_app_falha','funcionario',f.id,'funcionario',f.id,False,{'motivo':'senha_invalida'})
        return jsonify({'erro':'Credenciais invalidas'}),401
    if not pw_is_modern(f.app_senha):
        f.app_senha=pw_hash(senha)

    sess=_app_issue_session_tokens(f)
    db.session.commit()
    reg_auth_attempt('app',cpf,True,'ok')
    audit_event('auth_app_sucesso','funcionario',f.id,'funcionario',f.id,True,{'sessao_id':sess['sessao_id'],'modo':'senha'})
    return jsonify({'ok':True,**sess,'funcionario':{'id':f.id,'nome':f.nome,'cpf':f.cpf,'cargo':f.cargo,'setor':f.setor,'status':f.status}})

@app.route('/api/app/funcionario/auth/iniciar',methods=['POST'])
def api_app_funcionario_auth_iniciar():
    d=request.json or {}
    cpf=norm_cpf(d.get('cpf'))
    if not cpf or len(cpf)!=11:
        return jsonify({'erro':'CPF obrigatorio (11 digitos).'}),400
    if auth_blocked('app_otp',cpf,(request.remote_addr or '')):
        return jsonify({'erro':'Muitas tentativas. Aguarde alguns minutos.'}),429

    f=Funcionario.query.filter_by(cpf=cpf).first()
    if not f:
        reg_auth_attempt('app_otp',cpf,False,'nao_encontrado')
        return jsonify({'erro':'Funcionario nao encontrado para o CPF informado.'}),404
    if f.app_ativo is False:
        reg_auth_attempt('app_otp',cpf,False,'app_desativado')
        return jsonify({'erro':'Acesso do aplicativo desativado.'}),403

    codigo=_otp_new_code()
    f.app_otp_hash=token_hash(codigo)
    f.app_otp_expira_em=utcnow()+timedelta(minutes=10)
    f.app_otp_tentativas=0
    try:
        envio=_send_app_login_otp(codigo,f)
        db.session.commit()
        reg_auth_attempt('app_otp',cpf,True,'desafio_enviado')
        audit_event('auth_app_otp_enviado','funcionario',f.id,'funcionario',f.id,True,{'canal':envio.get('canal')})
        msg_ok='Codigo enviado com sucesso.' if envio.get('canal')!='email' else 'Codigo enviado para o e-mail cadastrado.'
        return jsonify({'ok':True,'mensagem':msg_ok,'destino':envio.get('destino'),'canal':envio.get('canal')})
    except Exception as ex:
        db.session.rollback()
        reg_auth_attempt('app_otp',cpf,False,'envio_falha')
        return jsonify({'erro':'Nao foi possivel enviar o codigo OTP por WhatsApp. Verifique telefone com o RH e as configuracoes do WhatsApp.','detalhe':str(ex)}),503

@app.route('/api/app/funcionario/auth/confirmar',methods=['POST'])
def api_app_funcionario_auth_confirmar():
    d=request.json or {}
    cpf=norm_cpf(d.get('cpf'))
    codigo=only_digits(d.get('codigo'))
    if not cpf or len(cpf)!=11 or not codigo:
        return jsonify({'erro':'CPF e codigo OTP sao obrigatorios.'}),400
    if auth_blocked('app_otp_confirm',cpf,(request.remote_addr or '')):
        return jsonify({'erro':'Muitas tentativas. Aguarde alguns minutos.'}),429

    f=Funcionario.query.filter_by(cpf=cpf).first()
    if not f:
        reg_auth_attempt('app_otp_confirm',cpf,False,'nao_encontrado')
        return jsonify({'erro':'Funcionario nao encontrado.'}),404
    if f.app_ativo is False:
        reg_auth_attempt('app_otp_confirm',cpf,False,'app_desativado')
        return jsonify({'erro':'Acesso do aplicativo desativado.'}),403

    if not (f.app_otp_hash or '').strip() or not f.app_otp_expira_em:
        return jsonify({'erro':'Solicite um novo codigo para acessar.'}),400
    if f.app_otp_expira_em<utcnow():
        return jsonify({'erro':'Codigo expirado. Solicite um novo codigo.'}),400

    tent=int(f.app_otp_tentativas or 0)
    if tent>=5:
        return jsonify({'erro':'Limite de tentativas excedido. Solicite novo codigo.'}),400
    if not hmac.compare_digest(token_hash(codigo),str(f.app_otp_hash or '')):
        f.app_otp_tentativas=tent+1
        db.session.commit()
        reg_auth_attempt('app_otp_confirm',cpf,False,'codigo_invalido')
        return jsonify({'erro':'Codigo OTP invalido.'}),401

    f.app_otp_hash=None
    f.app_otp_expira_em=None
    f.app_otp_tentativas=0
    sess=_app_issue_session_tokens(f)
    db.session.commit()
    reg_auth_attempt('app_otp_confirm',cpf,True,'ok')
    audit_event('auth_app_sucesso','funcionario',f.id,'funcionario',f.id,True,{'sessao_id':sess['sessao_id'],'modo':'otp'})
    return jsonify({'ok':True,**sess,'funcionario':{'id':f.id,'nome':f.nome,'cpf':f.cpf,'cargo':f.cargo,'setor':f.setor,'status':f.status}})

@app.route('/api/app/funcionario/refresh',methods=['POST'])
def api_app_funcionario_refresh():
    d=request.json or {}
    refresh=(d.get('refresh_token') or request.headers.get('X-Refresh-Token') or '').strip()
    if not refresh: return jsonify({'erro':'refresh_token obrigatorio'}),400
    sessao=FuncionarioAppSessao.query.filter_by(refresh_hash=token_hash(refresh),revogado=False).first()
    if not sessao or sessao.exp_refresh<utcnow(): return jsonify({'erro':'Refresh token invalido ou expirado'}),401
    f=Funcionario.query.get(sessao.funcionario_id)
    if not f or f.app_ativo is False: return jsonify({'erro':'Acesso desativado'}),403
    novo_refresh=app_issue_refresh_token()
    sessao.refresh_hash=token_hash(novo_refresh)
    access=app_issue_access_token(f.id,sessao.id,ttl=3600)
    db.session.commit()
    audit_event('auth_app_refresh','funcionario',f.id,'funcionario',f.id,True,{'sessao_id':sessao.id})
    return jsonify({'ok':True,'access_token':access,'refresh_token':novo_refresh,'token_type':'Bearer','expires_in':3600,'refresh_expires_in':max(0,int((sessao.exp_refresh-utcnow()).total_seconds()))})

@app.route('/api/app/funcionario/logout',methods=['POST'])
@app_func_required
def api_app_funcionario_logout():
    f=g.app_funcionario
    d=request.json or {}
    all_devices=bool(d.get('all_devices'))
    if all_devices:
        FuncionarioAppSessao.query.filter_by(funcionario_id=f.id,revogado=False).update({'revogado':True})
    else:
        g.app_sessao.revogado=True
    db.session.commit()
    audit_event('auth_app_logout','funcionario',f.id,'funcionario',f.id,True,{'all_devices':all_devices})
    return jsonify({'ok':True})

@app.route('/api/app/funcionario/me')
@app_func_required
def api_app_funcionario_me():
    f=g.app_funcionario
    ultimo_aso=FuncionarioArquivo.query.filter_by(funcionario_id=f.id,categoria='aso').order_by(
        FuncionarioArquivo.criado_em.desc(),FuncionarioArquivo.id.desc()
    ).first()
    emp=db.session.get(Empresa,f.empresa_id) if f.empresa_id else None
    foto_url='/api/app/funcionario/me/foto' if f.foto_perfil else None
    jornada_info=None
    if getattr(f,'jornada_id',None):
        j=JornadaTrabalho.query.get(f.jornada_id)
        if j:
            jornada_info={
                'id':j.id,'nome':j.nome,
                'hora_entrada':j.hora_entrada,'hora_saida':j.hora_saida,
                'hora_intervalo_inicio':j.hora_intervalo_inicio,'hora_intervalo_fim':j.hora_intervalo_fim,
                'dias_semana':j.dias_semana,'tolerancia_min':j.tolerancia_min,
            }
    return jsonify({'ok':True,'funcionario':{
        'id':f.id,
        'nome':f.nome,
        'cpf':f.cpf,
        'email':f.email,
        'telefone':f.telefone,
        'cargo':f.cargo,
        'setor':f.setor,
        'empresa_id':f.empresa_id,
        'empresa_nome':(emp.nome if emp else None),
        'posto_operacional':f.posto_operacional,
        'status':f.status,
        'foto_url':foto_url,
        'ultimo_aso_competencia':(ultimo_aso.competencia if ultimo_aso else None),
        'ultimo_aso_enviado_em':(ultimo_aso.criado_em.isoformat() if (ultimo_aso and ultimo_aso.criado_em) else None),
        'jornada':f.jornada,
        'jornada_info':jornada_info,
    }})

@app.route('/api/app/funcionario/me/foto',methods=['POST'])
@app_func_required
def api_app_funcionario_foto_upload():
    f=g.app_funcionario
    if 'foto' not in request.files:
        return jsonify({'erro':'Nenhuma foto enviada'}),400
    file=request.files['foto']
    ext=os.path.splitext(file.filename or 'foto.jpg')[1].lower() or '.jpg'
    if ext not in ('.jpg','.jpeg','.png','.webp'):
        return jsonify({'erro':'Formato inválido. Use JPG, PNG ou WEBP.'}),400
    # Limit size to 5MB
    file.seek(0,2)
    size=file.tell(); file.seek(0)
    if size>5*1024*1024:
        return jsonify({'erro':'Foto muito grande. Máximo 5MB.'}),400
    dir_path=os.path.join(UPLOAD_ROOT,'funcionarios',str(f.id),'foto')
    os.makedirs(dir_path,exist_ok=True)
    filename=f'perfil{ext}'
    abs_path=os.path.join(dir_path,filename)
    file.save(abs_path)
    rel_path=os.path.relpath(abs_path,UPLOAD_ROOT).replace('\\','/')
    f.foto_perfil=rel_path
    db.session.commit()
    return jsonify({'ok':True,'foto_url':'/api/app/funcionario/me/foto'})

@app.route('/api/app/funcionario/me/foto')
@app_func_required
def api_app_funcionario_foto_get():
    f=g.app_funcionario
    if not f.foto_perfil:
        return jsonify({'erro':'Sem foto'}),404
    abs_p=os.path.join(UPLOAD_ROOT,f.foto_perfil)
    if not os.path.exists(abs_p):
        return jsonify({'erro':'Arquivo não encontrado'}),404
    return send_file(abs_p)

@app.route('/api/app/funcionario/me/push-token',methods=['POST'])
@app_func_required
def api_app_funcionario_push_token():
    f=g.app_funcionario
    d=request.json or {}
    token=(d.get('token') or '').strip()
    if not token:
        return jsonify({'erro':'Token obrigatorio'}),400
    if len(token)>300:
        return jsonify({'erro':'Token invalido'}),400
    f.app_push_token=token
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/app/funcionario/me/push-token/teste',methods=['POST'])
@app_func_required
def api_app_funcionario_push_token_teste():
    f=g.app_funcionario
    tem_token=bool((f.app_push_token or '').strip())
    if not tem_token:
        return jsonify({'ok':False,'erro':'Funcionario sem token push salvo'}),400
    ok=_push_notify_funcionario(
        f.id,
        'Teste de notificacao',
        'Se voce recebeu esta mensagem, o push do app esta funcionando.',
        {'tipo':'documento_assinar','arquivo_id':'0','origem':'teste_push'}
    )
    return jsonify({'ok':ok,'tem_token':tem_token})

@app.route('/api/app/funcionario/me/localizacao',methods=['POST'])
@app_func_required
def api_app_funcionario_me_localizacao():
    f=g.app_funcionario
    d=request.json or {}
    lat=d.get('lat')
    lon=d.get('lon')
    if lat is None or lon is None:
        return jsonify({'erro':'lat e lon obrigatorios'}),400
    try:
        lat=float(lat); lon=float(lon)
    except Exception:
        return jsonify({'erro':'lat/lon invalidos'}),400
    if not (-90<=lat<=90) or not (-180<=lon<=180):
        return jsonify({'erro':'coordenadas fora do intervalo'}),400
    f.app_lat=lat
    f.app_lon=lon
    f.app_localizacao_em=utcnow()
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/app/funcionario/me/contato',methods=['PUT'])
@app_func_required
def api_app_funcionario_me_contato():
    f=g.app_funcionario
    d=request.json or {}
    mudou=False
    if 'email' in d:
        em=(d.get('email') or '').strip().lower()
        if em and not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+',em):
            return jsonify({'erro':'E-mail invalido.'}),400
        f.email=em
        mudou=True
    if 'telefone' in d:
        tel_raw=only_digits(d.get('telefone'))
        # No app, o usuário pode digitar sem DDI. Se vier com +55, removemos o país.
        if tel_raw.startswith('55') and len(tel_raw) in (12,13):
            tel=tel_raw[2:]
        else:
            tel=tel_raw[:11]
        if tel and len(tel) not in (10,11):
            return jsonify({'erro':'Telefone invalido. Informe DDD + numero.'}),400
        f.telefone=tel
        mudou=True
    if not mudou:
        return jsonify({'erro':'Informe ao menos um campo para atualizar (email/telefone).'}),400
    db.session.commit()
    audit_event('funcionario_app_atualizou_contato','funcionario',f.id,'funcionario',f.id,True,{})
    return jsonify({'ok':True,'funcionario':{'id':f.id,'email':f.email,'telefone':f.telefone}})

@app.route('/api/app/funcionario/me/solicitacoes-alteracao',methods=['GET'])
@app_func_required
def api_app_funcionario_minhas_solicitacoes_alteracao():
    itens=FuncionarioAlteracaoSolicitacao.query.filter_by(funcionario_id=g.app_funcionario.id).order_by(
        FuncionarioAlteracaoSolicitacao.solicitado_em.desc(),FuncionarioAlteracaoSolicitacao.id.desc()
    ).all()
    return jsonify({'ok':True,'items':[it.to_dict() for it in itens]})

@app.route('/api/app/funcionario/me/solicitacoes-alteracao',methods=['POST'])
@app_func_required
def api_app_funcionario_solicitar_alteracao():
    d=request.json or {}
    campos=d.get('campos') or {}
    if not isinstance(campos,dict):
        return jsonify({'erro':'Formato invalido para campos.'}),400
    permitidos={
        'nome','cargo','funcao','setor','endereco','endereco_numero','endereco_complemento',
        'endereco_bairro','cidade','estado','cep','banco_codigo','banco_nome',
        'banco_agencia','banco_conta','banco_tipo_conta','banco_pix'
    }
    payload={}
    for k,v in campos.items():
        if k not in permitidos:
            continue
        payload[k]=str(v or '').strip()
    if not payload:
        return jsonify({'erro':'Nenhum campo permitido foi informado para solicitacao.'}),400
    obs=(d.get('observacao') or '').strip()
    it=FuncionarioAlteracaoSolicitacao(
        funcionario_id=g.app_funcionario.id,
        payload=json.dumps(payload,ensure_ascii=False),
        observacao=obs,
        status='pendente'
    )
    db.session.add(it)
    db.session.commit()
    audit_event('funcionario_app_solicitou_alteracao','funcionario',g.app_funcionario.id,'funcionario',g.app_funcionario.id,True,{'solicitacao_id':it.id})
    return jsonify({'ok':True,'item':it.to_dict()}),201

@app.route('/api/funcionarios/<int:id>/solicitacoes-alteracao',methods=['GET'])
@lr
def api_funcionario_solicitacoes_alteracao(id):
    Funcionario.query.get_or_404(id)
    itens=FuncionarioAlteracaoSolicitacao.query.filter_by(funcionario_id=id).order_by(
        FuncionarioAlteracaoSolicitacao.solicitado_em.desc(),FuncionarioAlteracaoSolicitacao.id.desc()
    ).all()
    return jsonify([it.to_dict() for it in itens])

@app.route('/api/funcionarios/solicitacoes-alteracao/<int:id>/decidir',methods=['POST'])
@lr
def api_decidir_solicitacao_alteracao(id):
    it=FuncionarioAlteracaoSolicitacao.query.get_or_404(id)
    if (it.status or '')!='pendente':
        return jsonify({'erro':'Solicitacao ja foi analisada.'}),400
    d=request.json or {}
    acao=(d.get('acao') or '').strip().lower()
    motivo=(d.get('motivo') or '').strip()
    if acao not in ('aprovar','rejeitar'):
        return jsonify({'erro':'Acao invalida. Use aprovar ou rejeitar.'}),400
    f=Funcionario.query.get_or_404(it.funcionario_id)
    if acao=='aprovar':
        payload=jloads(it.payload,{})
        for k,v in (payload.items() if isinstance(payload,dict) else []):
            if not hasattr(f,k):
                continue
            if k=='cep':
                setattr(f,k,norm_cep(v))
            elif k=='estado':
                setattr(f,k,norm_uf(v))
            elif k=='banco_codigo':
                setattr(f,k,norm_bank_code(v))
            else:
                setattr(f,k,v)
        it.status='aprovada'
    else:
        it.status='rejeitada'
    it.motivo_admin=motivo
    it.analisado_por=session.get('uid')
    it.analisado_em=utcnow()
    db.session.commit()
    audit_event('funcionario_alteracao_solicitacao_decidida','usuario',session.get('uid'),'funcionario',f.id,True,{'solicitacao_id':it.id,'acao':acao})
    return jsonify({'ok':True,'item':it.to_dict(),'funcionario':f.to_dict()})

@app.route('/api/app/funcionario/me/documentos')
@app_func_required
def api_app_funcionario_me_documentos():
    resp,status=build_func_docs_response(g.app_funcionario.id)
    return jsonify(resp),status

@app.route('/api/app/funcionario/arquivos/<int:id>/download')
@app_func_required
def api_app_funcionario_download_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    if a.funcionario_id!=g.app_funcionario.id:
        return jsonify({'erro':'Acesso negado'}),403
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    audit_event('funcionario_app_arquivo_download','funcionario',g.app_funcionario.id,'funcionario',a.funcionario_id,True,{'arquivo_id':a.id,'caminho':a.caminho})
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

@app.route('/api/app/funcionario/arquivos/<int:id>/assinar',methods=['POST'])
@app_func_required
def api_app_funcionario_assinar_arquivo(id):
    f=g.app_funcionario
    a=FuncionarioArquivo.query.get_or_404(id)
    if a.funcionario_id!=f.id:
        return jsonify({'erro':'Acesso negado'}),403
    status_atual=(a.ass_status or '').strip().lower()
    if status_atual=='concluida':
        return jsonify({'ok':True,'mensagem':'Documento ja assinado.','item':a.to_dict()})

    a.ass_status='concluida'
    a.ass_nome=(f.nome or '').strip()
    a.ass_cpf=norm_cpf(f.cpf)
    a.ass_cargo=(f.cargo or '').strip()
    a.ass_ip=(request.remote_addr or '')[:60]
    a.ass_em=utcnow()
    a.ass_token=''
    a.ass_expira_em=None
    a.ass_otp_hash=''
    a.ass_otp_expira_em=None
    a.ass_otp_tentativas=0
    a.ass_codigo=(a.ass_codigo or secrets.token_urlsafe(16))

    db.session.commit()
    audit_event(
        'funcionario_app_arquivo_assinado',
        'funcionario',f.id,
        'arquivo',a.id,
        True,
        {'arquivo_id':a.id,'categoria':a.categoria,'competencia':a.competencia,'origem':'app'}
    )
    return jsonify({'ok':True,'mensagem':'Documento assinado com sucesso.','item':a.to_dict()})

@app.route('/api/app/funcionario/pendentes-assinatura')
@app_func_required
def api_app_funcionario_pendentes_assinatura():
    f=g.app_funcionario
    regs=FuncionarioArquivo.query.filter_by(funcionario_id=f.id,ass_status='pendente').order_by(FuncionarioArquivo.criado_em.desc()).all()
    itens=[]
    for a in regs:
        prazo_em=getattr(a,'ass_prazo_em',None)
        cat=norm_cat(a.categoria)
        itens.append({
            'id':a.id,
            'categoria':cat,
            'categoria_label':DOC_CAT_LABEL.get(cat,cat),
            'ano':arq_year_from_path(a.caminho),
            'nome_arquivo':a.nome_arquivo,
            'competencia':a.competencia,
            'ass_status':'pendente',
            'ass_em_fmt':'',
            'ass_prazo_em':prazo_em.isoformat() if prazo_em else None,
            'ass_prazo_fmt':prazo_em.strftime('%d/%m/%Y') if prazo_em else None,
            'can_assinar':True,
            'criado_em':a.criado_em.isoformat() if a.criado_em else '',
            'criado_fmt':a.criado_em.strftime('%d/%m/%Y %H:%M') if a.criado_em else '',
            'download_url':f'/api/funcionarios/arquivos/{a.id}/download',
            'app_download_url':f'/api/app/funcionario/arquivos/{a.id}/download',
        })
    return jsonify({'ok':True,'itens':itens})

@app.route('/api/app/funcionario/historico-assinaturas')
@app_func_required
def api_app_funcionario_historico_assinaturas():
    f=g.app_funcionario
    regs=FuncionarioArquivo.query.filter_by(funcionario_id=f.id,ass_status='concluida').order_by(FuncionarioArquivo.ass_em.desc()).all()
    itens=[]
    for a in regs:
        cat=norm_cat(a.categoria)
        ip_raw=a.ass_ip or ''
        # mask last octet for privacy: 192.168.1.100 -> 192.168.1.xxx
        ip_parts=ip_raw.split('.')
        ip_mask='.'.join(ip_parts[:-1]+['xxx']) if len(ip_parts)==4 else ip_raw
        itens.append({
            'id':a.id,
            'categoria':cat,
            'categoria_label':DOC_CAT_LABEL.get(cat,cat),
            'ano':arq_year_from_path(a.caminho),
            'nome_arquivo':a.nome_arquivo,
            'competencia':a.competencia or '',
            'ass_em':a.ass_em.isoformat() if a.ass_em else '',
            'ass_em_fmt':a.ass_em.strftime('%d/%m/%Y %H:%M') if a.ass_em else '',
            'ass_ip_mask':ip_mask,
            'ass_codigo':a.ass_codigo or '',
            'app_download_url':f'/api/app/funcionario/arquivos/{a.id}/download',
        })
    return jsonify({'ok':True,'itens':itens})

@app.route('/api/app/funcionario/me/senha',methods=['PUT'])
@app_func_required
def api_app_funcionario_me_senha():
    f=g.app_funcionario
    d=request.json or {}
    atual=(d.get('senha_atual') or '')
    nova=(d.get('nova_senha') or '')
    if not atual or not nova:
        return jsonify({'erro':'Senha atual e nova senha sao obrigatorias'}),400
    if not pw_check(f.app_senha,atual): return jsonify({'erro':'Senha atual invalida'}),401
    if len(nova)<6: return jsonify({'erro':'Nova senha deve ter ao menos 6 caracteres'}),400
    f.app_senha=pw_hash(nova)
    db.session.commit()
    audit_event('auth_app_troca_senha','funcionario',f.id,'funcionario',f.id,True,{})
    return jsonify({'ok':True})

# ============================================================
# COMUNICADOS - APP (funcionário lê comunicados do RH)
# ============================================================

@app.route('/api/app/funcionario/comunicados')
@app_func_required
def api_app_comunicados_lista():
    f=g.app_funcionario
    from sqlalchemy import or_
    itens=ComunicadoApp.query.filter(
        ComunicadoApp.ativo==True,
        or_(ComunicadoApp.funcionario_id==None, ComunicadoApp.funcionario_id==f.id),
        or_(ComunicadoApp.posto_operacional==None, ComunicadoApp.posto_operacional=='', ComunicadoApp.posto_operacional==f.posto_operacional)
    ).order_by(ComunicadoApp.criado_em.desc()).all()
    return jsonify([c.to_dict(funcionario_id=f.id) for c in itens])

@app.route('/api/app/funcionario/comunicados/<int:cid>/lido',methods=['POST'])
@app_func_required
def api_app_comunicado_marcar_lido(cid):
    f=g.app_funcionario
    c=ComunicadoApp.query.get_or_404(cid)
    c.marcar_lido(f.id)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/app/funcionario/comunicados/nao-lidos')
@app_func_required
def api_app_comunicados_nao_lidos():
    f=g.app_funcionario
    from sqlalchemy import or_
    itens=ComunicadoApp.query.filter(
        ComunicadoApp.ativo==True,
        or_(ComunicadoApp.funcionario_id==None, ComunicadoApp.funcionario_id==f.id),
        or_(ComunicadoApp.posto_operacional==None, ComunicadoApp.posto_operacional=='', ComunicadoApp.posto_operacional==f.posto_operacional)
    ).all()
    lidos=list(set(fid for c in itens for fid in c.lidos_por()))
    count=sum(1 for c in itens if f.id not in c.lidos_por())
    return jsonify({'nao_lidos':count})

# ============================================================
# MENSAGENS - APP (chat funcionário ↔ RH)
# ============================================================

@app.route('/api/app/funcionario/mensagens')
@app_func_required
def api_app_mensagens_lista():
    f=g.app_funcionario
    msgs=MensagemApp.query.filter_by(funcionario_id=f.id).order_by(MensagemApp.enviado_em.asc()).all()
    # Marcar mensagens do RH como lidas ao abrir
    for m in msgs:
        if m.de_rh and not m.lida:
            m.lida=True
    db.session.commit()
    return jsonify([m.to_dict() for m in msgs])

@app.route('/api/app/funcionario/mensagens',methods=['POST'])
@app_func_required
def api_app_mensagem_enviar():
    f=g.app_funcionario
    d=request.json or {}
    conteudo=(d.get('conteudo') or '').strip()
    if not conteudo: return jsonify({'erro':'Mensagem nao pode ser vazia'}),400
    if len(conteudo)>2000: return jsonify({'erro':'Mensagem muito longa'}),400
    m=MensagemApp(funcionario_id=f.id,de_rh=False,conteudo=conteudo,lida=False,enviado_por='funcionario',tipo='texto')
    db.session.add(m)
    db.session.commit()
    return jsonify(m.to_dict()),201

@app.route('/api/app/funcionario/mensagens/arquivo',methods=['POST'])
@app_func_required
def api_app_mensagem_enviar_arquivo():
    f=g.app_funcionario
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Nenhum arquivo enviado'}),400
    nome_orig=secure_filename(arq.filename or 'arquivo')
    if not nome_orig: return jsonify({'erro':'Nome de arquivo invalido'}),400
    ext=os.path.splitext(nome_orig)[1].lower()
    exts_permitidas={'.pdf','.jpg','.jpeg','.png','.doc','.docx','.xls','.xlsx','.txt','.zip'}
    if ext not in exts_permitidas:
        return jsonify({'erro':'Tipo de arquivo nao permitido'}),400
    conteudo=(request.form.get('conteudo') or '').strip()[:500]
    pasta=os.path.join(UPLOAD_ROOT,'funcionarios',str(f.id),'chat')
    os.makedirs(pasta,exist_ok=True)
    ts=datetime.utcnow().strftime('%Y%m%d%H%M%S')
    nome_final=f'{ts}_{nome_orig}'
    abs_p=os.path.join(pasta,nome_final)
    arq.save(abs_p)
    rel=os.path.relpath(abs_p,UPLOAD_ROOT)
    texto_msg=conteudo if conteudo else f'[Arquivo: {nome_orig}]'
    m=MensagemApp(
        funcionario_id=f.id,de_rh=False,conteudo=texto_msg,
        lida=False,enviado_por='funcionario',
        tipo='arquivo',arquivo_nome=nome_orig,arquivo_caminho=rel
    )
    db.session.add(m); db.session.commit()
    return jsonify(m.to_dict()),201

@app.route('/api/app/funcionario/mensagens/<int:mid>/arquivo')
@app_func_required
def api_app_mensagem_download_arquivo(mid):
    f=g.app_funcionario
    m=MensagemApp.query.get_or_404(mid)
    if m.funcionario_id!=f.id:
        return jsonify({'erro':'Acesso negado'}),403
    if not m.arquivo_caminho: return jsonify({'erro':'Mensagem sem arquivo'}),404
    abs_p=os.path.join(UPLOAD_ROOT,m.arquivo_caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    return send_file(abs_p,as_attachment=True,download_name=m.arquivo_nome or 'arquivo')

@app.route('/api/app/funcionario/mensagens/nao-lidas')
@app_func_required
def api_app_mensagens_nao_lidas():
    f=g.app_funcionario
    count=MensagemApp.query.filter_by(funcionario_id=f.id,de_rh=True,lida=False).count()
    return jsonify({'nao_lidas':count})

# ============================================================
# PONTO - APP (funcionário autenticado)
# ============================================================

_APP_PONTO_TIPOS=['entrada','saida_intervalo','retorno_intervalo','saida']

def _app_ponto_label(tipo):
    return {
        'entrada':'Entrada',
        'saida_intervalo':'Saída intervalo',
        'retorno_intervalo':'Retorno intervalo',
        'saida':'Saída',
    }.get((tipo or '').strip().lower(),(tipo or '').strip())

def _app_ponto_next_tipo(tipo):
    t=(tipo or '').strip().lower()
    if t not in _APP_PONTO_TIPOS:
        return 'entrada'
    return _APP_PONTO_TIPOS[(_APP_PONTO_TIPOS.index(t)+1)%len(_APP_PONTO_TIPOS)]

def _app_ponto_tipo_esperado(marcacoes):
    if not marcacoes:
        return 'entrada'
    return _app_ponto_next_tipo(marcacoes[-1].tipo)

def _app_ponto_parse_data_ref(v):
    s=(v or '').strip()
    if not s:
        return localnow().date()
    try:
        return datetime.strptime(s,'%Y-%m-%d').date()
    except Exception:
        return localnow().date()

def _app_ponto_marcacoes_dia(funcionario_id,data_ref):
    inicio=datetime.combine(data_ref,datetime.min.time())
    fim=inicio+timedelta(days=1)
    return (
        PontoMarcacao.query.filter(PontoMarcacao.funcionario_id==funcionario_id)
        .filter(PontoMarcacao.data_hora>=inicio)
        .filter(PontoMarcacao.data_hora<fim)
        .order_by(PontoMarcacao.data_hora.asc(),PontoMarcacao.id.asc())
        .all()
    )

def _app_ponto_min_esperado_jornada(funcionario):
    # Preferir jornada estruturada (JornadaTrabalho)
    if getattr(funcionario,'jornada_id',None):
        j=JornadaTrabalho.query.get(funcionario.jornada_id)
        if j:
            return j.carga_horaria_min()
    # Fallback: campo texto legado
    jornada=str(funcionario.jornada or '').strip().lower()
    if not jornada:
        return 8*60
    m=re.search(r'(\d{1,2})\s*[:h]\s*(\d{1,2})',jornada)
    if m:
        h=max(0,min(16,int(m.group(1))))
        mm=max(0,min(59,int(m.group(2))))
        return h*60+mm
    m=re.search(r'\b(\d{1,2})\b',jornada)
    if m:
        h=max(0,min(16,int(m.group(1))))
        return h*60
    return 8*60

def _app_ponto_fmt_minutos(total,signed=False):
    try:
        minutos=int(total or 0)
    except Exception:
        minutos=0
    sinal=''
    if signed and minutos<0:
        sinal='-'
    minutos=abs(minutos)
    return f'{sinal}{minutos//60:02d}:{minutos%60:02d}'

def _app_ponto_resumo_dia(funcionario,data_ref):
    marcacoes=_app_ponto_marcacoes_dia(funcionario.id,data_ref)
    inconsistencias=[]
    esperado='entrada'
    segundos_total=0
    aberta_em=None
    for m in marcacoes:
        if not getattr(m,'data_hora',None):
            inconsistencias.append('Marcação sem data/hora válida foi ignorada no cálculo.')
            esperado=_app_ponto_next_tipo(m.tipo)
            continue
        if m.tipo!=esperado:
            inconsistencias.append(
                f'Sequência inesperada: recebido {_app_ponto_label(m.tipo)}; esperado {_app_ponto_label(esperado)}.'
            )
        if m.tipo=='entrada':
            if aberta_em is not None:
                inconsistencias.append('Existe uma entrada sem fechamento antes desta nova entrada.')
            aberta_em=m.data_hora
        elif m.tipo=='saida_intervalo':
            if aberta_em is None:
                inconsistencias.append('Saída para intervalo sem entrada anterior.')
            else:
                segundos_total+=max(0,int((m.data_hora-aberta_em).total_seconds()))
                aberta_em=None
        elif m.tipo=='retorno_intervalo':
            if aberta_em is not None:
                inconsistencias.append('Retorno de intervalo sem saída anterior.')
            aberta_em=m.data_hora
        elif m.tipo=='saida':
            if aberta_em is None:
                inconsistencias.append('Saída final sem entrada anterior.')
            else:
                segundos_total+=max(0,int((m.data_hora-aberta_em).total_seconds()))
                aberta_em=None
        esperado=_app_ponto_next_tipo(m.tipo)
    if aberta_em is not None:
        inconsistencias.append('Jornada em aberto (faltou batida de fechamento).')

    min_trab=int(round(segundos_total/60.0))
    min_esp=0 if data_ref.weekday()>=5 else _app_ponto_min_esperado_jornada(funcionario)
    saldo=min_trab-min_esp

    itens=[]
    for m in marcacoes:
        dt=m.data_hora
        itens.append({
            'id':m.id,
            'tipo':m.tipo,
            'tipo_label':_app_ponto_label(m.tipo),
            'data_hora':dt.isoformat() if dt else '',
            'hora_fmt':dt.strftime('%H:%M') if dt else '',
            'origem':(m.origem or 'app'),
            'observacao':m.observacao or '',
        })

    prox=_app_ponto_tipo_esperado(marcacoes)
    return {
        'funcionario_id':funcionario.id,
        'funcionario_nome':funcionario.nome,
        'data_ref':data_ref.strftime('%Y-%m-%d'),
        'marcacoes':itens,
        'proximo_tipo':prox,
        'proximo_tipo_label':_app_ponto_label(prox),
        'horas_trabalhadas_min':min_trab,
        'horas_trabalhadas_fmt':_app_ponto_fmt_minutos(min_trab),
        'horas_esperadas_min':min_esp,
        'horas_esperadas_fmt':_app_ponto_fmt_minutos(min_esp),
        'saldo_min':saldo,
        'saldo_fmt':_app_ponto_fmt_minutos(saldo,signed=True),
        'status':'ok' if not inconsistencias else 'inconsistente',
        'inconsistencias':inconsistencias,
    }

def _geo_haversine_m(lat1,lon1,lat2,lon2):
    try:
        lat1=float(lat1); lon1=float(lon1); lat2=float(lat2); lon2=float(lon2)
    except (ValueError,TypeError):
        return None
    if not (-90<=lat1<=90 and -90<=lat2<=90 and -180<=lon1<=180 and -180<=lon2<=180):
        return None
    r=6371000.0
    p1=math.radians(lat1); p2=math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*r*math.atan2(math.sqrt(a),math.sqrt(1-a))

@app.route('/api/app/funcionario/me/ponto/dia')
@app_func_required
def api_app_ponto_dia_me():
    f=g.app_funcionario
    data_ref=_app_ponto_parse_data_ref(request.args.get('data'))
    return jsonify({'ok':True,'resumo':_app_ponto_resumo_dia(f,data_ref)})

@app.route('/api/app/funcionario/me/ponto/marcar',methods=['POST'])
@app_func_required
def api_app_ponto_marcar_me():
    f=g.app_funcionario
    if (f.status or '').strip().lower()!='ativo':
        return jsonify({'erro':'Somente funcionários ativos podem registrar ponto.'}),400
    dados=request.json or {}
    tipo=(dados.get('tipo') or '').strip().lower()
    observacao=(dados.get('observacao') or '').strip()[:500]
    data_hora=utcnow()
    if data_hora>(utcnow()+timedelta(minutes=1)):
        return jsonify({'erro':'Não é permitido registrar ponto em horário futuro.'}),400
    data_ref=data_hora.date()
    marcacoes_dia=_app_ponto_marcacoes_dia(f.id,data_ref)
    tipo=tipo or _app_ponto_tipo_esperado(marcacoes_dia)
    if tipo not in _APP_PONTO_TIPOS:
        return jsonify({'erro':'Tipo de marcação inválido.'}),400
    esperado=_app_ponto_tipo_esperado(marcacoes_dia)
    if tipo!=esperado:
        return jsonify({'erro':f'Ordem de marcação inválida. Agora é esperado: {_app_ponto_label(esperado)}.'}),400
    if any(abs((data_hora-m.data_hora).total_seconds())<60 for m in marcacoes_dia if getattr(m,'data_hora',None)):
        return jsonify({'erro':'Já existe marcação neste minuto para este funcionário.'}),400

    ip=(request.headers.get('X-Forwarded-For','') or request.remote_addr or '').split(',')[0].strip()[:60]
    lat=dados.get('lat')
    lon=dados.get('lon')
    precisao=dados.get('precisao')
    try: lat=float(lat) if lat is not None else None
    except (ValueError,TypeError): lat=None
    try: lon=float(lon) if lon is not None else None
    except (ValueError,TypeError): lon=None
    try: precisao=float(precisao) if precisao is not None else None
    except (ValueError,TypeError): precisao=None
    if lat is None or lon is None:
        return jsonify({'erro':'Localização obrigatória para registrar ponto. Ative o GPS e tente novamente.'}),400
    if not (-90<=lat<=90 and -180<=lon<=180):
        return jsonify({'erro':'Coordenadas de localização inválidas.'}),400

    localizacao={
        'status':'sem_referencia_posto',
        'distancia_m':None,
        'raio_m':None,
        'posto_cliente_id':f.posto_cliente_id,
    }
    if f.posto_cliente_id:
        cli=Cliente.query.get(f.posto_cliente_id)
        if cli and cli.geo_lat is not None and cli.geo_lon is not None:
            distancia=_geo_haversine_m(lat,lon,cli.geo_lat,cli.geo_lon)
            raio=float(cli.geofence_raio_m or 150)
            localizacao={
                'status':('no_posto' if (distancia is not None and distancia<=raio) else 'fora_posto'),
                'distancia_m':(round(distancia,1) if distancia is not None else None),
                'raio_m':raio,
                'posto_cliente_id':f.posto_cliente_id,
            }
        else:
            localizacao['status']='posto_sem_coordenada'
    m=PontoMarcacao(
        funcionario_id=f.id,
        tipo=tipo,
        data_hora=data_hora,
        origem='app',
        observacao=observacao,
        criado_por='funcionario-app',
        ip=ip,
        latitude=lat,
        longitude=lon,
        precisao_gps=precisao,
    )
    db.session.add(m)
    db.session.commit()
    audit_event('ponto_marcacao_app','funcionario',f.id,'funcionario',f.id,True,{
        'tipo':tipo,
        'data_ref':data_ref.strftime('%Y-%m-%d'),
        'origem':'app',
        'lat':lat,
        'lon':lon,
        'localizacao_status':localizacao.get('status'),
        'distancia_m':localizacao.get('distancia_m'),
    })
    return jsonify({
        'ok':True,
        'marcacao':{
            'id':m.id,
            'tipo':m.tipo,
            'tipo_label':_app_ponto_label(m.tipo),
            'hora_fmt':m.data_hora.strftime('%H:%M') if m.data_hora else '',
            'localizacao':localizacao,
        },
        'resumo':_app_ponto_resumo_dia(f,data_ref)
    })

# ============================================================
# JORNADAS DE TRABALHO
# ============================================================

@app.route('/api/jornadas',methods=['GET'])
@lr
def api_jornadas_listar():
    ativas=request.args.get('ativas','0').strip()
    q=JornadaTrabalho.query
    if ativas=='1':
        q=q.filter_by(ativo=True)
    return jsonify([j.to_dict() for j in q.order_by(JornadaTrabalho.nome).all()])

@app.route('/api/jornadas',methods=['POST'])
@lr
def api_jornadas_criar():
    d=request.json or {}
    nome=(d.get('nome') or '').strip()
    if not nome:
        return jsonify({'erro':'Nome é obrigatório'}),400
    j=JornadaTrabalho(
        nome=nome,
        descricao=(d.get('descricao') or '').strip()[:255],
        dias_semana=(d.get('dias_semana') or '1,2,3,4,5').strip(),
        hora_entrada=(d.get('hora_entrada') or '08:00').strip()[:5],
        hora_saida=(d.get('hora_saida') or '17:48').strip()[:5],
        hora_intervalo_inicio=(d.get('hora_intervalo_inicio') or '12:00').strip()[:5],
        hora_intervalo_fim=(d.get('hora_intervalo_fim') or '13:00').strip()[:5],
        tolerancia_min=max(0,min(60,int(d.get('tolerancia_min') or 10))),
        ativo=bool(d.get('ativo',True)),
    )
    db.session.add(j); db.session.commit()
    audit_event('jornada_criar','usuario',session.get('uid'),'jornada_trabalho',j.id,True,{'nome':nome})
    return jsonify(j.to_dict()),201

@app.route('/api/jornadas/<int:id>',methods=['GET'])
@lr
def api_jornada_detalhe(id):
    j=JornadaTrabalho.query.get_or_404(id)
    d=j.to_dict()
    d['funcionarios']=[{'id':f.id,'nome':f.nome,'cargo':f.cargo,'status':f.status} for f in Funcionario.query.filter_by(jornada_id=id).order_by(Funcionario.nome).all()]
    return jsonify(d)

@app.route('/api/jornadas/<int:id>',methods=['PUT'])
@lr
def api_jornada_editar(id):
    j=JornadaTrabalho.query.get_or_404(id)
    d=request.json or {}
    if 'nome' in d:
        n=(d['nome'] or '').strip()
        if not n: return jsonify({'erro':'Nome não pode ser vazio'}),400
        j.nome=n
    if 'descricao' in d: j.descricao=(d['descricao'] or '').strip()[:255]
    if 'dias_semana' in d: j.dias_semana=(d['dias_semana'] or '1,2,3,4,5').strip()
    if 'hora_entrada' in d: j.hora_entrada=(d['hora_entrada'] or '08:00').strip()[:5]
    if 'hora_saida' in d: j.hora_saida=(d['hora_saida'] or '17:48').strip()[:5]
    if 'hora_intervalo_inicio' in d: j.hora_intervalo_inicio=(d['hora_intervalo_inicio'] or '12:00').strip()[:5]
    if 'hora_intervalo_fim' in d: j.hora_intervalo_fim=(d['hora_intervalo_fim'] or '13:00').strip()[:5]
    if 'tolerancia_min' in d:
        try: j.tolerancia_min=max(0,min(60,int(d['tolerancia_min'])))
        except (ValueError,TypeError): pass
    if 'ativo' in d: j.ativo=bool(d['ativo'])
    db.session.commit()
    audit_event('jornada_editar','usuario',session.get('uid'),'jornada_trabalho',id,True,{})
    return jsonify(j.to_dict())

@app.route('/api/jornadas/<int:id>',methods=['DELETE'])
@lr
def api_jornada_excluir(id):
    j=JornadaTrabalho.query.get_or_404(id)
    count=Funcionario.query.filter_by(jornada_id=id).count()
    if count>0:
        return jsonify({'erro':f'Jornada está vinculada a {count} funcionário(s). Desvincule antes de excluir.'}),400
    db.session.delete(j); db.session.commit()
    audit_event('jornada_excluir','usuario',session.get('uid'),'jornada_trabalho',id,True,{'nome':j.nome})
    return jsonify({'ok':True})

@app.route('/api/jornadas/<int:id>/funcionarios',methods=['POST'])
@lr
def api_jornada_vincular_funcionarios(id):
    j=JornadaTrabalho.query.get_or_404(id)
    d=request.json or {}
    ids=d.get('funcionario_ids') or []
    if not isinstance(ids,list): return jsonify({'erro':'funcionario_ids deve ser lista'}),400
    vinculados=[]
    for fid in ids:
        f=Funcionario.query.get(fid)
        if f:
            f.jornada_id=id
            vinculados.append(fid)
    db.session.commit()
    audit_event('jornada_vincular','usuario',session.get('uid'),'jornada_trabalho',id,True,{'funcionarios':vinculados})
    return jsonify({'ok':True,'vinculados':len(vinculados)})

@app.route('/api/jornadas/<int:id>/funcionarios/<int:fid>',methods=['DELETE'])
@lr
def api_jornada_desvincular_funcionario(id,fid):
    f=Funcionario.query.get_or_404(fid)
    if f.jornada_id!=id:
        return jsonify({'erro':'Funcionário não está nesta jornada'}),400
    f.jornada_id=None
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/funcionarios/<int:id>/jornada',methods=['PUT'])
@lr
def api_funcionario_definir_jornada(id):
    f=Funcionario.query.get_or_404(id)
    d=request.json or {}
    jid=d.get('jornada_id')
    if jid is None or jid=='':
        f.jornada_id=None
    else:
        j=JornadaTrabalho.query.get(int(jid))
        if not j: return jsonify({'erro':'Jornada não encontrada'}),404
        f.jornada_id=j.id
    db.session.commit()
    return jsonify({'ok':True,'jornada_id':f.jornada_id})

# ============================================================
# COMUNICADOS - WEB RH (gestão)
# ============================================================

@app.route('/api/comunicados-app',methods=['GET'])
@lr
def api_rh_comunicados_lista():
    itens=ComunicadoApp.query.order_by(ComunicadoApp.criado_em.desc()).all()
    return jsonify([c.to_dict() for c in itens])

@app.route('/api/comunicados-app',methods=['POST'])
@lr
def api_rh_comunicado_criar():
    d=request.json or {}
    titulo=(d.get('titulo') or '').strip()
    conteudo=(d.get('conteudo') or '').strip()
    if not titulo: return jsonify({'erro':'Titulo obrigatorio'}),400
    if not conteudo: return jsonify({'erro':'Conteudo obrigatorio'}),400
    fid=d.get('funcionario_id')
    posto=(d.get('posto_operacional') or '').strip() or None
    c=ComunicadoApp(
        titulo=titulo,
        conteudo=conteudo,
        funcionario_id=int(fid) if fid else None,
        posto_operacional=posto,
        criado_por=session.get('nome') or session.get('email') or 'RH',
        ativo=True
    )
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()),201

@app.route('/api/comunicados-app/<int:cid>',methods=['PUT'])
@lr
def api_rh_comunicado_editar(cid):
    c=ComunicadoApp.query.get_or_404(cid)
    d=request.json or {}
    if 'titulo' in d: c.titulo=(d['titulo'] or '').strip()
    if 'conteudo' in d: c.conteudo=(d['conteudo'] or '').strip()
    if 'ativo' in d: c.ativo=bool(d['ativo'])
    db.session.commit()
    return jsonify(c.to_dict())

@app.route('/api/comunicados-app/<int:cid>',methods=['DELETE'])
@lr
def api_rh_comunicado_excluir(cid):
    c=ComunicadoApp.query.get_or_404(cid)
    c.ativo=False
    db.session.commit()
    return jsonify({'ok':True})

# ============================================================
# MENSAGENS - WEB RH (chat com funcionário)
# ============================================================

@app.route('/api/mensagens-app/funcionarios')
@lr
def api_rh_mensagens_funcionarios():
    """Lista funcionários que têm mensagens + contagem de não lidas."""
    from sqlalchemy import func as sqlfunc
    subq=db.session.query(
        MensagemApp.funcionario_id,
        sqlfunc.count(MensagemApp.id).label('total'),
        sqlfunc.sum(db.cast(db.and_(MensagemApp.de_rh==False,MensagemApp.lida==False),db.Integer)).label('nao_lidas'),
        sqlfunc.max(MensagemApp.enviado_em).label('ultima')
    ).group_by(MensagemApp.funcionario_id).all()
    result=[]
    for row in subq:
        f=Funcionario.query.get(row.funcionario_id)
        if not f: continue
        result.append({
            'funcionario_id':f.id,
            'nome':f.nome,
            'cargo':f.cargo,
            'total':row.total,
            'nao_lidas':int(row.nao_lidas or 0),
            'ultima':row.ultima.strftime('%d/%m/%Y %H:%M') if row.ultima else ''
        })
    result.sort(key=lambda x:x['nao_lidas'],reverse=True)
    return jsonify(result)

@app.route('/api/mensagens-app/<int:fid>')
@lr
def api_rh_mensagens_chat(fid):
    Funcionario.query.get_or_404(fid)
    msgs=MensagemApp.query.filter_by(funcionario_id=fid).order_by(MensagemApp.enviado_em.asc()).all()
    # Marcar mensagens do funcionário como lidas
    for m in msgs:
        if not m.de_rh and not m.lida:
            m.lida=True
    db.session.commit()
    return jsonify([m.to_dict() for m in msgs])

@app.route('/api/mensagens-app/<int:fid>',methods=['POST'])
@lr
def api_rh_mensagem_responder(fid):
    Funcionario.query.get_or_404(fid)
    d=request.json or {}
    conteudo=(d.get('conteudo') or '').strip()
    if not conteudo: return jsonify({'erro':'Mensagem nao pode ser vazia'}),400
    if len(conteudo)>2000: return jsonify({'erro':'Mensagem muito longa'}),400
    nome_rh=session.get('nome') or session.get('email') or 'RH'
    m=MensagemApp(funcionario_id=fid,de_rh=True,conteudo=conteudo,lida=False,enviado_por=nome_rh)
    db.session.add(m)
    db.session.commit()
    _push_notify_funcionario(fid,'Nova mensagem do RH',conteudo[:160],data={'tipo':'chat','funcionario_id':str(fid)})
    return jsonify(m.to_dict()),201

@app.route('/api/mensagens-app/nao-lidas-total')
@lr
def api_rh_mensagens_nao_lidas_total():
    count=MensagemApp.query.filter_by(de_rh=False,lida=False).count()
    return jsonify({'nao_lidas':count})

@app.route('/api/mensagens-app/broadcast',methods=['POST'])
@lr
def api_rh_mensagens_broadcast():
    d=request.json or {}
    conteudo=(d.get('conteudo') or '').strip()
    if not conteudo: return jsonify({'erro':'Mensagem nao pode ser vazia'}),400
    if len(conteudo)>2000: return jsonify({'erro':'Mensagem muito longa'}),400
    empresa_id=d.get('empresa_id')
    posto=(d.get('posto') or '').strip()
    nome_rh=session.get('nome') or session.get('email') or 'RH'
    q=Funcionario.query.filter_by(status='Ativo',app_ativo=True)
    if empresa_id:
        q=q.filter_by(empresa_id=int(empresa_id))
    if posto:
        q=q.filter(Funcionario.posto_operacional.ilike(f'%{posto}%'))
    funcs=q.all()
    if not funcs:
        return jsonify({'erro':'Nenhum colaborador encontrado com esse filtro'}),404
    for func in funcs:
        m=MensagemApp(funcionario_id=func.id,de_rh=True,conteudo=conteudo,lida=False,enviado_por=nome_rh)
        db.session.add(m)
    db.session.commit()
    for func in funcs:
        _push_notify_funcionario(func.id,'Nova mensagem do RH',conteudo[:160],data={'tipo':'chat_broadcast','funcionario_id':str(func.id)})
    return jsonify({'ok':True,'enviado_para':len(funcs)})

@app.route('/api/funcionarios/arquivos/<int:id>',methods=['DELETE'])
@lr
def api_funcionario_delete_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    fid=a.funcionario_id
    cam=a.caminho
    try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
    except: pass
    db.session.delete(a); db.session.commit()
    audit_event('funcionario_arquivo_delete','usuario',session.get('uid'),'funcionario',fid,True,{'arquivo_id':id,'caminho':cam})
    return jsonify({'ok':True})

@app.route('/api/funcionarios/arquivos/<int:id>/download')
@lr
def api_funcionario_download_arquivo(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    ator_tipo='funcionario_app' if getattr(g,'app_funcionario',None) else 'usuario'
    ator_id=(getattr(g,'app_funcionario',None).id if getattr(g,'app_funcionario',None) else session.get('uid'))
    audit_event('funcionario_arquivo_download',ator_tipo,ator_id,'funcionario',a.funcionario_id,True,{'arquivo_id':a.id,'caminho':a.caminho})
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

@app.route('/api/funcionarios/arquivos/<int:id>/assinatura/solicitar',methods=['POST'])
@lr
def api_func_arquivo_solicitar_assinatura(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    f=Funcionario.query.get_or_404(a.funcionario_id)
    d=request.json or {}
    canal_req=(d.get('canal') or '').strip().lower()
    canal_padrao=(a.ass_canal_envio or '').strip().lower()
    if not canal_padrao:
        if (a.ass_wa_status or '') in ('enviado','recebido') or bool(a.ass_wa_enviado_em):
            canal_padrao='whatsapp'
        elif (a.ass_email_status or '') in ('enviado','recebido') or bool(a.ass_email_enviado_em):
            canal_padrao='email'
        elif not (a.ass_token or '').strip():
            canal_padrao='app'
        else:
            canal_padrao='link'
    canal=_ass_track_channel(canal_req,canal_padrao)
    forcar_novo_token=bool(d.get('forcar_novo_token',True))
    rs=_solicitar_assinatura_arquivo_funcionario(
        a,
        f,
        canal=canal,
        commit_now=True,
        forcar_novo_token=forcar_novo_token,
        eh_lembrete=(not forcar_novo_token),
    )
    if not rs.get('ok'):
        return jsonify({'erro':rs.get('erro') or 'Falha ao solicitar assinatura.'}),400
    audit_event('func_arquivo_assinatura_solicitada','usuario',session.get('uid'),'funcionario',f.id,True,
                {'arquivo_id':id,'nome':a.nome_arquivo,'canal':rs.get('canal','link'),'lembrete':not forcar_novo_token})
    return jsonify({'ok':True,'arquivo_id':id,'link':rs.get('link',''),'link_curto':rs.get('link_curto',''),
                    'canal':rs.get('canal','link'),
                    'erro_envio':rs.get('erro_envio',''),
                    'expira_em':rs.get('expira_em','')})

@app.route('/api/funcionarios/arquivos/<int:id>/assinatura/rastreio')
@lr
def api_func_arquivo_assinatura_rastreio(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    return jsonify({
        'ok':True,
        'arquivo_id':a.id,
        'status_assinatura':a.ass_status or '',
        'canal_envio':a.ass_canal_envio or '',
        'enviado_em':(a.ass_enviado_em.isoformat() if a.ass_enviado_em else ''),
        'recebido_em':(a.ass_recebido_em.isoformat() if a.ass_recebido_em else ''),
        'aberto_em':(a.ass_aberto_em.isoformat() if a.ass_aberto_em else ''),
        'assinado_em':(a.ass_em.isoformat() if a.ass_em else ''),
        'prazo_em':(a.ass_prazo_em.isoformat() if a.ass_prazo_em else ''),
        'whatsapp':{
            'status':a.ass_wa_status or 'nao_enviado',
            'enviado_em':(a.ass_wa_enviado_em.isoformat() if a.ass_wa_enviado_em else ''),
            'recebido_em':(a.ass_wa_recebido_em.isoformat() if a.ass_wa_recebido_em else ''),
        },
        'email':{
            'status':a.ass_email_status or 'nao_enviado',
            'enviado_em':(a.ass_email_enviado_em.isoformat() if a.ass_email_enviado_em else ''),
            'recebido_em':(a.ass_email_recebido_em.isoformat() if a.ass_email_recebido_em else ''),
        }
    })

@app.route('/api/funcionarios/arquivos/<int:id>/assinatura/cancelar',methods=['POST'])
@lr
def api_func_arquivo_cancelar_assinatura(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    status_atual=(a.ass_status or '').strip().lower()
    if status_atual not in ('pendente','nao_solicitada','expirado',''):
        return jsonify({'erro':'Não é possível cancelar: assinatura já concluída ou em status inválido.'}),400
    a.ass_status='cancelada'
    a.ass_token=''
    a.ass_expira_em=None
    a.ass_prazo_em=None
    a.ass_lembretes_enviados=0
    db.session.commit()
    audit_event('func_arquivo_assinatura_cancelada','usuario',session.get('uid'),'funcionario',a.funcionario_id,True,
                {'arquivo_id':id,'nome':a.nome_arquivo})
    return jsonify({'ok':True,'mensagem':'Solicitação de assinatura cancelada.'})

@app.route('/api/rh/dashboard/assinaturas-pendentes')
@lr
def api_rh_dashboard_assinaturas_pendentes():
    """Dashboard: pendentes de assinatura por funcionário, ordenados por mais antigo."""
    from sqlalchemy import func as sqlfunc
    pendentes=FuncionarioArquivo.query.filter_by(ass_status='pendente').order_by(FuncionarioArquivo.criado_em.asc()).all()
    por_func={}
    agora=utcnow()
    for a in pendentes:
        fid=a.funcionario_id
        if fid not in por_func:
            f=Funcionario.query.get(fid)
            por_func[fid]={
                'funcionario_id':fid,
                'funcionario_nome':(f.nome if f else f'ID {fid}'),
                'funcionario_cargo':(f.cargo or '') if f else '',
                'total':0,
                'vencidos':0,
                'itens':[],
            }
        dias_pendente=(agora-a.criado_em).days if a.criado_em else 0
        vencido=bool(a.ass_prazo_em and a.ass_prazo_em<agora)
        por_func[fid]['total']+=1
        if vencido: por_func[fid]['vencidos']+=1
        por_func[fid]['itens'].append({
            'arquivo_id':a.id,
            'nome_arquivo':a.nome_arquivo,
            'categoria':DOC_CAT_LABEL.get(norm_cat(a.categoria),a.categoria),
            'competencia':a.competencia or '',
            'dias_pendente':dias_pendente,
            'prazo_em':(a.ass_prazo_em.isoformat() if a.ass_prazo_em else None),
            'prazo_fmt':(a.ass_prazo_em.strftime('%d/%m/%Y') if a.ass_prazo_em else None),
            'vencido':vencido,
            'lembretes_enviados':a.ass_lembretes_enviados or 0,
        })
    lista=sorted(por_func.values(),key=lambda x:x['vencidos'],reverse=True)
    return jsonify({'ok':True,'total_pendentes':len(pendentes),'funcionarios':lista})

@app.route('/api/rh/assinaturas/painel')
@lr
def api_rh_assinaturas_painel():
    """Painel completo de status de assinaturas com filtros."""
    status_filtro=(request.args.get('status') or 'todos').strip().lower()
    fid_filtro=to_num(request.args.get('funcionario_id') or 0)
    cat_filtro=(request.args.get('categoria') or '').strip().lower()
    comp_filtro=(request.args.get('competencia') or '').strip()
    data_ini_str=(request.args.get('data_ini') or '').strip()
    data_fim_str=(request.args.get('data_fim') or '').strip()

    q=FuncionarioArquivo.query.filter(
        FuncionarioArquivo.ass_status!='nao_solicitada',
        FuncionarioArquivo.ass_status!=None
    )
    if status_filtro!='todos':
        q=q.filter(FuncionarioArquivo.ass_status==status_filtro)
    if fid_filtro:
        q=q.filter(FuncionarioArquivo.funcionario_id==fid_filtro)
    if cat_filtro:
        q=q.filter(FuncionarioArquivo.categoria==cat_filtro)
    if comp_filtro:
        q=q.filter(FuncionarioArquivo.competencia==comp_filtro)
    try:
        if data_ini_str:
            q=q.filter(FuncionarioArquivo.criado_em>=datetime.strptime(data_ini_str,'%Y-%m-%d'))
        if data_fim_str:
            q=q.filter(FuncionarioArquivo.criado_em<datetime.strptime(data_fim_str,'%Y-%m-%d')+timedelta(days=1))
    except Exception:
        pass

    registros=q.order_by(FuncionarioArquivo.criado_em.desc()).limit(500).all()
    agora=utcnow()
    func_cache={}
    def _get_func(fid):
        if fid not in func_cache:
            func_cache[fid]=Funcionario.query.get(fid)
        return func_cache[fid]

    itens=[]
    totais={'pendente':0,'concluida':0,'cancelada':0,'expirado':0,'outros':0}
    for a in registros:
        f=_get_func(a.funcionario_id)
        st=a.ass_status or 'pendente'
        vencido=bool(st=='pendente' and a.ass_prazo_em and a.ass_prazo_em<agora)
        if st in totais: totais[st]+=1
        else: totais['outros']+=1
        itens.append({
            'arquivo_id':a.id,
            'funcionario_id':a.funcionario_id,
            'funcionario_nome':(f.nome if f else f'ID {a.funcionario_id}'),
            'funcionario_cargo':(f.cargo or '') if f else '',
            'nome_arquivo':a.nome_arquivo,
            'categoria':a.categoria or '',
            'categoria_label':DOC_CAT_LABEL.get(norm_cat(a.categoria),a.categoria or ''),
            'competencia':a.competencia or '',
            'ass_status':st,
            'ass_canal':a.ass_canal_envio or '',
            'criado_em':(a.criado_em.isoformat() if a.criado_em else None),
            'criado_fmt':(a.criado_em.strftime('%d/%m/%Y') if a.criado_em else ''),
            'ass_em':(a.ass_em.isoformat() if a.ass_em else None),
            'ass_em_fmt':(a.ass_em.strftime('%d/%m/%Y %H:%M') if a.ass_em else ''),
            'prazo_em':(a.ass_prazo_em.isoformat() if a.ass_prazo_em else None),
            'prazo_fmt':(a.ass_prazo_em.strftime('%d/%m/%Y') if a.ass_prazo_em else ''),
            'vencido':vencido,
            'lembretes':a.ass_lembretes_enviados or 0,
        })

    return jsonify({'ok':True,'total':len(itens),'totais':totais,'itens':itens})

@app.route('/api/rh/assinaturas/lembrete-pendentes',methods=['POST'])
@lr
def api_rh_assinaturas_lembrete_pendentes():
    """Reenvia lembrete para todos os documentos pendentes, respeitando o canal original."""
    fid_filtro=to_num(request.args.get('funcionario_id') or 0)
    cat_filtro=(request.args.get('categoria') or '').strip().lower()
    comp_filtro=(request.args.get('competencia') or '').strip()
    data_ini_str=(request.args.get('data_ini') or '').strip()
    data_fim_str=(request.args.get('data_fim') or '').strip()

    q=FuncionarioArquivo.query.filter(FuncionarioArquivo.ass_status=='pendente')
    if fid_filtro:
        q=q.filter(FuncionarioArquivo.funcionario_id==fid_filtro)
    if cat_filtro:
        q=q.filter(FuncionarioArquivo.categoria==cat_filtro)
    if comp_filtro:
        q=q.filter(FuncionarioArquivo.competencia==comp_filtro)
    try:
        if data_ini_str:
            q=q.filter(FuncionarioArquivo.criado_em>=datetime.strptime(data_ini_str,'%Y-%m-%d'))
        if data_fim_str:
            q=q.filter(FuncionarioArquivo.criado_em<datetime.strptime(data_fim_str,'%Y-%m-%d')+timedelta(days=1))
    except Exception:
        pass

    registros=q.order_by(FuncionarioArquivo.criado_em.desc()).limit(500).all()
    if not registros:
        return jsonify({'ok':True,'total':0,'enviados':0,'falhas':0,'mensagem':'Nenhum pendente encontrado para lembrete.'})

    enviados=0
    falhas=0
    canais={'whatsapp':0,'email':0,'app':0,'link':0}
    erros=[]
    func_cache={}

    def _canal_padrao_arquivo(a):
        ch=(a.ass_canal_envio or '').strip().lower()
        if ch:
            return ch
        if (a.ass_wa_status or '') in ('enviado','recebido') or bool(a.ass_wa_enviado_em):
            return 'whatsapp'
        if (a.ass_email_status or '') in ('enviado','recebido') or bool(a.ass_email_enviado_em):
            return 'email'
        if not (a.ass_token or '').strip():
            return 'app'
        return 'link'

    for a in registros:
        try:
            if a.funcionario_id not in func_cache:
                func_cache[a.funcionario_id]=Funcionario.query.get(a.funcionario_id)
            f=func_cache[a.funcionario_id]
            canal=_ass_track_channel('',_canal_padrao_arquivo(a))
            rs=_solicitar_assinatura_arquivo_funcionario(
                a,
                f,
                canal=canal,
                commit_now=False,
                forcar_novo_token=False,
                eh_lembrete=True,
            )
            if rs.get('ok'):
                enviados+=1
                canal_efetivo=_ass_track_channel(rs.get('canal') or canal,canal)
                canais[canal_efetivo]=canais.get(canal_efetivo,0)+1
                a.ass_lembretes_enviados=(a.ass_lembretes_enviados or 0)+1
                if rs.get('erro_envio'):
                    falhas+=1
                    erros.append(f"{a.nome_arquivo}: {rs.get('erro_envio')}")
            else:
                falhas+=1
                erros.append(f"{a.nome_arquivo}: {rs.get('erro') or 'falha no envio'}")
        except Exception as ex:
            falhas+=1
            erros.append(f"{a.nome_arquivo}: {str(ex)}")

    db.session.commit()
    audit_event(
        'rh_assinaturas_lembrete_massa',
        'usuario',
        session.get('uid'),
        'rh',
        0,
        True,
        {
            'total':len(registros),
            'enviados':enviados,
            'falhas':falhas,
            'canais':canais,
        }
    )
    return jsonify({
        'ok':True,
        'total':len(registros),
        'enviados':enviados,
        'falhas':falhas,
        'canais':canais,
        'erros':erros[:20],
    })

@app.route('/doc/assinar/<token>')
def func_doc_assinar_publica(token):
    a=FuncionarioArquivo.query.filter_by(ass_token=token).first()
    if not a:
        return render_template('doc_assinatura.html',ok=False,mensagem='Link de assinatura inválido.',arquivo=None,funcionario=None)
    if (a.ass_status or '')=='assinado':
        return render_template('doc_assinatura.html',ok=False,mensagem='Este documento já foi assinado.',arquivo=a,funcionario=Funcionario.query.get(a.funcionario_id))
    if a.ass_expira_em and a.ass_expira_em<utcnow():
        a.ass_status='expirado'; db.session.commit()
        return render_template('doc_assinatura.html',ok=False,mensagem='Link expirado. Solicite um novo link ao RH.',arquivo=a,funcionario=Funcionario.query.get(a.funcionario_id))
    src=request.args.get('src','')
    if _ass_track_mark_received(a,src):
        db.session.commit()
    return render_template('doc_assinatura.html',ok=True,mensagem='',arquivo=a,funcionario=Funcionario.query.get(a.funcionario_id))

@app.route('/doc/assinar/<token>/arquivo')
def func_doc_assinar_visualizar_arquivo(token):
    a=FuncionarioArquivo.query.filter_by(ass_token=token).first_or_404()
    if a.ass_expira_em and a.ass_expira_em<utcnow():
        return 'Link expirado.',400
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho or '')
    if not os.path.exists(abs_p):
        return 'Arquivo não encontrado.',404
    src=request.args.get('src','')
    if _ass_track_mark_opened(a,src):
        db.session.commit()
    return send_file(abs_p,as_attachment=False,download_name=a.nome_arquivo)

@app.route('/api/doc/assinar/<token>/enviar-otp',methods=['GET','POST'])
def api_func_doc_assinatura_enviar_otp(token):
    a=FuncionarioArquivo.query.filter_by(ass_token=token).first()
    if not a:
        return _assinatura_json_erro('Link inválido.',404)
    if (a.ass_status or '')=='assinado':
        return _assinatura_json_erro('Documento já assinado.',400)
    if a.ass_expira_em and a.ass_expira_em<utcnow():
        return _assinatura_json_erro('Link expirado.',400)
    f=Funcionario.query.get(a.funcionario_id)
    tel=(f.telefone if f else '') or ''
    email=(f.email if f else '') or ''
    if not (wa_norm_number(tel) or (email or '').strip()):
        return _assinatura_json_erro('Nenhum telefone ou e-mail cadastrado para envio do OTP.',400)
    codigo=_otp_new_code()
    a.ass_otp_hash=token_hash(codigo)
    a.ass_otp_expira_em=utcnow()+timedelta(minutes=10)
    a.ass_otp_tentativas=0
    try:
        envio=_send_signature_otp(codigo,nome_dest=(f.nome if f else ''),telefone=tel,email=email,contexto='documento')
        db.session.commit()
        return _assinatura_json_ok(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )
    except Exception as ex:
        db.session.rollback()
        return _assinatura_json_erro(f'Falha ao enviar OTP: {str(ex)}',500)

@app.route('/api/doc/assinar/<token>/confirmar',methods=['POST'])
def api_func_doc_assinatura_confirmar(token):
    a=FuncionarioArquivo.query.filter_by(ass_token=token).first()
    if not a:
        return _assinatura_json_erro('Link inválido.',404)
    if a.ass_expira_em and a.ass_expira_em<utcnow():
        a.ass_status='expirado'; db.session.commit()
        return _assinatura_json_erro('Link expirado. Solicite novo link ao RH.',400)
    d=request.json or {}
    nome=(d.get('nome') or '').strip()
    cargo=(d.get('cargo') or '').strip()
    cpf_info=only_digits(d.get('cpf') or '')
    otp=(only_digits(d.get('otp') or '') or '').strip()
    aceite=bool(d.get('aceite'))
    if not nome:
        return _assinatura_json_erro('Informe o nome completo para assinar.',400)
    if not cpf_info or len(cpf_info)!=11:
        return _assinatura_json_erro('Informe um CPF válido (11 dígitos) para assinar.',400)
    if not aceite:
        return _assinatura_json_erro('Confirme o aceite para concluir a assinatura.',400)

    # Regra RH: se CPF digitado for igual ao CPF cadastrado do funcionário, não bloquear.
    # Só bloquear por divergência quando existir CPF cadastrado diferente.
    f=Funcionario.query.get(a.funcionario_id)
    cpf_cadastrado=''
    if f and f.cpf:
        cpf_cadastrado=only_digits(f.cpf or '')
    cpf_confere_cadastro=bool(cpf_cadastrado and cpf_info==cpf_cadastrado)

    # Se não confere com cadastro, exigimos CPF válido por algoritmo.
    if not cpf_confere_cadastro and not _valida_cpf(cpf_info):
        return _assinatura_json_erro('Informe um CPF válido (11 dígitos) para assinar.',400)

    # Se existe CPF cadastrado e é diferente do informado, bloqueia para segurança.
    if cpf_cadastrado and cpf_info!=cpf_cadastrado:
        return _assinatura_json_erro('O CPF informado não confere com o funcionário vinculado ao documento.',400)

    if not otp:
        codigo=_otp_new_code()
        a.ass_otp_hash=token_hash(codigo)
        a.ass_otp_expira_em=utcnow()+timedelta(minutes=10)
        a.ass_otp_tentativas=0
        try:
            envio=_send_signature_otp(codigo,nome_dest=nome,telefone=(f.telefone if f else ''),email=(f.email if f else ''),contexto='documento')
        except Exception as ex:
            db.session.rollback()
            return _assinatura_json_erro(f'Falha ao enviar OTP de confirmação: {str(ex)}',400)
        db.session.commit()
        return _assinatura_json_otp(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )

    if not (a.ass_otp_hash or '').strip() or not a.ass_otp_expira_em:
        return _assinatura_json_erro('Solicite um novo código OTP para concluir a assinatura.',400)
    if a.ass_otp_expira_em<utcnow():
        return _assinatura_json_erro('Código OTP expirado. Solicite um novo código.',400)
    tent=int(a.ass_otp_tentativas or 0)
    if tent>=5:
        return _assinatura_json_erro('Limite de tentativas de OTP excedido. Solicite um novo código.',400)
    if not hmac.compare_digest(token_hash(otp),str(a.ass_otp_hash or '')):
        a.ass_otp_tentativas=tent+1
        db.session.commit()
        return _assinatura_json_erro('Código OTP inválido.',400)

    if not a.ass_codigo:
        a.ass_codigo=secrets.token_urlsafe(10)
    ip=(request.headers.get('X-Forwarded-For','') or request.remote_addr or '').split(',')[0].strip()[:60]
    a.ass_status='assinado'; a.ass_nome=nome; a.ass_cargo=cargo
    a.ass_cpf=cpf_info; a.ass_ip=ip; a.ass_em=utcnow(); a.ass_token=None
    a.ass_otp_hash=None; a.ass_otp_expira_em=None; a.ass_otp_tentativas=0
    if not a.ass_aberto_em:
        a.ass_aberto_em=utcnow()
    db.session.commit()
    copia_assinada=_salvar_pdf_assinado_em_arquivos_funcionario(a,f,request.url_root.rstrip('/'))

    if copia_assinada and copia_assinada.get('ok') and copia_assinada.get('abs_path'):
        rs_crypto=_try_sign_pdf_file_crypto(copia_assinada.get('abs_path'),empresa_id=(f.empresa_id if f else None),usuario_id=session.get('uid'))
        hash_final=_sha256_file(copia_assinada.get('abs_path'))
        novo=FuncionarioArquivo.query.get(copia_assinada.get('arquivo_id')) if copia_assinada.get('arquivo_id') else None
        if novo:
            novo.ass_doc_hash=hash_final
            novo.ass_crypto_ok=bool(rs_crypto.get('ok'))
            novo.ass_cert_subject=(rs_crypto.get('cert_subject') or '')[:255] if rs_crypto.get('ok') else None
        a.ass_doc_hash=hash_final
        a.ass_crypto_ok=bool(rs_crypto.get('ok'))
        a.ass_cert_subject=(rs_crypto.get('cert_subject') or '')[:255] if rs_crypto.get('ok') else None
        db.session.commit()

    audit_event('func_arquivo_assinatura_confirmada','externo',None,'funcionario',a.funcionario_id,True,
                {'arquivo_id':a.id,'nome':nome,'cpf_parcial':cpf_info[:3]+'***'+cpf_info[-2:]})
    validacao_link=f"{request.url_root.rstrip('/')}/doc/validar/{a.ass_codigo}"
    # Enviar cópia assinada ao funcionário via WhatsApp
    enviado_wa=False
    if f and (f.telefone or ''):
        tel=wa_norm_number(f.telefone or '')
        if wa_is_valid_number(tel):
            try:
                pdf_path=(copia_assinada.get('abs_path') if copia_assinada else '')
                pdf_nome=(copia_assinada.get('nome_arquivo') if copia_assinada else '') or f"doc_assinado_{a.ass_codigo}.pdf"
                if pdf_path and os.path.exists(pdf_path):
                    wa_send_pdf(tel,pdf_path,pdf_nome,
                        f"✅ Documento assinado: {a.nome_arquivo}\nCódigo: {a.ass_codigo}\nValidar: {validacao_link}")
                else:
                    pdf_buf=_build_doc_assinatura_pdf(a,f,request.url_root.rstrip('/'))
                    tmp_path=os.path.join(UPLOAD_ROOT,'tmp_ass')
                    os.makedirs(tmp_path,exist_ok=True)
                    tmp_file=os.path.join(tmp_path,pdf_nome)
                    with open(tmp_file,'wb') as fp: fp.write(pdf_buf)
                    wa_send_pdf(tel,tmp_file,pdf_nome,
                        f"✅ Documento assinado: {a.nome_arquivo}\nCódigo: {a.ass_codigo}\nValidar: {validacao_link}")
                    try: os.remove(tmp_file)
                    except: pass
                enviado_wa=True
            except Exception:
                pass
    signed_pdf_link=(f"{request.url_root.rstrip('/')}/doc/assinado/{a.ass_codigo}" if a.ass_codigo else '')
    return _assinatura_json_ok(
        mensagem='Assinatura concluída com sucesso.',
        validacao_link=validacao_link,
        signed_pdf_link=signed_pdf_link,
        whatsapp_enviado=enviado_wa,
        codigo=(a.ass_codigo or ''),
        arquivo_assinado_salvo=bool(copia_assinada and copia_assinada.get('ok'))
    )

@app.route('/doc/assinado/<codigo>')
def func_doc_assinado_publico(codigo):
    cod=(codigo or '').strip()
    if not cod:
        return 'Código inválido.',400
    arqs=(FuncionarioArquivo.query
          .filter_by(ass_codigo=cod,ass_status='assinado')
          .order_by(FuncionarioArquivo.id.desc())
          .all())
    if not arqs:
        return 'Documento assinado não encontrado.',404
    alvo=None
    for a in arqs:
        nm=(a.nome_arquivo or '').lower()
        if nm.endswith('_assinado.pdf') or '_assinado' in nm:
            alvo=a
            break
    if not alvo:
        alvo=arqs[0]
    raw=(alvo.caminho or '').strip()
    cands=[raw]
    if raw and not os.path.isabs(raw):
        cands.append(os.path.join(UPLOAD_ROOT,raw))
        cands.append(os.path.join(_get_uploads_base(),raw))
    abs_path=''
    for p in cands:
        if p and os.path.exists(p):
            abs_path=p
            break
    if not abs_path:
        return 'Arquivo assinado não encontrado.',404
    return send_file(abs_path,mimetype='application/pdf',as_attachment=False,download_name=alvo.nome_arquivo or 'documento_assinado.pdf')

@app.route('/doc/validar/<codigo>')
def func_doc_validar_publica(codigo):
    cod=(codigo or '').strip()
    if not cod:
        return render_template('doc_validacao.html',ok=False,mensagem='Código de validação inválido.',arquivo=None,funcionario=None)
    a=FuncionarioArquivo.query.filter_by(ass_codigo=cod).first()
    if not a:
        return render_template('doc_validacao.html',ok=False,mensagem='Assinatura não encontrada para o código informado.',arquivo=None,funcionario=None)
    f=Funcionario.query.get(a.funcionario_id)
    ok=(a.ass_status or '').strip().lower()=='assinado'
    msg='Assinatura válida.' if ok else ('Assinatura pendente ou não concluída.' if (a.ass_status or '')=='pendente' else 'Assinatura expirada ou inválida.')
    return render_template('doc_validacao.html',ok=ok,mensagem=msg,arquivo=a,funcionario=f)

def _build_doc_assinatura_pdf(arquivo,funcionario,url_root):
    """Gera PDF de auditoria da assinatura de documento de funcionário."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT,TA_RIGHT,TA_CENTER
    from reportlab.graphics.barcode import qr as qr_code
    from reportlab.graphics.shapes import Drawing

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.2*cm,rightMargin=1.2*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    W=A4[0]-2.4*cm
    AZ=colors.HexColor('#205d8a'); VD=colors.HexColor('#1a7a45'); CI=colors.HexColor('#f5f5f5')
    LJ=colors.HexColor('#f28e34')
    def ps(nm,**kw):
        b=dict(fontName='Helvetica',fontSize=10,leading=14,textColor=colors.HexColor('#020202'),spaceAfter=0,spaceBefore=0)
        b.update(kw); return ParagraphStyle(nm,**b)
    def _fmt(v):
        if isinstance(v,datetime): return v.strftime('%d/%m/%Y %H:%M')
        t=(v or '').strip()
        if not t: return '-'
        try: return datetime.fromisoformat(t.replace('Z','+00:00')).strftime('%d/%m/%Y %H:%M')
        except: return t

    emp=None
    if funcionario and funcionario.empresa_id:
        emp=Empresa.query.get(funcionario.empresa_id)
    if not emp:
        emp=Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem,Empresa.id).first()
    empresa_nome=(emp.nome if emp and emp.nome else 'RM Facilities')
    empresas_hdr=_pdf_companies_for_header(empresa_obj=emp,limit=2)

    def _logo_flowable(item):
        for cand in (item.get('logos') or []):
            try:
                if str(cand).lower().startswith('http'):
                    with urllib.request.urlopen(cand,timeout=8) as resp:
                        data=resp.read()
                    img=Image(io.BytesIO(data),width=3.2*cm,height=1.45*cm)
                    img.hAlign='LEFT'
                    return img
                if os.path.exists(cand):
                    img=Image(cand,width=3.2*cm,height=1.45*cm)
                    img.hAlign='LEFT'
                    return img
            except Exception:
                continue
            return Paragraph(f'<b>{item.get("nome") or empresa_nome}</b>',ps('lgfb',fontSize=11,textColor=colors.HexColor('#0f2b47')))

    validacao_link=f"{url_root}/doc/validar/{arquivo.ass_codigo}" if arquivo.ass_codigo else ''
    hash_comp=(arquivo.ass_doc_hash or '').strip()
    if not hash_comp:
        abs_src=''
        raw=(arquivo.caminho or '').strip()
        if raw:
            cands=[raw]
            if not os.path.isabs(raw):
                cands.append(os.path.join(UPLOAD_ROOT,raw))
                cands.append(os.path.join(_get_uploads_base(),raw))
            for p in cands:
                if p and os.path.exists(p):
                    abs_src=p
                    break
        if abs_src:
            try:
                hash_comp=_sha256_file(abs_src)
            except Exception:
                hash_comp=''
    if not hash_comp:
        trilha='|'.join([str(arquivo.id or ''),arquivo.nome_arquivo or '',arquivo.categoria or '',
                         arquivo.ass_nome or '',arquivo.ass_cpf or '',arquivo.ass_ip or '',
                         _fmt(arquivo.ass_em),arquivo.ass_codigo or ''])
        hash_comp=hashlib.sha256(trilha.encode('utf-8')).hexdigest().upper()
    story=[]

    # Cabeçalho profissional com identidade da empresa
    logo=_logo_flowable(empresas_hdr[0])
    hdr_right=Paragraph(
        f'<b>AUDITORIA E VALIDAÇÃO DE ASSINATURA ELETRÔNICA</b><br/>'
        f'<font size="9" color="#49607a">{empresa_nome}</font><br/>'
        f'<font size="8" color="#6f8093">Emitido em {_fmt(localnow())}</font>',
        ps('htr',fontSize=11,leading=13,textColor=colors.white)
    )
    hdr=Table([[logo,hdr_right]],colWidths=[W*0.26,W*0.74])
    hdr.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),AZ),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),10),
        ('RIGHTPADDING',(0,0),(-1,-1),10),
        ('TOPPADDING',(0,0),(-1,-1),8),
        ('BOTTOMPADDING',(0,0),(-1,-1),8),
    ]))
    story.append(hdr); story.append(Spacer(1,5))

    emp_cells=[]
    for i,item in enumerate(empresas_hdr[:2]):
        cell=Table([
            [_logo_flowable(item)],
            [Paragraph(f'<b>{item.get("nome") or "-"}</b><br/><font size="8" color="#4c6072">CNPJ: {item.get("cnpj") or "-"}</font>',ps(f'empc{i}',fontSize=8.2,leading=10))]
        ],colWidths=[W*0.49])
        cell.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#d0d7df')),
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fbff')),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        emp_cells.append(cell)
    while len(emp_cells)<2:
        emp_cells.append(Paragraph('',ps('empemptyd',fontSize=1)))
    emp_tbl=Table([emp_cells],colWidths=[W*0.495,W*0.495])
    emp_tbl.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(emp_tbl); story.append(Spacer(1,6))

    badge_cor=colors.HexColor('#ecf8f0'); badge_txt=VD
    st=Table([[Paragraph('<b>✔ DOCUMENTO ASSINADO ELETRONICAMENTE</b>',ps('bs',fontSize=10,textColor=badge_txt))]],colWidths=[W])
    st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),badge_cor),('BOX',(0,0),(-1,-1),0.7,colors.HexColor('#9ed3b1')),('LEFTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
    story.append(st); story.append(Spacer(1,5))

    fn_nome=(funcionario.nome if funcionario else '-')
    fn_cpf=(funcionario.cpf if funcionario else '-')
    fn_emp=(emp.nome if emp else '')
    detalhes=[
        ('Documento',arquivo.nome_arquivo or '-'),
        ('Categoria',arquivo.categoria or '-'),
        ('Competência',arquivo.competencia or '-'),
        ('Funcionário',fn_nome),
        ('CPF do funcionário',fn_cpf or '-'),
        ('Empresa',fn_emp or '-'),
        ('Assinante (nome)',arquivo.ass_nome or '-'),
        ('CPF informado na assinatura',arquivo.ass_cpf or '-'),
        ('Cargo informado',arquivo.ass_cargo or '-'),
        ('Data/Hora da assinatura',_fmt(arquivo.ass_em)),
        ('IP de origem',arquivo.ass_ip or '-'),
        ('Código de validação',arquivo.ass_codigo or '-'),
        ('Link de validação',validacao_link or '-'),
        ('Hash SHA-256',hash_comp[:32]+'...'),
    ]
    rows=[[Paragraph(f'<b>{k}</b>',ps('dk',fontSize=8,textColor=AZ)),Paragraph(v,ps('dv',fontSize=8,leading=11))] for k,v in detalhes]
    det=Table(rows,colWidths=[W*0.32,W*0.68])
    det.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('BOX',(0,0),(-1,-1),0.6,colors.HexColor('#d0d7df')),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#e3e8ef')),('LEFTPADDING',(0,0),(-1,-1),5),('RIGHTPADDING',(0,0),(-1,-1),5),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(det); story.append(Spacer(1,6))

    # Assinatura em estilo cursivo
    ass_nome=arquivo.ass_nome or '-'
    ass_cargo=arquivo.ass_cargo or 'Assinante'
    ass_data=_fmt(arquivo.ass_em)
    sig_title=Paragraph('<b>ASSINATURA ELETRÔNICA REGISTRADA</b>',ps('sgt',fontSize=9,textColor=AZ))
    sig_draw=Paragraph(ass_nome,ps('sgn',fontName='Times-Italic',fontSize=22,textColor=colors.HexColor('#1a2e42'),leading=24))
    sig_meta=Paragraph(f'{ass_cargo} · Data/hora: {ass_data}',ps('sgm',fontSize=8,textColor=colors.HexColor('#5d6f82')))
    sig=Table([[sig_title],[sig_draw],[Paragraph('<font color="#9db0c1">______________________________________________</font>',ps('sgl',fontSize=9))],[sig_meta]],colWidths=[W])
    sig.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fbff')),
        ('BOX',(0,0),(-1,-1),0.6,colors.HexColor('#d0d7df')),
        ('LEFTPADDING',(0,0),(-1,-1),10),
        ('RIGHTPADDING',(0,0),(-1,-1),10),
        ('TOPPADDING',(0,0),(-1,-1),6),
        ('BOTTOMPADDING',(0,0),(-1,-1),6),
    ]))
    story.append(sig); story.append(Spacer(1,6))

    if validacao_link:
        try:
            qr_widget=qr_code.QrCodeWidget(validacao_link)
            b=qr_widget.getBounds()
            bw=max(1,b[2]-b[0]); bh=max(1,b[3]-b[1])
            sz=80
            qr_draw=Drawing(sz,sz,transform=[sz/bw,0,0,sz/bh,0,0])
            qr_draw.add(qr_widget)
            qr_tbl=Table([[qr_draw,Paragraph('Escaneie o QR Code para validar esta assinatura em tempo real no portal RM Facilities.',ps('qrp',fontSize=9,leading=13,textColor=colors.HexColor('#4c6072')))]],colWidths=[W*0.22,W*0.78])
            qr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4)]))
            story.append(qr_tbl)
        except Exception:
            story.append(Paragraph(f'Link: {validacao_link}',ps('ql',fontSize=8)))
    story.append(Spacer(1,6))
    bar=Table([[' ']],colWidths=[W])
    bar.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(bar); story.append(Spacer(1,4))
    story.append(Paragraph(f'Gerado em {localnow().strftime("%d/%m/%Y %H:%M")} — RM Facilities',ps('rod',fontSize=7,textColor=colors.HexColor('#999'),alignment=TA_CENTER)))
    doc.build(story); return buf.getvalue()


def _montar_pdf_assinado_funcionario(arquivo,funcionario,url_root):
    """Retorna bytes do PDF assinado: páginas originais estampadas + página de auditoria."""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.graphics.barcode import qr as qr_code
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF

    # Resolve caminho do arquivo original
    raw=(arquivo.caminho or '').strip()
    abs_path=''
    if raw:
        cands=[raw]
        if not os.path.isabs(raw):
            cands.append(os.path.join(UPLOAD_ROOT,raw))
            cands.append(os.path.join(_get_uploads_base(),raw))
        for p in cands:
            if p and os.path.exists(p):
                abs_path=p; break

    validacao_link=f"{url_root}/doc/validar/{arquivo.ass_codigo}" if arquivo.ass_codigo else ''
    footer_text=(f"Assinado eletronicamente por {arquivo.ass_nome or '?'} em "
                 f"{(arquivo.ass_em.strftime('%d/%m/%Y %H:%M') if isinstance(arquivo.ass_em,datetime) else str(arquivo.ass_em or ''))} "
                 f"| Cód. {arquivo.ass_codigo or ''} | RM Facilities")

    writer=PdfWriter()
    pages_added=0

    if abs_path:
        try:
            with open(abs_path,'rb') as fh:
                reader=PdfReader(fh)
                for page in reader.pages:
                    try:
                        w=float(page.mediabox.width)
                        h=float(page.mediabox.height)
                        ov=io.BytesIO()
                        c=rl_canvas.Canvas(ov,pagesize=(w,h))
                        c.setFillColorRGB(0.93,0.95,0.99)
                        c.rect(0,0,w,22,fill=1,stroke=0)
                        text_x=6
                        if validacao_link:
                            try:
                                qr_widget=qr_code.QrCodeWidget(validacao_link)
                                b=qr_widget.getBounds()
                                bw=max(1,b[2]-b[0]); bh=max(1,b[3]-b[1])
                                sz=14
                                qr_draw=Drawing(sz,sz,transform=[sz/bw,0,0,sz/bh,0,0])
                                qr_draw.add(qr_widget)
                                renderPDF.draw(qr_draw,c,5,4)
                                text_x=22
                            except Exception:
                                text_x=6
                        c.setFillColorRGB(0.13,0.36,0.54)
                        c.setFont('Helvetica',6)
                        c.drawString(text_x,8,footer_text)
                        c.save(); ov.seek(0)
                        ov_page=PdfReader(ov).pages[0]
                        page.merge_page(ov_page)
                    except Exception:
                        pass
                    writer.add_page(page)
                    pages_added+=1
        except Exception:
            pass

    # Página de auditoria
    audit_bytes=_build_doc_assinatura_pdf(arquivo,funcionario,url_root)
    try:
        from pypdf import PdfReader as _PR
        audit_reader=_PR(io.BytesIO(audit_bytes))
    except Exception:
        try:
            from PyPDF2 import PdfReader as _PR2
            audit_reader=_PR2(io.BytesIO(audit_bytes))
        except Exception:
            audit_reader=None
    if audit_reader:
        for page in audit_reader.pages:
            writer.add_page(page)
            pages_added+=1

    if pages_added==0:
        raise ValueError('Nenhuma página gerada para o PDF assinado.')
    out=io.BytesIO(); writer.write(out); return out.getvalue()


def _salvar_pdf_assinado_em_arquivos_funcionario(arquivo,funcionario,url_root):
    if not arquivo or not funcionario:
        return {'ok':False,'erro':'Arquivo ou funcionario invalido.'}
    try:
        pdf_buf=_montar_pdf_assinado_funcionario(arquivo,funcionario,url_root)
        comp=(arquivo.competencia or '').strip()
        ano=infer_doc_year(comp)
        prepare_func_doc_dirs(funcionario.id,ano)
        subdir,cat=func_doc_subdir(funcionario.id,arquivo.categoria or 'outros',comp)
        base_nome=os.path.splitext((arquivo.nome_arquivo or 'documento').strip())[0] or 'documento'
        nome_ass=f'{base_nome}_ASSINADO.pdf'
        rel,abs_p,nome_final=unique_rel_filename(subdir,nome_ass)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out:
            out.write(pdf_buf)
        hash_pdf=_sha256_file(abs_p)
        # Mantém o mesmo registro do documento e troca o arquivo principal pelo PDF assinado.
        caminho_antigo=(arquivo.caminho or '').strip()
        arquivo.categoria=cat
        arquivo.competencia=arquivo.competencia
        arquivo.nome_arquivo=nome_final
        arquivo.caminho=rel
        arquivo.ass_status='assinado'
        arquivo.ass_doc_hash=hash_pdf
        db.session.commit()
        # Remove o arquivo anterior quando for diferente do novo assinado.
        try:
            if caminho_antigo and caminho_antigo!=rel:
                cand=[caminho_antigo]
                if not os.path.isabs(caminho_antigo):
                    cand.append(os.path.join(UPLOAD_ROOT,caminho_antigo))
                    cand.append(os.path.join(_get_uploads_base(),caminho_antigo))
                for p in cand:
                    if p and os.path.exists(p):
                        os.remove(p)
                        break
        except Exception:
            pass
        return {'ok':True,'arquivo_id':arquivo.id,'nome_arquivo':nome_final,'caminho':rel,'abs_path':abs_p}
    except Exception as e:
        db.session.rollback()
        return {'ok':False,'erro':str(e)}

@app.route('/api/funcionarios/arquivos/<int:id>/assinatura/pdf-auditoria')
@lr
def api_func_arquivo_assinatura_pdf(id):
    a=FuncionarioArquivo.query.get_or_404(id)
    if (a.ass_status or '')!='assinado':
        return jsonify({'erro':'Documento ainda não foi assinado.'}),400
    f=Funcionario.query.get(a.funcionario_id)
    buf=_build_doc_assinatura_pdf(a,f,request.url_root.rstrip('/'))
    return send_file(io.BytesIO(buf),mimetype='application/pdf',as_attachment=True,
                     download_name=f'auditoria_assinatura_{a.ass_codigo or a.id}.pdf')

# ══════════════════════════════════════════════════════
#  ASSINATURA DE ENVELOPES — helper PDF
# ══════════════════════════════════════════════════════

def _build_envelope_audit_pdf(envelope, signatarios, url_root):
    """Gera página de auditoria final do envelope (ReportLab)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT,TA_RIGHT,TA_CENTER
    from reportlab.graphics.barcode import qr as qr_code
    from reportlab.graphics.shapes import Drawing

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.2*cm,rightMargin=1.2*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    W=A4[0]-2.4*cm
    AZ=colors.HexColor('#205d8a'); VD=colors.HexColor('#1a7a45'); CI=colors.HexColor('#f5f5f5')
    LJ=colors.HexColor('#f28e34')
    def ps(nm,**kw):
        b=dict(fontName='Helvetica',fontSize=10,leading=14,textColor=colors.HexColor('#020202'),spaceAfter=0,spaceBefore=0)
        b.update(kw); return ParagraphStyle(nm,**b)

    emp=Empresa.query.get(envelope.empresa_id) if envelope.empresa_id else None
    if not emp:
        emp=Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem,Empresa.id).first()
    empresa_nome=(emp.nome if emp and emp.nome else 'RM Facilities')
    empresas_hdr=_pdf_companies_for_header(empresa_obj=emp,limit=2)

    def _logo_flowable(item):
        for cand in (item.get('logos') or []):
            try:
                if str(cand).lower().startswith('http'):
                    with urllib.request.urlopen(cand,timeout=8) as resp:
                        data=resp.read()
                    img=Image(io.BytesIO(data),width=3.2*cm,height=1.45*cm)
                    img.hAlign='LEFT'
                    return img
                if os.path.exists(cand):
                    img=Image(cand,width=3.2*cm,height=1.45*cm)
                    img.hAlign='LEFT'
                    return img
            except Exception:
                continue
            return Paragraph(f'<b>{item.get("nome") or empresa_nome}</b>',ps('lgfb2',fontSize=11,textColor=colors.HexColor('#0f2b47')))

    validacao_link=f"{url_root}/envelope/validar/{envelope.codigo}" if envelope.codigo else ''
    hash_comp=(envelope.assinatura_doc_hash or '').strip()
    if not hash_comp:
        try:
            abs_pdf,_=_envelope_signed_pdf_path(envelope)
            if abs_pdf and os.path.exists(abs_pdf):
                hash_comp=_sha256_file(abs_pdf)
        except Exception:
            hash_comp=''
    if not hash_comp:
        trilha_base='|'.join([str(envelope.id),envelope.titulo or '',envelope.codigo or '',
                              '|'.join([f"{s.nome}|{s.cpf or ''}|{s.ass_em.isoformat() if s.ass_em else ''}" for s in signatarios])])
        hash_comp=hashlib.sha256(trilha_base.encode('utf-8')).hexdigest().upper()

    story=[]
    logo=_logo_flowable(empresas_hdr[0])
    hdr_right=Paragraph(
        f'<b>AUDITORIA E VALIDAÇÃO DE ASSINATURA ELETRÔNICA</b><br/>'
        f'<font size="9" color="#49607a">{empresa_nome}</font><br/>'
        f'<font size="8" color="#6f8093">Documento: {envelope.titulo or "-"}</font>',
        ps('htr2',fontSize=11,leading=13,textColor=colors.white)
    )
    hdr=Table([[logo,hdr_right]],colWidths=[W*0.26,W*0.74])
    hdr.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),AZ),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),10),
        ('RIGHTPADDING',(0,0),(-1,-1),10),
        ('TOPPADDING',(0,0),(-1,-1),8),
        ('BOTTOMPADDING',(0,0),(-1,-1),8),
    ]))
    story.append(hdr); story.append(Spacer(1,5))

    emp_cells=[]
    for i,item in enumerate(empresas_hdr[:2]):
        cell=Table([
            [_logo_flowable(item)],
            [Paragraph(f'<b>{item.get("nome") or "-"}</b><br/><font size="8" color="#4c6072">CNPJ: {item.get("cnpj") or "-"}</font>',ps(f'empe{i}',fontSize=8.2,leading=10))]
        ],colWidths=[W*0.49])
        cell.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#d0d7df')),
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fbff')),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        emp_cells.append(cell)
    while len(emp_cells)<2:
        emp_cells.append(Paragraph('',ps('empemptye',fontSize=1)))
    emp_tbl=Table([emp_cells],colWidths=[W*0.495,W*0.495])
    emp_tbl.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(emp_tbl); story.append(Spacer(1,6))
    assinados=[s for s in signatarios if s.status=='assinado']
    pendentes=[s for s in signatarios if s.status!='assinado']
    if len(pendentes)==0:
        badge_cor=colors.HexColor('#ecf8f0'); badge_txt=VD; badge_label='✔ TODOS OS SIGNATÁRIOS ASSINARAM O DOCUMENTO'
    else:
        badge_cor=colors.HexColor('#fff8ec'); badge_txt=LJ; badge_label=f'⏳ {len(assinados)} de {len(signatarios)} SIGNATÁRIOS ASSINARAM'
    st=Table([[Paragraph(f'<b>{badge_label}</b>',ps('bs',fontSize=10,textColor=badge_txt))]],colWidths=[W])
    st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),badge_cor),('BOX',(0,0),(-1,-1),0.7,colors.HexColor('#9ed3b1')),('LEFTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
    story.append(st); story.append(Spacer(1,5))

    detalhes=[
        ('Título do Documento',envelope.titulo or '-'),
        ('Tipo',{'funcionario':'Funcionário','cliente':'Cliente','avulso':'Avulso'}.get(envelope.tipo,envelope.tipo or '-')),
        ('Status',envelope.status or '-'),
        ('Código de Validação',envelope.codigo or '-'),
        ('Criado por',envelope.criado_por or '-'),
        ('Data de Criação',envelope.criado_em.strftime('%d/%m/%Y %H:%M') if envelope.criado_em else '-'),
        ('Validade',envelope.expira_em.strftime('%d/%m/%Y') if envelope.expira_em else 'Sem prazo'),
        ('Hash SHA-256',hash_comp[:32]+'...'),
    ]
    rows=[[Paragraph(f'<b>{k}</b>',ps('dk',fontSize=8,textColor=AZ)),Paragraph(v,ps('dv',fontSize=8,leading=11))] for k,v in detalhes]
    det=Table(rows,colWidths=[W*0.32,W*0.68])
    det.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('BOX',(0,0),(-1,-1),0.6,colors.HexColor('#d0d7df')),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#e3e8ef')),('LEFTPADDING',(0,0),(-1,-1),5),('RIGHTPADDING',(0,0),(-1,-1),5),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(det); story.append(Spacer(1,6))

    # Tabela de signatários
    story.append(Paragraph('<b>Signatários</b>',ps('sh',fontSize=10,textColor=AZ,spaceAfter=6)))
    sig_header=[Paragraph(x,ps(f'sh{i}',fontSize=8,textColor=colors.white,fontName='Helvetica-Bold'))
                for i,x in enumerate(['Nome','Cargo','CPF','Data/Hora','IP','Status'])]
    sig_rows=[sig_header]
    for s in signatarios:
        cpf_m=''
        if s.ass_cpf_informado and len(s.ass_cpf_informado)>=11:
            c=s.ass_cpf_informado.replace('.','').replace('-','').replace(' ','')
            cpf_m=f'***{c[3:6]}***' if len(c)==11 else s.ass_cpf_informado
        elif s.cpf and len(s.cpf)>=3:
            cpf_m='***'+s.cpf[-3:]
        sc=VD if s.status=='assinado' else LJ
        sig_rows.append([
            Paragraph(s.nome or '-',ps('sn',fontSize=8)),
            Paragraph(s.cargo or '-',ps('sc',fontSize=8)),
            Paragraph(cpf_m or '-',ps('scpf',fontSize=8)),
            Paragraph(s.ass_em.strftime('%d/%m/%Y %H:%M') if s.ass_em else '-',ps('sem',fontSize=8)),
            Paragraph(s.ass_ip or '-',ps('sip',fontSize=8)),
            Paragraph(s.status.upper(),ps('sst',fontSize=8,textColor=sc,fontName='Helvetica-Bold')),
        ])
    sig_tbl=Table(sig_rows,colWidths=[W*0.22,W*0.15,W*0.14,W*0.18,W*0.16,W*0.15])
    sig_tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),AZ),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#d0d7df')),
        ('LINEBELOW',(0,0),(-1,-1),0.3,colors.HexColor('#e3e8ef')),
        ('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,CI]),
    ]))
    story.append(sig_tbl); story.append(Spacer(1,6))

    if assinados:
        story.append(Paragraph('<b>ASSINATURAS REGISTRADAS</b>',ps('asg1',fontSize=8,textColor=AZ,spaceAfter=3)))
        sig_rows_cursive=[
            [Paragraph(f'<i>{s.nome or "-"}</i>',ps(f'asn{i}',fontName='Times-Italic',fontSize=18,textColor=colors.HexColor('#1a2e42'),leading=20)),
             Paragraph(f'{s.cargo or "Signatário"}<br/><font color="#5d6f82">{s.ass_em.strftime("%d/%m/%Y %H:%M") if s.ass_em else "-"}</font>',ps(f'asd{i}',fontSize=7.5,leading=10))]
            for i,s in enumerate(assinados)
        ]
        sig_curs=Table(sig_rows_cursive,colWidths=[W*0.52,W*0.48])
        sig_curs.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fbff')),
            ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#d0d7df')),
            ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#e3e8ef')),
            ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]))
        story.append(sig_curs)
    story.append(Spacer(1,6))

    if validacao_link:
        try:
            qr_widget=qr_code.QrCodeWidget(validacao_link)
            b=qr_widget.getBounds()
            bw=max(1,b[2]-b[0]); bh=max(1,b[3]-b[1])
            sz=75
            qr_draw=Drawing(sz,sz,transform=[sz/bw,0,0,sz/bh,0,0])
            qr_draw.add(qr_widget)
            qr_tbl=Table([[qr_draw,Paragraph('Escaneie para validar esta assinatura no portal RM Facilities.',ps('qrp',fontSize=9,leading=13,textColor=colors.HexColor('#4c6072')))]],colWidths=[W*0.20,W*0.80])
            qr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),4)]))
            story.append(qr_tbl)
        except Exception:
            story.append(Paragraph(f'Link: {validacao_link}',ps('ql',fontSize=8)))
    story.append(Spacer(1,10))
    bar=Table([[' ']],colWidths=[W])
    bar.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(bar); story.append(Spacer(1,4))
    story.append(Paragraph(f'Gerado em {localnow().strftime("%d/%m/%Y %H:%M")} — RM Facilities',ps('rod',fontSize=7,textColor=colors.HexColor('#999'),alignment=TA_CENTER)))
    doc.build(story); return buf.getvalue()


def _stamp_envelope_pdfs(arquivos, footer_text, envelope, url_root):
    """Estampa rodapé + carimbo de signatários nos PDFs do envelope e adiciona página de auditoria."""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as rl_canvas

    from reportlab.graphics.barcode import qr as qr_code
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    import base64 as _b64

    writer=PdfWriter()
    pages_added=0
    footer_qr_link=f"{url_root}/envelope/validar/{envelope.codigo}" if getattr(envelope,'codigo',None) else ''
    stamp_habilitado=bool(getattr(envelope,'stamp_habilitado',False))
    stamp_pagina=int(getattr(envelope,'stamp_pagina',1) or 1)
    stamp_x_pct=float(getattr(envelope,'stamp_x_pct',60.0) or 60.0)
    stamp_y_pct=float(getattr(envelope,'stamp_y_pct',10.0) or 10.0)
    stamp_todas_paginas=bool(getattr(envelope,'stamp_todas_paginas',False))
    stamp_todos_arquivos=bool(getattr(envelope,'stamp_todos_arquivos',False))
    stamp_signatarios=[]
    assinatura_img_map={}
    if stamp_habilitado:
        stamp_signatarios=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=envelope.id,status='assinado').all()
        for s in stamp_signatarios:
            if getattr(s,'ass_assinatura_img',None):
                assinatura_img_map[s.id]=s.ass_assinatura_img

    def _resolve_abs_path(caminho):
        raw=(caminho or '').strip()
        if not raw:
            return ''
        cands=[raw]
        if not os.path.isabs(raw):
            cands.append(os.path.join(UPLOAD_ROOT,raw))
            cands.append(os.path.join(_get_uploads_base(),raw))
        for p in cands:
            if p and os.path.exists(p):
                return p
        return ''

    def _make_footer_overlay(w,h):
        ov=io.BytesIO()
        c=rl_canvas.Canvas(ov,pagesize=(w,h))
        c.setFillColorRGB(0.93,0.95,0.99)
        c.rect(0,0,w,40,fill=1,stroke=0)
        text_x=8
        if footer_qr_link:
            try:
                qr_widget=qr_code.QrCodeWidget(footer_qr_link)
                b=qr_widget.getBounds()
                bw=max(1,b[2]-b[0]); bh=max(1,b[3]-b[1])
                sz=34
                qr_draw=Drawing(sz,sz,transform=[sz/bw,0,0,sz/bh,0,0])
                qr_draw.add(qr_widget)
                renderPDF.draw(qr_draw,c,5,4)
                text_x=44
            except Exception:
                text_x=8
        c.setFillColorRGB(0.13,0.36,0.54)
        c.setFont('Helvetica',8)
        c.drawString(text_x,22,footer_text)
        c.save(); ov.seek(0)
        return ov.getvalue()

    def _make_stamp_overlay(w,h,sigs,img_map,x_pct,y_pct):
        """Gera overlay com carimbos de assinatura posicionados na página."""
        if not sigs:
            return None
        ov=io.BytesIO()
        c=rl_canvas.Canvas(ov,pagesize=(w,h))
        stamp_w=w*0.38
        stamp_h=62.0
        gap=5.0
        x=w*(x_pct/100.0)
        x=max(0,min(x,w-stamp_w-4))
        y_base=h*(y_pct/100.0)
        # empilha stamps para cima a partir de y_base
        for i,sig in enumerate(sigs):
            y_bottom=y_base+(i*(stamp_h+gap))
            if y_bottom+stamp_h>h:
                break
            # fundo branco com borda azul
            c.setFillColorRGB(0.97,0.98,1.0)
            c.setStrokeColorRGB(0.55,0.70,0.90)
            c.roundRect(x,y_bottom,stamp_w,stamp_h,5,fill=1,stroke=1)
            # barra lateral azul esquerda
            c.setFillColorRGB(0.18,0.40,0.72)
            c.rect(x,y_bottom,3,stamp_h,fill=1,stroke=0)
            # label SIGNATÁRIO top-right
            c.setFillColorRGB(0.18,0.40,0.72)
            c.setFont('Helvetica-Bold',5.5)
            lbl='SIGNATÁRIO'
            lw=c.stringWidth(lbl,'Helvetica-Bold',5.5)
            c.drawString(x+stamp_w-lw-5,y_bottom+stamp_h-9,lbl)
            # imagem de assinatura (se disponível)
            text_x=x+8
            img_b64=img_map.get(sig.id)
            if img_b64:
                try:
                    from reportlab.lib.utils import ImageReader
                    raw=img_b64.split(',')[-1] if ',' in img_b64 else img_b64
                    img_bytes=_b64.b64decode(raw)
                    img_reader=ImageReader(io.BytesIO(img_bytes))
                    img_h_pt=28; img_w_pt=62
                    c.drawImage(img_reader,x+5,y_bottom+stamp_h-42,width=img_w_pt,height=img_h_pt,
                                preserveAspectRatio=True,mask='auto')
                    text_x=x+8
                except Exception:
                    pass
            else:
                # Fallback visual: quando não houver assinatura manuscrita,
                # imprime o nome completo no bloco de assinatura.
                nome_ass=(sig.nome or '').strip()[:42] or 'Assinatura não informada'
                c.setFillColorRGB(0.15,0.29,0.56)
                c.setFont('Helvetica-Oblique',10.5)
                c.drawString(x+6,y_bottom+stamp_h-29,nome_ass)
                c.setStrokeColorRGB(0.70,0.76,0.86)
                c.setLineWidth(0.6)
                c.line(x+6,y_bottom+stamp_h-31,x+6+min(stamp_w-18,180),y_bottom+stamp_h-31)
            # texto do signatário
            nome=(sig.nome or '').strip()[:38]
            data_str=sig.ass_em.strftime('%d/%m/%Y %H:%M') if sig.ass_em else ''
            codigo=('#'+(sig.ass_codigo or '')[:16])
            c.setFillColorRGB(0.08,0.18,0.38)
            c.setFont('Helvetica',6.5)
            c.drawString(text_x,y_bottom+17,'Assinado eletronicamente por')
            c.setFont('Helvetica-Bold',7)
            c.drawString(text_x,y_bottom+9,nome)
            c.setFont('Helvetica',6.5)
            c.drawString(text_x,y_bottom+2,f'Data: {data_str}')
            c.setFillColorRGB(0.45,0.52,0.65)
            c.setFont('Helvetica',5.5)
            c.drawString(text_x+stamp_w*0.35,y_bottom+stamp_h-10,codigo)
        c.save(); ov.seek(0)
        return ov.getvalue()

    file_idx=0
    for arq in arquivos:
        abs_path=_resolve_abs_path(arq.caminho)
        if not abs_path:
            continue
        try:
            with open(abs_path,'rb') as fh:
                reader=PdfReader(fh)
                n_pages=len(reader.pages)
                stamp_page_idx=min(stamp_pagina,n_pages)-1  # 0-based
                for page_idx,page in enumerate(reader.pages):
                    try:
                        w=float(page.mediabox.width)
                        h=float(page.mediabox.height)
                        # footer overlay
                        footer_bytes=_make_footer_overlay(w,h)
                        footer_page=PdfReader(io.BytesIO(footer_bytes)).pages[0]
                        page.merge_page(footer_page)
                        # stamp overlay
                        apply_stamp=False
                        if stamp_habilitado and stamp_signatarios:
                            if stamp_todos_arquivos or file_idx==0:
                                if stamp_todas_paginas or page_idx==stamp_page_idx:
                                    apply_stamp=True
                        if apply_stamp:
                            stamp_bytes=_make_stamp_overlay(w,h,stamp_signatarios,assinatura_img_map,stamp_x_pct,stamp_y_pct)
                            if stamp_bytes:
                                stamp_page=PdfReader(io.BytesIO(stamp_bytes)).pages[0]
                                page.merge_page(stamp_page)
                    except Exception:
                        pass
                    writer.add_page(page)
                    pages_added+=1
            file_idx+=1
        except Exception:
            continue

    if pages_added==0:
        raise ValueError('Nenhum arquivo PDF válido foi encontrado no envelope para geração do documento assinado.')

    # Página de auditoria final
    signatarios=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=envelope.id).all()
    audit_bytes=_build_envelope_audit_pdf(envelope,signatarios,url_root)
    PdfReaderAudit=PdfReader
    try:
        audit_reader=PdfReaderAudit(io.BytesIO(audit_bytes))
    except Exception:
        try:
            from PyPDF2 import PdfReader as PdfReaderAudit2
            audit_reader=PdfReaderAudit2(io.BytesIO(audit_bytes))
        except Exception:
            audit_reader=None
    if audit_reader:
        for page in audit_reader.pages:
            writer.add_page(page)

    out=io.BytesIO(); writer.write(out); return out.getvalue()


def _get_uploads_base():
    return UPLOAD_ROOT


def _normalize_signed_pdf_name(name):
    raw=(name or '').strip()
    if not raw:
        return ''
    base=os.path.basename(raw)
    if not base.lower().endswith('.pdf'):
        base=f'{base}.pdf'
    safe=secure_filename(base)
    if not safe:
        return ''
    if not safe.lower().endswith('.pdf'):
        safe=f"{os.path.splitext(safe)[0]}.pdf"
    return safe


def _default_signed_pdf_name(envelope):
    arq=(AssinaturaEnvelopeArquivo.query
         .filter_by(envelope_id=envelope.id)
         .order_by(AssinaturaEnvelopeArquivo.id.asc())
         .first())
    base='documento'
    if arq and (arq.nome_arquivo or '').strip():
        base=os.path.splitext(os.path.basename(arq.nome_arquivo.strip()))[0] or 'documento'
    default_name=_normalize_signed_pdf_name(f'{base} ASSINADO.pdf')
    if default_name:
        return default_name
    return f'documento_assinado_{envelope.id}.pdf'


def _envelope_signed_pdf_path(envelope):
    base=os.path.join(_get_uploads_base(),'envelopes',str(envelope.id),'assinado')
    os.makedirs(base,exist_ok=True)
    fname=_normalize_signed_pdf_name(getattr(envelope,'nome_documento_assinado',None))
    if not fname:
        fname=_default_signed_pdf_name(envelope)
    return os.path.join(base,fname),fname


def _gerar_pdf_assinado_envelope(envelope,url_root):
    arquivos=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=envelope.id).all()
    if not arquivos:
        raise ValueError('Envelope sem arquivos para gerar PDF assinado.')
    footer_text=(f"RM Facilities | Assinatura Eletrônica | Cód: {envelope.codigo} | "
                 f"Valide em: {url_root}/envelope/validar/{envelope.codigo}")
    pdf_bytes=_stamp_envelope_pdfs(arquivos,footer_text,envelope,url_root)
    abs_path,fname=_envelope_signed_pdf_path(envelope)
    with open(abs_path,'wb') as out:
        out.write(pdf_bytes)
    return abs_path,fname


def _salvar_pdf_assinado_destino_envelope(envelope,abs_pdf_path,fname):
    """Salva o PDF assinado no destino configurado da aba Assinaturas.
    Retorna dict com ok, destino e detalhes/erro.
    """
    destino=(getattr(envelope,'destino_salvar_tipo',None) or 'envelope').strip().lower()
    if destino!='funcionario':
        return {'ok':True,'destino':'envelope'}

    fid=getattr(envelope,'destino_funcionario_id',None)
    if not fid and (envelope.tipo or '').strip().lower()=='funcionario':
        fid=envelope.ref_id
    if not fid:
        return {'ok':False,'destino':'funcionario','erro':'Destino funcionário selecionado, mas nenhum funcionário foi definido.'}

    func=Funcionario.query.get(fid)
    if not func:
        return {'ok':False,'destino':'funcionario','erro':'Funcionário de destino não encontrado.'}

    cat=norm_cat(getattr(envelope,'destino_categoria',None) or 'outros')
    comp=(getattr(envelope,'destino_competencia',None) or '').strip()
    ano=infer_doc_year(comp)
    prepare_func_doc_dirs(func.id,ano)
    subdir,_=func_doc_subdir(func.id,cat,comp)

    base_nome=os.path.splitext(os.path.basename(fname or envelope.nome_documento_assinado or 'documento_assinado.pdf'))[0] or 'documento_assinado'
    rel,abs_dest,nome_final=unique_rel_filename(subdir,f'{base_nome}.pdf')
    os.makedirs(os.path.dirname(abs_dest),exist_ok=True)
    with open(abs_pdf_path,'rb') as src:
        pdf_bytes=src.read()
    with open(abs_dest,'wb') as out:
        out.write(pdf_bytes)

    reg=FuncionarioArquivo(
        funcionario_id=func.id,
        categoria=cat,
        competencia=comp,
        nome_arquivo=nome_final,
        caminho=rel,
        ass_status='assinado',
        ass_codigo=envelope.codigo,
        ass_nome='Assinatura por envelope',
        ass_cargo='Fluxo digital',
        ass_em=localnow(),
    )
    db.session.add(reg)
    return {
        'ok':True,
        'destino':'funcionario',
        'funcionario_id':func.id,
        'funcionario_nome':func.nome,
        'arquivo_id':reg.id,
        'categoria':cat,
        'competencia':comp,
        'caminho':rel,
    }


# ══════════════════════════════════════════════════════
#  ASSINATURA DE ENVELOPES — rotas
# ══════════════════════════════════════════════════════

@app.route('/api/envelopes',methods=['GET'])
@lr
def api_envelopes_listar():
    tipo=request.args.get('tipo','')
    q=AssinaturaEnvelope.query
    if tipo:
        q=q.filter_by(tipo=tipo)
    envs=q.order_by(AssinaturaEnvelope.criado_em.desc()).all()
    result=[]
    for env in envs:
        d=env.to_dict()
        d['total_arquivos']=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=env.id).count()
        d['total_signatarios']=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=env.id).count()
        d['total_assinados']=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=env.id,status='assinado').count()
        result.append(d)
    return jsonify(result)


@app.route('/api/envelopes',methods=['POST'])
@lr
def api_envelopes_criar():
    data=request.get_json() or {}
    titulo=(data.get('titulo') or '').strip()
    if not titulo:
        return jsonify({'erro':'Título obrigatório'}),400
    env=AssinaturaEnvelope(
        titulo=titulo,
        descricao=(data.get('descricao') or '').strip() or None,
        tipo=data.get('tipo') or 'avulso',
        empresa_id=(int(data.get('empresa_id')) if str(data.get('empresa_id') or '').isdigit() else None),
        ref_id=data.get('ref_id') or None,
        status='rascunho',
        codigo=secrets.token_urlsafe(12),
        nome_documento_assinado=_normalize_signed_pdf_name(data.get('nome_documento_assinado')) or None,
        destino_salvar_tipo=((data.get('destino_salvar_tipo') or 'envelope').strip().lower() if (data.get('destino_salvar_tipo') or '').strip().lower() in ('envelope','funcionario') else 'envelope'),
        destino_funcionario_id=(int(data.get('destino_funcionario_id')) if str(data.get('destino_funcionario_id') or '').isdigit() else None),
        destino_categoria=norm_cat(data.get('destino_categoria') or 'outros'),
        destino_competencia=(data.get('destino_competencia') or '').strip() or None,
        criado_por=session.get('usuario',''),
        expira_em=datetime.fromisoformat(data['expira_em']) if data.get('expira_em') else None,
    )
    db.session.add(env); db.session.commit()
    return jsonify(env.to_dict()),201


@app.route('/api/envelopes/<int:id>',methods=['GET'])
@lr
def api_envelope_detalhe(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    d=env.to_dict()
    emp=Empresa.query.get(env.empresa_id) if env.empresa_id else None
    d['empresa_nome']=(emp.nome if emp else '')
    d['empresa_razao']=(emp.razao if emp else '')
    d['arquivos']=[a.to_dict() for a in AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=id).all()]
    d['signatarios']=[s.to_dict() for s in AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=id).order_by(AssinaturaEnvelopeSignatario.ordem,AssinaturaEnvelopeSignatario.id).all()]
    return jsonify(d)


@app.route('/api/envelopes/<int:id>',methods=['PUT'])
@lr
def api_envelope_atualizar(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    data=request.get_json() or {}
    if 'titulo' in data: env.titulo=(data['titulo'] or '').strip() or env.titulo
    if 'descricao' in data: env.descricao=(data['descricao'] or '').strip() or None
    if 'expira_em' in data: env.expira_em=datetime.fromisoformat(data['expira_em']) if data['expira_em'] else None
    if 'empresa_id' in data:
        env.empresa_id=(int(data.get('empresa_id')) if str(data.get('empresa_id') or '').isdigit() else None)
    if 'nome_documento_assinado' in data:
        env.nome_documento_assinado=_normalize_signed_pdf_name(data.get('nome_documento_assinado')) or None
    if 'destino_salvar_tipo' in data:
        tp=(data.get('destino_salvar_tipo') or 'envelope').strip().lower()
        env.destino_salvar_tipo=tp if tp in ('envelope','funcionario') else 'envelope'
    if 'destino_funcionario_id' in data:
        env.destino_funcionario_id=(int(data.get('destino_funcionario_id')) if str(data.get('destino_funcionario_id') or '').isdigit() else None)
    if 'destino_categoria' in data:
        env.destino_categoria=norm_cat(data.get('destino_categoria') or 'outros')
    if 'destino_competencia' in data:
        env.destino_competencia=(data.get('destino_competencia') or '').strip() or None
    db.session.commit()
    return jsonify(env.to_dict())


@app.route('/api/envelopes/<int:id>',methods=['DELETE'])
@lr
def api_envelope_deletar(id):
    try:
        env=AssinaturaEnvelope.query.get_or_404(id)
        # Permite excluir em qualquer status (inclusive concluído/assinado).
        # Remove também arquivos físicos para evitar órfãos em disco.
        arqs=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=id).all()
        for arq in arqs:
            raw=(arq.caminho or '').strip()
            if not raw:
                continue
            cands=[raw]
            if raw and not os.path.isabs(raw):
                cands.append(os.path.join(UPLOAD_ROOT,raw))
                cands.append(os.path.join(_get_uploads_base(),raw))
            for p in cands:
                try:
                    if p and os.path.isfile(p):
                        os.remove(p)
                        break
                except Exception:
                    pass
        # Remove diretório do envelope (uploads + assinado)
        try:
            env_dir=os.path.join(_get_uploads_base(),'envelopes',str(id))
            if os.path.isdir(env_dir):
                shutil.rmtree(env_dir,ignore_errors=True)
        except Exception:
            pass
        AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=id).delete()
        AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=id).delete()
        db.session.delete(env)
        db.session.commit()
        return jsonify({'ok':True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro':f'Falha ao excluir envelope: {str(e)}'}),500


@app.route('/api/envelopes/<int:id>/cancelar',methods=['POST'])
@lr
def api_envelope_cancelar(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    if env.status=='concluido':
        return jsonify({'erro':'Não é possível cancelar um envelope já concluído.'}),400
    env.status='cancelado'
    db.session.commit()
    return jsonify({'ok':True,'status':'cancelado'})


@app.route('/api/envelopes/<int:id>/reativar',methods=['POST'])
@lr
def api_envelope_reativar(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    if env.status!='cancelado':
        return jsonify({'erro':'Só é possível reativar envelopes cancelados.'}),400
    # Verifica se ainda há signatários pendentes
    pendentes=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=id,status='pendente').count()
    env.status='pendente' if pendentes else 'rascunho'
    db.session.commit()
    return jsonify({'ok':True,'status':env.status})


@app.route('/api/envelopes/<int:id>/stamp',methods=['PUT'])
@lr
def api_envelope_stamp(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    data=request.get_json() or {}
    env.stamp_habilitado=bool(data.get('stamp_habilitado',False))
    env.stamp_pagina=max(1,int(data.get('stamp_pagina',1) or 1))
    env.stamp_x_pct=max(0.0,min(100.0,float(data.get('stamp_x_pct',60.0) or 60.0)))
    env.stamp_y_pct=max(0.0,min(100.0,float(data.get('stamp_y_pct',10.0) or 10.0)))
    env.stamp_todas_paginas=bool(data.get('stamp_todas_paginas',False))
    env.stamp_todos_arquivos=bool(data.get('stamp_todos_arquivos',False))
    db.session.commit()
    return jsonify({'ok':True,'stamp_habilitado':env.stamp_habilitado,'stamp_pagina':env.stamp_pagina,'stamp_x_pct':env.stamp_x_pct,'stamp_y_pct':env.stamp_y_pct,'stamp_todas_paginas':env.stamp_todas_paginas,'stamp_todos_arquivos':env.stamp_todos_arquivos})


@app.route('/api/envelopes/<int:id>/arquivos',methods=['POST'])
@lr
def api_envelope_add_arquivo(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    # Se for arquivo do sistema (func_arquivo_id)
    data_json=request.get_json(silent=True) or {}
    func_arquivo_id=request.form.get('func_arquivo_id') or data_json.get('func_arquivo_id')
    if func_arquivo_id:
        fa=FuncionarioArquivo.query.get(int(func_arquivo_id))
        if not fa:
            return jsonify({'erro':'Arquivo não encontrado'}),404
        arq=AssinaturaEnvelopeArquivo(
            envelope_id=id, origem='sistema',
            func_arquivo_id=fa.id,
            nome_arquivo=fa.nome_arquivo,
            caminho=fa.caminho,
        )
        db.session.add(arq)
        if not (env.nome_documento_assinado or '').strip():
            base_nome=os.path.splitext(os.path.basename(arq.nome_arquivo or 'documento'))[0] or 'documento'
            env.nome_documento_assinado=_normalize_signed_pdf_name(f'{base_nome} ASSINADO.pdf') or _default_signed_pdf_name(env)
        db.session.commit()
        return jsonify(arq.to_dict()),201
    # Upload
    f=request.files.get('arquivo')
    if not f:
        return jsonify({'erro':'Arquivo não enviado'}),400
    base=os.path.join(_get_uploads_base(),'envelopes',str(id))
    os.makedirs(base,exist_ok=True)
    fname=f'{secrets.token_urlsafe(8)}_{secure_filename(f.filename)}'
    path=os.path.join(base,fname)
    f.save(path)
    arq=AssinaturaEnvelopeArquivo(
        envelope_id=id, origem='upload',
        nome_arquivo=f.filename,
        caminho=path,
    )
    db.session.add(arq)
    if not (env.nome_documento_assinado or '').strip():
        base_nome=os.path.splitext(os.path.basename(arq.nome_arquivo or 'documento'))[0] or 'documento'
        env.nome_documento_assinado=_normalize_signed_pdf_name(f'{base_nome} ASSINADO.pdf') or _default_signed_pdf_name(env)
    db.session.commit()
    return jsonify(arq.to_dict()),201


@app.route('/api/envelopes/<int:id>/arquivos/<int:arq_id>',methods=['DELETE'])
@lr
def api_envelope_del_arquivo(id,arq_id):
    arq=AssinaturaEnvelopeArquivo.query.filter_by(id=arq_id,envelope_id=id).first_or_404()
    if arq.origem=='upload' and arq.caminho and os.path.exists(arq.caminho):
        try: os.remove(arq.caminho)
        except Exception: pass
    db.session.delete(arq); db.session.commit()
    return jsonify({'ok':True})


@app.route('/api/envelopes/<int:id>/arquivos/<int:arq_id>/visualizar',methods=['GET'])
@lr
def api_envelope_visualizar_arquivo_admin(id,arq_id):
    AssinaturaEnvelope.query.get_or_404(id)
    arq=AssinaturaEnvelopeArquivo.query.filter_by(id=arq_id,envelope_id=id).first_or_404()
    raw=(arq.caminho or '').strip()
    cands=[raw]
    if raw and not os.path.isabs(raw):
        cands.append(os.path.join(UPLOAD_ROOT,raw))
        cands.append(os.path.join(_get_uploads_base(),raw))
    abs_path=''
    for p in cands:
        if p and os.path.exists(p):
            abs_path=p
            break
    if not abs_path:
        return jsonify({'erro':'Arquivo não encontrado'}),404
    return send_file(abs_path,as_attachment=False,download_name=arq.nome_arquivo or os.path.basename(abs_path))


@app.route('/api/envelopes/<int:id>/signatarios',methods=['POST'])
@lr
def api_envelope_add_signatario(id):
    AssinaturaEnvelope.query.get_or_404(id)
    data=request.get_json() or {}
    nome=(data.get('nome') or '').strip()
    if not nome:
        return jsonify({'erro':'Nome obrigatório'}),400
    sig=AssinaturaEnvelopeSignatario(
        envelope_id=id,
        nome=nome,
        email=(data.get('email') or '').strip() or None,
        telefone=(data.get('telefone') or '').strip() or None,
        cpf=(data.get('cpf') or '').strip() or None,
        cargo=(data.get('cargo') or '').strip() or None,
        tipo=data.get('tipo') or 'externo',
        ref_id=data.get('ref_id') or None,
        ordem=int(data.get('ordem') or 0),
        status='pendente',
    )
    db.session.add(sig); db.session.commit()
    return jsonify(sig.to_dict()),201


@app.route('/api/envelopes/<int:id>/signatarios/<int:sig_id>',methods=['DELETE'])
@lr
def api_envelope_del_signatario(id,sig_id):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(id=sig_id,envelope_id=id).first_or_404()
    db.session.delete(sig); db.session.commit()
    return jsonify({'ok':True})


@app.route('/api/envelopes/<int:id>/enviar',methods=['POST'])
@lr
def api_envelope_enviar(id):
    env=AssinaturaEnvelope.query.get_or_404(id)
    data=request.get_json(silent=True) or {}
    canal=(data.get('canal') or 'whatsapp').strip().lower()
    if canal not in ('whatsapp','email','link'):
        return jsonify({'erro':'Canal inválido. Use whatsapp, email ou link.'}),400
    signatarios=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=id).all()
    if not signatarios:
        return jsonify({'erro':'Adicione ao menos um signatário antes de enviar'}),400
    arquivos=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=id).all()
    if not arquivos:
        return jsonify({'erro':'Adicione ao menos um arquivo antes de enviar'}),400
    if not env.codigo:
        env.codigo=secrets.token_urlsafe(12)
    emp=Empresa.query.get(env.empresa_id) if env.empresa_id else None
    empresa_nome=(emp.razao or emp.nome) if emp else 'RM Facilities'
    url_root=request.url_root.rstrip('/')
    enviados=[]
    falhas=[]
    for sig in signatarios:
        if sig.status=='assinado':
            continue
        if not sig.token:
            sig.token=secrets.token_urlsafe(24)
        src_q=_ass_track_channel(canal)
        link=f"{url_root}/envelope/assinar/{sig.token}?src={src_q}"
        try:
            sc=_short_link_criar(link)
            link_curto=(f"{url_root}/s/{sc}" if sc else link)
        except Exception:
            link_curto=link
        if canal=='link':
            _ass_track_mark_sent(sig,'link')
            enviados.append({'id':sig.id,'nome':sig.nome,'canal':'link','link':link,'link_curto':link_curto})
            continue
        if canal=='whatsapp':
            if not (sig.telefone or '').strip():
                falhas.append({'id':sig.id,'nome':sig.nome,'canal':'whatsapp','erro':'Signatário sem telefone cadastrado.'})
                continue
            tel=sig.telefone.strip().replace(' ','').replace('-','').replace('(','').replace(')','')
            msg=(f"Olá {sig.nome}, você recebeu um documento para assinar eletronicamente.\n"
                  f"🏢 Empresa remetente: *{empresa_nome}*\n"
                 f"📄 Documento: *{env.titulo}*\n"
                 f"🔗 Acesse e assine aqui:\n{link_curto}")
            try:
                wa_send_text(tel,msg)
                _ass_track_mark_sent(sig,'whatsapp')
                enviados.append({'id':sig.id,'nome':sig.nome,'canal':'whatsapp','destino':tel,'link':link,'link_curto':link_curto})
            except Exception as e:
                falhas.append({'id':sig.id,'nome':sig.nome,'canal':'whatsapp','erro':str(e)})
            continue
        if canal=='email':
            if not (sig.email or '').strip():
                falhas.append({'id':sig.id,'nome':sig.nome,'canal':'email','erro':'Signatário sem e-mail cadastrado.'})
                continue
            try:
                smtp_send_link_assinatura(sig.email.strip(),sig.nome or 'Signatário',env.titulo or 'Documento',link_curto)
                _ass_track_mark_sent(sig,'email')
                enviados.append({'id':sig.id,'nome':sig.nome,'canal':'email','destino':sig.email.strip(),'link':link,'link_curto':link_curto})
            except Exception as e:
                falhas.append({'id':sig.id,'nome':sig.nome,'canal':'email','erro':str(e)})
    env.status='pendente'
    db.session.commit()
    return jsonify({'ok':True,'canal':canal,'enviados':enviados,'falhas':falhas})
# Rota para copiar/obter link
@app.route('/api/envelopes/<int:id>/signatarios/<int:sig_id>/link',methods=['GET'])
@lr
def api_envelope_sig_link(id,sig_id):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(id=sig_id,envelope_id=id).first_or_404()
    if sig.status=='assinado':
        return jsonify({'erro':'Este signatário já assinou.'}),400
    if not sig.token:
        sig.token=secrets.token_urlsafe(24)
        db.session.commit()
    url_root=request.url_root.rstrip('/')
    link=f"{url_root}/envelope/assinar/{sig.token}?src=link"
    try:
        sc=_short_link_criar(link)
        link_curto=(f"{url_root}/s/{sc}" if sc else link)
    except Exception:
        link_curto=link
    return jsonify({'link':link,'link_curto':link_curto,'nome':sig.nome})

@app.route('/api/envelopes/<int:id>/signatarios/<int:sig_id>/rastreio')
@lr
def api_envelope_sig_rastreio(id,sig_id):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(id=sig_id,envelope_id=id).first_or_404()
    return jsonify({
        'ok':True,
        'signatario_id':sig.id,
        'status_assinatura':sig.status or '',
        'canal_envio':sig.ass_canal_envio or '',
        'enviado_em':(sig.ass_enviado_em.isoformat() if sig.ass_enviado_em else ''),
        'recebido_em':(sig.ass_recebido_em.isoformat() if sig.ass_recebido_em else ''),
        'aberto_em':(sig.ass_aberto_em.isoformat() if sig.ass_aberto_em else ''),
        'assinado_em':(sig.ass_em.isoformat() if sig.ass_em else ''),
        'whatsapp':{
            'status':sig.ass_wa_status or 'nao_enviado',
            'enviado_em':(sig.ass_wa_enviado_em.isoformat() if sig.ass_wa_enviado_em else ''),
            'recebido_em':(sig.ass_wa_recebido_em.isoformat() if sig.ass_wa_recebido_em else ''),
        },
        'email':{
            'status':sig.ass_email_status or 'nao_enviado',
            'enviado_em':(sig.ass_email_enviado_em.isoformat() if sig.ass_email_enviado_em else ''),
            'recebido_em':(sig.ass_email_recebido_em.isoformat() if sig.ass_email_recebido_em else ''),
        }
    })


@app.route('/envelope/assinar/<token>/arquivo/<int:arq_id>')
def envelope_assinar_visualizar_arquivo(token,arq_id):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(token=token).first_or_404()
    arq=AssinaturaEnvelopeArquivo.query.filter_by(id=arq_id,envelope_id=sig.envelope_id).first_or_404()
    raw=(arq.caminho or '').strip()
    cands=[raw]
    if raw and not os.path.isabs(raw):
        cands.append(os.path.join(UPLOAD_ROOT,raw))
        cands.append(os.path.join(_get_uploads_base(),raw))
    abs_path=''
    for p in cands:
        if p and os.path.exists(p):
            abs_path=p
            break
    if not abs_path:
        return 'Arquivo não encontrado.',404
    src=request.args.get('src','')
    if _ass_track_mark_opened(sig,src):
        db.session.commit()
    return send_file(abs_path,as_attachment=False,download_name=arq.nome_arquivo or os.path.basename(abs_path))


# Página pública de assinatura do envelope
@app.route('/envelope/assinar/<token>')
def envelope_assinar_publica(token):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(token=token).first_or_404()
    env=AssinaturaEnvelope.query.get_or_404(sig.envelope_id)
    arquivos=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=env.id).all()
    empresa=Empresa.query.get(env.empresa_id) if env.empresa_id else Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem,Empresa.id).first()
    empresas=[empresa] if empresa else []
    src=request.args.get('src','')
    if _ass_track_mark_received(sig,src):
        db.session.commit()
    return render_template('envelope_assinar.html',sig=sig,env=env,arquivos=arquivos,empresas=empresas)

@app.route('/api/envelope/assinar/<token>/enviar-otp',methods=['GET','POST'])
def api_envelope_assinatura_enviar_otp(token):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(token=token).first_or_404()
    if sig.status=='assinado':
        return _assinatura_json_erro('Você já assinou este documento.',400)
    env=AssinaturaEnvelope.query.get_or_404(sig.envelope_id)
    if env.expira_em and datetime.utcnow() > env.expira_em:
        return _assinatura_json_erro('O prazo para assinatura deste documento expirou.',400)
    if not (wa_norm_number(sig.telefone or '') or (sig.email or '').strip()):
        return _assinatura_json_erro('Nenhum telefone ou e-mail cadastrado para envio do OTP.',400)
    codigo=_otp_new_code()
    sig.ass_otp_hash=token_hash(codigo)
    sig.ass_otp_expira_em=utcnow()+timedelta(minutes=10)
    sig.ass_otp_tentativas=0
    try:
        envio=_send_signature_otp(codigo,nome_dest=(sig.nome or ''),telefone=sig.telefone or '',email=sig.email or '',contexto='envelope')
        db.session.commit()
        return _assinatura_json_ok(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )
    except Exception as ex:
        db.session.rollback()
        return _assinatura_json_erro(f'Falha ao enviar OTP: {str(ex)}',500)


# API pública: confirmar assinatura
@app.route('/api/envelope/assinar/<token>/confirmar',methods=['POST'])
def api_envelope_assinatura_confirmar(token):
    sig=AssinaturaEnvelopeSignatario.query.filter_by(token=token).first_or_404()
    if sig.status=='assinado':
        return _assinatura_json_erro('Você já assinou este documento.',400)
    env=AssinaturaEnvelope.query.get_or_404(sig.envelope_id)
    if env.expira_em and datetime.utcnow() > env.expira_em:
        return _assinatura_json_erro('O prazo para assinatura deste documento expirou.',400)
    if not env.codigo:
        env.codigo=secrets.token_urlsafe(12)
    data=request.get_json() or {}
    nome=(data.get('nome') or '').strip()
    cargo=(data.get('cargo') or '').strip()
    cpf_inf=(data.get('cpf') or '').strip().replace('.','').replace('-','').replace(' ','')
    otp=(only_digits(data.get('otp') or '') or '').strip()
    aceite=data.get('aceite')
    if not nome or not cpf_inf or not aceite:
        return _assinatura_json_erro('Preencha nome, CPF e confirme o aceite.',400)
    if len(cpf_inf)<11 or not _valida_cpf(cpf_inf):
        return _assinatura_json_erro('CPF inválido. Verifique os dígitos e tente novamente.',400)
    cpf_base=only_digits(sig.cpf or '')
    if cpf_base and cpf_inf!=cpf_base:
        return _assinatura_json_erro('O CPF informado não confere com o CPF cadastrado para este signatário.',400)

    if not otp:
        codigo=_otp_new_code()
        sig.ass_otp_hash=token_hash(codigo)
        sig.ass_otp_expira_em=utcnow()+timedelta(minutes=10)
        sig.ass_otp_tentativas=0
        try:
            envio=_send_signature_otp(codigo,nome_dest=nome,telefone=sig.telefone or '',email=sig.email or '',contexto='envelope')
        except Exception as ex:
            db.session.rollback()
            return _assinatura_json_erro(f'Falha ao enviar OTP de confirmação: {str(ex)}',400)
        db.session.commit()
        return _assinatura_json_otp(
            mensagem=f"Código OTP enviado via {envio.get('canal','canal')} para {envio.get('destino','destino mascarado')}",
            canal=envio.get('canal',''),
            destino=envio.get('destino','')
        )

    if not (sig.ass_otp_hash or '').strip() or not sig.ass_otp_expira_em:
        return _assinatura_json_erro('Solicite um novo código OTP para concluir a assinatura.',400)
    if sig.ass_otp_expira_em<utcnow():
        return _assinatura_json_erro('Código OTP expirado. Solicite um novo código.',400)
    tent=int(sig.ass_otp_tentativas or 0)
    if tent>=5:
        return _assinatura_json_erro('Limite de tentativas de OTP excedido. Solicite um novo código.',400)
    if not hmac.compare_digest(token_hash(otp),str(sig.ass_otp_hash or '')):
        sig.ass_otp_tentativas=tent+1
        db.session.commit()
        return _assinatura_json_erro('Código OTP inválido.',400)

    ip=request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    sig.nome=nome
    sig.cargo=cargo
    sig.ass_cpf_informado=cpf_inf
    sig.ass_ip=ip
    sig.ass_em=datetime.utcnow()
    sig.ass_codigo=secrets.token_urlsafe(10)
    sig.ass_otp_hash=None
    sig.ass_otp_expira_em=None
    sig.ass_otp_tentativas=0
    sig.status='assinado'
    assinatura_img=(data.get('assinatura_img') or '').strip()
    if assinatura_img and assinatura_img.startswith('data:image/'):
        sig.ass_assinatura_img=assinatura_img[:200000]  # limita a ~150 KB base64
    if not sig.ass_aberto_em:
        sig.ass_aberto_em=utcnow()
    sig.token=None  # invalida o token após uso
    url_root=request.url_root.rstrip('/')
    signed_pdf_link=''
    emp=Empresa.query.get(env.empresa_id) if env.empresa_id else None
    empresa_nome=(emp.razao or emp.nome) if emp else 'RM Facilities'
    # Verifica se todos assinaram
    todos=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=env.id).all()
    destino_info={'ok':True,'destino':(env.destino_salvar_tipo or 'envelope')}
    if all(s.status=='assinado' for s in todos):
        env.status='concluido'
        try:
            abs_pdf,fname=_gerar_pdf_assinado_envelope(env,url_root)
            rs_crypto=_try_sign_pdf_file_crypto(abs_pdf,empresa_id=env.empresa_id,usuario_id=session.get('uid'))
            env.assinatura_doc_hash=_sha256_file(abs_pdf)
            env.assinatura_crypto_ok=bool(rs_crypto.get('ok'))
            env.assinatura_cert_subject=(rs_crypto.get('cert_subject') or '')[:255] if rs_crypto.get('ok') else None
            signed_pdf_link=f"{url_root}/envelope/baixar/{env.codigo}"
            destino_info=_salvar_pdf_assinado_destino_envelope(env,abs_pdf,fname)
            for s in todos:
                if s.telefone:
                    tel=s.telefone.strip().replace(' ','').replace('-','').replace('(','').replace(')','')
                    caption=(f"✅ Assinatura concluída por *{empresa_nome}*! "
                             f"Segue o documento *{env.titulo}* com a página de auditoria.")
                    try:
                        wa_send_pdf(tel,abs_pdf,fname,caption)
                        wa_send_text(tel,(f"🏢 Empresa remetente: *{empresa_nome}*\n"
                                          f"🔗 Link para download do documento assinado:\n{signed_pdf_link}"))
                    except Exception:
                        pass
        except Exception:
            pass
    else:
        env.status='parcial'
    db.session.commit()
    validacao_link=f"{url_root}/envelope/validar/{env.codigo}"
    enviado_wa=False
    if sig.telefone:
        try:
            tel_sig=sig.telefone.strip().replace(' ','').replace('-','').replace('(','').replace(')','')
            msg_sig=(f"✅ Assinatura registrada com sucesso.\n"
                     f"🏢 Empresa remetente: *{empresa_nome}*\n"
                     f"🔎 Validação: {validacao_link}")
            if signed_pdf_link:
                msg_sig+=f"\n⬇ Download do documento assinado: {signed_pdf_link}"
            wa_send_text(tel_sig,msg_sig)
            enviado_wa=True
        except Exception:
            pass
    return _assinatura_json_ok(
        mensagem='Assinatura concluída com sucesso.',
        codigo=(sig.ass_codigo or ''),
        validacao_link=validacao_link,
        signed_pdf_link=signed_pdf_link,
        whatsapp_enviado=enviado_wa,
        destino_salvamento=destino_info
    )


@app.route('/envelope/baixar/<codigo>')
def envelope_baixar_assinado_publico(codigo):
    env=AssinaturaEnvelope.query.filter_by(codigo=codigo).first_or_404()
    if env.status!='concluido':
        return 'Documento ainda não concluído para download.',400
    url_root=request.url_root.rstrip('/')
    abs_pdf,fname=_envelope_signed_pdf_path(env)
    if not os.path.exists(abs_pdf):
        try:
            abs_pdf,fname=_gerar_pdf_assinado_envelope(env,url_root)
        except Exception:
            app.logger.exception('Falha ao gerar PDF assinado do envelope %s',codigo)
            return 'Não foi possível gerar o PDF assinado neste momento.',500
    return send_file(abs_pdf,mimetype='application/pdf',as_attachment=False,download_name=fname)


# Página pública de validação do envelope
@app.route('/envelope/validar/<codigo>')
def envelope_validar_publica(codigo):
    env=AssinaturaEnvelope.query.filter_by(codigo=codigo).first_or_404()
    signatarios=AssinaturaEnvelopeSignatario.query.filter_by(envelope_id=env.id).order_by(
        AssinaturaEnvelopeSignatario.ordem,AssinaturaEnvelopeSignatario.id).all()
    arquivos=AssinaturaEnvelopeArquivo.query.filter_by(envelope_id=env.id).all()
    empresa=Empresa.query.get(env.empresa_id) if env.empresa_id else Empresa.query.filter_by(ativa=True).order_by(Empresa.ordem,Empresa.id).first()
    empresas=[empresa] if empresa else []
    return render_template('envelope_validar.html',env=env,signatarios=signatarios,arquivos=arquivos,empresas=empresas)


# Listar arquivos de um funcionário (para picker no envelope)
@app.route('/api/funcionarios/<int:fid>/arquivos-lista')
@lr
def api_funcionario_arquivos_lista(fid):
    arqs=FuncionarioArquivo.query.filter_by(funcionario_id=fid).order_by(FuncionarioArquivo.criado_em.desc()).all()
    return jsonify([{'id':a.id,'nome':a.nome_arquivo,'categoria':a.categoria,'competencia':a.competencia} for a in arqs])


@app.route('/api/rh/extrair-competencia',methods=['POST'])
@lr
def api_rh_extrair_competencia():
    """Extrai a competência (MM/YYYY) de um PDF enviado.
    Retorna a primeira competência encontrada ou o mês corrente como fallback."""
    from pypdf import PdfReader
    import io
    fs=request.files.get('arquivo')
    if not fs:
        return jsonify({'erro':'Arquivo não enviado'}),400
    if not _upload_is_pdf(fs):
        return jsonify({'erro':'Arquivo inválido. Envie um PDF válido.'}),400
    try:
        reader=PdfReader(io.BytesIO(fs.read()))
        texto=_extract_pdf_competencia_text(reader,max_pages=30)
    except Exception:
        texto=''
    comp,origem=_resolver_competencia_envio(comp_in='',texto=texto,nome_arquivo=(fs.filename or ''))
    return jsonify({'competencia':comp,'competencia_origem':origem,'encontrada_no_documento':origem in ('texto_pdf','nome_arquivo')})

@app.route('/api/funcionarios/holerites/upload',methods=['POST'])
@lr
def api_holerites_upload():
    fs=request.files.get('arquivo')
    comp_in=(request.form.get('competencia') or '').strip()
    if not fs: return jsonify({'erro':'PDF nao enviado'}),400
    if not _upload_is_pdf(fs): return jsonify({'erro':'Arquivo invalido. Envie um PDF valido.'}),400
    canal_ass=(request.form.get('canal_assinatura') or 'nao').strip().lower()
    if canal_ass not in ('nao','whatsapp','link','app'):
        canal_ass='nao'
    ids_ass_raw=(request.form.get('assinatura_funcionario_ids') or '').strip()
    ids_ass_sel=set()
    if ids_ass_raw:
        for p in ids_ass_raw.split(','):
            p=(p or '').strip()
            if p.isdigit():
                ids_ass_sel.add(int(p))
    try:
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500
    funcs=Funcionario.query.all()
    if not funcs: return jsonify({'erro':'Cadastre funcionarios antes do upload'}),400
    reader=PdfReader(fs)
    texto_amostra=_extract_pdf_competencia_text(reader,max_pages=30)
    comp,comp_origem=_resolver_competencia_envio(comp_in=comp_in,texto=texto_amostra,nome_arquivo=(fs.filename or ''))
    enviados=0; sem_match=[]; assinaturas_auto=0; sem_tel=[]; erro_ass=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=_extract_pdf_page_text(page)
        alvo=find_funcionario_in_text(txt,funcs)
        if not alvo:
            sem_match.append(idx)
            continue
        ano=infer_doc_year(comp)
        prepare_func_doc_dirs(alvo.id,ano)
        writer=PdfWriter(); writer.add_page(page)
        fake_name=holerite_batch_filename(alvo,comp)
        subdir,_=func_doc_subdir(alvo.id,'holerite',comp)
        rel,abs_p,fake_name=unique_rel_filename(subdir,fake_name)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out: writer.write(out)
        a=FuncionarioArquivo(funcionario_id=alvo.id,categoria='holerite',competencia=comp,nome_arquivo=fake_name,caminho=rel)
        db.session.add(a); db.session.commit(); enviados+=1
        solicitar_ass=(canal_ass!='nao') and (not ids_ass_sel or alvo.id in ids_ass_sel)
        if solicitar_ass and canal_ass=='app':
            a.ass_status='pendente'; db.session.commit(); assinaturas_auto+=1
            _push_notify_funcionario(alvo.id,'Documento para assinar',f'{fake_name} aguarda sua assinatura no app.',{'tipo':'documento_assinar','arquivo_id':str(a.id)})
        elif solicitar_ass and canal_ass=='whatsapp':
            tel=wa_norm_number(alvo.telefone or '')
            if wa_is_valid_number(tel):
                rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='whatsapp',commit_now=True)
                if rs.get('ok'):
                    assinaturas_auto+=1
                else:
                    erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao solicitar assinatura.'})
            else:
                sem_tel.append({'funcionario_id':alvo.id,'nome':alvo.nome})
        elif solicitar_ass and canal_ass=='link':
            rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='link',commit_now=True)
            if rs.get('ok'):
                assinaturas_auto+=1
            else:
                erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao gerar link de assinatura.'})
    return jsonify({'ok':True,'arquivos_gerados':enviados,'paginas_sem_funcionario':sem_match,'assinaturas_auto':assinaturas_auto,'sem_telefone':sem_tel,'falhas_assinatura':erro_ass,'canal_assinatura':canal_ass,'competencia':comp,'competencia_origem':comp_origem})

@app.route('/api/funcionarios/folhas-ponto/upload',methods=['POST'])
@lr
def api_folhas_ponto_upload():
    fs=request.files.get('arquivo')
    comp_in=(request.form.get('competencia') or '').strip()
    if not fs: return jsonify({'erro':'PDF nao enviado'}),400
    if not _upload_is_pdf(fs): return jsonify({'erro':'Arquivo invalido. Envie um PDF valido.'}),400
    canal_ass=(request.form.get('canal_assinatura') or 'nao').strip().lower()
    if canal_ass not in ('nao','whatsapp','link','app'):
        canal_ass='nao'
    ids_ass_raw=(request.form.get('assinatura_funcionario_ids') or '').strip()
    ids_ass_sel=set()
    if ids_ass_raw:
        for p in ids_ass_raw.split(','):
            p=(p or '').strip()
            if p.isdigit():
                ids_ass_sel.add(int(p))
    try:
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500
    funcs=Funcionario.query.all()
    if not funcs: return jsonify({'erro':'Cadastre funcionarios antes do upload'}),400
    reader=PdfReader(fs)
    texto_amostra=_extract_pdf_competencia_text(reader,max_pages=30)
    comp,comp_origem=_resolver_competencia_envio(comp_in=comp_in,texto=texto_amostra,nome_arquivo=(fs.filename or ''))
    enviados=0; sem_match=[]; assinaturas_auto=0; sem_tel=[]; erro_ass=[]; duplicadas=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=_extract_pdf_page_text(page)
        alvo=find_funcionario_in_text(txt,funcs)
        if not alvo:
            sem_match.append(idx)
            continue
        ano=infer_doc_year(comp)
        prepare_func_doc_dirs(alvo.id,ano)
        # verifica se ja existe arquivo desta competencia para nao duplicar
        existente=FuncionarioArquivo.query.filter_by(funcionario_id=alvo.id,categoria='folha_ponto',competencia=comp).first()
        if existente:
            duplicadas.append(f'{alvo.nome} ({comp})')
            continue
        writer=PdfWriter(); writer.add_page(page)
        fake_name=f'folha_ponto_{comp}_{alvo.id}.pdf'.replace('/','_').replace(' ','_')
        subdir,_=func_doc_subdir(alvo.id,'folha_ponto',comp)
        rel,abs_p,fake_name=unique_rel_filename(subdir,fake_name)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out: writer.write(out)
        a=FuncionarioArquivo(funcionario_id=alvo.id,categoria='folha_ponto',competencia=comp,nome_arquivo=fake_name,caminho=rel)
        db.session.add(a); db.session.commit(); enviados+=1
        solicitar_ass=(canal_ass!='nao') and (not ids_ass_sel or alvo.id in ids_ass_sel)
        if solicitar_ass and canal_ass=='app':
            a.ass_status='pendente'; db.session.commit(); assinaturas_auto+=1
            _push_notify_funcionario(alvo.id,'Documento para assinar',f'{fake_name} aguarda sua assinatura no app.',{'tipo':'documento_assinar','arquivo_id':str(a.id)})
        elif solicitar_ass and canal_ass=='whatsapp':
            tel=wa_norm_number(alvo.telefone or '')
            if wa_is_valid_number(tel):
                rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='whatsapp',commit_now=True)
                if rs.get('ok'):
                    assinaturas_auto+=1
                else:
                    erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao solicitar assinatura.'})
            else:
                sem_tel.append({'funcionario_id':alvo.id,'nome':alvo.nome})
        elif solicitar_ass and canal_ass=='link':
            rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='link',commit_now=True)
            if rs.get('ok'):
                assinaturas_auto+=1
            else:
                erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao gerar link de assinatura.'})
    return jsonify({'ok':True,'arquivos_gerados':enviados,'paginas_sem_funcionario':sem_match,'assinaturas_auto':assinaturas_auto,'sem_telefone':sem_tel,'falhas_assinatura':erro_ass,'canal_assinatura':canal_ass,'duplicadas':duplicadas,'competencia':comp,'competencia_origem':comp_origem})


@app.route('/api/funcionarios/documentos-rh/upload',methods=['POST'])
@lr
def api_documentos_rh_upload():
    fs=request.files.get('arquivo')
    comp_in=(request.form.get('competencia') or '').strip()
    cat_in=(request.form.get('categoria') or 'outros').strip().lower()
    if not fs:
        return jsonify({'erro':'PDF nao enviado'}),400
    if not _upload_is_pdf(fs):
        return jsonify({'erro':'Arquivo invalido. Envie um PDF valido.'}),400

    canal_ass=(request.form.get('canal_assinatura') or 'nao').strip().lower()
    if canal_ass not in ('nao','whatsapp','link','app'):
        canal_ass='nao'

    ids_ass_raw=(request.form.get('assinatura_funcionario_ids') or '').strip()
    ids_ass_sel=set()
    if ids_ass_raw:
        for p in ids_ass_raw.split(','):
            p=(p or '').strip()
            if p.isdigit():
                ids_ass_sel.add(int(p))

    categoria=norm_cat(cat_in)
    if categoria in ('holerite','folha_ponto'):
        categoria='outros'
    if categoria not in DOC_CAT_PATH:
        categoria='outros'

    try:
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500

    funcs=Funcionario.query.all()
    if not funcs:
        return jsonify({'erro':'Cadastre funcionarios antes do upload'}),400

    reader=PdfReader(fs)
    texto_amostra=_extract_pdf_competencia_text(reader,max_pages=30)
    comp,comp_origem=_resolver_competencia_envio(comp_in=comp_in,texto=texto_amostra,nome_arquivo=(fs.filename or ''))
    enviados=0
    sem_match=[]
    assinaturas_auto=0
    sem_tel=[]
    erro_ass=[]

    for idx,page in enumerate(reader.pages,start=1):
        txt=_extract_pdf_page_text(page)
        alvo=find_funcionario_in_text(txt,funcs)
        if not alvo:
            sem_match.append(idx)
            continue

        ano=infer_doc_year(comp)
        prepare_func_doc_dirs(alvo.id,ano)
        writer=PdfWriter()
        writer.add_page(page)

        doc_tipo={
            'aso':'ASO',
            'epi':'Ficha de EPI',
            'treinamento':'Treinamento',
            'contrato_trabalho':'Contrato de Trabalho',
            'vale_transporte':'Vale Transporte',
            'requisicao_vale_transporte':'Requisicao de Vale Transporte',
            'uniforme':'Uniforme/Fardamento',
            'atestado':'Atestado',
            'recibo_ferias':'Recibo de Ferias',
            'outros':'Documento RH'
        }.get(categoria,'Documento RH')
        nome_base=f'{doc_tipo} - {(alvo.nome or "Colaborador")} - {holerite_comp_label(comp)}.pdf'
        nome_base=_clean_file_part(nome_base,120,'documento_rh')
        if not nome_base.lower().endswith('.pdf'):
            nome_base=nome_base+'.pdf'

        subdir,_=func_doc_subdir(alvo.id,categoria,comp)
        rel,abs_p,nome_final=unique_rel_filename(subdir,nome_base)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out:
            writer.write(out)

        a=FuncionarioArquivo(
            funcionario_id=alvo.id,
            categoria=categoria,
            competencia=comp,
            nome_arquivo=nome_final,
            caminho=rel
        )
        db.session.add(a)
        db.session.commit()
        enviados+=1
        solicitar_ass=(canal_ass!='nao') and (not ids_ass_sel or alvo.id in ids_ass_sel)
        if solicitar_ass and canal_ass=='app':
            a.ass_status='pendente'; db.session.commit(); assinaturas_auto+=1
            _push_notify_funcionario(alvo.id,'Documento para assinar',f'{nome_final} aguarda sua assinatura no app.',{'tipo':'documento_assinar','arquivo_id':str(a.id)})
        elif solicitar_ass and canal_ass=='whatsapp':
            tel=wa_norm_number(alvo.telefone or '')
            if wa_is_valid_number(tel):
                rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='whatsapp',commit_now=True)
                if rs.get('ok'):
                    assinaturas_auto+=1
                else:
                    erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao solicitar assinatura.'})
            else:
                sem_tel.append({'funcionario_id':alvo.id,'nome':alvo.nome})
        elif solicitar_ass and canal_ass=='link':
            rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='link',commit_now=True)
            if rs.get('ok'):
                assinaturas_auto+=1
            else:
                erro_ass.append({'funcionario_id':alvo.id,'nome':alvo.nome,'erro':rs.get('erro') or 'Falha ao gerar link de assinatura.'})

    return jsonify({
        'ok':True,
        'arquivos_gerados':enviados,
        'paginas_sem_funcionario':sem_match,
        'assinaturas_auto':assinaturas_auto,
        'sem_telefone':sem_tel,
        'falhas_assinatura':erro_ass,
        'canal_assinatura':canal_ass,
        'categoria':categoria,
        'competencia':comp,
        'competencia_origem':comp_origem
    })


@app.route('/api/rh/preview-destinatarios',methods=['POST'])
@lr
def api_rh_preview_destinatarios():
    fs=request.files.get('arquivo')
    comp_in=(request.form.get('competencia') or '').strip()
    funcionario_id=to_num(request.form.get('funcionario_id'))
    if not fs:
        return jsonify({'erro':'PDF nao enviado'}),400
    if not _upload_is_pdf(fs):
        return jsonify({'erro':'Arquivo invalido. Envie um PDF valido.'}),400
    try:
        from pypdf import PdfReader
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500

    if funcionario_id:
        f=Funcionario.query.get(funcionario_id)
        if not f:
            return jsonify({'erro':'Funcionario selecionado nao encontrado.'}),404
        tel=wa_norm_number(f.telefone or '')
        return jsonify({
            'ok':True,
            'total_paginas':0,
            'paginas_sem_funcionario':[],
            'destinatarios':[{
                'funcionario_id':f.id,
                'nome':f.nome or '',
                'matricula':f.matricula or '',
                'email':f.email or '',
                'telefone':(tel if wa_is_valid_number(tel) else ''),
                'match_confianca':'manual',
                'match_score':100,
                'match_detalhe':'Selecionado manualmente.',
                'paginas':[]
            }],
            'competencia':(comp_in or _competencia_mes_atual()),
            'competencia_origem':('manual' if comp_in else 'mes_atual'),
            'preview_tipo':'funcionario_especifico'
        })

    funcs=Funcionario.query.all()
    if not funcs:
        return jsonify({'erro':'Cadastre funcionarios antes do upload'}),400

    try:
        reader=PdfReader(fs)
    except Exception:
        return jsonify({'erro':'PDF invalido ou corrompido.'}),400

    texto_amostra=_extract_pdf_competencia_text(reader,max_pages=30)
    comp,comp_origem=_resolver_competencia_envio(comp_in=comp_in,texto=texto_amostra,nome_arquivo=(fs.filename or ''))

    dest_idx={}
    sem_match=[]
    debug_paginas=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=_extract_pdf_page_text(page)
        alvo,match_meta=find_funcionario_in_text(txt,funcs,return_meta=True)
        if not alvo:
            sem_match.append(idx)
            rank=_rank_funcionarios_in_text(txt,funcs,limit=5)
            txt_limpo=' '.join((txt or '').split())
            debug_paginas.append({
                'pagina':idx,
                'motivo':('texto_vazio' if not (txt or '').strip() else 'sem_match'),
                'snippet':txt_limpo[:280],
                'indicadores':rank.get('indicadores') or {},
                'top_candidatos':rank.get('top') or []
            })
            continue
        if alvo.id not in dest_idx:
            tel=wa_norm_number(alvo.telefone or '')
            dest_idx[alvo.id]={
                'funcionario_id':alvo.id,
                'nome':alvo.nome or '',
                'matricula':alvo.matricula or '',
                'email':alvo.email or '',
                'telefone':(tel if wa_is_valid_number(tel) else ''),
                'match_scores':[],
                'paginas':[]
            }
        dest_idx[alvo.id]['paginas'].append(idx)
        dest_idx[alvo.id]['match_scores'].append(int(match_meta.get('score') or 0))

    for item in dest_idx.values():
        scores=[int(s) for s in item.pop('match_scores',[]) if isinstance(s,(int,float)) or str(s).isdigit()]
        if not scores:
            item['match_confianca']='baixa'
            item['match_score']=0
            item['match_detalhe']='Sem score de correspondencia.'
            continue
        media=int(round(sum(scores)/max(1,len(scores))))
        cont_alta=sum(1 for s in scores if _match_conf_level(s)=='alta')
        cont_media=sum(1 for s in scores if _match_conf_level(s)=='media')
        cont_baixa=sum(1 for s in scores if _match_conf_level(s)=='baixa')
        if cont_baixa>0:
            conf='baixa'
        elif cont_media>0:
            conf='media'
        else:
            conf='alta'
        item['match_confianca']=conf
        item['match_score']=media
        item['match_detalhe']=f'Alta: {cont_alta}, Media: {cont_media}, Baixa: {cont_baixa}.'

    destinatarios=sorted(dest_idx.values(),key=lambda x:(x.get('nome') or '').lower())
    return jsonify({
        'ok':True,
        'total_paginas':len(reader.pages),
        'paginas_sem_funcionario':sem_match,
        'debug_paginas':debug_paginas,
        'destinatarios':destinatarios,
        'competencia':comp,
        'competencia_origem':comp_origem,
        'preview_tipo':'separacao_automatica'
    })

@app.route('/api/ordens-compra',methods=['GET'])
@lr
def api_ordens_compra():
    return jsonify([o.to_dict() for o in OrdemCompra.query.order_by(OrdemCompra.criado_em.desc()).all()])

@app.route('/api/ordens-compra',methods=['POST'])
@lr
def api_criar_ordem_compra():
    d=request.json or {}
    num=d.get('numero') or f"OC-{localnow().strftime('%Y%m%d%H%M%S')}"
    o=OrdemCompra(numero=num,empresa_id=d.get('empresa_id'),solicitante=d.get('solicitante',''),fornecedor=d.get('fornecedor',''),descricao=d.get('descricao',''),valor=to_num(d.get('valor'),dec=True),status=d.get('status','Aberta'),data_emissao=d.get('data_emissao',''),criado_por=session.get('nome',''))
    db.session.add(o); db.session.commit(); return jsonify(o.to_dict()),201

@app.route('/api/ordens-compra/<int:id>',methods=['PUT'])
@lr
def api_atualizar_ordem_compra(id):
    o=OrdemCompra.query.get_or_404(id); d=request.json or {}
    for k in ['numero','empresa_id','solicitante','fornecedor','descricao','status','data_emissao']:
        if k in d: setattr(o,k,d[k])
    if 'valor' in d: o.valor=to_num(d.get('valor'),dec=True)
    db.session.commit(); return jsonify(o.to_dict())

@app.route('/api/ordens-compra/<int:id>',methods=['DELETE'])
@lr
def api_deletar_ordem_compra(id):
    db.session.delete(OrdemCompra.query.get_or_404(id)); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/operacional/documentos',methods=['GET'])
@lr
def api_oper_docs():
    return jsonify([d.to_dict() for d in OperacionalDocumento.query.order_by(OperacionalDocumento.criado_em.desc()).all()])

@app.route('/api/operacional/documentos',methods=['POST'])
@lr
def api_criar_oper_doc():
    titulo=(request.form.get('titulo') or '').strip()
    if not titulo: return jsonify({'erro':'Titulo obrigatorio'}),400
    fs=request.files.get('arquivo')
    rel=None
    if fs and fs.filename:
        rel,_=save_upload(fs,'operacional')
    d=OperacionalDocumento(empresa_id=to_num(request.form.get('empresa_id')) or None,tipo=(request.form.get('tipo') or 'Documento').strip(),titulo=titulo,descricao=(request.form.get('descricao') or '').strip(),nome_arquivo=(fs.filename if fs else None),caminho=rel,criado_por=session.get('nome',''))
    db.session.add(d); db.session.commit(); return jsonify(d.to_dict()),201

@app.route('/api/operacional/documentos/<int:id>',methods=['DELETE'])
@lr
def api_del_oper_doc(id):
    d=OperacionalDocumento.query.get_or_404(id)
    if d.caminho:
        try: os.remove(os.path.join(UPLOAD_ROOT,d.caminho))
        except: pass
    db.session.delete(d); db.session.commit(); return jsonify({'ok':True})

@app.route('/api/operacional/documentos/<int:id>/download')
@lr
def api_download_oper_doc(id):
    d=OperacionalDocumento.query.get_or_404(id)
    if not d.caminho: return jsonify({'erro':'Documento sem arquivo'}),404
    abs_p=os.path.join(UPLOAD_ROOT,d.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    return send_file(abs_p,as_attachment=True,download_name=d.nome_arquivo or 'documento.bin')

@app.route('/api/operacional/postos',methods=['GET'])
@lr
def api_operacional_postos():
    empresa_id=to_num(request.args.get('empresa_id')) or None
    q=Funcionario.query.filter_by(status='Ativo')
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    cls=Cliente.query.order_by(Cliente.nome).all()
    if empresa_id:
        cls=[c for c in cls if c.empresa_id==empresa_id]
    clientes=[{
        'id':c.id,
        'nome':c.nome or '',
        'empresa_id':c.empresa_id,
        'numero_contrato':c.numero_contrato or '',
        'qtd_funcionarios_posto':max(0,to_num(c.qtd_funcionarios_posto)),
        'ocupados':Funcionario.query.filter_by(posto_cliente_id=c.id,status='Ativo').count(),
        'posto_label':(c.nome or '').strip()
    } for c in cls]
    cli_map={c.id:c for c in cls}
    itens=[]
    for f in q.order_by(Funcionario.nome).all():
        emp=Empresa.query.get(f.empresa_id) if f.empresa_id else None
        cli=cli_map.get(f.posto_cliente_id) if f.posto_cliente_id else None
        posto_label=(f.posto_operacional or 'Reserva tecnica')
        if cli:
            posto_label=(cli.nome or '').strip() or 'Reserva tecnica'
        itens.append({
            'funcionario_id':f.id,
            'matricula':f.matricula or '',
            'nome':f.nome or '',
            'empresa_id':f.empresa_id,
            'empresa_nome':(emp.nome if emp else ''),
            'posto_operacional':f.posto_operacional or '',
            'posto_cliente_id':f.posto_cliente_id,
            'posto_label':posto_label
        })
    return jsonify({'ok':True,'itens':itens,'clientes':clientes})

@app.route('/api/operacional/postos',methods=['POST'])
@lr
def api_operacional_postos_salvar():
    d=request.json or {}
    fid=to_num(d.get('funcionario_id'))
    if not fid:
        return jsonify({'erro':'Funcionario obrigatorio'}),400
    f=Funcionario.query.get_or_404(fid)
    posto_cliente_id=to_num(d.get('posto_cliente_id')) or None
    if posto_cliente_id:
        cli=Cliente.query.get_or_404(posto_cliente_id)
        if f.empresa_id and cli.empresa_id and f.empresa_id!=cli.empresa_id:
            return jsonify({'erro':'O posto selecionado pertence a outra empresa.'}),400
        cap=max(0,to_num(cli.qtd_funcionarios_posto))
        if cap<=0:
            return jsonify({'erro':'Defina no cliente a quantidade de funcionarios por posto para permitir vinculo.'}),400
        ocupados=Funcionario.query.filter_by(posto_cliente_id=cli.id,status='Ativo').count()
        if f.posto_cliente_id!=cli.id and ocupados>=cap:
            return jsonify({'erro':f'Limite deste posto atingido ({ocupados}/{cap}).'}),400
        f.posto_cliente_id=cli.id
        f.posto_operacional=(cli.nome or '').strip() or 'Reserva tecnica'
    else:
        f.posto_cliente_id=None
        f.posto_operacional='Reserva tecnica'
    db.session.commit()
    return jsonify({'ok':True,'funcionario':f.to_dict()})

@app.route('/api/beneficios/lancamentos',methods=['GET'])
@lr
def api_beneficios_lancamentos():
    comp=norm_competencia(request.args.get('competencia'))
    empresa_id=to_num(request.args.get('empresa_id')) or None
    qf=Funcionario.query.filter_by(status='Ativo')
    if empresa_id:
        qf=qf.filter_by(empresa_id=empresa_id)
    funcs_ativos=qf.order_by(Funcionario.nome).all()
    qb=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        qb=qb.filter_by(empresa_id=empresa_id)
    mapa={b.funcionario_id:b for b in qb.all()}
    itens=[]
    def _benef_val(bm,func_obj,attr_name):
        if not bm:
            return getattr(func_obj,attr_name) or 0
        val=getattr(bm,attr_name)
        if val is None:
            return getattr(func_obj,attr_name) or 0
        return val
    for f in funcs_ativos:
        emp=Empresa.query.get(f.empresa_id) if f.empresa_id else None
        cli=Cliente.query.get(f.posto_cliente_id) if f.posto_cliente_id else None
        posto_nome=(cli.nome.strip() if cli and (cli.nome or '').strip() else (f.posto_operacional or 'Reserva tecnica'))
        b=mapa.get(f.id)
        itens.append({
            'funcionario_id':f.id,
            'matricula':f.matricula or '',
            're':f.re or '',
            'nome':f.nome or '',
            'posto_operacional':posto_nome,
            'posto_cliente_id':f.posto_cliente_id,
            'empresa_id':f.empresa_id,
            'empresa_nome':(emp.nome if emp else ''),
            'competencia':comp,
            'dias_trabalhados':(b.dias_trabalhados if b else 0),
            'dias_vt':(b.dias_vt if b else 0),
            'dias_vr':(b.dias_vr if b else 0),
            'dias_va':(b.dias_va if b else 0),
            'dias_vg':(b.dias_vg if b else 0),
            'salario':(b.salario if b else (f.salario or 0)),
            'vale_refeicao':_benef_val(b,f,'vale_refeicao'),
            'vale_alimentacao':_benef_val(b,f,'vale_alimentacao'),
            'vale_transporte':_benef_val(b,f,'vale_transporte'),
            'opta_vt': True if f.opta_vt is None else bool(f.opta_vt),
            'opta_vr': True if f.opta_vr is None else bool(f.opta_vr),
            'opta_va': True if f.opta_va is None else bool(f.opta_va),
            'opta_premio_prod': bool(f.opta_premio_prod),
            'opta_vale_gasolina': bool(f.opta_vale_gasolina),
            'opta_cesta_natal': bool(f.opta_cesta_natal),
            'pp_falta': bool(b.pp_falta) if b and (b.pp_falta is not None) else False,
            'premio_produtividade':_benef_val(b,f,'premio_produtividade'),
            'vale_gasolina':_benef_val(b,f,'vale_gasolina'),
            'cesta_natal':_benef_val(b,f,'cesta_natal'),
        })
    return jsonify({'ok':True,'competencia':comp,'itens':itens})

@app.route('/api/beneficios/lancamentos',methods=['DELETE'])
@lr
def api_beneficios_lancamentos_excluir():
    d=request.json or {}
    comp=norm_competencia(d.get('competencia'))
    tipo=(d.get('tipo') or 'todos').strip().lower()
    empresa_id=to_num(d.get('empresa_id')) or None
    func_ids=d.get('funcionarios') or []
    func_ids={int(x) for x in func_ids if str(x).isdigit()}
    if tipo not in {'vt','vr','va','pp','vg','cn','todos'}:
        return jsonify({'erro':'Tipo inválido para exclusão.'}),400

    q=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    regs=q.all()
    if func_ids:
        regs=[b for b in regs if b.funcionario_id in func_ids]

    excluidos=0
    for b in regs:
        if tipo=='todos':
            db.session.delete(b)
            excluidos+=1
        else:
            changed=False
            if tipo=='vt':
                if b.dias_vt!=0: b.dias_vt=0; changed=True
                if b.vale_transporte is not None: b.vale_transporte=None; changed=True
            elif tipo=='vr':
                if b.dias_vr!=0: b.dias_vr=0; changed=True
                if b.vale_refeicao is not None: b.vale_refeicao=None; changed=True
            elif tipo=='va':
                if b.dias_va!=0: b.dias_va=0; changed=True
                if b.vale_alimentacao is not None: b.vale_alimentacao=None; changed=True
            elif tipo=='pp':
                if bool(b.pp_falta): b.pp_falta=False; changed=True
                if b.premio_produtividade is not None: b.premio_produtividade=None; changed=True
            elif tipo=='vg':
                if b.dias_vg!=0: b.dias_vg=0; changed=True
                if b.vale_gasolina is not None: b.vale_gasolina=None; changed=True
            elif tipo=='cn':
                if b.cesta_natal is not None: b.cesta_natal=None; changed=True
            if changed: excluidos+=1

    db.session.commit()
    return jsonify({'ok':True,'competencia':comp,'tipo':tipo,'excluidos':excluidos})

@app.route('/api/beneficios/lancamentos/limpar',methods=['POST'])
@lr
def api_beneficios_lancamentos_limpar():
    d=request.json or {}
    comp=norm_competencia(d.get('competencia'))
    tipo=(d.get('tipo') or '').strip().lower()
    empresa_id=to_num(d.get('empresa_id')) or None
    func_ids=d.get('funcionarios') or []
    func_ids={int(x) for x in func_ids if str(x).isdigit()}
    if tipo not in {'vt','vr','va','pp','vg','cn','todos'}:
        return jsonify({'erro':'Tipo inválido para limpeza.'}),400

    q=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    regs=q.all()
    if func_ids:
        regs=[b for b in regs if b.funcionario_id in func_ids]

    alterados=0
    for b in regs:
        changed=False
        if tipo in {'vt','todos'}:
            if b.dias_vt!=0: b.dias_vt=0; changed=True
            if b.vale_transporte is not None: b.vale_transporte=None; changed=True
        if tipo in {'vr','todos'}:
            if b.dias_vr!=0: b.dias_vr=0; changed=True
            if b.vale_refeicao is not None: b.vale_refeicao=None; changed=True
        if tipo in {'va','todos'}:
            if b.dias_va!=0: b.dias_va=0; changed=True
            if b.vale_alimentacao is not None: b.vale_alimentacao=None; changed=True
        if tipo in {'pp','todos'}:
            if bool(b.pp_falta): b.pp_falta=False; changed=True
            if b.premio_produtividade is not None: b.premio_produtividade=None; changed=True
        if tipo in {'vg','todos'}:
            if b.dias_vg!=0: b.dias_vg=0; changed=True
            if b.vale_gasolina is not None: b.vale_gasolina=None; changed=True
        if tipo in {'cn','todos'}:
            if b.cesta_natal is not None: b.cesta_natal=None; changed=True
        if changed:
            alterados+=1

    db.session.commit()
    return jsonify({'ok':True,'competencia':comp,'tipo':tipo,'alterados':alterados})

@app.route('/api/beneficios/lancamentos',methods=['POST'])
@lr
def api_beneficios_lancamentos_salvar():
    d=request.json or {}
    comp=norm_competencia(d.get('competencia'))
    itens=d.get('itens') or []
    salvos=0
    for it in itens:
        fid=to_num(it.get('funcionario_id'))
        if not fid:
            continue
        f=Funcionario.query.get(fid)
        if not f:
            continue
        b=BeneficioMensal.query.filter_by(funcionario_id=fid,competencia=comp).first()
        if not b:
            b=BeneficioMensal(funcionario_id=fid,competencia=comp)
            db.session.add(b)
        b.empresa_id=f.empresa_id
        vt_optante=(f.opta_vt is not False)
        vr_optante=(f.opta_vr is not False)
        va_optante=(f.opta_va is not False)
        pp_optante=bool(f.opta_premio_prod)
        vg_optante=bool(f.opta_vale_gasolina)
        cn_optante=bool(f.opta_cesta_natal)

        dias_vt=max(0,to_num(it.get('dias_vt')))
        dias_vr=max(0,to_num(it.get('dias_vr')))
        dias_va=max(0,to_num(it.get('dias_va')))
        dias_vg=max(0,to_num(it.get('dias_vg')))
        pp_falta=to_bool(it.get('pp_falta')) if pp_optante else False

        b.dias_vt=dias_vt if vt_optante else 0
        b.dias_vr=dias_vr if vr_optante else 0
        b.dias_va=dias_va if va_optante else 0
        b.dias_vg=dias_vg if vg_optante else 0
        b.pp_falta=pp_falta
        b.dias_trabalhados=max(b.dias_vt,b.dias_vr)

        b.salario=to_num(it.get('salario'),dec=True)
        b.vale_transporte=(to_num(it.get('vale_transporte'),dec=True) if vt_optante else 0)
        b.vale_refeicao=(to_num(it.get('vale_refeicao'),dec=True) if vr_optante else 0)
        b.vale_alimentacao=(to_num(it.get('vale_alimentacao'),dec=True) if va_optante else 0)
        b.premio_produtividade=(to_num(it.get('premio_produtividade'),dec=True) if (pp_optante and not pp_falta) else 0)
        b.vale_gasolina=(to_num(it.get('vale_gasolina'),dec=True) if vg_optante else 0)
        b.cesta_natal=(to_num(it.get('cesta_natal'),dec=True) if cn_optante else 0)
        salvos+=1
    db.session.commit()
    return jsonify({'ok':True,'competencia':comp,'salvos':salvos})

@app.route('/api/beneficios/alerta-config',methods=['GET'])
@lr
def api_beneficios_alerta_config_get():
    dia_vt=to_num(gc('benef_alerta_vt_dia','25')) or 25
    dia_vrva=to_num(gc('benef_alerta_vrva_dia','26')) or 26
    dia_vt=max(1,min(31,dia_vt))
    dia_vrva=max(1,min(31,dia_vrva))
    return jsonify({'ok':True,'vt_dia':dia_vt,'vrva_dia':dia_vrva})

@app.route('/api/beneficios/alerta-config',methods=['POST'])
@lr
def api_beneficios_alerta_config_salvar():
    d=request.json or {}
    vt=to_num(d.get('vt_dia'))
    vrva=to_num(d.get('vrva_dia'))
    if vt<1 or vt>31:
        return jsonify({'erro':'Dia do alerta de VT deve ser entre 1 e 31.'}),400
    if vrva<1 or vrva>31:
        return jsonify({'erro':'Dia do alerta de VR/VA deve ser entre 1 e 31.'}),400
    sc_cfg('benef_alerta_vt_dia',vt)
    sc_cfg('benef_alerta_vrva_dia',vrva)
    return jsonify({'ok':True,'vt_dia':vt,'vrva_dia':vrva})

@app.route('/api/beneficios/vale-transporte/pdf')
@lr
def api_beneficios_vale_transporte_pdf():
    return _api_beneficios_pdf_tipo('vale_transporte')

@app.route('/api/beneficios/vale-refeicao/pdf')
@lr
def api_beneficios_vale_refeicao_pdf():
    return _api_beneficios_pdf_tipo('vale_refeicao')

@app.route('/api/beneficios/vale-alimentacao/pdf')
@lr
def api_beneficios_vale_alimentacao_pdf():
    return _api_beneficios_pdf_tipo('vale_alimentacao')

@app.route('/api/beneficios/premio-produtividade/pdf')
@lr
def api_beneficios_premio_produtividade_pdf():
    return _api_beneficios_pdf_tipo('premio_produtividade')

@app.route('/api/beneficios/vale-gasolina/pdf')
@lr
def api_beneficios_vale_gasolina_pdf():
    return _api_beneficios_pdf_tipo('vale_gasolina')

@app.route('/api/beneficios/cesta-natal/pdf')
@lr
def api_beneficios_cesta_natal_pdf():
    return _api_beneficios_pdf_tipo('cesta_natal')

@app.route('/api/beneficios/vale-transporte/xlsx')
@lr
def api_beneficios_vale_transporte_xlsx():
    return _api_beneficios_xlsx_tipo('vale_transporte')

@app.route('/api/beneficios/vale-refeicao/xlsx')
@lr
def api_beneficios_vale_refeicao_xlsx():
    return _api_beneficios_xlsx_tipo('vale_refeicao')

@app.route('/api/beneficios/vale-alimentacao/xlsx')
@lr
def api_beneficios_vale_alimentacao_xlsx():
    return _api_beneficios_xlsx_tipo('vale_alimentacao')

@app.route('/api/beneficios/premio-produtividade/xlsx')
@lr
def api_beneficios_premio_produtividade_xlsx():
    return _api_beneficios_xlsx_tipo('premio_produtividade')

@app.route('/api/beneficios/vale-gasolina/xlsx')
@lr
def api_beneficios_vale_gasolina_xlsx():
    return _api_beneficios_xlsx_tipo('vale_gasolina')

@app.route('/api/beneficios/cesta-natal/xlsx')
@lr
def api_beneficios_cesta_natal_xlsx():
    return _api_beneficios_xlsx_tipo('cesta_natal')

@app.route('/beneficios/relatorio')
@lr
def beneficios_relatorio_preview():
    tipo_slug=(request.args.get('tipo') or 'vale-transporte').strip().lower()
    mapa_tipo={
        'vale-transporte':'vale_transporte',
        'vale-refeicao':'vale_refeicao',
        'vale-alimentacao':'vale_alimentacao',
        'premio-produtividade':'premio_produtividade',
        'vale-gasolina':'vale_gasolina',
        'cesta-natal':'cesta_natal',
    }
    tipo=mapa_tipo.get(tipo_slug)
    if not tipo:
        return 'Tipo de benefício inválido.',400

    comp=norm_competencia(request.args.get('competencia'))
    empresa_id=to_num(request.args.get('empresa_id')) or None
    func_ids_raw=request.args.get('funcionarios','').strip()
    func_ids_filter={int(x) for x in func_ids_raw.split(',') if x.strip().isdigit()} if func_ids_raw else set()
    cfg={
        'vale_transporte':('Vale Transporte','vale_transporte','dias_vt','opta_vt'),
        'vale_refeicao':('Vale Refeição','vale_refeicao','dias_vr','opta_vr'),
        'vale_alimentacao':('Vale Alimentação','vale_alimentacao','dias_va','opta_va'),
        'premio_produtividade':('Prêmio Produtividade','premio_produtividade','','opta_premio_prod'),
        'vale_gasolina':('Vale Gasolina','vale_gasolina','dias_vg','opta_vale_gasolina'),
        'cesta_natal':('Cesta de Natal','cesta_natal','','opta_cesta_natal'),
    }
    tit,col_valor,col_dias,opta_col=cfg[tipo]
    is_va=(tipo in {'vale_alimentacao','premio_produtividade','cesta_natal'})

    q=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)

    funcs_map={f.id:f for f in Funcionario.query.all()}
    emps_map={e.id:e for e in Empresa.query.all()}

    def _is_optante(b):
        f=funcs_map.get(b.funcionario_id)
        return f is None or getattr(f,opta_col,True) is not False

    regs=[b for b in q.all() if float(getattr(b,col_valor) or 0)>0 and _is_optante(b) and (not func_ids_filter or b.funcionario_id in func_ids_filter)]

    grupos={}
    for r in regs:
        grupos.setdefault(r.empresa_id or 0,[]).append(r)

    empresas=[]
    total_geral=0.0
    qtd_geral=0
    for emp_id,items in sorted(grupos.items(),key=lambda kv:((emps_map.get(kv[0]).nome if emps_map.get(kv[0]) else 'ZZZ'),kv[0])):
        emp=emps_map.get(emp_id)
        nome_emp=(emp.nome if emp else 'Sem empresa')
        cnpj_emp=(emp.cnpj if emp and emp.cnpj else '')
        linhas=[]
        total_emp=0.0
        for r in sorted(items,key=lambda x:(funcs_map.get(x.funcionario_id).nome if funcs_map.get(x.funcionario_id) else '')):
            f=funcs_map.get(r.funcionario_id)
            valor=float(getattr(r,col_valor) or 0)
            dias=int(getattr(r,col_dias) or 0) if col_dias else 0
            total=(valor if is_va else (dias*valor if dias>0 else valor))
            total_emp+=total
            linhas.append({
                're':str((f.re if f and f.re else (f.matricula if f and f.matricula else '')) or ''),
                'nome':(f.nome if f else f'Funcionario {r.funcionario_id}'),
                'cpf':(f.cpf if f and f.cpf else ''),
                'dias':dias,
                'valor_fmt':fmt_brl(valor),
                'total_fmt':fmt_brl(total),
            })
        qtd_emp=len(items)
        total_geral+=total_emp
        qtd_geral+=qtd_emp
        empresas.append({
            'nome':nome_emp,
            'cnpj':cnpj_emp,
            'qtd':qtd_emp,
            'total_fmt':fmt_brl(total_emp),
            'linhas':linhas,
        })

    base_tipo=tipo.replace('_','-')
    query_base=f'competencia={comp}'
    if empresa_id:
        query_base+=f'&empresa_id={empresa_id}'
    if func_ids_raw:
        query_base+=f'&funcionarios={func_ids_raw}'
    pdf_url=f"/api/beneficios/{base_tipo}/pdf?{query_base}"
    xlsx_url=f"/api/beneficios/{base_tipo}/xlsx?{query_base}"

    return render_template(
        'beneficios_relatorio_preview.html',
        titulo=tit,
        competencia=comp,
        empresas=empresas,
        is_va=is_va,
        qtd_geral=qtd_geral,
        total_geral_fmt=fmt_brl(total_geral),
        pdf_url=pdf_url,
        xlsx_url=xlsx_url,
        voltar_url=url_for('index')
    )

def _api_beneficios_xlsx_tipo(tipo):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    comp=norm_competencia(request.args.get('competencia'))
    empresa_id=to_num(request.args.get('empresa_id')) or None
    func_ids_raw=request.args.get('funcionarios','').strip()
    func_ids_filter={int(x) for x in func_ids_raw.split(',') if x.strip().isdigit()} if func_ids_raw else set()
    cfg={
        'vale_transporte':('Vale Transporte','vale_transporte','dias_vt','vt','opta_vt'),
        'vale_refeicao':('Vale Refeição','vale_refeicao','dias_vr','vr','opta_vr'),
        'vale_alimentacao':('Vale Alimentação','vale_alimentacao','dias_va','va','opta_va'),
        'premio_produtividade':('Prêmio Produtividade','premio_produtividade','','pp','opta_premio_prod'),
        'vale_gasolina':('Vale Gasolina','vale_gasolina','dias_vg','vg','opta_vale_gasolina'),
        'cesta_natal':('Cesta de Natal','cesta_natal','','cn','opta_cesta_natal'),
    }
    if tipo not in cfg:
        return jsonify({'erro':'Tipo de beneficio invalido'}),400
    tit,col_valor,col_dias,sigla,opta_col=cfg[tipo]
    is_fixed=not col_dias or tipo=='vale_alimentacao'
    q=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    funcs_map_pre={f.id:f for f in Funcionario.query.all()}
    def _is_optante_x(b):
        f=funcs_map_pre.get(b.funcionario_id)
        return f is None or getattr(f,opta_col,True) is not False
    regs=[b for b in q.all() if float(getattr(b,col_valor) or 0)>0 and _is_optante_x(b) and (not func_ids_filter or b.funcionario_id in func_ids_filter)]
    if not regs:
        return jsonify({'erro':f'Nenhum lançamento de {tit.lower()} com valor para a competência informada.'}),400

    emps_map={e.id:e for e in Empresa.query.all()}
    funcs_map={f.id:f for f in Funcionario.query.all()}
    grupos={}
    for r in regs:
        grupos.setdefault(r.empresa_id or 0,[]).append(r)

    wb=Workbook()
    first=True
    header_fill=PatternFill('solid',fgColor='205D8A')
    total_fill=PatternFill('solid',fgColor='EAF2FB')
    header_font=Font(bold=True,color='FFFFFF',size=10)
    total_font=Font(bold=True,size=10)
    normal_font=Font(size=10)
    center=Alignment(horizontal='center',vertical='center')
    left=Alignment(horizontal='left',vertical='center')
    right=Alignment(horizontal='right',vertical='center')
    thin=Side(style='thin',color='D0D7DE')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    comp_fmt=f"{comp[5:7]}/{comp[:4]}" if isinstance(comp,str) and len(comp)>=7 and '-' in comp else str(comp)

    for emp_id,items in sorted(grupos.items(),key=lambda kv:((emps_map.get(kv[0]).nome if emps_map.get(kv[0]) else 'ZZZ'),kv[0])):
        emp=emps_map.get(emp_id)
        nome_emp=(emp.nome if emp else 'Sem empresa')
        ws=wb.active if first else wb.create_sheet()
        first=False
        ws.title=nome_emp[:31].replace('/','_').replace('\\','_').replace('?','').replace('*','').replace('[','').replace(']','').replace(':','')
        ws.append([f'Relatório de {tit} — {nome_emp}'])
        ws.append([f'Competência: {comp_fmt}'])
        ws.append([])
        if is_fixed:
            headers=['RE','Colaborador','CPF',f'{tit} (R$)','Total (R$)']
        else:
            headers=['RE','Colaborador','CPF','Dias',f'{tit} (R$)','Total (R$)']
        ws.append(headers)
        hrow=4
        for col_idx,_ in enumerate(headers,1):
            cell=ws.cell(row=hrow,column=col_idx)
            cell.fill=header_fill
            cell.font=header_font
            cell.alignment=center
            cell.border=border
        total_geral=0.0
        for r in sorted(items,key=lambda x:(funcs_map.get(x.funcionario_id).nome if funcs_map.get(x.funcionario_id) else '')):
            f=funcs_map.get(r.funcionario_id)
            valor=float(getattr(r,col_valor) or 0)
            dias=int(getattr(r,col_dias) or 0) if col_dias else 0
            total=(valor if is_fixed else (dias*valor if dias>0 else valor))
            total_geral+=total
            re_val=(f.re if f and f.re else (f.matricula if f and f.matricula else ''))
            nome_val=(f.nome if f else f'Funcionario {r.funcionario_id}')
            cpf_val=(f.cpf if f and f.cpf else '')
            if is_fixed:
                row=[re_val,nome_val,cpf_val,valor,total]
            else:
                row=[re_val,nome_val,cpf_val,dias,valor,total]
            ws.append(row)
            dr=ws.max_row
            for ci,val in enumerate(row,1):
                cell=ws.cell(row=dr,column=ci)
                cell.font=normal_font
                cell.border=border
                if isinstance(val,float):
                    cell.number_format='#,##0.00'
                    cell.alignment=right
                elif isinstance(val,int) and ci>4:
                    cell.alignment=right
                else:
                    cell.alignment=left
        # Quantidade de funcionários
        qr=ws.max_row+1
        qtd_funcs=len(items)
        if is_fixed:
            qc=ws.cell(row=qr,column=4,value='Funcionários:')
            qc.font=total_font; qc.alignment=right; qc.fill=total_fill
            qv=ws.cell(row=qr,column=5,value=qtd_funcs)
        else:
            qc=ws.cell(row=qr,column=5,value='Funcionários:')
            qc.font=total_font; qc.alignment=right; qc.fill=total_fill
            qv=ws.cell(row=qr,column=6,value=qtd_funcs)
        qv.font=total_font; qv.alignment=right; qv.fill=total_fill
        # Total row
        tr=ws.max_row+1
        if is_fixed:
            ws.cell(row=tr,column=4,value='Total da empresa:').font=total_font
            tc=ws.cell(row=tr,column=5,value=total_geral)
        else:
            ws.cell(row=tr,column=5,value='Total da empresa:').font=total_font
            tc=ws.cell(row=tr,column=6,value=total_geral)
        tc.font=total_font
        tc.number_format='#,##0.00'
        tc.alignment=right
        tc.fill=total_fill
        # Column widths: RE, Colaborador, CPF, [Dias,] Valor, Total
        col_widths=[10,35,18,8,16,16] if not is_fixed else [10,35,18,16,16]
        for i,w in enumerate(col_widths[:len(headers)],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.row_dimensions[1].height=16
        ws.cell(row=1,column=1).font=Font(bold=True,size=12)

    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    comp_nome=f"{comp[5:7]}-{comp[:4]}" if isinstance(comp,str) and len(comp)>=7 and '-' in comp else str(comp)
    nome=f"relatorio_{sigla}_competencia_{comp_nome}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=nome)

def _api_beneficios_pdf_tipo(tipo):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,Image
    from reportlab.lib.styles import ParagraphStyle

    comp=norm_competencia(request.args.get('competencia'))
    empresa_id=to_num(request.args.get('empresa_id')) or None
    func_ids_raw=request.args.get('funcionarios','').strip()
    func_ids_filter={int(x) for x in func_ids_raw.split(',') if x.strip().isdigit()} if func_ids_raw else set()
    cfg={
        'vale_transporte':('Vale Transporte','vale_transporte','dias_vt','opta_vt'),
        'vale_refeicao':('Vale Refeição','vale_refeicao','dias_vr','opta_vr'),
        'vale_alimentacao':('Vale Alimentação','vale_alimentacao','dias_va','opta_va'),
        'premio_produtividade':('Prêmio Produtividade','premio_produtividade','','opta_premio_prod'),
        'vale_gasolina':('Vale Gasolina','vale_gasolina','','opta_vale_gasolina'),
        'cesta_natal':('Cesta de Natal','cesta_natal','','opta_cesta_natal'),
    }
    if tipo not in cfg:
        return jsonify({'erro':'Tipo de beneficio invalido'}),400
    tit,col_valor,col_dias,opta_col=cfg[tipo]
    is_fixed=not col_dias or tipo=='vale_alimentacao'
    q=BeneficioMensal.query.filter_by(competencia=comp)
    if empresa_id:
        q=q.filter_by(empresa_id=empresa_id)
    funcs_map_pre={f.id:f for f in Funcionario.query.all()}
    def _is_optante(b):
        f=funcs_map_pre.get(b.funcionario_id)
        return f is None or getattr(f,opta_col,True) is not False
    regs=[b for b in q.all() if float(getattr(b,col_valor) or 0)>0 and _is_optante(b) and (not func_ids_filter or b.funcionario_id in func_ids_filter)]
    if not regs:
        return jsonify({'erro':f'Nenhum lançamento de {tit.lower()} com valor para a competência informada.'}),400

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.3*cm,rightMargin=1.3*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    W=A4[0]-2.6*cm
    st=ParagraphStyle('n',fontName='Helvetica',fontSize=9,leading=12)
    st_h=ParagraphStyle('h',fontName='Helvetica-Bold',fontSize=12,leading=14,textColor=colors.HexColor('#205d8a'))
    st_cell=ParagraphStyle('cell',fontName='Helvetica',fontSize=8.3,leading=10)
    st_num=ParagraphStyle('num',fontName='Helvetica',fontSize=8.3,leading=10,alignment=2)
    story=[]

    lp=get_logo()
    logo_flow=Paragraph('<b>RM Facilities</b>',ParagraphStyle('lgfb',fontName='Helvetica-Bold',fontSize=12,textColor=colors.HexColor('#205d8a')))
    for cand in [lp,LOGO_URL]:
        if not cand:
            continue
        try:
            if isinstance(cand,str) and cand.startswith(('http://','https://')):
                req=urllib.request.Request(cand,headers={'User-Agent':'Mozilla/5.0'})
                with urllib.request.urlopen(req,timeout=8) as r:
                    img_data=r.read()
                logo_flow=Image(io.BytesIO(img_data),width=3.8*cm,height=1.5*cm,kind='proportional')
            elif os.path.exists(cand):
                logo_flow=Image(cand,width=3.8*cm,height=1.5*cm,kind='proportional')
            break
        except Exception:
            continue

    def _esc(v):
        s=str(v or '')
        return s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    emps_map={e.id:e for e in Empresa.query.all()}
    funcs_map={f.id:f for f in Funcionario.query.all()}
    grupos={}
    for r in regs:
        grupos.setdefault(r.empresa_id or 0,[]).append(r)

    for emp_id,items in sorted(grupos.items(),key=lambda kv: ((emps_map.get(kv[0]).nome if emps_map.get(kv[0]) else 'ZZZ'),kv[0])):
        emp=emps_map.get(emp_id)
        nome_emp=(emp.nome if emp else 'Sem empresa')
        empresas_hdr=_pdf_companies_for_header(empresa_obj=emp,limit=2)

        def _logo_flowable_b(item):
            for cand in (item.get('logos') or []):
                try:
                    if isinstance(cand,str) and cand.startswith(('http://','https://')):
                        req=urllib.request.Request(cand,headers={'User-Agent':'Mozilla/5.0'})
                        with urllib.request.urlopen(req,timeout=8) as r:
                            img_data=r.read()
                        return Image(io.BytesIO(img_data),width=3.2*cm,height=1.3*cm,kind='proportional')
                    if os.path.exists(cand):
                        return Image(cand,width=3.2*cm,height=1.3*cm,kind='proportional')
                except Exception:
                    continue
            return Paragraph(f'<b>{item.get("nome") or nome_emp}</b>',ParagraphStyle('lgfbb',fontName='Helvetica-Bold',fontSize=10,textColor=colors.HexColor('#205d8a')))

        hdr=Table([[logo_flow,Paragraph(f'Relatório de {tit}<br/><font size="9">{nome_emp}</font>',ParagraphStyle('h2',fontName='Helvetica-Bold',fontSize=12,leading=14,textColor=colors.HexColor('#205d8a')))]],colWidths=[W*0.28,W*0.72])
        hdr.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),
            ('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),0),
            ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(hdr)
        cnpj_str=(emp.cnpj if emp and emp.cnpj else '')
        emp_info=Paragraph(
            f'<font color="#4c6072">CNPJ: {cnpj_str}</font>',
            ParagraphStyle('empinfo',fontName='Helvetica',fontSize=8.5,leading=11)
        )
        story.append(emp_info)
        story.append(Spacer(1,4))
        story.append(Paragraph(f'Competência: {comp}',st))
        story.append(Spacer(1,6))
        if is_fixed:
            rows=[['RE','Colaborador','CPF',f'{tit} (R$)',f'Total {tit} (R$)']]
        else:
            rows=[['RE','Colaborador','CPF','Dias',f'{tit} (R$)',f'Total {tit} (R$)']]
        total_emp=0.0
        for r in sorted(items,key=lambda x:(funcs_map.get(x.funcionario_id).nome if funcs_map.get(x.funcionario_id) else '')):
            f=funcs_map.get(r.funcionario_id)
            valor=float(getattr(r,col_valor) or 0)
            dias=int(getattr(r,col_dias) or 0) if col_dias else 0
            total=valor if is_fixed else (dias*valor if dias>0 else valor)
            total_emp+=total
            re_str=str(f.re if f and f.re else (f.matricula if f and f.matricula else ''))
            cpf_str=(f.cpf if f and f.cpf else '')
            if is_fixed:
                rows.append([
                    Paragraph(_esc(re_str),st_cell),
                    Paragraph(_esc(f.nome if f else f'Funcionario {r.funcionario_id}'),st_cell),
                    Paragraph(_esc(cpf_str),st_cell),
                    Paragraph(_esc(fmt_brl(valor)),st_num),
                    Paragraph(_esc(fmt_brl(total)),st_num)
                ])
            else:
                rows.append([
                    Paragraph(_esc(re_str),st_cell),
                    Paragraph(_esc(f.nome if f else f'Funcionario {r.funcionario_id}'),st_cell),
                    Paragraph(_esc(cpf_str),st_cell),
                    Paragraph(_esc(str(dias)),st_num),
                    Paragraph(_esc(fmt_brl(valor)),st_num),
                    Paragraph(_esc(fmt_brl(total)),st_num)
                ])
        qtd_funcs=len(items)
        if is_fixed:
            rows.append(['','','', Paragraph('Funcionários:',st_num), Paragraph(_esc(str(qtd_funcs)),st_num)])
            rows.append(['','','', Paragraph('Total da empresa:',st_num), Paragraph(_esc(fmt_brl(total_emp)),st_num)])
            tb=Table(rows,colWidths=[1.8*cm,5.8*cm,3.5*cm,3.2*cm,3.2*cm])
        else:
            rows.append(['','','','', Paragraph('Funcionários:',st_num), Paragraph(_esc(str(qtd_funcs)),st_num)])
            rows.append(['','','','', Paragraph('Total da empresa:',st_num), Paragraph(_esc(fmt_brl(total_emp)),st_num)])
            tb=Table(rows,colWidths=[1.8*cm,5.2*cm,3.3*cm,1.4*cm,2.8*cm,2.8*cm])
        tb.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#205d8a')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#d0d7de')),
            ('ALIGN',(3,1),(-1,-1),'RIGHT'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('FONTSIZE',(0,0),(-1,-1),8.5),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(tb)
        story.append(Spacer(1,10))

    doc.build(story)
    buf.seek(0)
    comp_nome=f"{comp[5:7]}-{comp[:4]}" if isinstance(comp,str) and len(comp)>=7 and '-' in comp else str(comp)
    sigla={'vale_transporte':'vt','vale_refeicao':'vr','vale_alimentacao':'va','premio_produtividade':'pp','vale_gasolina':'vg','cesta_natal':'cn'}.get(tipo,'beneficio')
    nome=f"relatorio_{sigla}_competencia_{comp_nome}.pdf"
    return send_file(buf,mimetype='application/pdf',as_attachment=False,download_name=nome)

@app.route('/api/dashboard')
@lr
def api_dashboard():
    ativos=Cliente.query.filter_by(status='Ativo').all()
    # sum revenue from Contrato table; fall back to Cliente fields for clients without contracts
    contratos_ativos=Contrato.query.filter_by(status='Ativo').all()
    clientes_com_contrato={ct.cliente_id for ct in Contrato.query.all()}
    receita_contratos=sum(
        (ct.limpeza or 0)+(ct.jardinagem or 0)+(ct.portaria or 0)+(ct.materiais_equip_locacao or 0)
        for ct in contratos_ativos
    )
    receita_legado=sum(
        to_num(c.limpeza,True)+to_num(c.jardinagem,True)+to_num(c.portaria,True)+to_num(c.materiais_equip_locacao,True)
        for c in ativos if c.id not in clientes_com_contrato
    )
    receita=receita_contratos+receita_legado
    total_ativos=len(ativos)
    mes=localnow().strftime('%Y-%m')
    try:
        medicoes_mes=Medicao.query.filter_by(mes_ref=mes).all()
        medicoes_validas=Medicao.query.filter(Medicao.status!='cancelada').all()
    except OperationalError as e:
        if not _is_missing_medicao_stamp_error(e):
            raise
        db.session.rollback()
        _ensure_medicao_stamp_cols_runtime(force=True)
        medicoes_mes=Medicao.query.filter_by(mes_ref=mes).all()
        medicoes_validas=Medicao.query.filter(Medicao.status!='cancelada').all()
    emitidos={m.cliente_id for m in medicoes_mes if m.cliente_id}
    emps_all={e.id:e for e in Empresa.query.all()}

    pendentes_clientes=[c for c in ativos if c.id not in emitidos]
    emitidos_qtd=max(0,total_ativos-len(pendentes_clientes))
    taxa_emissao=(emitidos_qtd*100.0/total_ativos) if total_ativos else 0.0
    ticket_medio=(receita/total_ativos) if total_ativos else 0.0

    inad_itens=[]
    total_inadimplencia=0.0
    hoje=date.today()
    for m in medicoes_validas:
        dt=(m.dt_vencimento or '').strip()
        if not dt:
            continue
        try:
            venc=datetime.strptime(dt,'%Y-%m-%d').date()
        except Exception:
            continue
        if venc>=hoje:
            continue
        dias=(hoje-venc).days
        valor=float(m.valor_bruto or 0)
        total_inadimplencia+=valor
        inad_itens.append({
            'medicao_id':m.id,
            'numero':m.numero or '',
            'cliente_nome':m.cliente_nome or '',
            'empresa_nome':m.empresa_nome or 'Sem empresa',
            'vencimento':venc.strftime('%d/%m/%Y'),
            'dias_atraso':dias,
            'valor':round(valor,2)
        })
    inad_itens=sorted(inad_itens,key=lambda x:(-x['dias_atraso'],x['cliente_nome'].lower()))
    alertas_faturamento=[{
        'cliente_id':c.id,
        'cliente_nome':c.nome or '',
        'empresa_nome':(emps_all.get(c.empresa_id).nome if emps_all.get(c.empresa_id) else 'Sem empresa')
    } for c in sorted(pendentes_clientes,key=lambda x:(x.nome or '').lower())]

    def _parse_aso_validade(raw):
        s=(raw or '').strip()
        if not s:
            return None
        for fmt in ('%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%d.%m.%Y'):
            try:
                return datetime.strptime(s,fmt).date()
            except Exception:
                pass
        m=re.match(r'^(\d{2})/(\d{4})$',s)
        if m:
            mm,yy=int(m.group(1)),int(m.group(2))
            if 1<=mm<=12:
                prox=date(yy+1,1,1) if mm==12 else date(yy,mm+1,1)
                return prox-timedelta(days=1)
        m=re.match(r'^(\d{4})-(\d{2})$',s)
        if m:
            yy,mm=int(m.group(1)),int(m.group(2))
            if 1<=mm<=12:
                prox=date(yy+1,1,1) if mm==12 else date(yy,mm+1,1)
                return prox-timedelta(days=1)
        m=re.match(r'^(19|20)\d{2}$',s)
        if m:
            yy=int(s)
            return date(yy,12,31)
        return None

    limite=hoje+timedelta(days=30)
    dia_alerta_vt=max(1,min(31,to_num(gc('benef_alerta_vt_dia','25')) or 25))
    dia_alerta_vrva=max(1,min(31,to_num(gc('benef_alerta_vrva_dia','26')) or 26))
    alertas_calculo=[]
    if hoje.day==dia_alerta_vt:
        alertas_calculo.append({
            'codigo':'vt_25',
            'titulo':'Cálculo de Vale Transporte',
            'descricao':f'Hoje é dia {dia_alerta_vt}: gerar e conferir cálculo de VT da competência.',
            'competencia':mes
        })
    if hoje.day==dia_alerta_vrva:
        alertas_calculo.append({
            'codigo':'vr_va_26',
            'titulo':'Cálculo de VR e VA',
            'descricao':f'Hoje é dia {dia_alerta_vrva}: gerar e conferir cálculo de VR/VA da competência.',
            'competencia':mes
        })

    aso_por_func={}
    for a in FuncionarioArquivo.query.filter_by(categoria='aso').all():
        validade=_parse_aso_validade(a.competencia)
        if not validade:
            continue
        atual=aso_por_func.get(a.funcionario_id)
        if not atual or validade>atual['validade']:
            aso_por_func[a.funcionario_id]={'validade':validade,'competencia':a.competencia}

    funcs_ativos={f.id:f for f in Funcionario.query.filter_by(status='Ativo').all()}

    lancados_benef=set(b.funcionario_id for b in BeneficioMensal.query.filter_by(competencia=mes).all() if b.funcionario_id)
    alertas_benef_pend=[]
    for f in sorted(funcs_ativos.values(),key=lambda x:(x.nome or '').lower()):
        if f.id not in lancados_benef:
            alertas_benef_pend.append({
                'funcionario_id':f.id,
                'matricula':f.matricula or '',
                'funcionario_nome':f.nome or '',
                'empresa_nome':(emps_all.get(f.empresa_id).nome if emps_all.get(f.empresa_id) else 'Sem empresa')
            })

    alertas_aso=[]
    for fid,meta in aso_por_func.items():
        f=funcs_ativos.get(fid)
        if not f:
            continue
        validade=meta['validade']
        if hoje<=validade<=limite:
            alertas_aso.append({
                'funcionario_id':f.id,
                'matricula':f.matricula or '',
                'funcionario_nome':f.nome or '',
                'validade':validade.strftime('%d/%m/%Y'),
                'dias_restantes':(validade-hoje).days
            })
    alertas_aso=sorted(alertas_aso,key=lambda x:x['dias_restantes'])

    alertas_sem_aso=[]
    for f in sorted(funcs_ativos.values(),key=lambda x:(x.nome or '').lower()):
        meta=aso_por_func.get(f.id)
        if not meta:
            alertas_sem_aso.append({
                'funcionario_id':f.id,
                'matricula':f.matricula or '',
                'funcionario_nome':f.nome or '',
                'empresa_nome':(emps_all.get(f.empresa_id).nome if emps_all.get(f.empresa_id) else 'Sem empresa'),
                'status':'Sem ASO cadastrado',
                'validade':'—',
                'dias_vencido':None
            })
            continue
        validade=meta['validade']
        if validade<hoje:
            alertas_sem_aso.append({
                'funcionario_id':f.id,
                'matricula':f.matricula or '',
                'funcionario_nome':f.nome or '',
                'empresa_nome':(emps_all.get(f.empresa_id).nome if emps_all.get(f.empresa_id) else 'Sem empresa'),
                'status':'ASO vencido',
                'validade':validade.strftime('%d/%m/%Y'),
                'dias_vencido':(hoje-validade).days
            })

    # Faturamento últimos 12 meses
    fat_por_mes = {}
    for _m in Medicao.query.filter(Medicao.status != 'cancelada').all():
        if _m.mes_ref:
            fat_por_mes[_m.mes_ref] = fat_por_mes.get(_m.mes_ref, 0) + float(_m.valor_bruto or 0)
    fat_mensal = []
    for _i in range(11, -1, -1):
        _yr = hoje.year; _mo = hoje.month - _i
        while _mo <= 0: _mo += 12; _yr -= 1
        _k = f"{_yr}-{_mo:02d}"
        fat_mensal.append({'mes': _k, 'total': round(fat_por_mes.get(_k, 0), 2)})

    # faturas vencidas
    venc_itens = []
    total_vencidas = 0.0
    for _m in medicoes_validas:
        _status = (_m.status or '').lower()
        if _status in ('cancelada', 'paga'):
            continue
        _dt = (_m.dt_vencimento or '').strip()
        if not _dt:
            continue
        try:
            _vd = datetime.strptime(_dt, '%Y-%m-%d').date()
        except Exception:
            continue
        if _vd >= hoje:
            continue
        _val = float(_m.valor_bruto or 0)
        total_vencidas += _val
        venc_itens.append({'id': _m.id, 'numero': _m.numero or '', 'cliente': _m.cliente_nome or '',
                           'vencimento': _vd.strftime('%d/%m/%Y'), 'valor': round(_val, 2),
                           'dias': (hoje - _vd).days})

    # alertas contratos vencendo (30 dias)
    alertas_contratos = []
    limite_cont = hoje + timedelta(days=30)
    for _c in ativos:
        _dtv = getattr(_c, 'dt_contrato_vencimento', None) or ''
        if not _dtv:
            continue
        try:
            _dv = datetime.strptime(_dtv.strip(), '%Y-%m-%d').date()
        except Exception:
            continue
        if hoje <= _dv <= limite_cont:
            alertas_contratos.append({'cliente_id': _c.id, 'cliente_nome': _c.nome or '',
                                      'vencimento': _dv.strftime('%d/%m/%Y'),
                                      'dias': (_dv - hoje).days})
        elif _dv < hoje:
            alertas_contratos.append({'cliente_id': _c.id, 'cliente_nome': _c.nome or '',
                                      'vencimento': _dv.strftime('%d/%m/%Y'),
                                      'dias': (_dv - hoje).days, 'vencido': True})
    alertas_contratos.sort(key=lambda x: x['dias'])

    return jsonify({'ativos':len(ativos),'receita':receita,'total_med':Medicao.query.count(),
        'med_mes':Medicao.query.filter_by(mes_ref=mes).count(),
        'pendentes':len(pendentes_clientes),
        'taxa_emissao':round(taxa_emissao,1),
        'ticket_medio':round(ticket_medio,2),
        'faturamento_mensal':fat_mensal,
        'total_vencidas':len(venc_itens), 'valor_vencidas':round(total_vencidas,2),
        'alerta_inadimplencia':{'qtd':len(inad_itens),'total':round(total_inadimplencia,2),'itens':inad_itens[:8]},
        'alerta_faturamento':{'qtd':len(alertas_faturamento),'itens':alertas_faturamento[:8]},
        'alerta_calculo_beneficios':{'qtd':len(alertas_calculo),'itens':alertas_calculo},
        'alerta_beneficios_pendentes':{'qtd':len(alertas_benef_pend),'itens':alertas_benef_pend[:8],'competencia':mes},
        'alerta_sem_aso_valido':{'qtd':len(alertas_sem_aso),'itens':alertas_sem_aso[:8]},
        'alerta_aso':{'qtd':len(alertas_aso),'itens':alertas_aso[:8],'janela_dias':30},
        'alerta_contratos':{'qtd':len(alertas_contratos),'itens':alertas_contratos[:8]},
        'total_cli':Cliente.query.count(),'proximo_num':prox_num(),
        'empresas':[{'id':e.id,'nome':e.nome,'cli':Cliente.query.filter_by(empresa_id=e.id,status='Ativo').count()} for e in Empresa.query.filter_by(ativa=True).all()]})

@app.route('/api/backup')
@lr
def api_backup():
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as z:
        db_p=DB_PATH
        if os.path.exists(db_p): z.write(db_p,'rmfacilities.db')
        z.writestr('clientes.json',json.dumps([c.to_dict() for c in Cliente.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('medicoes.json',json.dumps([m.to_dict() for m in Medicao.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('empresas.json',json.dumps([e.to_dict() for e in Empresa.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('funcionarios.json',json.dumps([f.to_dict() for f in Funcionario.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('config.json',json.dumps([{'chave':c.chave,'valor':c.valor} for c in Config.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('whatsapp_conversas.json',json.dumps([c.to_dict() for c in WhatsAppConversa.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('whatsapp_mensagens.json',json.dumps([m.to_dict() for m in WhatsAppMensagem.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('funcionario_arquivos.json',json.dumps([a.to_dict() for a in FuncionarioArquivo.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('beneficios_mensais.json',json.dumps([b.to_dict() for b in BeneficioMensal.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('ordens_compra.json',json.dumps([o.to_dict() for o in OrdemCompra.query.all()],default=str,ensure_ascii=False,indent=2))
        z.writestr('operacional_documentos.json',json.dumps([d.to_dict() for d in OperacionalDocumento.query.all()],default=str,ensure_ascii=False,indent=2))
        if os.path.isdir(UPLOAD_ROOT):
            for root,_,files in os.walk(UPLOAD_ROOT):
                for fn in files:
                    ap=os.path.join(root,fn)
                    rel=os.path.relpath(ap,UPLOAD_ROOT)
                    z.write(ap,os.path.join('uploads',rel))
        z.writestr('info.json',json.dumps({'data':localnow().isoformat(),'versao':'3.0'},ensure_ascii=False))
    buf.seek(0)
    return send_file(buf,mimetype='application/zip',as_attachment=True,download_name=f'backup_rm_{localnow().strftime("%Y%m%d_%H%M")}.zip')

@app.route('/api/backup/restore',methods=['POST'])
@lr
def api_backup_restore():
    arq=request.files.get('arquivo')
    if not arq: return jsonify({'erro':'Arquivo ZIP nao enviado'}),400
    try:
        z=zipfile.ZipFile(arq)
    except Exception:
        return jsonify({'erro':'Arquivo invalido. Envie um ZIP de backup'}),400

    try:
        # Garante colunas novas em instalações antigas antes de inserir os dados do backup.
        ensure_cols('funcionario_arquivo',[
            'ass_status VARCHAR(20) DEFAULT "nao_solicitada"',
            'ass_token VARCHAR(120)',
            'ass_expira_em DATETIME',
            'ass_codigo VARCHAR(120)',
            'ass_nome VARCHAR(200)',
            'ass_cargo VARCHAR(120)',
            'ass_cpf VARCHAR(20)',
            'ass_ip VARCHAR(60)',
            'ass_em DATETIME',
            'ass_canal_envio VARCHAR(20)',
            'ass_enviado_em DATETIME',
            'ass_recebido_em DATETIME',
            'ass_aberto_em DATETIME',
            'ass_wa_status VARCHAR(20) DEFAULT "nao_enviado"',
            'ass_wa_enviado_em DATETIME',
            'ass_wa_recebido_em DATETIME',
            'ass_email_status VARCHAR(20) DEFAULT "nao_enviado"',
            'ass_email_enviado_em DATETIME',
            'ass_email_recebido_em DATETIME',
        ])

        def jread(name,default):
            try: return json.loads(z.read(name).decode('utf-8'))
            except Exception: return default

        # Limpa dados operacionais antes de restaurar.
        for model in [WhatsAppMensagem,WhatsAppConversa,BeneficioMensal,FuncionarioArquivo,OperacionalDocumento,OrdemCompra,Medicao,Cliente,Funcionario,Empresa,Config]:
            model.query.delete()

        empresas=jread('empresas.json',[])
        clientes=jread('clientes.json',[])
        medicoes=jread('medicoes.json',[])
        funcs=jread('funcionarios.json',[])
        configs=jread('config.json',[])
        wa_convs=jread('whatsapp_conversas.json',[])
        wa_msgs=jread('whatsapp_mensagens.json',[])
        farqs=jread('funcionario_arquivos.json',[])
        bens=jread('beneficios_mensais.json',[])
        ocs=jread('ordens_compra.json',[])
        opdocs=jread('operacional_documentos.json',[])

        def _coerce_value(col,val):
            if val is None:
                return None
            type_name=(col.type.__class__.__name__ or '').lower()
            try:
                py_t=col.type.python_type
            except Exception:
                py_t=None

            # Backups antigos podem trazer objetos em colunas textuais (ex.: contexto da conversa).
            if isinstance(val,(dict,list)) and (py_t is str or 'text' in type_name or 'string' in type_name):
                try:
                    return json.dumps(val,ensure_ascii=False)
                except Exception:
                    return str(val)

            if isinstance(val,bytes) and (py_t is str or 'text' in type_name or 'string' in type_name):
                try:
                    return val.decode('utf-8')
                except Exception:
                    return val.decode('latin-1',errors='ignore')

            if py_t is datetime or 'datetime' in type_name:
                if isinstance(val,datetime):
                    return val
                if isinstance(val,str):
                    s=val.strip()
                    if not s:
                        return None
                    dt=_parse_dt_iso(s)
                    if not dt and ' ' in s and 'T' not in s:
                        dt=_parse_dt_iso(s.replace(' ','T'))
                    if dt:
                        return dt
                    for fmt in ('%Y-%m-%d %H:%M:%S.%f','%Y-%m-%d %H:%M:%S','%Y-%m-%dT%H:%M:%S.%f','%Y-%m-%dT%H:%M:%S'):
                        try:
                            return datetime.strptime(s,fmt)
                        except Exception:
                            pass
                return None

            if py_t is date or (type_name=='date'):
                if isinstance(val,date):
                    return val
                if isinstance(val,str):
                    s=val.strip()
                    if not s:
                        return None
                    for fmt in ('%Y-%m-%d','%d/%m/%Y'):
                        try:
                            return datetime.strptime(s,fmt).date()
                        except Exception:
                            pass
                return None

            if py_t is bool or 'boolean' in type_name:
                if isinstance(val,bool):
                    return val
                if isinstance(val,str):
                    return val.strip().lower() in ('1','true','yes','on','sim')
                return bool(val)

            if py_t is int or 'integer' in type_name:
                if val=='':
                    return None
                try:
                    return int(val)
                except Exception:
                    return None

            if py_t is float or 'float' in type_name:
                if val=='':
                    return None
                try:
                    return float(val)
                except Exception:
                    return None

            return val

        def add_rows(model,rows,conv=None):
            cols_map={c.name:c for c in model.__table__.columns}
            table_name=model.__table__.name
            db_cols={r[1] for r in db.session.execute(text(f'PRAGMA table_info({table_name})')).fetchall()}
            for r in rows:
                d={}
                for k,v in r.items():
                    if k in cols_map and k in db_cols:
                        d[k]=_coerce_value(cols_map[k],v)
                if conv: d=conv(d)
                db.session.add(model(**d))

        add_rows(Empresa,empresas)
        add_rows(Cliente,clientes)
        add_rows(Medicao,medicoes,lambda d: ({**d,'servicos':json.dumps(d.get('svcs',[]),ensure_ascii=False)} if 'svcs' in d and 'servicos' not in d else d))
        add_rows(Funcionario,funcs,lambda d: ({**d,'areas':json.dumps(d.get('areas',[]),ensure_ascii=False)} if isinstance(d.get('areas'),list) else d))
        add_rows(Config,configs)
        add_rows(WhatsAppConversa,wa_convs)
        add_rows(WhatsAppMensagem,wa_msgs)
        add_rows(FuncionarioArquivo,farqs)
        add_rows(BeneficioMensal,bens)
        add_rows(OrdemCompra,ocs)
        add_rows(OperacionalDocumento,opdocs)
        db.session.commit()

        # Restaura uploads.
        if os.path.isdir(UPLOAD_ROOT): shutil.rmtree(UPLOAD_ROOT)
        os.makedirs(UPLOAD_ROOT,exist_ok=True)
        for n in z.namelist():
            if not n.startswith('uploads/') or n.endswith('/'): continue
            rel=n[len('uploads/'):]
            ap=os.path.join(UPLOAD_ROOT,rel)
            os.makedirs(os.path.dirname(ap),exist_ok=True)
            with open(ap,'wb') as out: out.write(z.read(n))

        return jsonify({'ok':True,'restaurado':{'empresas':len(empresas),'clientes':len(clientes),'medicoes':len(medicoes),'funcionarios':len(funcs),'configs':len(configs),'wa_conversas':len(wa_convs),'wa_mensagens':len(wa_msgs)}})
    except Exception as e:
        app.logger.exception('Falha ao restaurar backup completo')
        db.session.rollback()
        return jsonify({'erro':f'Falha ao restaurar backup: {str(e)}'}),500
    finally:
        try:
            z.close()
        except Exception:
            pass

@app.route('/api/pdf/<int:id>')
@lr
def api_pdf(id):
    m=Medicao.query.get_or_404(id)
    emp=Empresa.query.get(m.empresa_id) if m.empresa_id else None
    d=m.to_dict(); d['empresa']=emp.to_dict() if emp else {}
    return _build_pdf(d)

@app.route('/api/pdf/preview',methods=['POST'])
@lr
def api_pdf_preview():
    d=request.json
    if d.get('empresa_id'):
        emp=Empresa.query.get(d['empresa_id'])
        d['empresa']=emp.to_dict() if emp else {}
    else: d['empresa']={}
    return _build_pdf(d)

def _build_pdf(d):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,HRFlowable,Image,PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT,TA_RIGHT,TA_CENTER
    from reportlab.graphics.barcode import qr as qr_code
    from reportlab.graphics.shapes import Drawing

    emp=d.get('empresa',{}); nmed=d.get('numero','001'); tipo=d.get('tipo','Medição de Serviços')
    mes=d.get('mes_ref',''); dtem=d.get('dt_emissao',''); dtvenc=d.get('dt_vencimento','')
    enome=emp.get('razao') or emp.get('nome') or d.get('empresa_nome','RM Facilities')
    eend=emp.get('end_fmt','') or ''; esite=norm_url(emp.get('site','')); epix=emp.get('pix','')
    ebanco=emp.get('banco',''); eagencia=emp.get('agencia',''); econta=emp.get('conta','')
    eboleto=emp.get('boleto','')
    cname=d.get('cliente_nome',''); ccnpj=d.get('cliente_cnpj','')
    cend=d.get('cliente_end',''); cresp=d.get('cliente_resp','')
    obs=d.get('observacoes',''); a1=d.get('ass_empresa',''); a2=d.get('ass_cliente','')
    ass_status=(d.get('assinatura_status') or '').strip().lower()
    ass_nome=(d.get('assinatura_nome') or '').strip()
    ass_cpf=(d.get('assinatura_cpf') or '').strip()
    ass_cargo=(d.get('assinatura_cargo') or '').strip()
    ass_ip=(d.get('assinatura_ip') or '').strip()
    ass_em=d.get('assinatura_em')
    ass_expira=d.get('assinatura_expira_em')
    ass_codigo=(d.get('assinatura_codigo') or '').strip()
    ass_doc_hash=(d.get('assinatura_doc_hash') or '').strip()
    ass_crypto_ok=bool(d.get('assinatura_crypto_ok'))
    ass_cert_subject=(d.get('assinatura_cert_subject') or '').strip()
    validacao_link=f"{request.url_root.rstrip('/')}/assinatura/validar/{ass_codigo}" if ass_codigo else ''
    svcs=d.get('svcs',d.get('servicos',[]));
    if isinstance(svcs,str):
        try: svcs=json.loads(svcs)
        except: svcs=[]
    sub=sum(float(s.get('vtot',0)) for s in svcs)
    por=d.get('criado_por',session.get('nome','')); now=localnow()

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=2*cm)
    W=A4[0]-3*cm

    AZ=colors.HexColor('#205d8a'); LJ=colors.HexColor('#f28e34')
    VD=colors.HexColor('#1a7a45'); VDC=colors.HexColor('#d8f0e5'); CI=colors.HexColor('#f5f5f5')

    def ps(nm,**kw):
        b=dict(fontName='Helvetica',fontSize=10,leading=14,textColor=colors.HexColor('#020202'),spaceAfter=0,spaceBefore=0)
        b.update(kw); return ParagraphStyle(nm,**b)

    story=[]

    empresas_hdr=_pdf_companies_for_header(empresa_dict=emp,limit=2)
    def _logo_flowable_hdr(item,w=3.6*cm,h=1.4*cm):
        for cand in (item.get('logos') or []):
            try:
                if isinstance(cand,str) and cand.startswith(('http://','https://')):
                    req=urllib.request.Request(cand,headers={'User-Agent':'Mozilla/5.0'})
                    with urllib.request.urlopen(req,timeout=8) as r:
                        img_data=r.read()
                    return Image(io.BytesIO(img_data),width=w,height=h,kind='proportional')
                if os.path.exists(cand):
                    return Image(cand,width=w,height=h,kind='proportional')
            except Exception:
                continue
        return Paragraph(f'<b>{item.get("nome") or enome}</b>',ps('lgfbm',fontSize=10,textColor=AZ))

    # Cabeçalho
    lp=get_logo()
    lc=Paragraph(f'<b>{enome}</b>',ps('lg',fontSize=13,textColor=AZ))
    for cand in [emp.get('logo_url'),lp,LOGO_URL]:
        if not cand: continue
        try:
            if isinstance(cand,str) and cand.startswith(('http://','https://')):
                req=urllib.request.Request(cand,headers={'User-Agent':'Mozilla/5.0'})
                with urllib.request.urlopen(req,timeout=8) as r: img_data=r.read()
                lc=Image(io.BytesIO(img_data),width=4.5*cm,height=1.8*cm,kind='proportional')
            elif os.path.exists(cand):
                lc=Image(cand,width=4.5*cm,height=1.8*cm,kind='proportional')
            break
        except Exception:
            continue

    hdr=Table([[lc,Paragraph(f'<font size="13" color="white"><b>{tipo.upper()}</b></font><br/><font size="9" color="#AADDFF">Nº {nmed}</font>',ps('hr',alignment=TA_CENTER))]],colWidths=[W*0.55,W*0.45])
    hdr.setStyle(TableStyle([('BACKGROUND',(1,0),(1,0),AZ),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10),('LEFTPADDING',(1,0),(1,-1),12),('LEFTPADDING',(0,0),(0,-1),0)]))
    story.append(hdr)

    if ass_status=='assinado':
        ass_quando='' if not ass_em else (ass_em.strftime('%d/%m/%Y %H:%M') if isinstance(ass_em,datetime) else str(ass_em))
        selo_txt=f'<b>ASSINADO DIGITALMENTE</b>'
        if ass_quando:
            selo_txt+=f' · {ass_quando}'
        if ass_codigo:
            selo_txt+=f' · Código {ass_codigo}'
        selo=Table([[Paragraph(selo_txt,ps('sigstamp',fontSize=8,textColor=colors.HexColor('#155f3b'),alignment=TA_CENTER))]],colWidths=[W])
        selo.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#ecf8f0')),('BOX',(0,0),(-1,-1),0.8,colors.HexColor('#9ed3b1')),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8)]))
        story.append(selo)
        story.append(Spacer(1,4))

    bar=Table([[' ']],colWidths=[W])
    bar.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(bar); story.append(Spacer(1,6))

    emp_cells=[]
    for i,item in enumerate(empresas_hdr[:2]):
        cell=Table([
            [_logo_flowable_hdr(item)],
            [Paragraph(f'<b>{item.get("nome") or "-"}</b><br/><font size="8" color="#4c6072">CNPJ: {item.get("cnpj") or "-"}</font>',ps(f'empm{i}',fontSize=8.2,leading=10))]
        ],colWidths=[W*0.49])
        cell.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#d0d7df')),
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fbff')),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        emp_cells.append(cell)
    while len(emp_cells)<2:
        emp_cells.append(Paragraph('',ps('empemptym',fontSize=1)))
    emp_tbl=Table([emp_cells],colWidths=[W*0.495,W*0.495])
    emp_tbl.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(emp_tbl); story.append(Spacer(1,6))

    def campo(lbl,val):
        return [Paragraph(f'<b><font color="#205d8a">{lbl}</font></b>',ps('lb',fontSize=8)),Paragraph(str(val or '—'),ps('vl',fontSize=9))]

    flat=[]
    for row in [[campo('Cliente:',cname),campo('Mês ref.:',fmt_mes(mes))],[campo('CNPJ/CPF:',ccnpj),campo('Emissão:',fmt_data(dtem))],[campo('Endereço:',cend),campo('Vencimento:',fmt_data(dtvenc))],[campo('Responsável:',cresp),campo('Empresa prestadora:',enome)]]:
        flat.append([row[0][0],row[0][1],row[1][0],row[1][1]])

    inf=Table(flat,colWidths=[W*0.17,W*0.33,W*0.17,W*0.33])
    inf.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('BACKGROUND',(2,0),(2,-1),CI),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
    story.append(inf); story.append(Spacer(1,8))

    story.append(Paragraph('<b><font color="white">  SERVIÇOS REALIZADOS NO PERÍODO</font></b>',ps('sh',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
    story.append(Spacer(1,4))

    th=[Paragraph(f'<b>{h}</b>',ps('th',textColor=colors.white,fontSize=9,alignment=a)) for h,a in [('Descrição',TA_LEFT),('Unid.',TA_CENTER),('Qtd.',TA_CENTER),('Valor unit.',TA_RIGHT),('Valor total',TA_RIGHT)]]
    sr=[th]
    for i,s in enumerate(svcs):
        desc=(s.get('desc','') or '').strip()
        tags=[]
        if bool(s.get('fornece_materiais')):
            tags.append('com materiais')
        if bool(s.get('fornece_equipamentos')):
            tags.append('com equipamentos')
        if tags:
            desc=f"{desc} ({', '.join(tags)})"
        sr.append([Paragraph(desc,ps('td',fontSize=9)),Paragraph(str(s.get('unid','')),ps('tc',fontSize=9,alignment=TA_CENTER)),Paragraph(str(s.get('qtd',1)),ps('tc2',fontSize=9,alignment=TA_CENTER)),Paragraph(fmt_brl(s.get('vun',0)),ps('tr',fontSize=9,alignment=TA_RIGHT)),Paragraph(fmt_brl(s.get('vtot',0)),ps('tv',fontSize=9,alignment=TA_RIGHT,textColor=VD))])

    st2=TableStyle([('BACKGROUND',(0,0),(-1,0),AZ),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),('LINEBELOW',(0,0),(-1,-1),0.3,colors.HexColor('#DDDDDD')),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))])
    for i in range(1,len(sr)):
        if i%2==0: st2.add('BACKGROUND',(0,i),(-1,i),CI)

    svc_tbl=Table(sr,colWidths=[W*0.46,W*0.09,W*0.09,W*0.18,W*0.18])
    svc_tbl.setStyle(st2); story.append(svc_tbl); story.append(Spacer(1,6))

    tot=Table([[Paragraph('<b>VALOR TOTAL A RECEBER:</b>',ps('tl',fontSize=13,textColor=VD)),Paragraph(f'<b>{fmt_brl(sub)}</b>',ps('tv2',fontSize=15,alignment=TA_RIGHT,textColor=VD))]],colWidths=[W*0.68,W*0.32])
    tot.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),VDC),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10),('LEFTPADDING',(0,0),(0,-1),12),('RIGHTPADDING',(-1,0),(-1,-1),12),('BOX',(0,0),(-1,-1),1.5,VD)]))
    story.append(tot); story.append(Spacer(1,10))

    pags=[]
    if epix: pags.append(('Chave Pix:',epix))
    if ebanco: pags.append(('Banco:',f'{ebanco}{" — Ag: "+eagencia if eagencia else ""}{" / Cc: "+econta if econta else ""}'))
    if pags:
        story.append(Paragraph('<b><font color="white">  DADOS PARA PAGAMENTO</font></b>',ps('ph',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        pr=[[Paragraph(f'<b>{l}</b>',ps('pl',fontSize=9,textColor=AZ)),Paragraph(v,ps('pv',fontSize=10))] for l,v in pags]
        pt=Table(pr,colWidths=[W*0.22,W*0.78])
        pt.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD'))]))
        story.append(pt)
        # QR Code PIX
        if epix:
            try:
                qr_pix=qr_code.QrCodeWidget(epix)
                b_pix=qr_pix.getBounds()
                bw_pix=max(1,(b_pix[2]-b_pix[0])); bh_pix=max(1,(b_pix[3]-b_pix[1]))
                qr_sz=80
                qr_pix_draw=Drawing(qr_sz,qr_sz,transform=[qr_sz/bw_pix,0,0,qr_sz/bh_pix,0,0])
                qr_pix_draw.add(qr_pix)
                txt_pix=Paragraph('Escaneie para<br/>pagar via PIX',ps('qrpixtxt',fontSize=8,leading=11,alignment=TA_CENTER,textColor=AZ))
                qr_row=Table([[qr_pix_draw,txt_pix]],colWidths=[qr_sz*0.5*cm,W-qr_sz*0.5*cm])
                qr_row.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),8)]))
                story.append(qr_row)
            except Exception:
                pass
        story.append(Spacer(1,10))

    if eboleto:
        story.append(Paragraph('<b><font color="white">  INFORMAÇÕES DE BOLETO</font></b>',ps('bh',fontSize=10,backColor=AZ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        bt=Table([[Paragraph(eboleto,ps('bv',fontSize=9,leading=14))]],colWidths=[W])
        bt.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
        story.append(bt); story.append(Spacer(1,10))

    if obs:
        story.append(Paragraph('<b><font color="white">  OBSERVAÇÕES</font></b>',ps('oh',fontSize=10,backColor=LJ,leading=22,leftIndent=6)))
        story.append(Spacer(1,4))
        ot=Table([[Paragraph(obs,ps('ov',fontSize=9,leading=14))]],colWidths=[W])
        ot.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(-1,-1),8),('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC'))]))
        story.append(ot); story.append(Spacer(1,16))

    # Assinaturas
    story.append(Spacer(1,24))
    at=Table([[Paragraph(f'_________________________________<br/><font size="8" color="#888888">{a1 or "Empresa prestadora"}</font>',ps('a1',alignment=TA_CENTER,leading=18)),Paragraph('',ps('ae')),Paragraph(f'_________________________________<br/><font size="8" color="#888888">{a2 or "Cliente / Tomador"}</font>',ps('a2',alignment=TA_CENTER,leading=18))]],colWidths=[W*0.42,W*0.16,W*0.42])
    at.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'BOTTOM')]))
    story.append(at); story.append(Spacer(1,14))

    # Evidencia de assinatura eletronica
    def _fmt_dt(v):
        if isinstance(v,datetime):
            return v.strftime('%d/%m/%Y %H:%M')
        if isinstance(v,str):
            t=v.strip()
            if not t:
                return ''
            try:
                return datetime.fromisoformat(t.replace('Z','+00:00')).strftime('%d/%m/%Y %H:%M')
            except Exception:
                return t
        return ''

    if ass_status=='assinado':
        ass_quem=(ass_nome or a2 or 'Assinante')
        if ass_cargo:
            ass_quem=f'{ass_quem} - {ass_cargo}'
        ass_quando=_fmt_dt(ass_em)
        linhas=[
            f'<b>Documento assinado eletronicamente</b>',
            f'Assinante: {ass_quem}',
            f'CPF: {ass_cpf or "-"}',
            f'Data/hora: {ass_quando or "-"}',
            f'IP: {ass_ip or "-"}',
            f'Codigo de validacao: {ass_codigo or "-"}',
            (f'Validacao: {validacao_link}' if validacao_link else 'Validacao: indisponivel'),
        ]
        qr_flow=None
        if validacao_link:
            try:
                qr_widget=qr_code.QrCodeWidget(validacao_link)
                b=qr_widget.getBounds()
                bw=max(1,(b[2]-b[0])); bh=max(1,(b[3]-b[1]))
                qr_size=80
                qr_flow=Drawing(qr_size,qr_size,transform=[qr_size/bw,0,0,qr_size/bh,0,0])
                qr_flow.add(qr_widget)
            except Exception:
                qr_flow=None
        texto=Paragraph('<br/>'.join(linhas),ps('asev',fontSize=8,leading=12,textColor=colors.HexColor('#155f3b')))
        if qr_flow is not None:
            ass_box=Table([[texto,qr_flow]],colWidths=[W*0.78,W*0.22])
        else:
            ass_box=Table([[texto]],colWidths=[W])
        ass_box.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#ecf8f0')),('BOX',(0,0),(-1,-1),0.7,colors.HexColor('#9ed3b1')),('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7)]))
        story.append(ass_box); story.append(Spacer(1,10))
    elif ass_status in ('pendente','expirado'):
        info='Assinatura eletronica pendente.' if ass_status=='pendente' else 'Assinatura eletronica expirada.'
        if ass_status=='pendente' and ass_expira:
            info=f'{info} Expira em {_fmt_dt(ass_expira) or "-"}.'
        ass_box=Table([[Paragraph(info,ps('aspd',fontSize=8,leading=11,textColor=colors.HexColor('#9a5a00')))]],colWidths=[W])
        ass_box.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#fff6e8')),('BOX',(0,0),(-1,-1),0.7,colors.HexColor('#f0c98a')),('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
        story.append(ass_box); story.append(Spacer(1,10))

    # Rodapé
    bar2=Table([[' ']],colWidths=[W])
    bar2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LJ),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(bar2); story.append(Spacer(1,4))
    if eend or esite:
        rodape_end=f'<b>{enome}</b>'
        if eend: rodape_end+=f' — {eend}'
        if esite: rodape_end+=f' | {esite}'
        story.append(Paragraph(rodape_end,ps('re',fontSize=8,textColor=colors.HexColor('#555'),alignment=TA_CENTER)))
        story.append(Spacer(1,3))
    story.append(Paragraph(f'Emitido em {now.strftime("%d/%m/%Y")} às {now.strftime("%H:%M")} por {por} — {enome} · {tipo} Nº {nmed} · {fmt_mes(mes)}',ps('rod',fontSize=7,textColor=colors.HexColor('#999'),alignment=TA_CENTER)))

    # Pagina extra de auditoria da assinatura
    if ass_status in ('assinado','pendente','expirado') or ass_codigo:
        def _status_label(v):
            s=(v or '').strip().lower()
            if s=='assinado':
                return 'Assinado'
            if s=='pendente':
                return 'Pendente'
            if s=='expirado':
                return 'Expirado'
            return 'Nao solicitado'

        status_txt=_status_label(ass_status)
        trilha='|'.join([
            str(nmed or ''),str(tipo or ''),str(cname or ''),str(ccnpj or ''),str(mes or ''),
            str(ass_status or ''),str(ass_nome or ''),str(ass_cpf or ''),str(ass_cargo or ''),str(ass_ip or ''),
            str(_fmt_dt(ass_em) or ''),str(ass_codigo or '')
        ])
        hash_comp=ass_doc_hash or hashlib.sha256(trilha.encode('utf-8')).hexdigest().upper()

        story.append(PageBreak())
        story.append(Paragraph('AUDITORIA DA ASSINATURA ELETRONICA',ps('aud_tit',fontSize=14,textColor=AZ,alignment=TA_CENTER)))
        story.append(Spacer(1,8))
        story.append(Paragraph('Esta pagina registra os metadados da assinatura para conferencia e validacao.',ps('aud_sub',fontSize=9,textColor=colors.HexColor('#4c6072'),alignment=TA_CENTER)))
        story.append(Spacer(1,10))

        badge_cor=colors.HexColor('#ecf8f0') if ass_status=='assinado' else (colors.HexColor('#fff6e8') if ass_status=='pendente' else colors.HexColor('#fff1f1'))
        badge_txt=colors.HexColor('#155f3b') if ass_status=='assinado' else (colors.HexColor('#9a5a00') if ass_status=='pendente' else colors.HexColor('#9f1f1f'))
        st=Table([[Paragraph(f'<b>Status da assinatura: {status_txt}</b>',ps('aud_st',fontSize=10,textColor=badge_txt))]],colWidths=[W])
        st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),badge_cor),('BOX',(0,0),(-1,-1),0.7,colors.HexColor('#d0d7df')),('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7)]))
        story.append(st); story.append(Spacer(1,10))

        detalhes=[
            ('Documento',f'{tipo} Nº {nmed}'),
            ('Cliente',cname or '-'),
            ('Empresa',enome or '-'),
            ('Competencia',fmt_mes(mes) if mes else '-'),
            ('Assinante',ass_nome or '-'),
            ('CPF do assinante',ass_cpf or '-'),
            ('Cargo',ass_cargo or '-'),
            ('Data/Hora assinatura',_fmt_dt(ass_em) or '-'),
            ('IP de origem',ass_ip or '-'),
            ('Codigo de validacao',ass_codigo or '-'),
            ('Link de validacao',validacao_link or '-'),
            ('Assinatura criptográfica PDF',('Ativa' if ass_crypto_ok else 'Não aplicada')),
            ('Certificado digital',ass_cert_subject or '-'),
            ('Hash de comprovacao (SHA-256)',hash_comp),
            ('Emitido por',por or '-'),
            ('Data/Hora da emissao',now.strftime('%d/%m/%Y %H:%M')),
        ]
        rows=[[Paragraph(f'<b>{k}</b>',ps('aud_k',fontSize=8,textColor=AZ)),Paragraph(v,ps('aud_v',fontSize=8,leading=11))] for k,v in detalhes]
        det=Table(rows,colWidths=[W*0.30,W*0.70])
        det.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),CI),('BOX',(0,0),(-1,-1),0.6,colors.HexColor('#d0d7df')),('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#e3e8ef')),('LEFTPADDING',(0,0),(-1,-1),7),('RIGHTPADDING',(0,0),(-1,-1),7),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
        story.append(det); story.append(Spacer(1,10))

        if validacao_link:
            try:
                qr_widget=qr_code.QrCodeWidget(validacao_link)
                b=qr_widget.getBounds()
                bw=max(1,(b[2]-b[0])); bh=max(1,(b[3]-b[1]))
                qr_size=130
                qr_draw=Drawing(qr_size,qr_size,transform=[qr_size/bw,0,0,qr_size/bh,0,0])
                qr_draw.add(qr_widget)
                qr_tbl=Table([[qr_draw,Paragraph('Escaneie o QR Code para validar esta assinatura em tempo real no portal RM Facilities.',ps('aud_qr',fontSize=9,leading=13,textColor=colors.HexColor('#4c6072')))]],colWidths=[W*0.28,W*0.72])
                qr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4)]))
                story.append(qr_tbl)
            except Exception:
                story.append(Paragraph(f'Link de validacao: {validacao_link}',ps('aud_qrf',fontSize=8,textColor=colors.HexColor('#4c6072'))))

    doc.build(story)
    pdf_bytes=buf.getvalue()
    crypto_ok=False
    cert_subject=''
    if ass_status=='assinado':
        rs_crypto=_try_sign_pdf_bytes_crypto(pdf_bytes,empresa_id=(to_num(d.get('empresa_id')) or None),usuario_id=session.get('uid'))
        pdf_bytes=rs_crypto.get('bytes') or pdf_bytes
        crypto_ok=bool(rs_crypto.get('ok'))
        cert_subject=(rs_crypto.get('cert_subject') or '')[:255] if crypto_ok else ''
        mid=to_num(d.get('id'))
        if mid:
            med=Medicao.query.get(mid)
            if med:
                med.assinatura_doc_hash=_sha256_bytes(pdf_bytes)
                med.assinatura_crypto_ok=crypto_ok
                med.assinatura_cert_subject=(cert_subject or None)
                db.session.commit()
    buf=io.BytesIO(pdf_bytes)
    buf.seek(0)
    slug=(cname or 'cliente').replace(' ','_')[:20]
    return send_file(buf,mimetype='application/pdf',as_attachment=False,download_name=f'{tipo.replace(" ","_")}_{nmed.replace("/","-")}_{slug}_{mes}.pdf')

@app.route('/api/rh/holerites/processar',methods=['POST'])
@lr
def api_rh_holerites_processar():
    fs=request.files.get('arquivo')
    comp=(request.form.get('competencia') or '').strip()
    if not fs: return jsonify({'erro':'PDF nao enviado'}),400
    canal_ass=(request.form.get('canal_assinatura') or 'nao').strip().lower()
    if canal_ass not in ('nao','whatsapp','link','app'):
        canal_ass='nao'
    prazo_dias_lote_raw=request.form.get('prazo_dias') or ''
    try:
        prazo_dias_lote=max(1,min(90,int(prazo_dias_lote_raw))) if prazo_dias_lote_raw.strip() else None
    except (ValueError, TypeError):
        prazo_dias_lote=None
    nome_arq=(fs.filename or '').strip().lower()
    if nome_arq and not nome_arq.endswith('.pdf'):
        return jsonify({'erro':'Arquivo invalido. Envie um PDF (.pdf).'}),400
    try:
        from pypdf import PdfReader,PdfWriter
    except Exception:
        return jsonify({'erro':'Dependencia pypdf nao instalada'}),500
    funcs_todos=Funcionario.query.all()
    if not funcs_todos: return jsonify({'erro':'Cadastre funcionarios primeiro'}),400
    pdf_bytes=fs.read()
    if not pdf_bytes or len(pdf_bytes)<8:
        return jsonify({'erro':'Arquivo vazio ou corrompido.'}),400
    if pdf_bytes[:4]!=b'%PDF':
        return jsonify({'erro':'Arquivo invalido. O conteúdo enviado nao parece ser um PDF.'}),400
    try:
        reader=PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return jsonify({'erro':'PDF invalido ou corrompido. Gere/exporte o arquivo novamente e tente de novo.'}),400
    def _norm_comp(v):
        return ''.join(ch for ch in (v or '') if ch.isalnum()).lower()

    def _folha_ponto_for_func(func_id, competencia):
        q=FuncionarioArquivo.query.filter_by(funcionario_id=func_id,categoria='folha_ponto').order_by(FuncionarioArquivo.id.desc()).all()
        if not q:
            return None
        if competencia:
            alvo=_norm_comp(competencia)
            for a in q:
                if _norm_comp(a.competencia)==alvo:
                    return a
        return q[0]

    itens=[]; sem_match=[]
    for idx,page in enumerate(reader.pages,start=1):
        txt=(page.extract_text() or '')
        alvo=find_funcionario_in_text(txt,funcs_todos)
        if not alvo: sem_match.append(idx); continue
        ano=infer_doc_year(comp); prepare_func_doc_dirs(alvo.id,ano)
        writer=PdfWriter(); writer.add_page(page)
        fake_name=holerite_batch_filename(alvo,comp)
        subdir,_=func_doc_subdir(alvo.id,'holerite',comp)
        rel,abs_p,fake_name=unique_rel_filename(subdir,fake_name)
        os.makedirs(os.path.dirname(abs_p),exist_ok=True)
        with open(abs_p,'wb') as out: writer.write(out)
        a=FuncionarioArquivo(funcionario_id=alvo.id,categoria='holerite',competencia=comp,nome_arquivo=fake_name,caminho=rel)
        db.session.add(a); db.session.flush()
        assinatura_auto={'status':'nao_solicitada','link':'','erro':''}
        if canal_ass=='app':
            a.ass_status='pendente'
            if prazo_dias_lote:
                a.ass_prazo_em=utcnow()+timedelta(days=prazo_dias_lote)
            db.session.flush()
            _push_notify_funcionario(
                alvo.id,
                'Documento para assinar',
                f'{fake_name} aguarda sua assinatura no app.',
                {'tipo':'documento_assinar','arquivo_id':str(a.id)}
            )
            assinatura_auto={'status':'app_pendente','link':'','erro':''}
        elif canal_ass=='whatsapp':
            tel=wa_norm_number(alvo.telefone or '')
            if wa_is_valid_number(tel):
                rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='whatsapp',commit_now=True)
                if rs.get('ok'):
                    assinatura_auto={'status':'solicitada','link':rs.get('link',''),'erro':''}
                else:
                    assinatura_auto={'status':'erro','link':rs.get('link',''),'erro':rs.get('erro') or 'Falha ao solicitar assinatura.'}
            elif (alvo.telefone or '').strip():
                assinatura_auto={'status':'telefone_invalido','link':'','erro':'Telefone cadastrado sem formato WhatsApp valido.'}
            else:
                assinatura_auto={'status':'sem_telefone','link':'','erro':'Funcionario sem telefone cadastrado.'}
        elif canal_ass=='link':
            rs=_solicitar_assinatura_arquivo_funcionario(a,alvo,canal='link',commit_now=True)
            if rs.get('ok'):
                assinatura_auto={'status':'link_gerado','link':rs.get('link',''),'erro':''}
            else:
                assinatura_auto={'status':'erro','link':rs.get('link',''),'erro':rs.get('erro') or 'Falha ao gerar link de assinatura.'}
        folha_ponto=_folha_ponto_for_func(alvo.id,comp)
        whatsapp_num=wa_norm_number(alvo.telefone or '')
        wa_habilitado=funcionario_docs_whatsapp_habilitado(alvo)
        itens.append({'pagina':idx,'funcionario_id':alvo.id,'funcionario_nome':alvo.nome,'arquivo_id':a.id,'nome_arquivo':fake_name,'caminho':rel,'abs_caminho':abs_p,'email':alvo.email or '','whatsapp':(whatsapp_num if wa_is_valid_number(whatsapp_num) else ''),'whatsapp_habilitado':wa_habilitado,'folha_ponto_id':(folha_ponto.id if folha_ponto else None),'folha_ponto_nome':(folha_ponto.nome_arquivo if folha_ponto else ''),'folha_ponto_caminho':(folha_ponto.caminho if folha_ponto else ''),'status_envio':None,'erro_envio':None,'assinatura_auto_status':assinatura_auto.get('status'),'assinatura_auto_link':assinatura_auto.get('link'),'assinatura_auto_erro':assinatura_auto.get('erro')})
    db.session.commit()
    job_id=secrets.token_hex(16)
    _holerite_jobs[job_id]={'id':job_id,'status':'pronto','total_paginas':len(reader.pages),'itens':itens,'sem_match':sem_match,'competencia':comp,'criado_em':utcnow().isoformat()}
    itens_resp=[{k:v for k,v in it.items() if k!='abs_caminho'} for it in itens]
    return jsonify({'ok':True,'job_id':job_id,'total_paginas':len(reader.pages),'separados':len(itens),'sem_match':sem_match,'itens':itens_resp,'canal_assinatura':canal_ass})

@app.route('/api/rh/holerites/job/<job_id>')
@lr
def api_rh_holerites_job(job_id):
    job=_holerite_jobs.get(job_id)
    if not job: return jsonify({'erro':'Job nao encontrado'}),404
    out=dict(job); out['itens']=[{k:v for k,v in it.items() if k!='abs_caminho'} for it in out.get('itens',[])]
    return jsonify({'ok':True,'job':out})

@app.route('/api/rh/holerites/enviar/<job_id>',methods=['POST'])
@lr
def api_rh_holerites_enviar(job_id):
    job=_holerite_jobs.get(job_id)
    if not job: return jsonify({'erro':'Job nao encontrado. Processe o PDF novamente.'}),404
    d=request.json or {}; canal=d.get('canal','email')
    incluir_folha_ponto=bool(d.get('incluir_folha_ponto',False))
    ids_sel=set(int(x) for x in (d.get('arquivo_ids') or []))
    def do_send():
        with app.app_context():
            for item in job['itens']:
                if ids_sel and item.get('arquivo_id') not in ids_sel: continue
                try:
                    abs_p=item.get('abs_caminho') or os.path.join(UPLOAD_ROOT,item['caminho'])
                    fp_caminho=(item.get('folha_ponto_caminho') or '').strip()
                    fp_abs=os.path.join(UPLOAD_ROOT,fp_caminho) if fp_caminho else ''
                    tem_fp=bool(incluir_folha_ponto and fp_abs and os.path.exists(fp_abs))
                    comp=job.get('competencia',''); fn=item.get('funcionario_nome','')
                    s_e=False; s_w=False
                    if canal in ('email','ambos') and item.get('email'):
                        smtp_send_pdf(item['email'],fn,abs_p,item['nome_arquivo'],comp)
                        if tem_fp:
                            smtp_send_pdf(item['email'],fn,fp_abs,item.get('folha_ponto_nome') or 'folha_ponto.pdf',comp)
                        s_e=True
                    if canal in ('whatsapp','ambos') and item.get('whatsapp_habilitado') and item.get('whatsapp'):
                        wa_send_pdf(item['whatsapp'],abs_p,item['nome_arquivo'],f"Holerite {comp} - {fn}")
                        if tem_fp:
                            wa_send_pdf(item['whatsapp'],fp_abs,item.get('folha_ponto_nome') or 'folha_ponto.pdf',f"Folha de ponto {comp} - {fn}")
                        s_w=True
                    if s_e and s_w: item['status_envio']='enviado_ambos'
                    elif s_e: item['status_envio']='enviado_email'
                    elif s_w: item['status_envio']='enviado_wa'
                    elif canal=='link': item['status_envio']='link_disponivel'
                    else: item['status_envio']='sem_contato'
                    item['erro_envio']=None
                except Exception as e:
                    item['erro_envio']=str(e); item['status_envio']='erro'
            job['status']='concluido'
    job['status']='enviando'
    threading.Thread(target=do_send,daemon=True).start()
    return jsonify({'ok':True,'job_id':job_id,'status':'enviando'})

@app.route('/api/config/smtp',methods=['GET'])
@dr
def api_smtp_get():
    senha=(gc('smtp_senha','') or '').strip()
    return jsonify({
        'host':gc('smtp_host',''),
        'port':gc('smtp_port','587'),
        'user':gc('smtp_user',''),
        'senha':'',
        'has_senha':bool(senha),
        'de':gc('smtp_de',''),
        'tls':gc('smtp_tls','1')
    })

@app.route('/api/config/smtp',methods=['POST'])
@dr
def api_smtp_save():
    d=request.json or {}
    for k in ['host','port','user','de','tls']:
        if k in d: sc_cfg(f'smtp_{k}',str(d[k]))
    if str(d.get('senha_clear','0')).strip().lower() in ('1','true','yes','on'):
        sc_cfg('smtp_senha','')
    elif 'senha' in d and str(d.get('senha','')).strip():
        sc_cfg('smtp_senha',str(d.get('senha','')).strip())
    return jsonify({'ok':True})

@app.route('/api/config/smtp/testar',methods=['POST'])
@dr
def api_smtp_testar():
    d=request.json or {}; dest=(d.get('email') or gc('smtp_user','')).strip()
    if not dest: return jsonify({'erro':'Informe o e-mail de destino'}),400
    try:
        cfg=smtp_cfg()
        if not cfg['host'] or not cfg['user']: return jsonify({'erro':'SMTP nao configurado'}),400
        msg=MIMEMultipart(); msg['From']=cfg['de'] or cfg['user']; msg['To']=dest
        msg['Subject']='Teste de envio — RM Facilities'
        msg.attach(MIMEText('E-mail de teste enviado com sucesso pelo sistema RM Facilities!','plain','utf-8'))
        port=int(cfg['port'] or 587)
        if str(cfg['tls']) in ('1','true','True'):
            with smtplib.SMTP(cfg['host'],port,timeout=20) as s: s.starttls(); s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
        else:
            with smtplib.SMTP_SSL(cfg['host'],port,timeout=20) as s: s.login(cfg['user'],cfg['senha']); s.sendmail(cfg['de'] or cfg['user'],dest,msg.as_string())
        return jsonify({'ok':True,'mensagem':f'E-mail enviado para {dest}'})
    except Exception as e: return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp',methods=['GET'])
@lr
def api_wa_cfg_get():
    b=wa_backup_cfg()
    token=(gc('wa_token','') or '').strip()
    return jsonify({
        'url':gc('wa_url',''),
        'instancia':gc('wa_instancia',''),
        'token':'',
        'has_token':bool(token),
        'backup_enabled':b.get('enabled','1'),
        'backup_email':b.get('email',''),
        'backup_interval_hours':b.get('interval_hours','2'),
        'backup_window_hours':b.get('window_hours','8'),
        'backup_max_conversas':b.get('max_conversas','10'),
    })

@app.route('/api/config/whatsapp',methods=['POST'])
@lr
def api_wa_cfg_save():
    d=request.json or {}
    for k in ['url','instancia']:
        if k in d: sc_cfg(f'wa_{k}',str(d[k]))
    if str(d.get('token_clear','0')).strip().lower() in ('1','true','yes','on'):
        sc_cfg('wa_token','')
    elif 'token' in d and str(d.get('token','')).strip():
        sc_cfg('wa_token',str(d.get('token','')).strip())
    if 'backup_enabled' in d:
        sc_cfg('wa_backup_enabled','1' if str(d.get('backup_enabled','0')).strip().lower() in ('1','true','yes','on') else '0')
    if 'backup_email' in d:
        sc_cfg('wa_backup_email',str(d.get('backup_email','')).strip())
    if 'backup_interval_hours' in d:
        sc_cfg('wa_backup_interval_hours',str(max(1,min(168,_to_int(d.get('backup_interval_hours'),2)))))
    if 'backup_window_hours' in d:
        sc_cfg('wa_backup_window_hours',str(max(1,min(168,_to_int(d.get('backup_window_hours'),8)))))
    if 'backup_max_conversas' in d:
        sc_cfg('wa_backup_max_conversas',str(max(1,min(50,_to_int(d.get('backup_max_conversas'),10)))))
    return jsonify({'ok':True})

@app.route('/api/config/whatsapp/backup/testar',methods=['POST'])
@lr
def api_wa_backup_testar():
    try:
        r=wa_backup_maybe_send(force=True)
        if not r.get('ok'):
            return jsonify({'erro':f"Backup nao enviado: {r.get('skip','falha')}"}),400
        return jsonify({'ok':True,'mensagem':f"Backup enviado para {r.get('email','')}",'arquivo':r.get('arquivo',''),'total_conversas':r.get('total_conversas',0),'total_mensagens':r.get('total_mensagens',0)})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/backup/gerar',methods=['POST'])
@lr
def api_wa_backup_gerar():
    try:
        cfg=wa_backup_cfg()
        payload=_wa_backup_collect(cfg.get('window_hours','8'),cfg.get('max_conversas','10'))
        arq=_wa_backup_store(payload)
        return jsonify({'ok':True,'arquivo':arq,'payload':payload})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/backup/restaurar',methods=['POST'])
@lr
def api_wa_backup_restaurar():
    arq=request.files.get('arquivo')
    if not arq:
        return jsonify({'erro':'Arquivo JSON do backup nao enviado'}),400
    try:
        payload=json.loads(arq.read().decode('utf-8'))
    except Exception:
        return jsonify({'erro':'JSON invalido'}),400
    restore_config=str(request.form.get('restore_config','1')).strip().lower() in ('1','true','yes','on')
    restore_conversas=str(request.form.get('restore_conversas','1')).strip().lower() in ('1','true','yes','on')
    try:
        st=wa_backup_restore_payload(payload,restore_config=restore_config,restore_conversas=restore_conversas)
        return jsonify({'ok':True,'restaurado':st})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro':str(e)}),500

@app.route('/api/config/whatsapp/testar',methods=['POST'])
@lr
def api_wa_testar():
    d=request.json or {}; numero=(d.get('numero') or '').strip()
    if not numero: return jsonify({'erro':'Informe o numero de destino'}),400
    try:
        wa_send_text(numero,'Teste de conexao WhatsApp — RM Facilities')
        return jsonify({'ok':True})
    except Exception as e: return jsonify({'erro':str(e)}),500

@app.route('/api/config/ia-whatsapp',methods=['GET'])
@lr
def api_ia_wa_cfg_get():
    d=ai_wa_cfg()
    return jsonify({
        'enabled':d['enabled'],
        'provider':d['provider'],
        'api_key':'',
        'has_api_key':bool((d.get('api_key') or '').strip()),
        'model':d['model'],
        'prompt':(d['prompt'] or DEFAULT_IA_WA_PROMPT),
        'temperature':d['temperature'],
        'max_tokens':d['max_tokens']
    })

@app.route('/api/config/ia-whatsapp',methods=['POST'])
@lr
def api_ia_wa_cfg_save():
    d=request.json or {}
    enabled='1' if str(d.get('enabled','0')).strip().lower() in ('1','true','yes','on') else '0'
    provider=ai_provider_norm(d.get('provider','gemini'))
    raw_model=str(d.get('model','') or '').strip()
    model=ai_model_norm(provider,raw_model) if raw_model else (gc('ia_wa_model','') or ai_model_norm(provider,''))
    try:
        temp=max(0.0,min(1.0,float(d.get('temperature',0.3))))
    except Exception:
        temp=0.3
    try:
        max_tokens=max(50,min(2000,int(float(d.get('max_tokens',350)))))
    except Exception:
        max_tokens=350
    sc_cfg('ia_wa_enabled',enabled)
    sc_cfg('ia_wa_provider',provider)
    if str(d.get('api_key_clear','0')).strip().lower() in ('1','true','yes','on'):
        sc_cfg('ia_wa_api_key','')
    elif 'api_key' in d and str(d.get('api_key','')).strip():
        sc_cfg('ia_wa_api_key',str(d.get('api_key','')).strip())
    sc_cfg('ia_wa_model',model)
    sc_cfg('ia_wa_prompt',str(d.get('prompt','')))
    sc_cfg('ia_wa_temperature',str(temp))
    sc_cfg('ia_wa_max_tokens',str(max_tokens))
    warn=''
    if raw_model and raw_model!=model:
        warn=f'Modelo ajustado automaticamente para: {model}'
    return jsonify({'ok':True,'provider':provider,'model':model,'temperature':temp,'max_tokens':max_tokens,'warning':warn})

@app.route('/api/whatsapp/ia/testar',methods=['POST'])
@lr
def api_wa_ia_testar():
    d=request.json or {}
    txt=(d.get('texto') or '').strip()
    numero=wa_norm_number(d.get('numero') or '5511999999999')
    if not txt: return jsonify({'erro':'Informe o texto para teste'}),400
    try:
        resp=ai_wa_reply(numero,txt)
        if not resp: return jsonify({'erro':'A IA nao retornou resposta'}),400
        return jsonify({'ok':True,'resposta':resp})
    except Exception as e:
        return jsonify({'erro':str(e)}),500

@app.route('/api/whatsapp/ia/retomar',methods=['POST'])
@lr
def api_wa_ia_retomar():
    d=request.json or {}
    numero=wa_norm_number(d.get('numero') or '')
    if not numero:
        return jsonify({'erro':'numero obrigatorio'}),400
    if not wa_is_valid_number(numero):
        return jsonify({'erro':'numero invalido'}),400
    wa_ai_resume(numero)
    return jsonify({'ok':True,'numero':numero,'ia_pausada':False})

@app.route('/api/whatsapp/ia/retomar-todos',methods=['POST'])
@lr
def api_wa_ia_retomar_todos():
    agora=utcnow()
    pausados=Config.query.filter(Config.chave.like('wa_ai_pause_until_%')).all()
    count=0
    for cfg in pausados:
        until=_parse_dt_iso(cfg.valor or '')
        if until and until>agora:
            numero=cfg.chave.replace('wa_ai_pause_until_','')
            wa_ai_resume(numero)
            count+=1
    return jsonify({'ok':True,'retomados':count,'mensagem':f'{count} número(s) reativado(s).' if count else 'Nenhum agente estava pausado.'})

@app.route('/api/whatsapp/ia/status')
@lr
def api_wa_ia_status():
    agora=utcnow()
    pausados=Config.query.filter(Config.chave.like('wa_ai_pause_until_%')).all()
    ativos=[]
    for cfg in pausados:
        until=_parse_dt_iso(cfg.valor or '')
        if until and until>agora:
            numero=cfg.chave.replace('wa_ai_pause_until_','')
            c=WhatsAppConversa.query.filter_by(numero=numero).first()
            ativos.append({'numero':numero,'nome':c.nome if c else numero,'pausado_ate':until.isoformat()})
    return jsonify({'pausados':ativos,'total_pausados':len(ativos)})

@app.route('/api/whatsapp/ia/pausar',methods=['POST'])
@lr
def api_wa_ia_pausar():
    d=request.json or {}
    numero=wa_norm_number(d.get('numero') or '')
    horas=_to_int(d.get('horas'),8)
    if not numero:
        return jsonify({'erro':'numero obrigatorio'}),400
    if not wa_is_valid_number(numero):
        return jsonify({'erro':'numero invalido'}),400
    pausa_ate=wa_ai_pause_set(numero,max(1,min(168,horas)))
    return jsonify({'ok':True,'numero':numero,'ia_pausada':True,'ia_pausada_ate':(pausa_ate.isoformat() if pausa_ate else '')})

@app.route('/api/whatsapp/conversas')
@lr
def api_wa_conversas():
    lst=WhatsAppConversa.query.order_by(WhatsAppConversa.ultima_msg.desc()).all()
    return jsonify([c.to_dict() for c in lst])

@app.route('/api/whatsapp/conversas/<numero>')
@lr
def api_wa_conversa_msgs(numero):
    c=WhatsAppConversa.query.filter_by(numero=numero).first()
    if not c: return jsonify({'conversa':None,'mensagens':[],'ia_pausada':False,'ia_pausada_ate':''})
    msgs=WhatsAppMensagem.query.filter_by(conversa_id=c.id).order_by(WhatsAppMensagem.criado_em).all()
    pausa_ate=wa_ai_pause_until(numero)
    return jsonify({
        'conversa':c.to_dict(),
        'mensagens':[m.to_dict() for m in msgs],
        'ia_pausada':bool(pausa_ate and pausa_ate>utcnow()),
        'ia_pausada_ate':pausa_ate.isoformat() if pausa_ate else '',
    })

@app.route('/api/whatsapp/send',methods=['POST'])
@lr
def api_wa_send():
    d=request.json or {}; numero=(d.get('numero') or '').strip(); texto=(d.get('texto') or '').strip()
    if not numero or not texto: return jsonify({'erro':'numero e texto obrigatorios'}),400
    try: wa_send_text(numero,texto)
    except Exception as e: return jsonify({'erro':str(e)}),500
    c=WhatsAppConversa.query.filter_by(numero=numero).first()
    if not c:
        c=WhatsAppConversa(numero=numero,nome=d.get('nome') or numero)
        db.session.add(c); db.session.flush()
    c.ultima_msg=utcnow()
    db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=texto))
    db.session.commit()
    wa_ai_pause_for(numero, 2)
    return jsonify({'ok':True})

@app.route('/api/whatsapp/send-colaboradores',methods=['POST'])
@lr
def api_wa_send_colaboradores():
    is_form=bool(request.files) or 'multipart/form-data' in (request.content_type or '').lower()
    d=request.form if is_form else (request.json or {})
    texto=(d.get('texto') or '').strip()
    ids_raw=request.form.getlist('funcionario_ids') if is_form else (d.get('funcionario_ids') or [])
    fs=request.files.get('arquivo') if is_form else None
    if not texto and not fs:
        return jsonify({'erro':'Informe uma mensagem ou anexe um arquivo.'}),400
    if not isinstance(ids_raw,list) or not ids_raw:
        return jsonify({'erro':'funcionario_ids obrigatorio'}),400

    arquivo_bytes=b''
    arquivo_nome=''
    arquivo_tipo=''
    arquivo_mimetype=''
    if fs and (fs.filename or '').strip():
        arquivo_nome=secure_filename(fs.filename or 'arquivo') or 'arquivo'
        arquivo_bytes=fs.read()
        if not arquivo_bytes:
            return jsonify({'erro':'Arquivo anexado está vazio.'}),400
        if len(arquivo_bytes)>16*1024*1024:
            return jsonify({'erro':'Arquivo excede o limite de 16 MB.'}),400
        try:
            arquivo_tipo,arquivo_mimetype=wa_media_meta(arquivo_nome,fs.mimetype or '')
        except Exception as e:
            return jsonify({'erro':str(e)}),400

    ids=[]
    for x in ids_raw:
        try:
            ids.append(int(x))
        except Exception:
            continue
    ids=sorted(set(ids))
    if not ids:
        return jsonify({'erro':'Nenhum colaborador valido informado.'}),400

    remetente=(session.get('nome') or 'RM Facilities').strip() or 'RM Facilities'
    prefixo=f'Mensagem enviada por {remetente}:'
    texto_envio=f'{prefixo}\n\n{texto}' if texto else prefixo

    enviados=[]
    falhas=[]
    for func_id in ids:
        f=Funcionario.query.get(func_id)
        if not f:
            falhas.append({'funcionario_id':func_id,'erro':'Colaborador não encontrado.'})
            continue
        if (f.status or '').strip().lower()!='ativo':
            falhas.append({'funcionario_id':f.id,'nome':f.nome,'erro':'Colaborador não está ativo.'})
            continue
        tel=wa_norm_number(f.telefone or '')
        if not wa_is_valid_number(tel):
            falhas.append({'funcionario_id':f.id,'nome':f.nome,'erro':'Telefone cadastrado inválido ou ausente.'})
            continue
        try:
            if arquivo_bytes:
                wa_send_media_bytes(tel,arquivo_bytes,arquivo_nome,arquivo_mimetype,texto_envio)
            else:
                wa_send_text(tel,texto_envio)
            c=WhatsAppConversa.query.filter_by(numero=tel).first()
            if not c:
                c=WhatsAppConversa(numero=tel,nome=f.nome or tel)
                db.session.add(c)
                db.session.flush()
            c.ultima_msg=utcnow()
            conteudo_log=texto_envio
            if arquivo_bytes and not texto:
                conteudo_log=f'{prefixo}\n\n[{arquivo_tipo}] {arquivo_nome}'
            elif arquivo_bytes:
                conteudo_log=f'{texto_envio}\n\n[{arquivo_tipo}] {arquivo_nome}'
            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=tel,direcao='out',tipo=(arquivo_tipo or 'texto'),conteudo=conteudo_log,nome_arquivo=(arquivo_nome or None)))
            db.session.commit()
            wa_ai_pause_for(tel,2)
            enviados.append({'funcionario_id':f.id,'nome':f.nome,'telefone':tel,'arquivo':arquivo_nome})
        except Exception as e:
            db.session.rollback()
            falhas.append({'funcionario_id':f.id,'nome':f.nome,'erro':str(e)})

    return jsonify({'ok':True,'enviados':enviados,'enviados_count':len(enviados),'falhas':falhas})

@app.route('/webhook/whatsapp',methods=['GET','POST'])
def webhook_whatsapp():
    if request.method=='GET':
        return jsonify({'ok':True,'endpoint':'/webhook/whatsapp','metodos':['POST'],'status':'ativo'})
    debug=str(request.args.get('debug','0')).strip().lower() in ('1','true','yes','on')
    diag={'evento':None,'mensagens_recebidas':0,'mensagens_processadas':0,'ia_ativa':ai_wa_enabled(),'respostas_enviadas':0,'erros':[]}
    data=request.json or {}
    try:
        evento=(data.get('event') or '').lower()
        diag['evento']=evento
        raw=data.get('data',{})
        is_message_payload=False
        if isinstance(raw,dict):
            is_message_payload=bool(raw.get('message') or raw.get('body') or raw.get('text'))
        elif isinstance(raw,list):
            for _m in raw:
                if isinstance(_m,dict) and (_m.get('message') or _m.get('body') or _m.get('text')):
                    is_message_payload=True
                    break
        if 'message' in evento or 'upsert' in evento or is_message_payload:
            msgs=[raw] if isinstance(raw,dict) else (raw if isinstance(raw,list) else [])
            diag['mensagens_recebidas']=len(msgs)
            for msg_data in msgs:
                if bool(msg_data.get('key',{}).get('fromMe')) or bool(msg_data.get('fromMe')):
                    continue
                jid=(msg_data.get('key',{}).get('remoteJid') or msg_data.get('sender') or msg_data.get('from') or '')
                if not jid.endswith('@s.whatsapp.net'): continue
                numero=(jid.split('@')[0] if jid else '')
                if not numero: continue
                numero=only_digits(numero) or numero
                if not wa_is_valid_number(numero):
                    diag['erros'].append(f'Numero invalido no webhook: {numero}')
                    continue

                msg_obj=msg_data.get('message',{}) if isinstance(msg_data.get('message',{}),dict) else {}
                audio_obj=msg_obj.get('audioMessage',{}) if isinstance(msg_obj.get('audioMessage',{}),dict) else {}
                tipo_in='audio' if audio_obj else 'texto'
                conteudo=(
                    msg_obj.get('conversation') or
                    msg_obj.get('extendedTextMessage',{}).get('text') or
                    msg_data.get('body') or
                    msg_data.get('text') or
                    msg_data.get('transcription') or
                    msg_data.get('transcript') or
                    audio_obj.get('transcription') or
                    audio_obj.get('transcript') or
                    audio_obj.get('text') or
                    ''
                )
                conteudo=str(conteudo or '').strip()

                if tipo_in=='audio' and not conteudo:
                    try:
                        conteudo=wa_transcribe_audio(msg_data)
                    except Exception as e:
                        diag['erros'].append(f'Falha transcricao audio: {str(e)}')
                        conteudo=''

                if tipo_in=='audio' and not conteudo:
                    # Audio chegou sem transcricao no payload: salva no historico e pede texto.
                    c=WhatsAppConversa.query.filter_by(numero=numero).first()
                    if not c:
                        nome=(msg_data.get('pushName') or msg_data.get('notifyName') or numero)
                        c=WhatsAppConversa(numero=numero,nome=nome)
                        db.session.add(c)
                        db.session.flush()
                    c.ultima_msg=utcnow()
                    db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='in',tipo='audio',conteudo='[audio sem transcricao]'))
                    db.session.commit()

                    cfg_wa=wa_cfg()
                    wa_ready=bool((cfg_wa.get('url') or '').strip() and (cfg_wa.get('instancia') or '').strip())
                    if ai_wa_enabled() and wa_ready and not wa_ai_pause_active(numero):
                        try:
                            aviso='Recebi seu audio, mas ainda nao consegui transcrever automaticamente. Pode me enviar em texto?'
                            wa_send_text(numero,aviso)
                            diag['respostas_enviadas']+=1
                            c.ultima_msg=utcnow()
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=aviso))
                            db.session.commit()
                        except Exception as e:
                            diag['erros'].append(str(e))
                            db.session.rollback()
                    continue

                if not conteudo:
                    continue
                diag['mensagens_processadas']+=1
                c=WhatsAppConversa.query.filter_by(numero=numero).first()
                if not c:
                    nome=(msg_data.get('pushName') or msg_data.get('notifyName') or numero)
                    c=WhatsAppConversa(numero=numero,nome=nome)
                    db.session.add(c); db.session.flush()
                agora=utcnow()
                ultimo_msg_anterior=c.ultima_msg
                # Após 2h sem interação, reinicia o estado do fluxo para começar nova conversa.
                if ultimo_msg_anterior and (agora-ultimo_msg_anterior)>timedelta(hours=2):
                    try:
                        ctx_reset=json.loads(c.contexto or '{}')
                    except Exception:
                        ctx_reset={}
                    ctx_reset['holerite_estado']=None
                    ctx_reset['holerite_tentativas']=0
                    ctx_reset['_reiniciou_inatividade']=True
                    c.contexto=json.dumps(ctx_reset,ensure_ascii=False)
                c.ultima_msg=agora
                db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='in',tipo=tipo_in,conteudo=conteudo))
                db.session.commit()
                # A IA considera apenas as mensagens das últimas 2 horas.
                corte_hist=agora-timedelta(hours=2)
                historico_db=(
                    WhatsAppMensagem.query
                    .filter_by(conversa_id=c.id)
                    .filter(WhatsAppMensagem.criado_em>=corte_hist)
                    .order_by(WhatsAppMensagem.criado_em.desc())
                    .limit(20)
                    .all()
                )
                historico_db.reverse()
                cfg_wa=wa_cfg()
                wa_ready=bool((cfg_wa.get('url') or '').strip() and (cfg_wa.get('instancia') or '').strip())
                if ai_wa_enabled() and not wa_ready:
                    msg_cfg='Auto-reply desativado: WhatsApp nao configurado (url/instancia).'
                    if msg_cfg not in diag['erros']:
                        diag['erros'].append(msg_cfg)
                    app.logger.warning('Auto-reply WhatsApp inativo para %s: configuracao incompleta',numero)
                if ai_wa_enabled() and wa_ready and not wa_ai_pause_active(numero):
                    try:
                        resposta=ai_wa_reply(numero,conteudo,historico=historico_db)
                        if resposta:
                            wa_send_text(numero,resposta)
                            diag['respostas_enviadas']+=1
                            c.ultima_msg=utcnow()
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='texto',conteudo=resposta))
                            db.session.commit()
                        else:
                            diag['erros'].append('IA nao retornou resposta')
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='erro',conteudo='IA nao retornou resposta.'))
                            db.session.commit()
                    except Exception as e:
                        diag['erros'].append(str(e))
                        db.session.rollback()
                        try:
                            app.logger.exception('Falha no auto-reply WhatsApp para %s',numero)
                            db.session.add(WhatsAppMensagem(conversa_id=c.id,numero=numero,direcao='out',tipo='erro',conteudo=('Falha auto-reply: '+str(e))[:700]))
                            db.session.commit()
                        except Exception:
                            db.session.rollback()
        elif debug:
            diag['erros'].append(f'Evento ignorado: {evento or "sem_evento"}')
    except Exception:
        app.logger.exception('Falha no processamento do webhook WhatsApp')
        diag['erros'].append('Falha no processamento do webhook')
        db.session.rollback()
    try:
        bk=wa_backup_maybe_send(force=False)
        if debug and bk.get('ok'):
            diag['backup']=f"enviado para {bk.get('email','')}"
    except Exception as e:
        app.logger.exception('Falha no backup automatico WhatsApp')
        if debug:
            diag['erros'].append(f'Backup automatico: {str(e)}')
    return jsonify({'ok':True,'debug':diag} if debug else {'ok':True})

@app.route('/api/medicoes/<int:id>/status',methods=['PUT'])
@lr
def api_medicao_status(id):
    m=Medicao.query.get_or_404(id)
    d=request.json or {}
    novo=(d.get('status') or '').strip().lower()
    validos=['rascunho','emitida','cancelada','paga']
    if novo not in validos: return jsonify({'erro':f'Status invalido. Use: {validos}'}),400
    m.status=novo
    if novo=='paga':
        m.dt_pagamento=(d.get('dt_pagamento') or '').strip() or localnow().strftime('%Y-%m-%d')
        m.forma_pagamento=(d.get('forma_pagamento') or '').strip()
    db.session.commit()
    audit_event('medicao_status_alterado','usuario',session.get('uid'),'medicao',m.id,True,{'status':novo,'numero':m.numero})
    return jsonify({'ok':True,'id':m.id,'status':novo})

@app.route('/api/medicoes/<int:id>/duplicar',methods=['POST'])
@lr
def api_medicao_duplicar(id):
    orig=Medicao.query.get_or_404(id)
    novo_num=prox_num()
    copia=Medicao(
        numero=novo_num, tipo=orig.tipo,
        cliente_id=orig.cliente_id, cliente_nome=orig.cliente_nome,
        cliente_cnpj=orig.cliente_cnpj, cliente_end=orig.cliente_end,
        cliente_resp=orig.cliente_resp,
        empresa_id=orig.empresa_id, empresa_nome=orig.empresa_nome,
        mes_ref=localnow().strftime('%Y-%m'),
        dt_emissao=localnow().strftime('%Y-%m-%d'),
        dt_vencimento=orig.dt_vencimento,
        servicos=orig.servicos, valor_bruto=orig.valor_bruto,
        observacoes=orig.observacoes,
        ass_empresa=orig.ass_empresa, ass_cliente=orig.ass_cliente,
        status='rascunho', desconto=orig.desconto, impostos=orig.impostos,
        criado_por=session.get('nome','')
    )
    db.session.add(copia); db.session.commit()
    audit_event('medicao_duplicada','usuario',session.get('uid'),'medicao',copia.id,True,{'original_id':orig.id,'numero':novo_num})
    return jsonify({'ok':True,'id':copia.id,'numero':novo_num})

@app.route('/api/medicoes/<int:id>/enviar-email',methods=['POST'])
@lr
def api_medicao_enviar_email(id):
    m=Medicao.query.get_or_404(id)
    d=request.json or {}
    destino=(d.get('email') or '').strip()
    if not destino:
        cli=Cliente.query.get(m.cliente_id) if m.cliente_id else None
        destino=(cli.email or '').strip() if cli else ''
    if not destino: return jsonify({'erro':'E-mail do destinatário não informado e cliente não possui e-mail cadastrado.'}),400
    emp=Empresa.query.get(m.empresa_id) if m.empresa_id else None
    md=m.to_dict(); md['empresa']=emp.to_dict() if emp else {}
    try:
        from flask import current_app
        from reportlab.lib.pagesizes import A4
        import io as _io
        # gera PDF em memória
        buf=_io.BytesIO()
        _pdf_data=md.copy()
        # build via _build_pdf retorna Response; precisa gerar bytes direto
        # Chamamos a mesma lógica mas capturamos os bytes
        resp=_build_pdf(_pdf_data)
        pdf_bytes=resp.get_data()
        # salva temporário
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False,suffix='.pdf') as tf:
            tf.write(pdf_bytes); tmp_path=tf.name
        tipo=m.tipo or 'Medição'
        numero=m.numero or str(m.id)
        slug=(m.cliente_nome or 'cliente').replace(' ','_')[:20]
        nome_arq=f'{tipo.replace(" ","_")}_{numero.replace("/","-")}_{slug}_{m.mes_ref or ""}.pdf'
        assunto=f'{tipo} Nº {numero} — {m.cliente_nome or ""}'
        corpo=(f'Prezado(a),\n\nSegue em anexo a {tipo} Nº {numero} referente ao período {m.mes_ref or ""}.\n\n'
               f'Valor: R$ {m.valor_bruto:,.2f}\nVencimento: {m.dt_vencimento or "-"}\n\n'
               f'Em caso de dúvidas, entre em contato conosco.\n\nAtenciosamente,\n{emp.nome if emp else "RM Facilities"}')
        smtp_send_text(destino, assunto, corpo, anexos=[{'path':tmp_path,'name':nome_arq}])
        try: os.remove(tmp_path)
        except: pass
        audit_event('medicao_email_enviado','usuario',session.get('uid'),'medicao',m.id,True,{'dest':destino,'numero':numero})
        return jsonify({'ok':True,'enviado_para':destino})
    except Exception as ex:
        return jsonify({'erro':str(ex)}),500

# ============================================================
# JUROS E MULTA AUTOMÁTICOS
# ============================================================
@app.route('/api/medicoes/<int:id>/calcular-encargos',methods=['POST'])
@lr
def api_medicao_calcular_encargos(id):
    m=Medicao.query.get_or_404(id)
    if m.status not in ('emitida','rascunho'): return jsonify({'erro':'Somente medições emitidas podem ter encargos calculados.'}),400
    if not m.dt_vencimento: return jsonify({'erro':'Medição sem data de vencimento.'}),400
    from datetime import date as _date
    try:
        venc=_date.fromisoformat(m.dt_vencimento)
    except Exception:
        return jsonify({'erro':'Data de vencimento inválida.'}),400
    hoje=_date.today()
    dias_atraso=max(0,(hoje-venc).days)
    taxa_juros_dia=float(gc('taxa_juros_dia','0.033'))   # % ao dia
    taxa_multa=float(gc('taxa_multa','2.0'))              # % única
    base=m.valor_bruto or 0
    multa=round(base*taxa_multa/100,2) if dias_atraso>0 else 0
    juros=round(base*taxa_juros_dia/100*dias_atraso,2)
    d=request.json or {}
    if d.get('aplicar'):
        m.valor_multa=multa; m.valor_juros=juros
        db.session.commit()
        audit_event('medicao_encargos_aplicados','usuario',session.get('uid'),'medicao',m.id,True,
                    {'dias_atraso':dias_atraso,'multa':multa,'juros':juros})
    return jsonify({'dias_atraso':dias_atraso,'valor_multa':multa,'valor_juros':juros,
                    'total_encargos':multa+juros,'total_com_encargos':base+multa+juros,
                    'aplicado':bool(d.get('aplicar'))})

@app.route('/api/medicoes/<int:id>/encargos',methods=['GET'])
@lr
def api_medicao_encargos(id):
    m=Medicao.query.get_or_404(id)
    from datetime import date as _date
    venc=None
    if m.dt_vencimento:
        try: venc=_date.fromisoformat(m.dt_vencimento)
        except: pass
    dias_atraso=max(0,(_date.today()-venc).days) if venc else 0
    return jsonify({'valor_juros':m.valor_juros or 0,'valor_multa':m.valor_multa or 0,
                    'total_encargos':(m.valor_juros or 0)+(m.valor_multa or 0),
                    'total_com_encargos':(m.valor_bruto or 0)+(m.valor_juros or 0)+(m.valor_multa or 0),
                    'dias_atraso':dias_atraso})

# ============================================================
# RÉGUA DE COBRANÇA
# ============================================================
def _cobranca_enviar_lembrete(m,tipo,emp=None):
    """Envia e-mail de cobrança e registra no log. Retorna (ok, erro_msg)."""
    cli=Cliente.query.get(m.cliente_id) if m.cliente_id else None
    dest=(cli.email or '').strip() if cli else ''
    if not dest: return False,'Cliente sem e-mail'
    if not emp and m.empresa_id: emp=Empresa.query.get(m.empresa_id)
    remetente=emp.nome if emp else 'RM Facilities'
    if tipo in ('D-5','D-1'):
        assunto=f'Lembrete de vencimento — {m.tipo or "Medição"} Nº {m.numero}'
        corpo=(f'Prezado(a) {m.cliente_resp or m.cliente_nome or "Cliente"},\n\n'
               f'Este é um lembrete de que a {m.tipo or "Medição"} Nº {m.numero} vence em {m.dt_vencimento}.\n'
               f'Valor: R$ {(m.valor_bruto or 0):,.2f}\n\n'
               f'Favor efetuar o pagamento até a data de vencimento.\n\nAtenciosamente,\n{remetente}')
    else:
        dias=((__import__('datetime').date.today()-__import__('datetime').date.fromisoformat(m.dt_vencimento)).days
              if m.dt_vencimento else 0)
        encargos=(m.valor_juros or 0)+(m.valor_multa or 0)
        total=( m.valor_bruto or 0)+encargos
        assunto=f'Cobrança em atraso — {m.tipo or "Medição"} Nº {m.numero}'
        corpo=(f'Prezado(a) {m.cliente_resp or m.cliente_nome or "Cliente"},\n\n'
               f'A {m.tipo or "Medição"} Nº {m.numero} encontra-se em atraso há {dias} dia(s).\n'
               f'Valor original: R$ {(m.valor_bruto or 0):,.2f}\n'
               +(f'Encargos (juros/multa): R$ {encargos:,.2f}\nTotal com encargos: R$ {total:,.2f}\n' if encargos else '')+
               f'\nPor favor entre em contato para regularizar.\n\nAtenciosamente,\n{remetente}')
    try:
        smtp_send_text(dest,assunto,corpo)
        log=CobrangaLog(medicao_id=m.id,tipo=tipo,status='ok',dest_email=dest)
        db.session.add(log); db.session.commit()
        return True,None
    except Exception as ex:
        log=CobrangaLog(medicao_id=m.id,tipo=tipo,status='erro',dest_email=dest,erro=str(ex)[:500])
        db.session.add(log); db.session.commit()
        return False,str(ex)

@app.route('/api/cobrancas/processar',methods=['POST'])
@lr
def api_cobrancas_processar():
    from datetime import date as _date, timedelta as _td
    hoje=_date.today()
    d=request.json or {}
    tipos_solicitados=d.get('tipos',['D-5','D-1','D+3'])
    medicoes_emitidas=Medicao.query.filter(Medicao.status=='emitida').all()
    resultados=[]; processadas=0; erros=0
    for m in medicoes_emitidas:
        if not m.dt_vencimento: continue
        try: venc=_date.fromisoformat(m.dt_vencimento)
        except: continue
        dias=(hoje-venc).days  # positivo = atrasado
        tipo=None
        if 'D-5' in tipos_solicitados and (venc-hoje).days==5: tipo='D-5'
        elif 'D-1' in tipos_solicitados and (venc-hoje).days==1: tipo='D-1'
        elif 'D+3' in tipos_solicitados and dias==3: tipo='D+3'
        if not tipo: continue
        # evita reenvio no mesmo dia
        ja_enviado=CobrangaLog.query.filter_by(medicao_id=m.id,tipo=tipo).filter(
            CobrangaLog.enviado_em>=__import__('datetime').datetime.combine(hoje,__import__('datetime').time.min)
        ).first()
        if ja_enviado: continue
        ok,erro=_cobranca_enviar_lembrete(m,tipo)
        resultados.append({'medicao_id':m.id,'numero':m.numero,'tipo':tipo,'ok':ok,'erro':erro})
        if ok: processadas+=1
        else: erros+=1
    return jsonify({'processadas':processadas,'erros':erros,'detalhes':resultados})

@app.route('/api/cobrancas/logs')
@lr
def api_cobrancas_logs():
    page=max(1,int(request.args.get('page',1)))
    per=min(100,int(request.args.get('per_page',50)))
    qr=CobrangaLog.query.order_by(CobrangaLog.enviado_em.desc())
    total=qr.count()
    items=[l.to_dict() for l in qr.offset((page-1)*per).limit(per).all()]
    for it in items:
        m=db.session.get(Medicao,it['medicao_id'])
        it['numero']=m.numero if m else ''; it['cliente']=m.cliente_nome if m else ''
    return jsonify({'items':items,'total':total,'page':page,'per_page':per})

@app.route('/api/cobrancas/<int:medicao_id>/manual',methods=['POST'])
@lr
def api_cobranca_manual(medicao_id):
    m=Medicao.query.get_or_404(medicao_id)
    ok,erro=_cobranca_enviar_lembrete(m,'manual')
    return jsonify({'ok':ok,'erro':erro})

# ============================================================
# CONCILIAÇÃO BANCÁRIA OFX
# ============================================================
def _parse_ofx(texto):
    """Parser simples de OFX (SGML/XML) — retorna lista de transações."""
    import re
    transacoes=[]
    # Remove cabeçalho SGML antes do primeiro <OFX> ou <STMTTRN>
    blocos=re.findall(r'<STMTTRN>(.*?)</STMTTRN>',texto,re.DOTALL|re.IGNORECASE)
    if not blocos:
        # tenta formato linha-a-linha (OFX 1.x sem fechamento)
        ttrn=None
        for linha in texto.splitlines():
            linha=linha.strip()
            if re.match(r'<STMTTRN>',linha,re.I): ttrn={}
            elif ttrn is not None:
                if re.match(r'</STMTTRN>',linha,re.I):
                    transacoes.append(ttrn); ttrn=None
                else:
                    m2=re.match(r'<([A-Z]+)>([^<]*)',linha,re.I)
                    if m2: ttrn[m2.group(1).upper()]=m2.group(2).strip()
    else:
        for bloco in blocos:
            ttrn={}
            for m2 in re.finditer(r'<([A-Z]+)>([^<\n]*)',bloco,re.I):
                ttrn[m2.group(1).upper()]=m2.group(2).strip()
            transacoes.append(ttrn)
    resultado=[]
    for t in transacoes:
        dtraw=t.get('DTPOSTED','')[:8]
        try: dt=f'{dtraw[:4]}-{dtraw[4:6]}-{dtraw[6:8]}'
        except: dt=''
        try: valor=float(t.get('TRNAMT','0').replace(',','.'))
        except: valor=0.0
        tipo='C' if valor>=0 else 'D'
        resultado.append({'data_mov':dt,'valor':abs(valor),'historico':t.get('MEMO',t.get('NAME','')),'num_doc':t.get('FITID',''),'tipo':tipo})
    return resultado

@app.route('/api/conciliacao/importar',methods=['POST'])
@lr
def api_conciliacao_importar():
    empresa_id=to_num(request.form.get('empresa_id'))
    fs=request.files.get('arquivo')
    if not fs: return jsonify({'erro':'Arquivo OFX não enviado'}),400
    nome=fs.filename or 'extrato.ofx'
    texto=fs.read().decode('latin-1','replace')
    try:
        txns=_parse_ofx(texto)
    except Exception as ex:
        return jsonify({'erro':f'Erro ao processar OFX: {ex}'}),400
    if not txns: return jsonify({'erro':'Nenhuma transação encontrada no arquivo OFX.'}),400
    lote=ConciliacaoLote(empresa_id=empresa_id,arquivo_nome=nome,total_transacoes=len(txns),
                         importado_por=session.get('nome',''))
    db.session.add(lote); db.session.flush()
    for t in txns:
        db.session.add(ConciliacaoTransacao(lote_id=lote.id,**t))
    db.session.commit()
    audit_event('conciliacao_importada','usuario',session.get('uid'),'conciliacao_lote',lote.id,True,
                {'arquivo':nome,'total':len(txns)})
    return jsonify({'ok':True,'lote_id':lote.id,'total':len(txns)})

@app.route('/api/conciliacao/lotes')
@lr
def api_conciliacao_lotes():
    lotes=ConciliacaoLote.query.order_by(ConciliacaoLote.importado_em.desc()).limit(50).all()
    return jsonify([l.to_dict() for l in lotes])

@app.route('/api/conciliacao/transacoes')
@lr
def api_conciliacao_transacoes():
    lote_id=to_num(request.args.get('lote_id'))
    status=request.args.get('status','')  # pendente|conciliado
    page=max(1,int(request.args.get('page',1)))
    per=min(200,int(request.args.get('per_page',50)))
    qr=ConciliacaoTransacao.query
    if lote_id: qr=qr.filter_by(lote_id=lote_id)
    if status=='pendente': qr=qr.filter(ConciliacaoTransacao.medicao_id==None)
    elif status=='conciliado': qr=qr.filter(ConciliacaoTransacao.medicao_id!=None)
    qr=qr.order_by(ConciliacaoTransacao.data_mov.desc(),ConciliacaoTransacao.id.desc())
    total=qr.count()
    items=[]
    for t in qr.offset((page-1)*per).limit(per).all():
        d=t.to_dict()
        if t.medicao_id:
            m=db.session.get(Medicao,t.medicao_id)
            d['medicao_numero']=m.numero if m else ''; d['medicao_cliente']=m.cliente_nome if m else ''
        else:
            d['medicao_numero']=''; d['medicao_cliente']=''
        items.append(d)
    return jsonify({'items':items,'total':total,'page':page,'per_page':per})

@app.route('/api/conciliacao/<int:transacao_id>/conciliar',methods=['POST'])
@lr
def api_conciliacao_conciliar(transacao_id):
    t=ConciliacaoTransacao.query.get_or_404(transacao_id)
    d=request.json or {}
    medicao_id=to_num(d.get('medicao_id'))
    if medicao_id:
        m=Medicao.query.get_or_404(medicao_id)
        t.medicao_id=medicao_id; t.conciliado_em=utcnow()
        audit_event('conciliacao_conciliada','usuario',session.get('uid'),'conciliacao_transacao',t.id,True,
                    {'medicao_id':medicao_id,'numero':m.numero})
    else:
        t.medicao_id=None; t.conciliado_em=None
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/conciliacao/auto',methods=['POST'])
@lr
def api_conciliacao_auto():
    """Auto-conciliação: cruza transações de crédito com medições pelo valor exato."""
    d=request.json or {}
    lote_id=to_num(d.get('lote_id'))
    pendentes=ConciliacaoTransacao.query.filter(ConciliacaoTransacao.medicao_id==None,
                                                ConciliacaoTransacao.tipo=='C')
    if lote_id: pendentes=pendentes.filter_by(lote_id=lote_id)
    pendentes=pendentes.all()
    conciliadas=0
    for t in pendentes:
        # busca medições emitidas com valor_bruto igual (tolerância R$ 0,01)
        candidatas=Medicao.query.filter(Medicao.status=='emitida',
            db.func.abs(Medicao.valor_bruto-t.valor)<0.02).all()
        # refina por data próxima ao vencimento (±15 dias)
        from datetime import date as _d, timedelta as _td
        melhor=None
        if t.data_mov:
            try: dt_mov=_d.fromisoformat(t.data_mov)
            except: dt_mov=None
            if dt_mov:
                for c in candidatas:
                    if c.dt_vencimento:
                        try:
                            venc=_d.fromisoformat(c.dt_vencimento)
                            if abs((dt_mov-venc).days)<=15: melhor=c; break
                        except: pass
        if melhor is None and len(candidatas)==1: melhor=candidatas[0]
        if melhor:
            t.medicao_id=melhor.id; t.conciliado_em=utcnow(); conciliadas+=1
    db.session.commit()
    return jsonify({'ok':True,'conciliadas':conciliadas,'pendentes_total':len(pendentes)})

@app.route('/api/medicoes/<int:id>/anexos',methods=['GET'])
@lr
def api_medicao_anexos(id):
    Medicao.query.get_or_404(id)
    return jsonify([a.to_dict() for a in MedicaoAnexo.query.filter_by(medicao_id=id).order_by(MedicaoAnexo.criado_em.desc()).all()])

@app.route('/api/medicoes/<int:id>/anexos',methods=['POST'])
@lr
def api_medicao_add_anexo(id):
    Medicao.query.get_or_404(id)
    fs=request.files.get('arquivo')
    if not fs: return jsonify({'erro':'Arquivo nao enviado'}),400
    rel,_=save_upload(fs,f'medicoes/{id}')
    a=MedicaoAnexo(medicao_id=id,nome_arquivo=fs.filename,caminho=rel,criado_por=session.get('nome',''))
    db.session.add(a); db.session.commit()
    return jsonify(a.to_dict()),201

@app.route('/api/medicoes/anexos/<int:id>',methods=['DELETE'])
@lr
def api_medicao_del_anexo(id):
    a=MedicaoAnexo.query.get_or_404(id)
    try: os.remove(os.path.join(UPLOAD_ROOT,a.caminho))
    except: pass
    db.session.delete(a); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/medicoes/anexos/<int:id>/download')
@lr
def api_medicao_download_anexo(id):
    a=MedicaoAnexo.query.get_or_404(id)
    abs_p=os.path.join(UPLOAD_ROOT,a.caminho)
    if not os.path.exists(abs_p): return jsonify({'erro':'Arquivo nao encontrado'}),404
    return send_file(abs_p,as_attachment=True,download_name=a.nome_arquivo)

def seed():
    from werkzeug.security import generate_password_hash
    if Usuario.query.count()==0:
        db.session.add(Usuario(
            nome='Administrador',
            email='admin@rmfacilities.com.br',
            senha=generate_password_hash('naoseinao', method='scrypt'),
            perfil='admin',
            ativo=True,
            twofa_ativo=False
        ))
    if Empresa.query.count()==0:
        db.session.add(Empresa(nome='RM Facilities',razao='RM CONSERVAÇÃO E SERVIÇOS LTDA',site='https://rmfacilities.com.br',cidade='São José dos Campos',estado='SP',ordem=1))
    if not Config.query.filter_by(chave='num_base').first(): db.session.add(Config(chave='num_base',valor='100'))
    if not Config.query.filter_by(chave='num_ultima').first(): db.session.add(Config(chave='num_ultima',valor='0'))
    db.session.commit()

with app.app_context():
    os.makedirs(DATA_DIR,exist_ok=True)
    os.makedirs(UPLOAD_ROOT,exist_ok=True)
    db.create_all()
    ensure_cols('usuario',[
        'areas TEXT',
        'permissoes TEXT DEFAULT "{}"',
        'telefone VARCHAR(30)',
        'twofa_ativo BOOLEAN DEFAULT 1',
        'cert_arquivo VARCHAR(500)',
        'cert_nome_arquivo VARCHAR(255)',
        'cert_senha VARCHAR(255)',
        'cert_ativo BOOLEAN DEFAULT 0',
        'cert_assunto VARCHAR(255)',
        'cert_validade_fim VARCHAR(30)'
    ])
    ensure_cols('empresa',[
        'contato_nome VARCHAR(150)',
        'contato_email VARCHAR(150)',
        'contato_telefone VARCHAR(30)',
        'logo_url VARCHAR(500)',
        'boleto TEXT',
        'cert_arquivo VARCHAR(500)',
        'cert_nome_arquivo VARCHAR(255)',
        'cert_senha VARCHAR(255)',
        'cert_ativo BOOLEAN DEFAULT 0',
        'cert_assunto VARCHAR(255)',
        'cert_validade_fim VARCHAR(30)'
    ])
    ensure_cols('funcionario',[
        're INTEGER',
        'matricula VARCHAR(30)',
        'funcao VARCHAR(150)',
        'cbo VARCHAR(20)',
        'tipo_contrato VARCHAR(60)',
        'jornada VARCHAR(80)',
        'vale_refeicao FLOAT DEFAULT 0',
        'vale_alimentacao FLOAT DEFAULT 0',
        'vale_transporte FLOAT DEFAULT 0',
        'rg VARCHAR(30)',
        'orgao_emissor VARCHAR(30)',
        'pis VARCHAR(30)',
        'ctps VARCHAR(30)',
        'titulo_eleitor VARCHAR(30)',
        'cert_reservista VARCHAR(30)',
        'cnh VARCHAR(30)',
        'exame_admissional_data VARCHAR(10)',
        'docs_admissao_ok BOOLEAN DEFAULT 0',
        'docs_admissao_obs TEXT',
        'app_senha VARCHAR(256)',
        'app_ativo BOOLEAN DEFAULT 1',
        'app_ultimo_acesso DATETIME',
        'app_otp_hash VARCHAR(256)',
        'app_otp_expira_em DATETIME',
        'app_otp_tentativas INTEGER DEFAULT 0',
        'app_push_token VARCHAR(300)',
        'app_lat FLOAT',
        'app_lon FLOAT',
        'app_localizacao_em DATETIME',
        'endereco_numero VARCHAR(20)',
        'endereco_complemento VARCHAR(120)',
        'endereco_bairro VARCHAR(120)',
        'banco_codigo VARCHAR(3)',
        'banco_nome VARCHAR(150)',
        'banco_agencia VARCHAR(30)',
        'banco_conta VARCHAR(40)',
        'banco_tipo_conta VARCHAR(20)',
        'banco_pix VARCHAR(150)',
        'posto_operacional VARCHAR(150)',
        'posto_cliente_id INTEGER',
        'opta_vt BOOLEAN DEFAULT 1',
        'opta_vr BOOLEAN DEFAULT 1',
        'opta_va BOOLEAN DEFAULT 1',
        'opta_premio_prod BOOLEAN DEFAULT 0',
        'opta_vale_gasolina BOOLEAN DEFAULT 0',
        'opta_cesta_natal BOOLEAN DEFAULT 0',
        'premio_produtividade FLOAT DEFAULT 0',
        'vale_gasolina FLOAT DEFAULT 0',
        'cesta_natal FLOAT DEFAULT 0',
        'foto_perfil VARCHAR(500)'
    ])
    ensure_cols('comunicado_app',[
        'posto_operacional VARCHAR(150)'
    ])
    ensure_cols('cliente',[
        'numero_contrato VARCHAR(60)',
        'qtd_funcionarios_posto INTEGER DEFAULT 0',
        'materiais_equip_locacao FLOAT DEFAULT 0',
        'dia_faturamento INTEGER DEFAULT 1',
        'dias_faturamento INTEGER DEFAULT 30',
        'dt_contrato_vencimento VARCHAR(10)',
        'reajuste_percentual FLOAT DEFAULT 0',
        'reajuste_data_base VARCHAR(10)',
        'ultimo_reajuste_em VARCHAR(10)',
        'geo_lat FLOAT',
        'geo_lon FLOAT',
        'geofence_raio_m FLOAT DEFAULT 150',
    ])
    ensure_cols('beneficio_mensal',[
        'dias_vt INTEGER DEFAULT 0',
        'dias_vr INTEGER DEFAULT 0',
        'dias_va INTEGER DEFAULT 0',
        'dias_vg INTEGER DEFAULT 0',
        'pp_falta BOOLEAN DEFAULT 0',
        'premio_produtividade FLOAT DEFAULT 0',
        'vale_gasolina FLOAT DEFAULT 0',
        'cesta_natal FLOAT DEFAULT 0'
    ])
    ensure_cols('ponto_marcacao',[
        'latitude FLOAT',
        'longitude FLOAT',
        'precisao_gps FLOAT',
    ])
    ensure_cols('jornada_trabalho',[
        'descricao VARCHAR(255)',
        'tolerancia_min INTEGER DEFAULT 10',
        'ativo BOOLEAN DEFAULT 1',
    ])
    ensure_cols('funcionario',[
        'jornada_id INTEGER',
    ])
    ensure_cols('funcionario_arquivo',[
        'ass_status VARCHAR(20) DEFAULT "nao_solicitada"',
        'ass_token VARCHAR(120)',
        'ass_expira_em DATETIME',
        'ass_codigo VARCHAR(120)',
        'ass_nome VARCHAR(200)',
        'ass_cargo VARCHAR(120)',
        'ass_cpf VARCHAR(20)',
        'ass_ip VARCHAR(60)',
        'ass_em DATETIME',
        'ass_otp_hash VARCHAR(256)',
        'ass_otp_expira_em DATETIME',
        'ass_otp_tentativas INTEGER DEFAULT 0',
        'ass_doc_hash VARCHAR(128)',
        'ass_crypto_ok BOOLEAN DEFAULT 0',
        'ass_cert_subject VARCHAR(255)',
        'ass_canal_envio VARCHAR(20)',
        'ass_enviado_em DATETIME',
        'ass_recebido_em DATETIME',
        'ass_aberto_em DATETIME',
        'ass_wa_status VARCHAR(20) DEFAULT "nao_enviado"',
        'ass_wa_enviado_em DATETIME',
        'ass_wa_recebido_em DATETIME',
        'ass_email_status VARCHAR(20) DEFAULT "nao_enviado"',
        'ass_email_enviado_em DATETIME',
        'ass_email_recebido_em DATETIME',
        'ass_lembretes_enviados INTEGER DEFAULT 0',
        'ass_ultimo_lembrete_em DATETIME',
        'ass_prazo_em DATETIME',
    ])
    db.session.execute(text('CREATE UNIQUE INDEX IF NOT EXISTS ix_funcionario_re ON funcionario(re)'))
    db.session.commit()
    ensure_cols('whats_app_conversa',[
        "contexto TEXT DEFAULT '{}'"
    ])
    ensure_cols('medicao',[
        'status VARCHAR(20) DEFAULT "emitida"',
        'desconto FLOAT DEFAULT 0',
        'impostos TEXT',
        'assinatura_status VARCHAR(20) DEFAULT "nao_solicitada"',
        'assinatura_token VARCHAR(120)',
        'assinatura_expira_em DATETIME',
        'assinatura_nome VARCHAR(200)',
        'assinatura_cpf VARCHAR(20)',
        'assinatura_cargo VARCHAR(120)',
        'assinatura_ip VARCHAR(60)',
        'assinatura_em DATETIME',
        'assinatura_codigo VARCHAR(120)',
        'assinatura_otp_hash VARCHAR(256)',
        'assinatura_otp_expira_em DATETIME',
        'assinatura_otp_tentativas INTEGER DEFAULT 0',
        'assinatura_doc_hash VARCHAR(128)',
        'assinatura_crypto_ok BOOLEAN DEFAULT 0',
        'assinatura_cert_subject VARCHAR(255)',
        'dt_pagamento VARCHAR(10)',
        'forma_pagamento VARCHAR(50)',
        'valor_juros FLOAT DEFAULT 0',
        'valor_multa FLOAT DEFAULT 0',
    ])
    ensure_cols('assinatura_envelope',[
        'id INTEGER PRIMARY KEY AUTOINCREMENT',
        'titulo VARCHAR(200)',
        'descricao TEXT',
        'tipo VARCHAR(20) DEFAULT "avulso"',
        'empresa_id INTEGER',
        'ref_id INTEGER',
        'status VARCHAR(20) DEFAULT "rascunho"',
        'codigo VARCHAR(120)',
        'nome_documento_assinado VARCHAR(255)',
        'destino_salvar_tipo VARCHAR(30) DEFAULT "envelope"',
        'destino_funcionario_id INTEGER',
        'destino_categoria VARCHAR(40) DEFAULT "outros"',
        'destino_competencia VARCHAR(20)',
        'criado_por VARCHAR(100)',
        'criado_em DATETIME',
        'expira_em DATETIME',
        'assinatura_doc_hash VARCHAR(128)',
        'assinatura_crypto_ok BOOLEAN DEFAULT 0',
        'assinatura_cert_subject VARCHAR(255)',
        'stamp_habilitado BOOLEAN DEFAULT 0',
        'stamp_pagina INTEGER DEFAULT 1',
        'stamp_x_pct REAL DEFAULT 60.0',
        'stamp_y_pct REAL DEFAULT 10.0',
        'stamp_todas_paginas BOOLEAN DEFAULT 0',
        'stamp_todos_arquivos BOOLEAN DEFAULT 0',
        'stamp_todas_paginas BOOLEAN DEFAULT 0',
        'stamp_todos_arquivos BOOLEAN DEFAULT 0',
    ])
    ensure_cols('assinatura_envelope_arquivo',[
        'id INTEGER PRIMARY KEY AUTOINCREMENT',
        'envelope_id INTEGER',
        'origem VARCHAR(10) DEFAULT "upload"',
        'func_arquivo_id INTEGER',
        'nome_arquivo VARCHAR(250)',
        'caminho VARCHAR(500)',
        'criado_em DATETIME',
    ])
    ensure_cols('assinatura_envelope_signatario',[
        'id INTEGER PRIMARY KEY AUTOINCREMENT',
        'envelope_id INTEGER',
        'nome VARCHAR(200)',
        'email VARCHAR(150)',
        'telefone VARCHAR(30)',
        'cpf VARCHAR(20)',
        'cargo VARCHAR(120)',
        'tipo VARCHAR(20) DEFAULT "externo"',
        'ref_id INTEGER',
        'token VARCHAR(120)',
        'status VARCHAR(20) DEFAULT "pendente"',
        'ass_ip VARCHAR(60)',
        'ass_em DATETIME',
        'ass_codigo VARCHAR(120)',
        'ass_cpf_informado VARCHAR(20)',
        'ass_otp_hash VARCHAR(256)',
        'ass_otp_expira_em DATETIME',
        'ass_otp_tentativas INTEGER DEFAULT 0',
        'ass_canal_envio VARCHAR(20)',
        'ass_enviado_em DATETIME',
        'ass_recebido_em DATETIME',
        'ass_aberto_em DATETIME',
        'ass_wa_status VARCHAR(20) DEFAULT "nao_enviado"',
        'ass_wa_enviado_em DATETIME',
        'ass_wa_recebido_em DATETIME',
        'ass_email_status VARCHAR(20) DEFAULT "nao_enviado"',
        'ass_email_enviado_em DATETIME',
        'ass_email_recebido_em DATETIME',
        'ordem INTEGER DEFAULT 0',
        'criado_em DATETIME',
        'ass_assinatura_img TEXT',
    ])
    ensure_cols('ordem_compra',[
        'empresa_id INTEGER',
        'solicitante VARCHAR(200)',
        'fornecedor VARCHAR(200)',
        'descricao TEXT',
        'valor REAL DEFAULT 0',
        'status VARCHAR(50) DEFAULT "Aberta"',
        'data_emissao VARCHAR(10)',
        'criado_por VARCHAR(100)',
        'criado_em DATETIME',
        'ass_assinatura_img TEXT',
    ])
    ensure_cols('mensagem_app',[
        'tipo VARCHAR(20) DEFAULT "texto"',
        'arquivo_nome VARCHAR(300)',
        'arquivo_caminho VARCHAR(500)',
    ])
    seed(); get_logo()

if __name__=='__main__':
    app.run(host='0.0.0.0',port=5000,debug=False)

# ============================================================
# BACKUP AUTOMÁTICO DIÁRIO
# ============================================================
AUTO_BACKUP_DIR = os.path.join(DATA_DIR, 'backups')
AUTO_BACKUP_KEEP = 7
AUTO_BACKUP_HOUR = 3
_auto_backup_status = {'ultimo': None, 'proximo': None, 'ok': None, 'msg': ''}

def _auto_backup_gerar():
    os.makedirs(AUTO_BACKUP_DIR, exist_ok=True)
    nome = f"auto_backup_{localnow().strftime('%Y%m%d_%H%M%S')}.zip"
    dest = os.path.join(AUTO_BACKUP_DIR, nome)
    with app.app_context():
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
            if os.path.exists(DB_PATH): z.write(DB_PATH, 'rmfacilities.db')
            z.writestr('clientes.json', json.dumps([c.to_dict() for c in Cliente.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('medicoes.json', json.dumps([m.to_dict() for m in Medicao.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('empresas.json', json.dumps([e.to_dict() for e in Empresa.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('funcionarios.json', json.dumps([f.to_dict() for f in Funcionario.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('config.json', json.dumps([{'chave': c.chave, 'valor': c.valor} for c in Config.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('whatsapp_conversas.json', json.dumps([c.to_dict() for c in WhatsAppConversa.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('whatsapp_mensagens.json', json.dumps([m.to_dict() for m in WhatsAppMensagem.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('funcionario_arquivos.json', json.dumps([a.to_dict() for a in FuncionarioArquivo.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('beneficios_mensais.json', json.dumps([b.to_dict() for b in BeneficioMensal.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('ordens_compra.json', json.dumps([o.to_dict() for o in OrdemCompra.query.all()], default=str, ensure_ascii=False, indent=2))
            z.writestr('operacional_documentos.json', json.dumps([d.to_dict() for d in OperacionalDocumento.query.all()], default=str, ensure_ascii=False, indent=2))
            if os.path.isdir(UPLOAD_ROOT):
                for root, _, files in os.walk(UPLOAD_ROOT):
                    for fn in files:
                        ap = os.path.join(root, fn)
                        rel = os.path.relpath(ap, UPLOAD_ROOT)
                        z.write(ap, os.path.join('uploads', rel))
            z.writestr('info.json', json.dumps({'data': localnow().isoformat(), 'tipo': 'auto', 'versao': '3.0'}, ensure_ascii=False))
        buf.seek(0)
        with open(dest, 'wb') as f:
            f.write(buf.read())
    arqs = sorted([a for a in os.listdir(AUTO_BACKUP_DIR) if a.startswith('auto_backup_') and a.endswith('.zip')], reverse=True)
    for antigo in arqs[AUTO_BACKUP_KEEP:]:
        try: os.remove(os.path.join(AUTO_BACKUP_DIR, antigo))
        except: pass
    return dest

def _auto_backup_segundos_ate_proxima():
    agora = datetime.now(APP_TZ)
    proxima = agora.replace(hour=AUTO_BACKUP_HOUR, minute=0, second=0, microsecond=0)
    if agora >= proxima:
        proxima += timedelta(days=1)
    return max(60, (proxima - agora).total_seconds()), proxima.replace(tzinfo=None)

def _auto_backup_loop():
    while True:
        secs, proxima_dt = _auto_backup_segundos_ate_proxima()
        _auto_backup_status['proximo'] = proxima_dt.strftime('%d/%m/%Y %H:%M')
        time.sleep(secs)
        try:
            dest = _auto_backup_gerar()
            _auto_backup_status['ultimo'] = localnow().strftime('%d/%m/%Y %H:%M')
            _auto_backup_status['ok'] = True
            _auto_backup_status['msg'] = os.path.basename(dest)
        except Exception as e:
            _auto_backup_status['ok'] = False
            _auto_backup_status['msg'] = str(e)

threading.Thread(target=_auto_backup_loop, daemon=True, name='auto-backup').start()

def _lembrete_assinatura_loop():
    """Envia lembretes automáticos para documentos pendentes de assinatura.
    Intervalo configurável via env LEMBRETE_ASSINATURA_INTERVALO_HORAS (padrão: 2h).
    Usa o canal original de cada documento (whatsapp / email / app / link)."""
    intervalo_horas = max(1, min(168, _to_int(os.environ.get('LEMBRETE_ASSINATURA_INTERVALO_HORAS'), 2)))
    intervalo_seg = intervalo_horas * 3600

    def _canal_padrao(a):
        ch = (a.ass_canal_envio or '').strip().lower()
        if ch:
            return ch
        if (a.ass_wa_status or '') in ('enviado', 'recebido') or bool(a.ass_wa_enviado_em):
            return 'whatsapp'
        if (a.ass_email_status or '') in ('enviado', 'recebido') or bool(a.ass_email_enviado_em):
            return 'email'
        if not (a.ass_token or '').strip():
            return 'app'
        return 'link'

    # Aguarda 2 minutos no boot antes do primeiro ciclo
    time.sleep(120)
    while True:
        try:
            with app.app_context():
                agora = utcnow()
                pendentes = FuncionarioArquivo.query.filter_by(ass_status='pendente').all()
                func_cache = {}
                for a in pendentes:
                    if not a.criado_em:
                        continue
                    # Só envia se nenhum lembrete foi enviado ainda OU
                    # se já passaram intervalo_horas desde o último lembrete
                    ultimo = a.ass_ultimo_lembrete_em
                    if ultimo is not None:
                        horas_desde = (agora - ultimo).total_seconds() / 3600
                        if horas_desde < intervalo_horas:
                            continue
                    else:
                        # Primeiro lembrete: aguarda ao menos intervalo_horas após criação
                        horas_desde_criacao = (agora - a.criado_em).total_seconds() / 3600
                        if horas_desde_criacao < intervalo_horas:
                            continue
                    try:
                        if a.funcionario_id not in func_cache:
                            func_cache[a.funcionario_id] = Funcionario.query.get(a.funcionario_id)
                        f = func_cache[a.funcionario_id]
                        if not f:
                            continue
                        canal = _canal_padrao(a)
                        rs = _solicitar_assinatura_arquivo_funcionario(
                            a, f,
                            canal=canal,
                            commit_now=False,
                            forcar_novo_token=False,
                            eh_lembrete=True,
                        )
                        a.ass_lembretes_enviados = (a.ass_lembretes_enviados or 0) + 1
                        a.ass_ultimo_lembrete_em = agora
                        db.session.commit()
                        app.logger.info(
                            f'[lembrete-auto] funcionario={a.funcionario_id} arquivo={a.id} '
                            f'canal={canal} ok={rs.get("ok")}'
                        )
                    except Exception as ex:
                        app.logger.error(f'[lembrete-auto] arquivo={a.id} erro={ex}')
                        db.session.rollback()
        except Exception as e:
            app.logger.error(f'[lembrete-assinatura] erro geral: {e}')
        time.sleep(intervalo_seg)

threading.Thread(target=_lembrete_assinatura_loop, daemon=True, name='lembrete-assinatura').start()

@app.route('/api/backup/auto/status')
@lr
def api_auto_backup_status():
    arqs = []
    if os.path.isdir(AUTO_BACKUP_DIR):
        arqs = sorted([a for a in os.listdir(AUTO_BACKUP_DIR) if a.startswith('auto_backup_') and a.endswith('.zip')], reverse=True)
    return jsonify({'ultimo': _auto_backup_status['ultimo'], 'proximo': _auto_backup_status['proximo'], 'ok': _auto_backup_status['ok'], 'msg': _auto_backup_status['msg'], 'arquivos': arqs})

@app.route('/api/backup/auto/agora', methods=['POST'])
@lr
def api_auto_backup_agora():
    try:
        dest = _auto_backup_gerar()
        _auto_backup_status['ultimo'] = localnow().strftime('%d/%m/%Y %H:%M')
        _auto_backup_status['ok'] = True
        _auto_backup_status['msg'] = os.path.basename(dest)
        return jsonify({'ok': True, 'arquivo': os.path.basename(dest)})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)}), 500

@app.route('/api/backup/auto/<nome>/download')
@lr
def api_auto_backup_download(nome):
    if '/' in nome or '\\' in nome or not nome.startswith('auto_backup_') or not nome.endswith('.zip'):
        return jsonify({'erro': 'Nome inválido'}), 400
    p = os.path.join(AUTO_BACKUP_DIR, nome)
    if not os.path.exists(p): return jsonify({'erro': 'Arquivo não encontrado'}), 404
    return send_file(p, mimetype='application/zip', as_attachment=True, download_name=nome)

# ============================================================
